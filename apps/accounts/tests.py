from django.test import TestCase
from django.contrib.auth.models import User
from django.urls import reverse
from django.core.exceptions import PermissionDenied
from apps.accounts.models import UserProfile, SMSTemplate
from apps.core.models import AuditLog
from unittest.mock import patch, MagicMock

class UserRoleAndPermissionTests(TestCase):
    def setUp(self):
        # Create standard test users for different roles
        self.admin_user = User.objects.create_user(
            username='admin_test',
            password='testpassword123',
            email='admin@test.com',
            first_name='Admin',
            last_name='User'
        )
        self.admin_user.profile.role = UserProfile.ROLE_ADMIN
        self.admin_user.profile.save()

        self.director_user = User.objects.create_user(
            username='director_test',
            password='testpassword123',
            email='director@test.com'
        )
        self.director_user.profile.role = UserProfile.ROLE_RECRUITMENT_DIRECTOR
        self.director_user.profile.save()

        self.interviewer_user = User.objects.create_user(
            username='interviewer_test',
            password='testpassword123',
            email='interviewer@test.com'
        )
        self.interviewer_user.profile.role = UserProfile.ROLE_INTERVIEWER
        self.interviewer_user.profile.save()

        self.external_assessor_user = User.objects.create_user(
            username='external_test',
            password='testpassword123',
            email='external@test.com'
        )
        self.external_assessor_user.profile.role = UserProfile.ROLE_EXTERNAL_ASSESSOR
        self.external_assessor_user.profile.save()

    def test_user_profile_creation_on_signal(self):
        """تست ایجاد خودکار پروفایل کاربر پس از ساخت اکانت کاربر جنگو"""
        new_user = User.objects.create_user(username='new_test_user', password='password123')
        self.assertTrue(hasattr(new_user, 'profile'))
        self.assertEqual(new_user.profile.role, UserProfile.ROLE_INTERVIEWER)  # نقش پیش‌فرض

    def test_external_assessor_is_external_flag(self):
        """تست اینکه ارزیاب‌های خارجی به طور خودکار فلگ is_external=True می‌گیرند"""
        # Test database model save level
        user = User.objects.create_user(username='ext_assessor_test', password='password123')
        profile = user.profile
        profile.role = UserProfile.ROLE_EXTERNAL_ASSESSOR
        profile.save()
        self.assertTrue(profile.is_external)

    def test_rbac_view_access(self):
        """تست کنترل دسترسی نقش‌ها (RBAC) به صفحات مدیریت کاربران"""
        # User list view only allowed for ADMIN
        user_list_url = reverse('user_list')
        
        # Test non-admin user
        self.client.login(username='director_test', password='testpassword123')
        response = self.client.get(user_list_url)
        self.assertEqual(response.status_code, 403)  # Permission Denied

        # Test admin user
        self.client.login(username='admin_test', password='testpassword123')
        response = self.client.get(user_list_url)
        self.assertEqual(response.status_code, 200)

    def test_candidate_exclusion_from_user_management(self):
        """تست عدم نمایش متقاضیان در لیست کاربران و فرم‌های ثبت‌نام/ویرایش"""
        # Create a user with candidate role
        candidate_user = User.objects.create_user(
            username='candidate_test',
            password='testpassword123',
            email='candidate@test.com'
        )
        candidate_user.profile.role = UserProfile.ROLE_CANDIDATE
        candidate_user.profile.save()

        # 1. Verify candidate is excluded from user list queryset
        self.client.login(username='admin_test', password='testpassword123')
        user_list_url = reverse('user_list')
        response = self.client.get(user_list_url)
        self.assertEqual(response.status_code, 200)
        users = response.context['users']
        self.assertNotIn(candidate_user, users)

        # 2. Verify candidate is excluded from UserCreationForm choices
        from apps.accounts.forms import UserCreationForm, UserUpdateForm
        creation_form = UserCreationForm()
        creation_roles = [choice[0] for choice in creation_form.fields['role'].choices]
        self.assertNotIn(UserProfile.ROLE_CANDIDATE, creation_roles)

        # 3. Verify candidate is excluded from UserUpdateForm choices
        update_form = UserUpdateForm(instance=self.admin_user)
        update_roles = [choice[0] for choice in update_form.fields['role'].choices]
        self.assertNotIn(UserProfile.ROLE_CANDIDATE, update_roles)

    def test_audit_log_immutability(self):
        """تست اینکه رکوردهای جدول AuditLog غیرقابل ویرایش و حذف هستند"""
        # Create an audit log entry
        log = AuditLog.objects.create(
            user=self.admin_user,
            action_type=AuditLog.ACTION_CREATE,
            model_name='userprofile',
            object_id=str(self.admin_user.profile.pk),
            changes={'role': 'ADMIN'}
        )
        self.assertIsNotNone(log.pk)

        # Try to modify log
        log.action_type = AuditLog.ACTION_UPDATE
        with self.assertRaises(PermissionDenied):
            log.save()

        # Try to delete log
        with self.assertRaises(PermissionDenied):
            log.delete()

    def test_management_dashboard_metrics(self):
        """تست محاسبات آماری، میانگین زمان مراحل و متقاضیان تاخیردار در داشبورد مدیریتی"""
        from apps.jobs.models import JobOpportunity, WorkflowTemplate, WorkflowStageTemplate
        from apps.candidates.models import Candidate, JobApplication, ApplicationStageState
        from django.utils import timezone
        from datetime import timedelta

        # ۱. ایجاد موقعیت شغلی و فرآیند ارزیابی
        workflow = WorkflowTemplate.objects.create(name='فرآیند ارزیابی تستی')
        stage1 = WorkflowStageTemplate.objects.create(workflow=workflow, name='آزمون کتبی تستی', stage_type='EXAM', default_weight=50, sequence=1)
        stage2 = WorkflowStageTemplate.objects.create(workflow=workflow, name='مصاحبه فنی تستی', stage_type='INTERVIEW', default_weight=50, sequence=2)
        
        job = JobOpportunity.objects.create(
            request_number='REQ-TEST-100', title='برنامه‌نویس', code='DEV-100',
            department='فناوری', workflow=workflow, status=JobOpportunity.STATUS_PUBLISHED
        )

        # ۲. ایجاد متقاضی فعال
        cand1 = Candidate.objects.create(
            first_name='مهدی', last_name='حسینی', email='m.h@example.com',
            phone_number='09129999999', national_id='9999999999'
        )
        app1 = JobApplication.objects.create(job=job, candidate=cand1)
        
        # ۳. شبیه‌سازی متقاضی تاخیردار (بیش از ۷ روز توقف در مرحله اول)
        # برای تغییر فیلدهای created_at و updated_at که auto_now هستند، از متد update استفاده می‌کنیم
        JobApplication.objects.filter(id=app1.id).update(created_at=timezone.now() - timedelta(days=10))
        exam_state = app1.stage_states.filter(stage__name='آزمون کتبی تستی').first()
        ApplicationStageState.objects.filter(id=exam_state.id).update(
            created_at=timezone.now() - timedelta(days=10),
            updated_at=timezone.now() - timedelta(days=10)
        )

        # ۴. ایجاد متقاضی دوم که مرحله اول را با موفقیت پشت سر گذاشته
        cand2 = Candidate.objects.create(
            first_name='نیلوفر', last_name='کریمی', email='n.k@example.com',
            phone_number='09128888888', national_id='8888888888'
        )
        app2 = JobApplication.objects.create(job=job, candidate=cand2)
        
        # ثبت موفقیت‌آمیز آزمون کتبی برای کاندیدای دوم به صورت ۵ روز پیش
        exam_state2 = app2.stage_states.filter(stage__name='آزمون کتبی تستی').first()
        ApplicationStageState.objects.filter(id=exam_state2.id).update(
            status='COMPLETED',
            score=80.0,
            created_at=timezone.now() - timedelta(days=10),
            updated_at=timezone.now() - timedelta(days=5)
        )
        stage2_job = job.stages.filter(name='مصاحبه فنی تستی').first()
        JobApplication.objects.filter(id=app2.id).update(
            created_at=timezone.now() - timedelta(days=10),
            current_stage=stage2_job
        )
        job.update_status()

        # ۵. درخواست صفحه داشبورد توسط کاربر ادمین
        self.client.login(username='admin_test', password='testpassword123')
        url = reverse('dashboard')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

        # بررسی وجود متغیرهای آماری در کانتکست
        self.assertIn('total_jobs', response.context)
        self.assertIn('active_jobs_count', response.context)
        self.assertIn('total_candidates', response.context)
        self.assertIn('in_progress_apps', response.context)
        self.assertIn('selected_candidates', response.context)
        self.assertIn('active_stages', response.context)
        self.assertIn('dept_stats', response.context)
        self.assertIn('unit_stats', response.context)
        self.assertIn('category_stats', response.context)
        self.assertIn('avg_job_finalization_days', response.context)
        self.assertIn('avg_app_finalization_days', response.context)

        # ۶. تایید صحت شناسایی متقاضی تاخیردار (باید کاندیدای اول به عنوان تاخیردار حضور داشته باشد)
        delayed_list = response.context['delayed_candidates']
        self.assertTrue(any(item['candidate'].id == cand1.id for item in delayed_list))
        # کاندیدای دوم در مرحله دوم (مصاحبه فنی تستی) معلق است اما کمتر از ۷ روز پیش این مرحله برایش فعال شده، پس نباید جزو تاخیردارها باشد
        self.assertFalse(any(item['candidate'].id == cand2.id for item in delayed_list))

        # ۷. تایید محاسبه میانگین زمان سپری شده
        avg_days_list = response.context['avg_stage_days']
        self.assertTrue(any(item['stage_name'] == 'آزمون کتبی' for item in avg_days_list))
        test_stage_avg = next(item for item in avg_days_list if item['stage_name'] == 'آزمون کتبی')
        # کاندیدای دوم بعد از ۵ روز آزمونش ثبت شده (۱۰ روز پیش ثبت‌نام، ۵ روز پیش ثبت نمره)، پس میانگین حدود ۵ روز است
        self.assertEqual(test_stage_avg['avg_days'], 5.0)

    def test_role_based_stage_access_control(self):
        """تست کنترل دسترسی جامع نقش‌ها به مراحل ارزیابی و ثبت نمرات"""
        from apps.jobs.models import JobOpportunity, WorkflowTemplate, WorkflowStageTemplate, JobStageInterviewer
        from apps.candidates.models import Candidate, JobApplication, ApplicationStageState
        from apps.accounts.permissions import check_stage_access

        # ایجاد الگو با مراحل متنوع (مصاحبه و کانون ارزیابی)
        workflow = WorkflowTemplate.objects.create(name='فرآیند جامع تستی')
        stage_interview = WorkflowStageTemplate.objects.create(workflow=workflow, name='مصاحبه فنی تستی', default_weight=50, sequence=1)
        stage_assessment = WorkflowStageTemplate.objects.create(workflow=workflow, name='کانون ارزیابی تخصصی', default_weight=50, sequence=2)
        
        job = JobOpportunity.objects.create(
            request_number='REQ-ROLE-TEST', title='برنامه‌نویس ارشد', code='DEV-ROLE-TEST',
            department='فناوری', workflow=workflow, status=JobOpportunity.STATUS_PUBLISHED
        )

        # ایجاد متقاضی
        cand = Candidate.objects.create(
            first_name='علی', last_name='علوی', email='ali@example.com',
            phone_number='09127777777', national_id='7777777777'
        )
        app = JobApplication.objects.create(job=job, candidate=cand)
        
        # دریافت آبجکت‌های مراحل شغل
        job_stage_interview = job.stages.filter(name='مصاحبه فنی تستی').first()
        job_stage_assessment = job.stages.filter(name='کانون ارزیابی تخصصی').first()
        
        # ۱. بررسی تابع کمکی check_stage_access
        # مصاحبه‌گر (ROLE_INTERVIEWER) نباید به کانون ارزیابی دسترسی داشته باشد
        self.assertTrue(check_stage_access(self.interviewer_user, job_stage_interview))
        self.assertFalse(check_stage_access(self.interviewer_user, job_stage_assessment))
        
        # ارزیاب کانون (ROLE_EXTERNAL_ASSESSOR) نباید به مصاحبه فنی دسترسی داشته باشد
        self.assertTrue(check_stage_access(self.external_assessor_user, job_stage_assessment))
        self.assertFalse(check_stage_access(self.external_assessor_user, job_stage_interview))

        # ۲. بررسی فیلتر شدن پنل مصاحبه‌ها (InterviewsPanelView)
        # انتساب مصاحبه‌گر به هر دو مرحله در دیتابیس (هرچند که از لحاظ نقش نباید به کانون دسترسی داشته باشد)
        JobStageInterviewer.objects.create(job=job, stage=job_stage_interview, user=self.interviewer_user)
        JobStageInterviewer.objects.create(job=job, stage=job_stage_assessment, user=self.interviewer_user)
        
        # لاگین با مصاحبه‌گر
        self.client.login(username='interviewer_test', password='testpassword123')
        url_interviews = reverse('candidate_interviews')
        response = self.client.get(url_interviews)
        self.assertEqual(response.status_code, 200)
        
        # با اینکه مصاحبه‌گر به هر دو مرحله اساین شده، اما فقط باید مصاحبه فنی تستی را در لیست ببیند و کانون ارزیابی فیلتر شود
        stage_states = response.context['stage_states']
        self.assertTrue(any(state.stage.id == job_stage_interview.id for state in stage_states))
        self.assertFalse(any(state.stage.id == job_stage_assessment.id for state in stage_states))

        # ۳. بررسی بلاک کردن ثبت نمرات غیرمجاز (SubmitInterviewerScoreView)
        state_assessment = app.stage_states.filter(stage=job_stage_assessment).first()
        url_submit_score = reverse('candidate_submit_interviewer_score', kwargs={'pk': state_assessment.id})
        
        # ارسال نمره کانون ارزیابی توسط مصاحبه‌گر عادی باید با خطای ۴۰۳ (PermissionDenied) مواجه شود
        response = self.client.post(url_submit_score, {'score': '80', 'notes': 'تست'})
        self.assertEqual(response.status_code, 403)

    def test_audit_log_list_view_access_and_filtering(self):
        """تست دسترسی به لیست لاگ‌های ممیزی و فیلتر کردن آن‌ها"""
        # Create test audit logs
        log_create = AuditLog.objects.create(
            user=self.admin_user,
            action_type=AuditLog.ACTION_CREATE,
            model_name='jobopportunity',
            object_id='100',
            changes={'title': 'Developer'}
        )
        log_update = AuditLog.objects.create(
            user=self.director_user,
            action_type=AuditLog.ACTION_UPDATE,
            model_name='candidate',
            object_id='200',
            changes={'first_name': 'Ali'}
        )

        url = reverse('audit_log_list')

        # 1. Non-authorized user should get 403
        self.client.login(username='interviewer_test', password='testpassword123')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 403)

        # 2. Authorized user (ADMIN) should get 200 and see the logs
        self.client.login(username='admin_test', password='testpassword123')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        logs = response.context['logs']
        # Note: the test setup creates other logs (like in signal profile creation, user profile saves, etc.),
        # so we assert that our logs exist in the returned list.
        log_ids = [l.id for l in logs]
        self.assertIn(log_create.id, log_ids)
        self.assertIn(log_update.id, log_ids)

        # 3. Filtering by action_type
        response = self.client.get(url, {'action_type': 'CREATE'})
        self.assertEqual(response.status_code, 200)
        logs = response.context['logs']
        for l in logs:
            self.assertEqual(l.action_type, 'CREATE')

        # 4. Searching by query q
        response = self.client.get(url, {'q': 'Ali'})
        self.assertEqual(response.status_code, 200)
        logs = response.context['logs']
        self.assertTrue(any(l.id == log_update.id for l in logs))


class SystemBackupAndRestoreTests(TestCase):
    def setUp(self):
        # Create admin and non-admin users
        self.admin_user = User.objects.create_user(
            username='admin_test_backup',
            password='testpassword123',
            email='admin@test.com'
        )
        self.admin_user.profile.role = UserProfile.ROLE_ADMIN
        self.admin_user.profile.save()

        self.interviewer_user = User.objects.create_user(
            username='interviewer_test_backup',
            password='testpassword123',
            email='interviewer@test.com'
        )
        self.interviewer_user.profile.role = UserProfile.ROLE_INTERVIEWER
        self.interviewer_user.profile.save()

    def test_backup_restore_view_rbac(self):
        """تست کنترل دسترسی نقش‌ها به صفحات پشتیبان‌گیری و بازگردانی"""
        urls = [
            reverse('system_backup'),
            reverse('download_backup'),
            reverse('restore_backup')
        ]
        
        # Test non-admin user
        self.client.login(username='interviewer_test_backup', password='testpassword123')
        for url in urls:
            if url == reverse('system_backup'):
                response = self.client.get(url)
            else:
                response = self.client.post(url)
            self.assertEqual(response.status_code, 403)

        # Test admin user
        self.client.login(username='admin_test_backup', password='testpassword123')
        response = self.client.get(reverse('system_backup'))
        self.assertEqual(response.status_code, 200)

    def test_backup_generation(self):
        """تست تهیه نسخه پشتیبان و تولید فایل ZIP حاوی دیتابیس"""
        self.client.login(username='admin_test_backup', password='testpassword123')
        response = self.client.post(reverse('download_backup'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/zip')
        self.assertTrue(response['Content-Disposition'].startswith('attachment; filename='))
        
        # Read response content as ZIP
        import zipfile
        import io
        zip_content = io.BytesIO(response.content)
        with zipfile.ZipFile(zip_content, 'r') as zipf:
            file_list = zipf.namelist()
            self.assertIn('db.sqlite3', file_list)

    def test_invalid_restore_zip(self):
        """تست بازگردانی فایل نامعتبر (فایل زیپ بدون db.sqlite3)"""
        self.client.login(username='admin_test_backup', password='testpassword123')
        
        # Create an empty zip file in memory
        import zipfile
        import io
        from django.core.files.uploadedfile import SimpleUploadedFile
        
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zipf:
            zipf.writestr('dummy.txt', 'This is a test file')
            
        zip_buffer.seek(0)
        uploaded_file = SimpleUploadedFile(
            "invalid_backup.zip",
            zip_buffer.read(),
            content_type="application/zip"
        )
        
        response = self.client.post(reverse('restore_backup'), {'backup_file': uploaded_file})
        self.assertEqual(response.status_code, 302)  # Redirects back to system_backup
        
        # Verify error message is in session
        response = self.client.get(reverse('system_backup'))
        messages = list(response.context['messages'])
        self.assertTrue(any("فایل پشتیبان معتبر نیست" in str(m) for m in messages))

    def test_invalid_restore_sqlite(self):
        """تست بازگردانی فایل دیتابیس نامعتبر"""
        self.client.login(username='admin_test_backup', password='testpassword123')
        
        from django.core.files.uploadedfile import SimpleUploadedFile
        uploaded_file = SimpleUploadedFile(
            "invalid_db.sqlite3",
            b"not a database content",
            content_type="application/octet-stream"
        )
        
        response = self.client.post(reverse('restore_backup'), {'backup_file': uploaded_file})
        self.assertEqual(response.status_code, 302)
        
        response = self.client.get(reverse('system_backup'))
        messages = list(response.context['messages'])
        self.assertTrue(any("معتبر نیست یا آسیب دیده است" in str(m) for m in messages))

    def test_system_backup_view_context_version(self):
        """تست وجود متغیر نسخه جاری در کانتکست صفحه پشتیبان‌گیری"""
        self.client.login(username='admin_test_backup', password='testpassword123')
        response = self.client.get(reverse('system_backup'))
        self.assertEqual(response.status_code, 200)
        self.assertIn('current_version', response.context)
        from django.conf import settings
        self.assertEqual(response.context['current_version'], getattr(settings, 'APP_VERSION', '1.0.0'))

    @patch('urllib.request.urlopen')
    def test_update_check_up_to_date(self, mock_urlopen):
        """تست بررسی به‌روزرسانی در صورتی که سیستم به‌روز باشد"""
        self.client.login(username='admin_test_backup', password='testpassword123')
        
        from django.conf import settings
        current_version = getattr(settings, 'APP_VERSION', '1.0.0')
        
        mock_response = MagicMock()
        mock_response.read.return_value = f"{current_version}\n".encode('utf-8')
        mock_urlopen.return_value.__enter__.return_value = mock_response
        
        response = self.client.get(reverse('system_update_check'))
        self.assertEqual(response.status_code, 200)
        self.assertIn("سامانه شما کاملاً به‌روز است", response.content.decode('utf-8'))

    @patch('urllib.request.urlopen')
    def test_update_check_new_version(self, mock_urlopen):
        """تست بررسی به‌روزرسانی در صورتی که نسخه جدید موجود باشد"""
        self.client.login(username='admin_test_backup', password='testpassword123')
        
        from django.conf import settings
        current_version = getattr(settings, 'APP_VERSION', '1.0.0')
        parts = current_version.split('.')
        next_version = f"{parts[0]}.{int(parts[1])+1}.{parts[2]}"
        
        mock_response = MagicMock()
        mock_response.read.return_value = f"{next_version}\n".encode('utf-8')
        mock_urlopen.return_value.__enter__.return_value = mock_response
        
        response = self.client.get(reverse('system_update_check'))
        self.assertEqual(response.status_code, 200)
        self.assertIn("نسخه جدید در دسترس است", response.content.decode('utf-8'))

    @patch('urllib.request.urlopen')
    def test_update_check_failed(self, mock_urlopen):
        """تست رفتار سیستم در صورت ناموفق بودن ارتباط با گیت‌هاب"""
        self.client.login(username='admin_test_backup', password='testpassword123')
        
        mock_urlopen.side_effect = Exception("Network fail")
        
        response = self.client.get(reverse('system_update_check'))
        self.assertEqual(response.status_code, 200)
        self.assertIn("خطا در ارتباط با سرور به‌روزرسانی", response.content.decode('utf-8'))

    @patch('os.path.exists')
    @patch('urllib.request.urlopen')
    @patch('subprocess.run')
    def test_update_run_zip_success(self, mock_run, mock_urlopen, mock_exists):
        """تست اجرای موفق به‌روزرسانی از طریق ZIP در صورت عدم وجود پوشه .git"""
        self.client.login(username='admin_test_backup', password='testpassword123')
        
        # os.path.exists returns False for git_dir, True for manage.py
        mock_exists.side_effect = lambda path: False if '.git' in str(path) else True
        
        # Mock download response containing a zip with a dummy version file
        import io
        import zipfile
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zf:
            zf.writestr('Payjoo-ATS-main/version.txt', '1.3.1')
        zip_buffer.seek(0)
        
        mock_urlopen.return_value.__enter__.return_value = zip_buffer
        
        mock_res_migrate = MagicMock()
        mock_res_migrate.returncode = 0
        mock_run.return_value = mock_res_migrate
        
        response = self.client.post(reverse('system_update_run'))
        self.assertEqual(response.status_code, 200)
        self.assertIn("سیستم با موفقیت به آخرین نسخه به‌روزرسانی شد", response.content.decode('utf-8'))

    @patch('os.path.exists')
    @patch('subprocess.run')
    def test_update_run_success(self, mock_run, mock_exists):
        """تست اجرای موفق به‌روزرسانی کدهای سیستم از طریق Git"""
        self.client.login(username='admin_test_backup', password='testpassword123')
        
        mock_exists.return_value = True
        
        mock_res_git = MagicMock()
        mock_res_git.returncode = 0
        mock_res_migrate = MagicMock()
        mock_res_migrate.returncode = 0
        
        # Git: set-url, fetch, reset, migrate -> 4 calls
        mock_run.side_effect = [mock_res_git, mock_res_git, mock_res_git, mock_res_migrate]
        
        response = self.client.post(reverse('system_update_run'))
        self.assertEqual(response.status_code, 200)
        self.assertIn("به‌روزرسانی شد. سرور در حال راه‌اندازی مجدد است", response.content.decode('utf-8'))

    def test_system_health_check_view(self):
        """تست عملکرد نمای بررسی سلامت سیستم"""
        response = self.client.get(reverse('system_health_check'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['status'], 'ok')

    @patch('os.path.exists')
    def test_system_restart_view(self, mock_exists):
        """تست عملکرد نمای راه‌اندازی مجدد سیستم"""
        self.client.login(username='admin_test_backup', password='testpassword123')
        mock_exists.return_value = True
        
        with patch('os.utime') as mock_utime:
            response = self.client.post(reverse('system_restart'))
            self.assertEqual(response.status_code, 200)
            self.assertIn("راه‌اندازی مجدد سرور با موفقیت صادر شد", response.content.decode('utf-8'))
            mock_utime.assert_called_once()


class SMSTemplateAndPanelTests(TestCase):
    def setUp(self):
        # Create admin and non-admin users
        self.admin_user = User.objects.create_user(
            username='admin_test_sms',
            password='testpassword123',
            email='admin@test.com'
        )
        self.admin_user.profile.role = UserProfile.ROLE_ADMIN
        self.admin_user.profile.save()

        self.interviewer_user = User.objects.create_user(
            username='interviewer_test_sms',
            password='testpassword123',
            email='interviewer@test.com'
        )
        self.interviewer_user.profile.role = UserProfile.ROLE_INTERVIEWER
        self.interviewer_user.profile.save()

        # Create template
        self.template = SMSTemplate.objects.create(
            name="الگوی دعوت",
            body="سلام {نام} {نام_خانوادگی}. شما برای شغل {عنوان_شغل} در مرحله {نام_مرحله} نمره {نمره} گرفتید. کد ملی شما {کد_ملی} است."
        )

        # Create job opportunity
        from apps.jobs.models import JobOpportunity, WorkflowTemplate, WorkflowStageTemplate
        self.workflow = WorkflowTemplate.objects.create(name='فرآیند پیامک')
        self.stage = WorkflowStageTemplate.objects.create(workflow=self.workflow, name='آزمون اول', default_weight=100, sequence=1)
        
        self.job = JobOpportunity.objects.create(
            request_number='REQ-SMS-101', title='کارشناس سیستم', code='SYS-101',
            department='فناوری', workflow=self.workflow, status=JobOpportunity.STATUS_PUBLISHED
        )

        # Create candidates
        from apps.candidates.models import Candidate, JobApplication
        self.cand = Candidate.objects.create(
            first_name='حمید', last_name='رضایی', email='hamid@example.com',
            phone_number='09121111111', national_id='1111111111'
        )
        self.app = JobApplication.objects.create(job=self.job, candidate=self.cand)

    def test_sms_panel_rbac(self):
        """تست کنترل دسترسی نقش‌ها به پنل مدیریت پیامک‌ها"""
        urls = [
            reverse('sms_panel'),
            reverse('sms_panel_stages'),
            reverse('sms_panel_preview'),
            reverse('sms_panel_export')
        ]
        
        # Test non-admin user
        self.client.login(username='interviewer_test_sms', password='testpassword123')
        for url in urls:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 403)
            
            response = self.client.post(url)
            self.assertEqual(response.status_code, 403)

        # Test admin user
        self.client.login(username='admin_test_sms', password='testpassword123')
        response = self.client.get(reverse('sms_panel'))
        self.assertEqual(response.status_code, 200)

    def test_sms_template_crud(self):
        """تست ساخت، ویرایش و حذف قالب پیامک"""
        self.client.login(username='admin_test_sms', password='testpassword123')

        # 1. Create
        response = self.client.post(reverse('sms_panel'), {
            'action': 'create',
            'template_name': 'جدید',
            'body': 'سلام {نام}'
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(SMSTemplate.objects.filter(name='جدید').exists())

        # 2. Edit
        tmpl = SMSTemplate.objects.create(name='قالب موقت', body='تست')
        response = self.client.post(reverse('sms_panel'), {
            'action': 'edit',
            'template_id': tmpl.id,
            'template_name': 'قالب اصلاح‌شده',
            'body': 'متن اصلاح‌شده'
        })
        self.assertEqual(response.status_code, 302)
        tmpl.refresh_from_db()
        self.assertEqual(tmpl.name, 'قالب اصلاح‌شده')

        # 3. Delete
        response = self.client.post(reverse('sms_panel'), {
            'action': 'delete',
            'template_id': tmpl.id
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(SMSTemplate.all_objects.get(id=tmpl.id).is_deleted)

    def test_sms_live_preview_and_substitution(self):
        """تست رندر پیش‌نویس متن پیامک و جایگذاری فیلدها"""
        self.client.login(username='admin_test_sms', password='testpassword123')
        
        # Set stage state score
        state = self.app.stage_states.first()
        state.score = 85.5
        state.status = 'COMPLETED'
        state.save()

        url = reverse('sms_panel_preview')
        response = self.client.get(url, {
            'template_id': self.template.id,
            'job_id': self.job.id,
            'stage_id': state.stage.id
        })
        self.assertEqual(response.status_code, 200)
        
        # Check rendered message text in HTML
        content = response.content.decode('utf-8')
        expected_text = "سلام حمید رضایی. شما برای شغل کارشناس سیستم در مرحله آزمون اول نمره 85.5 گرفتید. کد ملی شما 1111111111 است."
        self.assertIn(expected_text, content)

    def test_sms_live_preview_stage_filtering(self):
        """تست فیلتر شدن متقاضیان بر اساس مرحله فعلی آن‌ها در پیش‌نمایش"""
        self.client.login(username='admin_test_sms', password='testpassword123')

        # Create a second stage for the job
        from apps.jobs.models import JobOpportunityStage
        from apps.candidates.models import Candidate, JobApplication
        
        stage2 = JobOpportunityStage.objects.create(
            job=self.job,
            name="مصاحبه فنی دوم",
            weight=0,
            sequence=2
        )

        # Create second candidate and set their current stage to stage2
        cand2 = Candidate.objects.create(
            first_name='سعید',
            last_name='احمدی',
            email='saeed@example.com',
            phone_number='09122222222',
            national_id='2222222222'
        )
        app2 = JobApplication.objects.create(job=self.job, candidate=cand2)
        app2.current_stage = stage2
        app2._bypass_stage_recalculation = True
        app2.save()

        # The first candidate (self.cand) has current_stage = self.job.stages.first() (sequence 1)
        first_stage = self.job.stages.all().order_by('sequence').first()

        url = reverse('sms_panel_preview')

        # 1. Filter by first stage
        response = self.client.get(url, {
            'template_id': self.template.id,
            'job_id': self.job.id,
            'stage_id': first_stage.id
        })
        self.assertEqual(response.status_code, 200)
        preview_data = response.context['preview_data']
        # Should only contain first candidate, not the second
        candidate_ids = [item['candidate'].id for item in preview_data]
        self.assertIn(self.cand.id, candidate_ids)
        self.assertNotIn(cand2.id, candidate_ids)

        # 2. Filter by second stage
        response = self.client.get(url, {
            'template_id': self.template.id,
            'job_id': self.job.id,
            'stage_id': stage2.id
        })
        self.assertEqual(response.status_code, 200)
        preview_data = response.context['preview_data']
        # Should only contain second candidate, not the first
        candidate_ids = [item['candidate'].id for item in preview_data]
        self.assertIn(cand2.id, candidate_ids)
        self.assertNotIn(self.cand.id, candidate_ids)

    def test_sms_export_excel(self):
        """تست تولید خروجی فایل اکسل با سرستون‌های فارسی"""
        self.client.login(username='admin_test_sms', password='testpassword123')
        
        response = self.client.post(reverse('sms_panel_export'), {
            'template_id': self.template.id,
            'candidate_ids': [self.cand.id],
            'job_id': self.job.id
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Verify content with openpyxl
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "شماره همراه")
        self.assertEqual(ws.cell(row=1, column=2).value, "متن پیامک")
        self.assertEqual(ws.cell(row=1, column=3).value, "نام متقاضی")
        
        # Verify row data
        self.assertEqual(ws.cell(row=2, column=1).value, "09121111111")
        self.assertIn("حمید رضایی", ws.cell(row=2, column=2).value)
        self.assertEqual(ws.cell(row=2, column=3).value, "حمید رضایی")
