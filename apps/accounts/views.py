from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.views import LoginView as DjangoLoginView
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, TemplateView
from django.urls import reverse_lazy, reverse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.http import HttpResponse

from .forms import PersianLoginForm, UserCreationForm, UserUpdateForm
from .models import UserProfile
from .permissions import RoleRequiredMixin

class CustomLoginView(DjangoLoginView):
    form_class = PersianLoginForm
    template_name = 'accounts/login.html'
    redirect_authenticated_user = True

    def get_success_url(self):
        user = self.request.user
        if hasattr(user, 'profile') and user.profile.role == UserProfile.ROLE_CANDIDATE:
            return reverse('candidate_dashboard')
        return reverse('dashboard')


class CustomLogoutView(View):
    def post(self, request):
        logout(request)
        return redirect('login')

    def get(self, request):
        logout(request)
        return redirect('login')


class UserListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    model = User
    template_name = 'accounts/user_list.html'
    context_object_name = 'users'
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def get_queryset(self):
        # فقط کاربرانی که پروفایل آن‌ها حذف نرم نشده و نقش متقاضی ندارند بازگردانده شوند
        return User.objects.filter(
            profile__is_deleted=False
        ).exclude(
            profile__role=UserProfile.ROLE_CANDIDATE
        ).select_related('profile')


class UserCreateView(LoginRequiredMixin, RoleRequiredMixin, CreateView):
    model = User
    form_class = UserCreationForm
    template_name = 'accounts/user_form.html'
    success_url = reverse_lazy('user_list')
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.headers.get('HX-Request'):
            # در صورتی که درخواست از طریق HTMX باشد، ردیف جدول جدید را بازمی‌گردانیم
            user = self.object
            edit_url = reverse('user_edit', kwargs={'pk': user.pk})
            delete_url = reverse('user_delete', kwargs={'pk': user.pk})
            external_text = "خارجی" if user.profile.is_external else "داخلی"
            return HttpResponse(
                f'<tr id="user-{user.id}">'
                f'<td>{user.username}</td>'
                f'<td>{user.get_full_name() or user.username}</td>'
                f'<td>{user.email}</td>'
                f'<td>{user.profile.get_role_display()}</td>'
                f'<td>{external_text}</td>'
                f'<td>'
                f'<a href="{edit_url}" class="btn btn-sm btn-outline-primary me-1">ویرایش</a>'
                f'<button hx-delete="{delete_url}" hx-target="#user-{user.id}" hx-confirm="آیا از حذف این کاربر مطمئن هستید؟" class="btn btn-sm btn-outline-danger">حذف</button>'
                f'</td>'
                f'</tr>'
            )
        return response


class UserUpdateView(LoginRequiredMixin, RoleRequiredMixin, UpdateView):
    model = User
    form_class = UserUpdateForm
    template_name = 'accounts/user_form.html'
    success_url = reverse_lazy('user_list')
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def get_initial(self):
        initial = super().get_initial()
        profile = self.object.profile
        initial['role'] = profile.role
        initial['is_external'] = profile.is_external
        initial['phone_number'] = profile.phone_number
        return initial


class UserDeleteView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def delete(self, request, pk):
        user = get_object_or_404(User, pk=pk)
        
        # غیرفعال کردن اکانت جنگو
        user.is_active = False
        user.save()

        # حذف نرم پروفایل کاربر
        profile = user.profile
        profile.delete()

        # بازگرداندن پاسخ خالی جهت حذف ردیف توسط HTMX
        return HttpResponse("")


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'index.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return super().dispatch(request, *args, **kwargs)
        if hasattr(request.user, 'profile') and request.user.profile.role == UserProfile.ROLE_CANDIDATE:
            return redirect('candidate_dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        from apps.jobs.models import JobOpportunity
        from apps.candidates.models import Candidate, JobApplication, ApplicationStageState
        from apps.core.models import AuditLog
        from django.utils import timezone
        from datetime import timedelta
        
        # ۱. آمارهای کلی
        total_jobs = JobOpportunity.objects.filter(is_deleted=False).count()
        active_jobs = JobOpportunity.objects.filter(is_deleted=False).exclude(status__in=['CLOSED', 'CANCELLED']).count()
        total_candidates = Candidate.objects.filter(is_deleted=False).count()
        in_progress_apps = JobApplication.objects.filter(status='IN_PROGRESS', is_deleted=False).exclude(job__status__in=['CLOSED', 'CANCELLED']).count()
        selected_candidates = JobApplication.objects.filter(status='SELECTED', is_deleted=False).count()
        
        data.update({
            'total_jobs': total_jobs,
            'active_jobs_count': active_jobs,
            'total_candidates': total_candidates,
            'in_progress_apps': in_progress_apps,
            'selected_candidates': selected_candidates,
        })
        
        # ۲. توزیع وضعیت فرصت‌های شغلی
        job_status_distribution = []
        for status_key, status_label in JobOpportunity.STATUS_CHOICES:
            count = JobOpportunity.objects.filter(status=status_key, is_deleted=False).count()
            if count > 0:
                job_status_distribution.append({
                    'label': status_label,
                    'count': count,
                    'percentage': round((count / total_jobs) * 100, 1) if total_jobs > 0 else 0
                })
        data['job_status_distribution'] = job_status_distribution

        # ۳. توزیع متقاضیان در مراحل ارزیابی فعال (بر اساس stage_type استاندارد)
        from django.db.models import Count
        STAGE_TYPE_LABELS = {
            'EXAM':      'آزمون کتبی',
            'INTERVIEW': 'مصاحبه حضوری',
            'SKILL_TEST':'آزمون مهارتی',
            'ASSESSMENT':'کانون ارزیابی',
        }
        STAGE_TYPE_ORDER = ['EXAM', 'INTERVIEW', 'SKILL_TEST', 'ASSESSMENT']

        # تعداد متقاضیان IN_PROGRESS در هر stage_type (بر اساس وضعیت فعلی فرصت شغلی job.status)
        pending_counts = JobApplication.objects.filter(
            status='IN_PROGRESS',
            is_deleted=False,
        ).exclude(job__status__in=['CLOSED', 'CANCELLED']).values(
            'job__status'
        ).annotate(count=Count('id'))

        stage_type_counts = {row['job__status']: row['count'] for row in pending_counts}
        stage_people_stats = []
        for stype in STAGE_TYPE_ORDER:
            cnt = stage_type_counts.get(stype, 0)
            stage_people_stats.append({
                'stage_type': stype,
                'stage_name': STAGE_TYPE_LABELS.get(stype, stype),
                'count': cnt
            })
        data['stage_people_stats'] = stage_people_stats

        # ۴. محاسبه میانگین زمان حضور در هر مرحله (روز)
        # پارامتر include_closed: اگر True باشد، متقاضیان اتمام‌یافته هم در محاسبه لحاظ می‌شوند
        include_closed_param = self.request.GET.get('include_closed', '')
        include_closed = include_closed_param in ('1', 'true', 'on')
        data['include_closed'] = include_closed

        completed_states_qs = ApplicationStageState.objects.filter(
            status__in=['COMPLETED', 'FAILED'],
            is_deleted=False,
        ).select_related('application__job', 'stage')
        if not include_closed:
            completed_states_qs = completed_states_qs.exclude(
                application__job__status__in=['CLOSED', 'CANCELLED']
            )

        stage_times = {}
        for state in completed_states_qs:
            app = state.application
            stage = state.stage
            days = None

            if stage.sequence == 1:
                if state.evaluation_date:
                    start_date = app.job.start_date if (app.job and app.job.start_date) else app.created_at.date()
                    days = float((state.evaluation_date - start_date).days)
                else:
                    duration = state.updated_at - app.created_at
                    days = duration.total_seconds() / (24 * 3600)
            else:
                prev_state = app.stage_states.filter(
                    stage__sequence__lt=stage.sequence,
                    status__in=['COMPLETED', 'FAILED'],
                    is_deleted=False
                ).order_by('-stage__sequence').first()
                
                if prev_state:
                    prev_time = prev_state.evaluation_date if prev_state.evaluation_date else prev_state.updated_at.date()
                    curr_time = state.evaluation_date if state.evaluation_date else state.updated_at.date()
                    days = float((curr_time - prev_time).days)
                else:
                    curr_time = state.evaluation_date if state.evaluation_date else state.updated_at.date()
                    start_time = app.job.start_date if (app.job and app.job.start_date) else app.created_at.date()
                    days = float((curr_time - start_time).days)

            if days is not None:
                days = max(days, 0.1)
                stype = stage.stage_type
                if stype not in stage_times:
                    stage_times[stype] = []
                stage_times[stype].append(days)

        avg_stage_days = []
        for stype in STAGE_TYPE_ORDER:
            durations = stage_times.get(stype)
            if durations:
                avg_stage_days.append({
                    'stage_type': stype,
                    'stage_name': STAGE_TYPE_LABELS.get(stype, stype),
                    'avg_days': round(sum(durations) / len(durations), 1),
                    'count': len(durations)
                })
            else:
                avg_stage_days.append({
                    'stage_type': stype,
                    'stage_name': STAGE_TYPE_LABELS.get(stype, stype),
                    'avg_days': 0,
                    'count': 0
                })
        data['avg_stage_days'] = avg_stage_days

        # ۵. شناسایی متقاضیان تاخیردار (بر اساس SLA تعریف شده برای هر مرحله)
        from apps.recruitment_planning.models import JobStagePlan, StageTypeConfiguration
        from datetime import datetime
        
        delayed_candidates = []
        in_progress_apps = JobApplication.objects.filter(
            status='IN_PROGRESS',
            is_deleted=False
        ).exclude(job__status__in=['CLOSED', 'CANCELLED']).select_related('candidate', 'job')
        
        for app in in_progress_apps:
            job = app.job
            # پیدا کردن مرحله ارزیابی متناظر با وضعیت شغل
            stage = job.stages.filter(stage_type=job.status, is_deleted=False).first()
            if not stage:
                stage = app.current_stage or job.stages.filter(is_deleted=False).order_by('sequence').first()
            
            if not stage:
                continue
                
            # دریافت تعداد روزهای مجاز طبق SLA
            stage_plan = JobStagePlan.objects.filter(
                plan__job=job,
                stage=stage,
                plan__is_deleted=False,
                is_deleted=False
            ).first()
            if stage_plan:
                sla_days = stage_plan.sla_days
            else:
                config = StageTypeConfiguration.objects.filter(
                    stage_type=stage.stage_type,
                    is_deleted=False
                ).first()
                sla_days = config.default_sla_days if config else 5
                
            # تعیین تاریخ شروع حضور در مرحله فعلی
            active_since = None
            if stage.sequence == 1:
                if job.start_date:
                    active_since = timezone.make_aware(datetime.combine(job.start_date, datetime.min.time()))
                else:
                    active_since = app.created_at
            else:
                prev_state = app.stage_states.filter(
                    stage__sequence__lt=stage.sequence,
                    status__in=['COMPLETED', 'FAILED'],
                    is_deleted=False
                ).order_by('-stage__sequence').first()
                
                if prev_state:
                    if prev_state.evaluation_date:
                        active_since = timezone.make_aware(datetime.combine(prev_state.evaluation_date, datetime.min.time()))
                    else:
                        active_since = prev_state.updated_at
                else:
                    if job.start_date:
                        active_since = timezone.make_aware(datetime.combine(job.start_date, datetime.min.time()))
                    else:
                        active_since = app.created_at
                        
            if active_since:
                days_waiting = (timezone.now() - active_since).days
                if days_waiting > sla_days:
                    delayed_candidates.append({
                        'candidate': app.candidate,
                        'job': job,
                        'stage_name': STAGE_TYPE_LABELS.get(stage.stage_type, stage.name),
                        'days_waiting': days_waiting,
                        'sla_days': sla_days,
                        'overdue_days': days_waiting - sla_days
                    })
                    
        # مرتب‌سازی بر اساس میزان تأخیر نزولی
        delayed_candidates.sort(key=lambda x: x['overdue_days'], reverse=True)
        data['delayed_candidates_count'] = len(delayed_candidates)
        data['delayed_candidates'] = delayed_candidates[:5] # نمایش حداکثر ۵ مورد بحرانی‌تر

        # ۶. لاگ فعالیت‌های اخیر سیستم
        data['recent_activities'] = AuditLog.objects.all().select_related('user').order_by('-timestamp')[:5]

        # ۷. آمارهای تکمیلی: توزیع بر اساس stage_type استاندارد (نه نام مرحله)
        # ستون‌های ثابت برای جداول
        active_stages = [STAGE_TYPE_LABELS.get(st, st) for st in STAGE_TYPE_ORDER]
        data['active_stages'] = active_stages
        data['active_stage_types'] = STAGE_TYPE_ORDER  # برای استفاده در view

        def build_stage_counts(apps_qs, stage_types):
            """برای یک queryset از applications، تعداد متقاضیان در هر stage_type را بر اساس وضعیت فرصت شغلی (job.status) برمی‌گرداند"""
            rows = apps_qs.values('job__status').annotate(cnt=Count('id'))
            return {row['job__status']: row['cnt'] for row in rows}

        # آمار دپارتمان‌ها
        departments = list(JobOpportunity.objects.filter(is_deleted=False).exclude(
            status__in=['CLOSED', 'CANCELLED']
        ).values_list('department', flat=True).order_by().distinct())
        dept_stats = []
        for dept in departments:
            if not dept:
                continue
            total_jobs_dept = JobOpportunity.objects.filter(department=dept, is_deleted=False).exclude(status__in=['CLOSED', 'CANCELLED']).count()
            dept_apps = JobApplication.objects.filter(
                job__department=dept, status='IN_PROGRESS', is_deleted=False
            ).exclude(job__status__in=['CLOSED', 'CANCELLED'])
            counts = build_stage_counts(dept_apps, STAGE_TYPE_ORDER)
            dept_row = {
                'department': dept,
                'total_jobs': total_jobs_dept,
                'active_jobs': total_jobs_dept,
                'stages': {STAGE_TYPE_LABELS.get(st, st): counts.get(st, 0) for st in STAGE_TYPE_ORDER},
                'total_candidates': sum(counts.get(st, 0) for st in STAGE_TYPE_ORDER),
            }
            dept_stats.append(dept_row)
        data['dept_stats'] = dept_stats

        # آمار واحدها
        units = list(JobOpportunity.objects.filter(is_deleted=False).exclude(
            status__in=['CLOSED', 'CANCELLED']
        ).exclude(unit='').exclude(unit=None).values_list('unit', flat=True).order_by().distinct())
        unit_stats = []
        for unit in units:
            if not unit:
                continue
            total_jobs_unit = JobOpportunity.objects.filter(unit=unit, is_deleted=False).exclude(status__in=['CLOSED', 'CANCELLED']).count()
            unit_apps = JobApplication.objects.filter(
                job__unit=unit, status='IN_PROGRESS', is_deleted=False
            ).exclude(job__status__in=['CLOSED', 'CANCELLED'])
            counts = build_stage_counts(unit_apps, STAGE_TYPE_ORDER)
            unit_row = {
                'unit': unit,
                'total_jobs': total_jobs_unit,
                'active_jobs': total_jobs_unit,
                'stages': {STAGE_TYPE_LABELS.get(st, st): counts.get(st, 0) for st in STAGE_TYPE_ORDER},
                'total_candidates': sum(counts.get(st, 0) for st in STAGE_TYPE_ORDER),
            }
            unit_stats.append(unit_row)
        data['unit_stats'] = unit_stats

        # آمار رده‌های شغلی
        categories = list(JobOpportunity.objects.filter(is_deleted=False).exclude(
            status__in=['CLOSED', 'CANCELLED']
        ).exclude(job_category='').exclude(job_category=None).values_list('job_category', flat=True).order_by().distinct())
        category_stats = []
        for cat in categories:
            if not cat:
                continue
            total_jobs_cat = JobOpportunity.objects.filter(job_category=cat, is_deleted=False).exclude(status__in=['CLOSED', 'CANCELLED']).count()
            cat_apps = JobApplication.objects.filter(
                job__job_category=cat, status='IN_PROGRESS', is_deleted=False
            ).exclude(job__status__in=['CLOSED', 'CANCELLED'])
            counts = build_stage_counts(cat_apps, STAGE_TYPE_ORDER)
            cat_row = {
                'job_category': cat,
                'total_jobs': total_jobs_cat,
                'active_jobs': total_jobs_cat,
                'stages': {STAGE_TYPE_LABELS.get(st, st): counts.get(st, 0) for st in STAGE_TYPE_ORDER},
                'total_candidates': sum(counts.get(st, 0) for st in STAGE_TYPE_ORDER),
            }
            category_stats.append(cat_row)
        data['category_stats'] = category_stats

        # ۸. میانگین زمان تعیین تکلیف فرصت شغلی و درخواست‌ها
        # میانگین زمان بستن فرصت شغلی (از تاریخ شروع/ایجاد تا انتخاب نهایی کاندیدا یا بسته‌شدن فرصت)
        closed_jobs = JobOpportunity.objects.filter(status='CLOSED', is_deleted=False)
        job_durations = []
        for job in closed_jobs:
            pub_date = job.start_date or job.created_at.date()
            # پیدا کردن اولین اپلیکیشن قبول نهایی شده برای این موقعیت شغلی
            selected_app = job.applications.filter(status='SELECTED', is_deleted=False).order_by('updated_at').first()
            if selected_app:
                close_date = selected_app.updated_at.date()
            else:
                close_date = job.updated_at.date()
            duration = (close_date - pub_date).days
            job_durations.append(max(duration, 0))
        
        data['avg_job_finalization_days'] = round(sum(job_durations) / len(job_durations), 1) if job_durations else 0

        # میانگین زمان تعیین تکلیف درخواست‌های متقاضیان (SELECTED یا REJECTED)
        finalized_apps = JobApplication.objects.filter(status__in=['SELECTED', 'REJECTED'], is_deleted=False)
        app_durations = []
        for app in finalized_apps:
            duration = (app.updated_at.date() - app.created_at.date()).days
            app_durations.append(max(duration, 0))
            
        data['avg_app_finalization_days'] = round(sum(app_durations) / len(app_durations), 1) if app_durations else 0

        # ۹. آمار تکمیلی برای داشبورد عملیاتی
        # تعداد فرصت‌های شغلی به تفکیک رده شغلی
        from django.db.models import Count as DCount
        category_job_counts = JobOpportunity.objects.filter(
            is_deleted=False
        ).exclude(job_category='').exclude(job_category=None).values('job_category').annotate(
            total=DCount('id'),
            active=DCount('id', filter=Q(status__in=['SCREENING', 'EXAM', 'SKILL_TEST', 'INTERVIEW', 'ASSESSMENT', 'REGISTRATION_CLOSED', 'PUBLISHED']))
        ).order_by('job_category')
        data['category_job_counts'] = list(category_job_counts)

        # ادغام آمار رده‌های شغلی برای نمایش آسان در قالب و نمودار
        unified_category_stats = []
        for cat_count in category_job_counts:
            cat_name = cat_count['job_category']
            cat_stat = next((cs for cs in category_stats if cs['job_category'] == cat_name), None)
            total_cand = cat_stat['total_candidates'] if cat_stat else 0
            
            total_jobs = cat_count['total']
            active_jobs = cat_count['active']
            percentage = round((active_jobs / total_jobs) * 100, 1) if total_jobs > 0 else 0
            
            unified_category_stats.append({
                'name': cat_name,
                'active_jobs': active_jobs,
                'total_jobs': total_jobs,
                'total_candidates': total_cand,
                'percentage': percentage
            })
        data['unified_category_stats'] = unified_category_stats

        # خلاصه pipeline: فرصت‌هایی که آماده تصمیم‌گیری نهایی هستند
        from apps.jobs.models import JobOpportunityStage
        ready_for_decision = []
        assessment_jobs = JobOpportunity.objects.filter(
            status__in=['ASSESSMENT', 'INTERVIEW', 'FINAL_SELECTION'], is_deleted=False
        ).order_by('-updated_at')[:10]
        for j in assessment_jobs:
            pending_count = ApplicationStageState.objects.filter(
                application__job=j,
                application__status='IN_PROGRESS',
                status='PENDING',
                is_deleted=False
            ).count()
            completed_count = ApplicationStageState.objects.filter(
                application__job=j,
                application__status='IN_PROGRESS',
                status__in=['COMPLETED', 'FAILED'],
                is_deleted=False
            ).count()
            ready_for_decision.append({
                'job': j,
                'pending': pending_count,
                'completed': completed_count,
                'total': pending_count + completed_count,
            })
        data['ready_for_decision'] = ready_for_decision

        # جداول واحد: فقط واحدهایی که متقاضی دارند
        data['unit_stats'] = sorted(
            [u for u in data['unit_stats'] if u['total_candidates'] > 0],
            key=lambda x: x['total_candidates'], reverse=True
        )

        return data



from django.core.exceptions import PermissionDenied
from django.db.models import Q
from apps.core.models import AuditLog

class AuditLogListView(LoginRequiredMixin, ListView):
    model = AuditLog
    template_name = 'accounts/audit_log_list.html'
    context_object_name = 'logs'
    paginate_by = 20

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not hasattr(request.user, 'profile') or not request.user.profile.can_view_audit_logs:
            raise PermissionDenied("شما دسترسی به مشاهده لاگ‌های ممیزی ندارید.")
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = AuditLog.objects.all().select_related('user').order_by('-timestamp')
        
        # Action type filter
        action_type = self.request.GET.get('action_type')
        if action_type:
            queryset = queryset.filter(action_type=action_type)
            
        # Search query
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(user__username__icontains=q) |
                Q(user__first_name__icontains=q) |
                Q(user__last_name__icontains=q) |
                Q(model_name__icontains=q) |
                Q(object_id__icontains=q) |
                Q(changes__icontains=q)
            )
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['selected_action_type'] = self.request.GET.get('action_type', '')
        context['search_q'] = self.request.GET.get('q', '')
        context['action_choices'] = AuditLog.ACTION_CHOICES
        return context


import os
import shutil
import zipfile
from django.conf import settings
from django.utils import timezone
from django.http import FileResponse, Http404

class SystemBackupView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    allowed_roles = [UserProfile.ROLE_ADMIN]
    template_name = 'accounts/system_backup.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            with open(settings.BASE_DIR / 'version.txt', 'r') as _f:
                context['current_version'] = _f.read().strip()
        except Exception:
            context['current_version'] = getattr(settings, 'APP_VERSION', '1.0.0')
        return context


class DownloadBackupView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def post(self, request):
        temp_dir = settings.BASE_DIR / 'scratch_backup'
        os.makedirs(temp_dir, exist_ok=True)
        
        # Clean older backups in temp_dir first
        for item in os.listdir(temp_dir):
            item_path = temp_dir / item
            try:
                if os.path.isfile(item_path):
                    os.remove(item_path)
            except Exception:
                pass
        
        timestamp = timezone.now().strftime('%Y-%m-%d_%H-%M-%S')
        zip_filename = f"backup_{timestamp}.zip"
        zip_path = temp_dir / zip_filename
        
        db_path = settings.DATABASES['default']['NAME']
        media_path = settings.MEDIA_ROOT
        
        # Copy db to temp copy
        temp_db_copy = temp_dir / 'db.sqlite3'
        try:
            if os.path.exists(db_path):
                shutil.copy2(db_path, temp_db_copy)
            else:
                # Create a dummy/empty sqlite3 database for testing (for in-memory databases)
                import sqlite3
                conn = sqlite3.connect(str(temp_db_copy))
                conn.execute("CREATE TABLE dummy (id INTEGER PRIMARY KEY);")
                conn.close()
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Add database
                zipf.write(temp_db_copy, arcname='db.sqlite3')
                
                # Add media files
                if os.path.exists(media_path):
                    for root, dirs, files in os.walk(media_path):
                        for file in files:
                            file_full_path = os.path.join(root, file)
                            # Get path relative to BASE_DIR so we can extract it cleanly
                            relative_path = os.path.relpath(file_full_path, settings.BASE_DIR)
                            zipf.write(file_full_path, arcname=relative_path)
                            
            # Read zip file in memory
            with open(zip_path, 'rb') as f:
                file_data = f.read()

            # Clean up files immediately
            if os.path.exists(zip_path):
                os.remove(zip_path)
            if os.path.exists(temp_db_copy):
                os.remove(temp_db_copy)
                
            response = HttpResponse(file_data, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
            return response
        except Exception as e:
            from django.contrib import messages
            messages.error(request, f"خطا در ایجاد فایل پشتیبان: {str(e)}")
            return redirect('system_backup')

class RestoreBackupView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def post(self, request):
        from django.contrib import messages
        backup_file = request.FILES.get('backup_file')
        if not backup_file:
            messages.error(request, "لطفاً یک فایل معتبر انتخاب کنید.")
            return redirect('system_backup')
            
        file_ext = os.path.splitext(backup_file.name)[1].lower()
        if file_ext not in ['.zip', '.sqlite3']:
            messages.error(request, "فرمت فایل باید zip یا sqlite3 باشد.")
            return redirect('system_backup')
            
        temp_dir = settings.BASE_DIR / 'scratch_backup'
        os.makedirs(temp_dir, exist_ok=True)
        uploaded_file_path = temp_dir / f'uploaded_restore{file_ext}'
        
        # Save uploaded file
        with open(uploaded_file_path, 'wb+') as destination:
            for chunk in backup_file.chunks():
                destination.write(chunk)
                
        # Perform restore
        try:
            if file_ext == '.zip':
                self.restore_system_zip(uploaded_file_path)
            else:
                self.restore_system_sqlite(uploaded_file_path)
            messages.success(request, "اطلاعات سیستم با موفقیت بازگردانی شد.")
        except Exception as e:
            messages.error(request, str(e))
            
        # Clean up uploaded file
        if os.path.exists(uploaded_file_path):
            os.remove(uploaded_file_path)
            
        return redirect('system_backup')
        
    def restore_system_zip(self, zip_file_path):
        import zipfile
        import os
        import shutil
        from django.db import connections, connection
        from django.conf import settings
        
        db_path = settings.DATABASES['default']['NAME']
        media_path = settings.MEDIA_ROOT
        
        is_memory = (str(db_path) == ':memory:')
        
        # 1. Validate ZIP file structure
        with zipfile.ZipFile(zip_file_path, 'r') as zipf:
            file_list = zipf.namelist()
            if 'db.sqlite3' not in file_list:
                raise Exception("فایل پشتیبان معتبر نیست. دیتابیس در فایل یافت نشد.")
                
        # 2. Backup current database and media for rollback
        db_backup = str(db_path) + '.backup' if not is_memory else None
        media_backup = str(media_path) + '_backup'
        
        db_backed_up = False
        media_backed_up = False
        
        try:
            if not is_memory:
                if os.path.exists(db_path):
                    connections.close_all()
                    shutil.copy2(db_path, db_backup)
                    db_backed_up = True
                
            if os.path.exists(media_path):
                shutil.copytree(media_path, media_backup)
                media_backed_up = True
                
            # 3. Clean current database and media files
            if not is_memory:
                connections.close_all()
                for ext in ['-wal', '-shm']:
                    p = str(db_path) + ext
                    if os.path.exists(p):
                        os.remove(p)
                if os.path.exists(db_path):
                    os.remove(db_path)
            
            if os.path.exists(media_path):
                shutil.rmtree(media_path)
                
            # 4. Extract from ZIP
            with zipfile.ZipFile(zip_file_path, 'r') as zipf:
                if not is_memory:
                    db_data = zipf.read('db.sqlite3')
                    with open(db_path, 'wb') as f:
                        f.write(db_data)
                    
                # Extract media files
                for member in zipf.infolist():
                    if member.filename.startswith('media/'):
                        zipf.extract(member, settings.BASE_DIR)
                        
            # 5. Verify database connection works (only for physical files)
            if not is_memory:
                connections.close_all()
                with connection.cursor() as cursor:
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
                
            # 6. Success! Remove backups
            if db_backed_up and db_backup and os.path.exists(db_backup):
                os.remove(db_backup)
            if media_backed_up and os.path.exists(media_backup):
                shutil.rmtree(media_backup)
                
        except Exception as e:
            # Rollback!
            connections.close_all()
            
            # Restore database
            if db_backed_up and db_backup:
                if os.path.exists(db_path):
                    os.remove(db_path)
                shutil.move(db_backup, db_path)
                
            # Restore media
            if media_backed_up:
                if os.path.exists(media_path):
                    shutil.rmtree(media_path)
                shutil.move(media_backup, media_path)
                
            raise Exception(f"خطا در بازگردانی پشتیبان: {str(e)}")

    def restore_system_sqlite(self, sqlite_file_path):
        import os
        import shutil
        from django.db import connections, connection
        from django.conf import settings
        
        db_path = settings.DATABASES['default']['NAME']
        is_memory = (str(db_path) == ':memory:')
        
        # 1. Validate SQLite file
        try:
            import sqlite3
            conn = sqlite3.connect(str(sqlite_file_path))
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            conn.close()
        except Exception:
            raise Exception("فایل دیتابیس معتبر نیست یا آسیب دیده است.")
            
        # 2. Backup current database for rollback
        db_backup = str(db_path) + '.backup' if not is_memory else None
        db_backed_up = False
        
        try:
            if not is_memory:
                if os.path.exists(db_path):
                    connections.close_all()
                    shutil.copy2(db_path, db_backup)
                    db_backed_up = True
                    
            # 3. Clean current database
            if not is_memory:
                connections.close_all()
                for ext in ['-wal', '-shm']:
                    p = str(db_path) + ext
                    if os.path.exists(p):
                        os.remove(p)
                if os.path.exists(db_path):
                    os.remove(db_path)
                    
            # 4. Copy new database
            if not is_memory:
                shutil.copy2(sqlite_file_path, db_path)
                
            # 5. Verify database connection works
            if not is_memory:
                connections.close_all()
                with connection.cursor() as cursor:
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
                
            # 6. Success! Remove backup
            if db_backed_up and db_backup and os.path.exists(db_backup):
                os.remove(db_backup)
                
        except Exception as e:
            # Rollback!
            connections.close_all()
            if db_backed_up and db_backup:
                if os.path.exists(db_path):
                    os.remove(db_path)
                shutil.move(db_backup, db_path)
            raise Exception(f"خطا در بازگردانی دیتابیس: {str(e)}")


import urllib.request
import sys
import subprocess

class SystemUpdateCheckView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def get(self, request):
        # خواندن مستقیم از فایل به جای settings کش‌شده
        def _read_local_version():
            try:
                with open(settings.BASE_DIR / 'version.txt', 'r') as _f:
                    return _f.read().strip()
            except Exception:
                return getattr(settings, 'APP_VERSION', '1.0.0')

        local_version = _read_local_version()
        remote_version = None
        error_msg = None
        
        try:
            url = 'https://raw.githubusercontent.com/omid516/Payjoo-ATS/main/version.txt'
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=5) as response:
                remote_version = response.read().decode('utf-8').strip()
        except urllib.error.HTTPError as e:
            if e.code == 404:
                remote_version = local_version
            else:
                error_msg = f"خطای سرور گیت‌هاب ({e.code})"
        except Exception:
            error_msg = "خطا در ارتباط با سرور به‌روزرسانی"

        update_available = False
        if remote_version and not error_msg:
            def parse_version(v_str):
                try:
                    return tuple(int(x) for x in v_str.strip().split('.'))
                except ValueError:
                    return (0, 0, 0)
            update_available = parse_version(remote_version) > parse_version(local_version)

        context = {
            'local_version': local_version,
            'remote_version': remote_version,
            'update_available': update_available,
            'error_msg': error_msg,
        }
        return render(request, 'accounts/partials/update_status.html', context)


class SystemUpdateRunView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def post(self, request):
        git_dir = settings.BASE_DIR / '.git'
        git_updated = False
        error_logs = []

        # 1. Try Git update first if .git folder exists
        if os.path.exists(git_dir):
            try:
                # Update remote URL to new Payjoo-ATS location
                subprocess.run(
                    ['git', 'remote', 'set-url', 'origin', 'https://github.com/omid516/Payjoo-ATS.git'],
                    cwd=settings.BASE_DIR,
                    capture_output=True,
                    timeout=10
                )

                # Fetch code from GitHub
                fetch_res = subprocess.run(
                    ['git', 'fetch', 'origin', 'main'],
                    cwd=settings.BASE_DIR,
                    capture_output=True,
                    text=True,
                    timeout=30
                )
                if fetch_res.returncode == 0:
                    # Hard reset local codes to origin/main (overwrites local edits safely)
                    reset_res = subprocess.run(
                        ['git', 'reset', '--hard', 'origin/main'],
                        cwd=settings.BASE_DIR,
                        capture_output=True,
                        text=True,
                        timeout=30
                    )
                    if reset_res.returncode == 0:
                        git_updated = True
                    else:
                        error_logs.append(f"git reset failed: {reset_res.stderr}")
                else:
                    error_logs.append(f"git fetch failed: {fetch_res.stderr}")
            except Exception as e:
                error_logs.append(f"Git update exception: {str(e)}")

        # 2. Fallback to ZIP download update if Git was not used or failed
        if not git_updated:
            try:
                import tempfile
                import zipfile
                import shutil

                url = 'https://github.com/omid516/Payjoo-ATS/archive/refs/heads/main.zip'
                req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
                
                with tempfile.TemporaryDirectory() as tmpdir:
                    zip_path = os.path.join(tmpdir, 'update.zip')
                    
                    # Download the ZIP file
                    with urllib.request.urlopen(req, timeout=30) as response, open(zip_path, 'wb') as out_file:
                        shutil.copyfileobj(response, out_file)
                        
                    # Extract zip contents
                    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                        namelist = zip_ref.namelist()
                        if not namelist:
                            raise Exception("فایل فشرده دریافتی از گیت‌هاب خالی است.")
                        
                        root_prefix = namelist[0].split('/')[0] + '/'
                        
                        for member in zip_ref.infolist():
                            if not member.filename.startswith(root_prefix):
                                continue
                            
                            rel_path = member.filename[len(root_prefix):]
                            if not rel_path:
                                continue
                            
                            # Protect database, virtualenv, and user media files
                            parts = rel_path.split('/')
                            if parts[0] in ['db.sqlite3', '.venv', '.git', 'media', '.env', 'scratch_backup']:
                                continue
                            if rel_path.endswith('.sqlite3'):
                                continue
                            
                            target_path = settings.BASE_DIR / rel_path
                            
                            if member.is_dir():
                                os.makedirs(target_path, exist_ok=True)
                            else:
                                os.makedirs(target_path.parent, exist_ok=True)
                                with zip_ref.open(member) as source, open(target_path, 'wb') as target:
                                    shutil.copyfileobj(source, target)
                                    
                # ZIP update succeeded
                git_updated = True
            except Exception as e:
                error_logs.append(f"ZIP update exception: {str(e)}")

        if not git_updated:
            errors_str = "<br>".join(error_logs)
            return HttpResponse(
                f'<div class="alert alert-danger font-semibold text-xs text-right mb-0">خطا در به‌روزرسانی خودکار سیستم:<br><pre dir="ltr" class="text-left mt-2 mb-0">{errors_str}</pre></div>',
                status=500
            )

        # 3. Migrate database
        try:
            migrate_res = subprocess.run(
                [sys.executable, 'manage.py', 'migrate'],
                cwd=settings.BASE_DIR,
                capture_output=True,
                text=True,
                timeout=30
            )
            if migrate_res.returncode != 0:
                return HttpResponse(
                    f'<div class="alert alert-warning font-semibold text-xs text-right mb-0">کدها با موفقیت دریافت شدند، اما اعمال مهاجرت‌ها با خطا مواجه شد:<br><pre dir="ltr" class="text-left mt-2 mb-0">{migrate_res.stderr}</pre></div>',
                    status=200
                )
        except Exception as e:
            return HttpResponse(
                f'<div class="alert alert-warning font-semibold text-xs text-right mb-0">کدها با موفقیت دریافت شدند، اما خطا در اجرای مهاجرت‌ها رخ داد: {str(e)}</div>',
                status=200
            )

        # 4. Touch manage.py to trigger django auto-reloader restart
        try:
            manage_py = settings.BASE_DIR / 'manage.py'
            if os.path.exists(manage_py):
                os.utime(manage_py, None)
        except Exception:
            pass

        return HttpResponse(
            '<div class="alert alert-success font-semibold text-xs text-right mb-0">سیستم با موفقیت به آخرین نسخه به‌روزرسانی شد. سرور در حال راه‌اندازی مجدد است...</div>',
            status=200
        )


class SystemHealthCheckView(View):
    def get(self, request):
        from django.http import JsonResponse
        try:
            with open(settings.BASE_DIR / 'version.txt', 'r') as _f:
                ver = _f.read().strip()
        except Exception:
            ver = getattr(settings, 'APP_VERSION', '1.0.0')
        return JsonResponse({'status': 'ok', 'version': ver})


class SystemRestartView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def post(self, request):
        try:
            manage_py = settings.BASE_DIR / 'manage.py'
            if os.path.exists(manage_py):
                os.utime(manage_py, None)
            return HttpResponse(
                '<div class="alert alert-success font-semibold text-xs text-right mb-0">فرمان راه‌اندازی مجدد سرور با موفقیت صادر شد. لطفاً چند لحظه منتظر بمانید...</div>',
                status=200
            )
        except Exception as e:
            return HttpResponse(
                f'<div class="alert alert-danger font-semibold text-xs text-right mb-0">خطا در راه‌اندازی مجدد سرور: {str(e)}</div>',
                status=500
            )


# ==========================================
# SMS Panel & Templates Feature (Phase 1)
# ==========================================
from django.contrib import messages
from django.db import models
from django.db.models import Q
from apps.accounts.models import SMSTemplate

def render_template_text(template_body, candidate, job=None, stage=None, stage_state=None, application=None):
    from apps.candidates.models import JobApplication
    
    if not application and job:
        application = candidate.applications.filter(job=job, is_deleted=False).first()
    if not application:
        application = candidate.applications.filter(is_deleted=False).order_by('-created_at').first()

    job_title = application.job.title if (application and application.job) else "- "

    if not stage_state and stage and application:
        stage_state = application.stage_states.filter(stage=stage, is_deleted=False).first()
    if not stage_state and application:
        stage_state = application.stage_states.filter(is_deleted=False).order_by('-stage__sequence').first()

    stage_name = stage_state.stage.name if (stage_state and stage_state.stage) else "- "

    if stage_state:
        score = str(stage_state.score)
    elif application:
        score = str(application.final_score)
    else:
        score = "- "

    context = {
        'نام': candidate.first_name or "",
        'نام_خانوادگی': candidate.last_name or "",
        'عنوان_شغل': job_title,
        'نام_مرحله': stage_name,
        'نمره': score,
        'کد_ملی': candidate.national_id or "",
    }

    rendered = template_body
    for key, val in context.items():
        placeholder = f"{{{key}}}"
        rendered = rendered.replace(placeholder, str(val))
    return rendered


class SMSPanelDashboardView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    allowed_roles = [UserProfile.ROLE_ADMIN]
    template_name = 'accounts/sms_panel.html'

    def get_context_data(self, **kwargs):
        from apps.jobs.models import JobOpportunity
        from apps.candidates.models import JobApplication, ApplicationStageState

        context = super().get_context_data(**kwargs)
        context['templates'] = SMSTemplate.objects.filter(is_deleted=False)
        context['jobs'] = JobOpportunity.objects.filter(is_deleted=False)
        context['app_statuses'] = JobApplication.STATUS_CHOICES
        context['stage_statuses'] = ApplicationStageState.STATUS_CHOICES
        
        edit_id = self.request.GET.get('edit_template')
        if edit_id:
            try:
                context['edit_template'] = SMSTemplate.objects.get(id=edit_id, is_deleted=False)
            except SMSTemplate.DoesNotExist:
                pass
        return context

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        name = request.POST.get('template_name') or request.POST.get('name')
        body = request.POST.get('body')
        
        if action == 'create':
            if name and body:
                SMSTemplate.objects.create(name=name, body=body)
                messages.success(request, "قالب پیامک با موفقیت ایجاد شد.")
            else:
                messages.error(request, "لطفاً تمامی فیلدها را پر کنید.")
        elif action == 'edit':
            template_id = request.POST.get('template_id')
            if template_id and name and body:
                try:
                    tmpl = SMSTemplate.objects.get(id=template_id, is_deleted=False)
                    tmpl.name = name
                    tmpl.body = body
                    tmpl.save()
                    messages.success(request, "قالب پیامک با موفقیت ویرایش شد.")
                except SMSTemplate.DoesNotExist:
                    messages.error(request, "قالب یافت نشد.")
            else:
                messages.error(request, "لطفاً تمامی فیلدها را پر کنید.")
        elif action == 'delete':
            template_id = request.POST.get('template_id')
            if template_id:
                try:
                    tmpl = SMSTemplate.objects.get(id=template_id, is_deleted=False)
                    tmpl.delete()
                    messages.success(request, "قالب پیامک با موفقیت حذف شد.")
                except SMSTemplate.DoesNotExist:
                    messages.error(request, "قالب یافت نشد.")
                    
        return redirect('sms_panel')


class JobStagesOptionsView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def get(self, request):
        from apps.jobs.models import JobOpportunityStage
        job_id = request.GET.get('job_id')
        html = '<option value="">همه مراحل</option>'
        if job_id:
            stages = JobOpportunityStage.objects.filter(job_id=job_id, is_deleted=False).order_by('sequence')
            for stage in stages:
                html += f'<option value="{stage.id}">{stage.name}</option>'
        return HttpResponse(html)


class SMSCandidatesPreviewView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def get(self, request):
        from apps.candidates.models import Candidate, JobApplication, ApplicationStageState
        
        template_id = request.GET.get('template_id')
        job_id = request.GET.get('job_id')
        stage_id = request.GET.get('stage_id')
        app_status = request.GET.get('app_status')
        stage_status = request.GET.get('stage_status')
        search_query = request.GET.get('q')

        # Fetch template
        template_obj = None
        if template_id:
            try:
                template_obj = SMSTemplate.objects.get(id=template_id, is_deleted=False)
            except SMSTemplate.DoesNotExist:
                pass

        # Base queryset is Candidates who have at least one application (or all active candidates)
        candidates = Candidate.objects.filter(is_deleted=False)

        # Apply job and other application filters
        if job_id or stage_id or app_status or stage_status:
            app_query = Q(is_deleted=False)
            if job_id:
                app_query &= Q(job_id=job_id)
            if app_status:
                app_query &= Q(status=app_status)
            
            if stage_id:
                if stage_status:
                    app_query &= Q(stage_states__stage_id=stage_id, stage_states__status=stage_status, stage_states__is_deleted=False)
                else:
                    app_query &= Q(current_stage_id=stage_id)
            elif stage_status:
                app_query &= Q(stage_states__status=stage_status, stage_states__is_deleted=False)

            candidates = candidates.filter(applications__in=JobApplication.objects.filter(app_query)).distinct()
        
        # Text Search
        if search_query:
            candidates = candidates.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(national_id__icontains=search_query) |
                Q(phone_number__icontains=search_query)
            ).distinct()

        # Render preview messages
        preview_data = []
        template_body = template_obj.body if template_obj else ""

        # Prefetch for performance
        candidates = candidates.prefetch_related('applications__job', 'applications__stage_states__stage')

        for candidate in candidates:
            job = None
            stage = None
            stage_state = None
            application = None

            if job_id:
                application = candidate.applications.filter(job_id=job_id, is_deleted=False).first()
            if not application:
                application = candidate.applications.filter(is_deleted=False).order_by('-created_at').first()

            if application:
                job = application.job
                if stage_id:
                    stage_state = application.stage_states.filter(stage_id=stage_id, is_deleted=False).first()
                else:
                    stage_state = application.stage_states.filter(stage=application.current_stage, is_deleted=False).first()
                if not stage_state:
                    stage_state = application.stage_states.filter(is_deleted=False).order_by('-stage__sequence').first()
                
                if stage_state:
                    stage = stage_state.stage

            # Personalized message
            sms_text = ""
            if template_body:
                sms_text = render_template_text(
                    template_body,
                    candidate=candidate,
                    job=job,
                    stage=stage,
                    stage_state=stage_state,
                    application=application
                )

            char_count = len(sms_text)
            if char_count == 0:
                sms_parts = 0
            elif char_count <= 70:
                sms_parts = 1
            else:
                import math
                sms_parts = math.ceil(char_count / 67)

            preview_data.append({
                'candidate': candidate,
                'sms_text': sms_text,
                'char_count': char_count,
                'sms_parts': sms_parts,
                'job': job,
                'stage': stage,
                'stage_state': stage_state,
                'application': application,
            })

        context = {
            'preview_data': preview_data,
            'template_selected': bool(template_obj),
        }
        return render(request, 'accounts/partials/sms_candidates_preview.html', context)


class SMSExportExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def post(self, request):
        from apps.candidates.models import Candidate
        
        template_body = request.POST.get('template_body', '')
        candidate_ids = request.POST.getlist('candidate_ids')
        job_id = request.POST.get('job_id')
        stage_id = request.POST.get('stage_id')

        if not template_body:
            template_id = request.POST.get('template_id')
            if template_id:
                try:
                    template_body = SMSTemplate.objects.get(id=template_id, is_deleted=False).body
                except SMSTemplate.DoesNotExist:
                    pass

        if not template_body:
            messages.error(request, "قالب پیامک مشخص نشده یا خالی است.")
            return redirect('sms_panel')

        if not candidate_ids:
            messages.error(request, "هیچ متقاضی انتخاب نشده است.")
            return redirect('sms_panel')

        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ارسال پیامک"

        ws.views.sheetView[0].showGridLines = True
        ws.sheet_properties.tabColor = "1072BA"
        ws.sheet_view.rightToLeft = True

        font_header = Font(name='Vazirmatn', size=11, bold=True, color='FFFFFF')
        font_body = Font(name='Vazirmatn', size=11)
        align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')

        headers = ["شماره همراه", "متن پیامک", "نام متقاضی"]
        ws.append(headers)

        for col_idx in range(1, 4):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = header_fill
            cell.alignment = align_center

        candidates = Candidate.objects.filter(id__in=candidate_ids, is_deleted=False).prefetch_related(
            'applications__job', 'applications__stage_states__stage'
        )
        
        for candidate in candidates:
            job = None
            stage = None
            stage_state = None
            application = None

            if job_id:
                application = candidate.applications.filter(job_id=job_id, is_deleted=False).first()
            if not application:
                application = candidate.applications.filter(is_deleted=False).order_by('-created_at').first()

            if application:
                job = application.job
                if stage_id:
                    stage_state = application.stage_states.filter(stage_id=stage_id, is_deleted=False).first()
                else:
                    stage_state = application.stage_states.filter(stage=application.current_stage, is_deleted=False).first()
                if not stage_state:
                    stage_state = application.stage_states.filter(is_deleted=False).order_by('-stage__sequence').first()
                if stage_state:
                    stage = stage_state.stage

            sms_text = render_template_text(
                template_body,
                candidate=candidate,
                job=job,
                stage=stage,
                stage_state=stage_state,
                application=application
            )

            row = [
                candidate.phone_number or "",
                sms_text,
                f"{candidate.first_name} {candidate.last_name}"
            ]
            ws.append(row)

        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=1).alignment = align_center
            ws.cell(row=row, column=1).font = font_body
            
            ws.cell(row=row, column=2).alignment = align_right
            ws.cell(row=row, column=2).font = font_body
            
            ws.cell(row=row, column=3).alignment = align_right
            ws.cell(row=row, column=3).font = font_body

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 30

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sms_export.xlsx"'
        wb.save(response)
        return response



