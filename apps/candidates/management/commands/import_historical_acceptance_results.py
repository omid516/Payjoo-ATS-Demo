"""
Management Command: import_historical_acceptance_results
======================================================
فاز 3-5 ایمپورت تاریخی — ثبت پذیرش نهایی متقاضیان و خاتمه فرصت‌های شغلی مربوطه

استفاده:
    python manage.py import_historical_acceptance_results <path_to_excel>
                       [--sheet قبولی6]
                       [--dry-run]
"""

import re
import sys
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q


def normalize_digits(s):
    """تبدیل ارقام فارسی/عربی به انگلیسی"""
    if s is None:
        return ''
    s = str(s).strip()
    for fa, en in zip('۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩', '01234567890123456789'):
        s = s.replace(fa, en)
    return s


def find_col_idx(headers, keywords):
    """یافتن ایندکس ستون بر اساس کلمات کلیدی"""
    for i, h in enumerate(headers):
        h_lower = str(h).strip().lower()
        for kw in keywords:
            if kw.lower() in h_lower:
                return i
    return None


class Command(BaseCommand):
    help = 'فاز 3-5 ایمپورت تاریخی: ثبت پذیرش نهایی متقاضیان و خاتمه فرصت‌های شغلی مربوطه'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='مسیر فایل اکسل'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default=None,
            help='نام شیت (پیش‌فرض: جستجوی خودکار شیت قبولی)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='فقط تحلیل کن، هیچ رکوردی در دیتابیس تغییر نخواهد کرد'
        )

    def handle(self, *args, **options):
        excel_path = options['excel_file']
        dry_run = options['dry_run']
        sheet_name_arg = options.get('sheet')

        self.stdout.write('')
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(self.style.HTTP_INFO('  ایمپورت نتایج قبولی نهایی و خاتمه مشاغل — فاز 3-5'))
        if dry_run:
            self.stdout.write(self.style.WARNING('  [DRY-RUN] هیچ تغییری در دیتابیس ایجاد نخواهد شد'))
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write('')

        # --- بارگذاری اکسل ---
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'فایل یافت نشد: {excel_path}')
        except Exception as e:
            raise CommandError(f'خطا در بارگذاری فایل: {e}')

        sheet_names = wb.sheetnames
        self.stdout.write(f'شیت‌های موجود: {", ".join(sheet_names)}')

        # --- انتخاب شیت ---
        main_sheet = None
        if sheet_name_arg:
            if sheet_name_arg not in sheet_names:
                raise CommandError(f"شیت '{sheet_name_arg}' یافت نشد.")
            main_sheet = sheet_name_arg
        else:
            # جستجوی خودکار شیت حاوی کلمه قبولی
            for name in sheet_names:
                if 'قبولی' in name:
                    main_sheet = name
                    # اگر قبولی6 وجود داشت، آن را ترجیح بده چون حاوی داده‌های کامل‌تر است
                    if 'قبولی6' in name:
                        break
            if not main_sheet:
                raise CommandError("شیت حاوی نتایج 'قبولی نهایی' یافت نشد.")

        self.stdout.write(f'شیت انتخاب شده: {self.style.SUCCESS(main_sheet)}')
        self.stdout.write('')

        ws = wb[main_sheet]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise CommandError('شیت انتخاب شده خالی است.')

        headers = [str(h).strip() if h is not None else '' for h in all_rows[0]]
        data_rows = [r for r in all_rows[1:] if any(c is not None for c in r)]

        self.stdout.write(f'تعداد ستون‌ها: {len(headers)}')
        self.stdout.write(f'تعداد ردیف‌های داده: {len(data_rows)}')
        self.stdout.write('')

        # --- تشخیص ستون‌ها ---
        col_map = {}
        col_map['exam_code'] = find_col_idx(headers, ['examcode', 'شماره آزمون', 'کد آزمون', 'کد فرصت', 'کد شغل'])
        col_map['national_id'] = find_col_idx(headers, ['nationcode', 'national', 'کد ملی', 'کدملی'])

        self.stdout.write('نگاشت ستون‌ها:')
        field_labels = {
            'exam_code': 'شماره آزمون (کد شغل)',
            'national_id': 'کد ملی متقاضی',
        }
        for field, label in field_labels.items():
            idx = col_map.get(field)
            col_name = f'"{headers[idx]}"' if idx is not None else '— یافت نشد (اختیاری)'
            symbol = '✓' if idx is not None else '?'
            style = self.style.SUCCESS if idx is not None else self.style.WARNING
            self.stdout.write(f'  {symbol} {label}: {style(col_name)}')

        if col_map.get('national_id') is None:
            raise CommandError('ستون حیاتی کد ملی متقاضی یافت نشد.')

        self.stdout.write('')

        # --- اجرای ایمپورت ---
        self._run_import(
            data_rows=data_rows,
            col_map=col_map,
            dry_run=dry_run
        )

    @transaction.atomic
    def _run_import(self, data_rows, col_map, dry_run):
        from apps.candidates.models import JobApplication
        from apps.jobs.models import JobOpportunity

        stats = {
            'processed': 0,
            'apps_accepted': 0,
            'jobs_closed': 0,
            'warnings': [],
            'errors': [],
        }

        # برای ردیابی تعداد مشاغل خاتمه‌یافته منحصر به فرد
        closed_jobs_set = set()

        def get_cell(row, field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row_num, row in enumerate(data_rows, start=2):
            raw_national_id = get_cell(row, 'national_id')
            national_id = normalize_digits(raw_national_id).zfill(10) if raw_national_id is not None else ''

            if not national_id or len(national_id) != 10 or not national_id.isdigit():
                stats['warnings'].append(f'ردیف {row_num}: کد ملی نامعتبر است "{raw_national_id}" — نادیده گرفته شد')
                continue

            # استخراج کد شغل (اگر وجود داشت)
            exam_code = ''
            if col_map.get('exam_code') is not None:
                raw_exam_code = get_cell(row, 'exam_code')
                exam_code = normalize_digits(raw_exam_code).strip() if raw_exam_code is not None else ''

            stats['processed'] += 1

            # یافتن JobApplication
            app = None
            if exam_code:
                app = JobApplication.objects.filter(
                    job__code=exam_code,
                    candidate__national_id=national_id,
                    is_deleted=False
                ).first()
            else:
                # جستجو بر اساس کدملی در صورت نبود کد آزمون در شیت
                apps = JobApplication.objects.filter(
                    candidate__national_id=national_id,
                    is_deleted=False
                )
                if apps.count() == 1:
                    app = apps.first()
                elif apps.exists():
                    stats['warnings'].append(
                        f'ردیف {row_num}: نامزد {national_id} دارای چندین درخواست فعال است و شیت فاقد کد شغل برای ابهام‌زدایی است — رد شد'
                    )
                    continue

            if not app:
                exam_msg = f' در آزمون "{exam_code}"' if exam_code else ''
                stats['warnings'].append(
                    f'ردیف {row_num}: درخواست همکاری برای متقاضی {national_id}{exam_msg} یافت نشد — رد شد'
                )
                continue

            job = app.job

            # ذخیره‌سازی یا بروزرسانی در دیتابیس
            if not dry_run:
                try:
                    # تغییر وضعیت درخواست به قبولی نهایی
                    if app.status != JobApplication.STATUS_SELECTED:
                        app.status = JobApplication.STATUS_SELECTED
                        app.save()
                        stats['apps_accepted'] += 1

                    # خاتمه دادن به فرصت شغلی
                    if job.status != JobOpportunity.STATUS_CLOSED:
                        job.status = JobOpportunity.STATUS_CLOSED
                        job.save(update_fields=['status'])
                        if job.id not in closed_jobs_set:
                            closed_jobs_set.add(job.id)
                            stats['jobs_closed'] += 1
                except Exception as e:
                    stats['errors'].append(f'ردیف {row_num} (کد ملی: {national_id}): خطا در ثبت پذیرش نهایی — {e}')
            else:
                # در حالت dry-run تخمین می‌زنیم
                if app.status != JobApplication.STATUS_SELECTED:
                    stats['apps_accepted'] += 1
                if job.status != JobOpportunity.STATUS_CLOSED and job.id not in closed_jobs_set:
                    closed_jobs_set.add(job.id)
                    stats['jobs_closed'] += 1

        # در صورت dry-run تراکنش را rollback می‌کنیم
        if dry_run:
            transaction.set_rollback(True)

        # --- گزارش نهایی ---
        self.stdout.write('')
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(self.style.HTTP_INFO('  گزارش نهایی ایمپورت قبولی نهایی و خاتمه مشاغل'))
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(f'  تعداد رکوردهای پردازش شده        : {stats["processed"]}')
        self.stdout.write(f'  تعداد درخواست‌های قبول نهایی شده : {self.style.SUCCESS(str(stats["apps_accepted"]))}')
        self.stdout.write(f'  تعداد فرصت‌های شغلی بسته‌شده       : {self.style.SUCCESS(str(stats["jobs_closed"]))}')
        self.stdout.write(f'  هشدارها                          : {self.style.WARNING(str(len(stats["warnings"])))}')
        self.stdout.write(f'  خطاهای سیستم                     : {self.style.ERROR(str(len(stats["errors"])))}')

        if stats['warnings']:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING('برخی هشدارها (حداکثر ۵۰ مورد):'))
            for w in stats['warnings'][:50]:
                self.stdout.write(f'  ⚠  {w}')
            if len(stats['warnings']) > 50:
                self.stdout.write(self.style.WARNING(f'  ... و {len(stats["warnings"])-50} هشدار دیگر.'))

        if stats['errors']:
            self.stdout.write('')
            self.stdout.write(self.style.ERROR('خطاهای سیستم:'))
            for e in stats['errors']:
                self.stdout.write(f'  ✗  {e}')

        self.stdout.write('')
        if dry_run:
            self.stdout.write(self.style.WARNING('[DRY-RUN] هیچ تغییری اعمال نشد. برای اجرای واقعی --dry-run را حذف کنید.'))
        else:
            self.stdout.write(self.style.SUCCESS('ایمپورت پذیرفته‌شدگان نهایی و بسته‌شدن فرصت‌های شغلی با موفقیت به پایان رسید.'))
        self.stdout.write('')
