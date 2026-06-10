"""
Management Command: import_historical_assessment_results
=========================================================
فاز 3-4 ایمپورت تاریخی — ثبت نتایج کانون ارزیابی متقاضیان

استفاده:
    python manage.py import_historical_assessment_results <path_to_excel>
                       [--sheet کانون5]
                       [--dry-run]
"""

import re
import sys
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q


def normalize_digits(s):
    """تبدیل ارقام فارسی/عربی به انگلیسی"""
    if s is None:
        return ''
    s = str(s).strip()
    for fa, en in zip('۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩', '01234567890123456789'):
        s = s.replace(fa, en)
    return s


def find_col_idx(headers, keywords):
    """یافتن ایندکس ستون بر اساس کلمات کلیدی"""
    for i, h in enumerate(headers):
        h_lower = str(h).strip().lower()
        for kw in keywords:
            if kw.lower() in h_lower:
                return i
    return None


class Command(BaseCommand):
    help = 'فاز 3-4 ایمپورت تاریخی: ثبت نتایج کانون ارزیابی متقاضیان از فایل اکسل'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='مسیر فایل اکسل'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default=None,
            help='نام شیت (پیش‌فرض: جستجوی خودکار شیت کانون)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='فقط تحلیل کن، هیچ رکوردی در دیتابیس تغییر نخواهد کرد'
        )

    def handle(self, *args, **options):
        excel_path = options['excel_file']
        dry_run = options['dry_run']
        sheet_name_arg = options.get('sheet')

        self.stdout.write('')
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(self.style.HTTP_INFO('  ایمپورت نتایج کانون ارزیابی متقاضیان — فاز 3-4'))
        if dry_run:
            self.stdout.write(self.style.WARNING('  [DRY-RUN] هیچ تغییری در دیتابیس ایجاد نخواهد شد'))
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write('')

        # --- بارگذاری اکسل ---
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'فایل یافت نشد: {excel_path}')
        except Exception as e:
            raise CommandError(f'خطا در بارگذاری فایل: {e}')

        sheet_names = wb.sheetnames
        self.stdout.write(f'شیت‌های موجود: {", ".join(sheet_names)}')

        # --- انتخاب شیت ---
        main_sheet = None
        if sheet_name_arg:
            if sheet_name_arg not in sheet_names:
                raise CommandError(f"شیت '{sheet_name_arg}' یافت نشد.")
            main_sheet = sheet_name_arg
        else:
            # جستجوی خودکار شیت حاوی کلمه کانون
            for name in sheet_names:
                if 'کانون' in name:
                    main_sheet = name
                    # اگر کانون۵ وجود داشت، کانون۵ را ترجیح بده چون حاوی داده‌های کامل‌تر است
                    if 'کانون5' in name:
                        break
            if not main_sheet:
                raise CommandError("شیت حاوی نتایج آزمون 'کانون' یافت نشد.")

        self.stdout.write(f'شیت انتخاب شده: {self.style.SUCCESS(main_sheet)}')
        self.stdout.write('')

        ws = wb[main_sheet]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise CommandError('شیت انتخاب شده خالی است.')

        headers = [str(h).strip() if h is not None else '' for h in all_rows[0]]
        data_rows = [r for r in all_rows[1:] if any(c is not None for c in r)]

        self.stdout.write(f'تعداد ستون‌ها: {len(headers)}')
        self.stdout.write(f'تعداد ردیف‌های داده: {len(data_rows)}')
        self.stdout.write('')

        # --- تشخیص ستون‌ها ---
        col_map = {}
        col_map['exam_code'] = find_col_idx(headers, ['examcode', 'شماره آزمون', 'کد آزمون', 'کد فرصت', 'کد شغل'])
        col_map['national_id'] = find_col_idx(headers, ['nationcode', 'national', 'کد ملی', 'کدملی'])
        col_map['score'] = find_col_idx(headers, ['scoreac', 'score_ac', 'نمره کانون', 'نمره'])
        col_map['cscore'] = find_col_idx(headers, ['cscoreac', 'cscore_ac', 'امتیاز تجمعی', 'نمره تجمعی'])
        col_map['result'] = find_col_idx(headers, ['result4', 'result', 'وضعیت قبولی کانون', 'وضعیت کانون', 'وضعیت قبولی', 'وضعیت'])

        self.stdout.write('نگاشت ستون‌ها:')
        field_labels = {
            'exam_code': 'شماره آزمون (کد شغل)',
            'national_id': 'کد ملی متقاضی',
            'score': 'نمره کانون ارزیابی (ScoreAC)',
            'cscore': 'امتیاز تجمعی (CScoreAC)',
            'result': 'وضعیت نتیجه کانون (Result4)',
        }
        for field, label in field_labels.items():
            idx = col_map.get(field)
            col_name = f'"{headers[idx]}"' if idx is not None else '— یافت نشد (اختیاری)'
            symbol = '✓' if idx is not None else '?'
            style = self.style.SUCCESS if idx is not None else self.style.WARNING
            self.stdout.write(f'  {symbol} {label}: {style(col_name)}')

        if any(col_map[f] is None for f in ['national_id', 'score']):
            raise CommandError('برخی از ستون‌های حیاتی کانون ارزیابی (کد ملی یا نمره کانون) یافت نشد.')

        self.stdout.write('')

        # --- اجرای ایمپورت ---
        self._run_import(
            data_rows=data_rows,
            col_map=col_map,
            dry_run=dry_run
        )

    @transaction.atomic
    def _run_import(self, data_rows, col_map, dry_run):
        from apps.candidates.models import JobApplication, ApplicationStageState
        from apps.recruitment_planning.models import JobStagePlan

        stats = {
            'processed': 0,
            'stage_created': 0,
            'stage_updated': 0,
            'warnings': [],
            'errors': [],
        }

        # Cache برای مراحل کانون فرصت‌های شغلی جهت سرعت بیشتر
        stage_cache = {}

        def get_cell(row, field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row_num, row in enumerate(data_rows, start=2):
            raw_national_id = get_cell(row, 'national_id')
            raw_score = get_cell(row, 'score')
            raw_cscore = get_cell(row, 'cscore')
            raw_result = get_cell(row, 'result')

            national_id = normalize_digits(raw_national_id).zfill(10) if raw_national_id is not None else ''
            result_str = str(raw_result).strip() if raw_result is not None else ''

            if not national_id or len(national_id) != 10 or not national_id.isdigit():
                stats['warnings'].append(f'ردیف {row_num}: کد ملی نامعتبر است "{raw_national_id}" — نادیده گرفته شد')
                continue

            # استخراج کد شغل (اگر وجود داشت)
            exam_code = ''
            if col_map.get('exam_code') is not None:
                raw_exam_code = get_cell(row, 'exam_code')
                exam_code = normalize_digits(raw_exam_code).strip() if raw_exam_code is not None else ''

            # تبدیل نمره
            score_val = 0.0
            if raw_score is not None:
                try:
                    score_val = float(normalize_digits(raw_score))
                except ValueError:
                    stats['warnings'].append(f'ردیف {row_num} (کد ملی: {national_id}): نمره نامعتبر "{raw_score}" — مقدار ۰ تنظیم شد')

            # تبدیل امتیاز تجمعی تاریخی
            cscore_val = None
            if raw_cscore is not None and str(raw_cscore).strip() != '':
                try:
                    cscore_val = float(normalize_digits(raw_cscore))
                except ValueError:
                    pass

            stats['processed'] += 1

            # یافتن JobApplication
            app = None
            if exam_code:
                app = JobApplication.objects.filter(
                    job__code=exam_code,
                    candidate__national_id=national_id,
                    is_deleted=False
                ).first()
            else:
                # جستجو بر اساس کدملی در صورت نبود کد آزمون در شیت
                apps = JobApplication.objects.filter(
                    candidate__national_id=national_id,
                    is_deleted=False
                )
                if apps.count() == 1:
                    app = apps.first()
                elif apps.exists():
                    # فیلتر برای فرصت‌هایی که مرحله ASSESSMENT دارند
                    apps_with_assessment = [a for a in apps if a.job.stages.filter(stage_type='ASSESSMENT', is_deleted=False).exists()]
                    if len(apps_with_assessment) == 1:
                        app = apps_with_assessment[0]
                    else:
                        stats['warnings'].append(
                            f'ردیف {row_num}: نامزد {national_id} دارای چندین درخواست فعال است و شیت فاقد کد شغل برای ابهام‌زدایی است — رد شد'
                        )
                        continue

            if not app:
                exam_msg = f' در آزمون "{exam_code}"' if exam_code else ''
                stats['warnings'].append(
                    f'ردیف {row_num}: درخواست همکاری برای متقاضی {national_id}{exam_msg} یافت نشد — رد شد'
                )
                continue

            job = app.job

            # پیدا کردن مرحله کانون ارزیابی مربوط به این شغل
            if job.id not in stage_cache:
                # مرحله‌ای از نوع ASSESSMENT
                assessment_stage = job.stages.filter(stage_type='ASSESSMENT', is_deleted=False).first()
                stage_cache[job.id] = assessment_stage

            stage = stage_cache[job.id]

            if not stage:
                stats['warnings'].append(
                    f'ردیف {row_num}: هیچ مرحله ارزیابی از نوع "کانون ارزیابی" (ASSESSMENT) برای فرصت شغلی {job.code} یافت نشد — رد شد'
                )
                continue

            # تعیین وضعیت مرحله ارزیابی
            # ۱. غایب -> FAILED
            # ۲. مجاز -> COMPLETED
            # ۳. غیر مجاز -> FAILED
            # ۴. خالی/سایر -> مقایسه نمره با کف قبولی مرحله
            notes_list = []
            notes_str = ''
            if 'غایب' in result_str:
                status_val = ApplicationStageState.STATUS_FAILED
                notes_str = 'غایب در کانون ارزیابی'
                score_val = 0.0
            elif 'غير' in result_str or 'غیر' in result_str or 'ineligible' in result_str.lower():
                status_val = ApplicationStageState.STATUS_FAILED
            elif 'مجاز' in result_str or 'eligible' in result_str.lower():
                status_val = ApplicationStageState.STATUS_COMPLETED
            else:
                # مقایسه نمره با کف قبولی مرحله
                if score_val >= stage.passing_score:
                    status_val = ApplicationStageState.STATUS_COMPLETED
                else:
                    status_val = ApplicationStageState.STATUS_FAILED

            if notes_str:
                notes_list.append(notes_str)
            if cscore_val is not None:
                notes_list.append(f"امتیاز تجمعی تاریخی: {cscore_val}")
            
            final_notes = " | ".join(notes_list)

            # یافتن تاریخ ارزیابی از روی JobStagePlan برنامه‌ریزی شده
            eval_date = None
            plan_stage = JobStagePlan.objects.filter(plan__job=job, stage=stage, is_deleted=False).first()
            if plan_stage:
                eval_date = plan_stage.planned_start_date

            # ذخیره‌سازی یا بروزرسانی در دیتابیس
            if not dry_run:
                try:
                    state, created = ApplicationStageState.objects.update_or_create(
                        application=app,
                        stage=stage,
                        defaults={
                            'score': score_val,
                            'status': status_val,
                            'notes': final_notes,
                            'evaluation_date': eval_date,
                        }
                    )
                    if created:
                        stats['stage_created'] += 1
                    else:
                        stats['stage_updated'] += 1
                except Exception as e:
                    stats['errors'].append(f'ردیف {row_num} (کد ملی: {national_id}): خطا در ثبت نتیجه — {e}')
            else:
                # در حالت dry-run تخمین می‌زنیم
                existing_state = ApplicationStageState.objects.filter(application=app, stage=stage).exists()
                if existing_state:
                    stats['stage_updated'] += 1
                else:
                    stats['stage_created'] += 1

        # در صورت dry-run تراکنش را rollback می‌کنیم
        if dry_run:
            transaction.set_rollback(True)

        # --- گزارش نهایی ---
        self.stdout.write('')
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(self.style.HTTP_INFO('  گزارش نهایی ایمپورت نتایج کانون ارزیابی'))
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(f'  تعداد رکوردهای پردازش شده : {stats["processed"]}')
        self.stdout.write(f'  Stage State های ایجاد شده  : {self.style.SUCCESS(str(stats["stage_created"]))}')
        self.stdout.write(f'  Stage State های بروزرسانی شده: {self.style.SUCCESS(str(stats["stage_updated"]))}')
        self.stdout.write(f'  هشدارها                      : {self.style.WARNING(str(len(stats["warnings"])))}')
        self.stdout.write(f'  خطاهای سیستم                 : {self.style.ERROR(str(len(stats["errors"])))}')

        if stats['warnings']:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING('برخی هشدارها (حداکثر ۵۰ مورد):'))
            for w in stats['warnings'][:50]:
                self.stdout.write(f'  ⚠  {w}')
            if len(stats['warnings']) > 50:
                self.stdout.write(self.style.WARNING(f'  ... و {len(stats["warnings"])-50} هشدار دیگر.'))

        if stats['errors']:
            self.stdout.write('')
            self.stdout.write(self.style.ERROR('خطاهای سیستم:'))
            for e in stats['errors']:
                self.stdout.write(f'  ✗  {e}')

        self.stdout.write('')
        if dry_run:
            self.stdout.write(self.style.WARNING('[DRY-RUN] هیچ تغییری اعمال نشد. برای اجرای واقعی --dry-run را حذف کنید.'))
        else:
            self.stdout.write(self.style.SUCCESS('ایمپورت نتایج کانون ارزیابی با موفقیت به پایان رسید.'))
        self.stdout.write('')
