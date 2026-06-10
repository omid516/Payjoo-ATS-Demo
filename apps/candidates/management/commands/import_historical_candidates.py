"""
Management Command: import_historical_candidates
================================================
فاز دوم ایمپورت تاریخی — ایجاد متقاضیان و درخواست‌های همکاری

استفاده:
    python manage.py import_historical_candidates <path_to_excel>
                       [--sheet ثبت_نام 1]
                       [--dry-run]
"""

import re
import sys
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


def normalize_digits(s):
    """تبدیل ارقام فارسی/عربی به انگلیسی"""
    if s is None:
        return ''
    s = str(s).strip()
    for fa, en in zip('۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩', '01234567890123456789'):
        s = s.replace(fa, en)
    return s


def clean_phone_number(val):
    """پاکسازی و استانداردسازی شماره تلفن همراه"""
    if val is None:
        return ''
    s = normalize_digits(val)
    if s.endswith('.0'):
        s = s[:-2]
    s = re.sub(r'\D', '', s)  # فقط ارقام حفظ شود
    
    # تصحیح پیشوندها
    if s.startswith('989') and len(s) == 12:
        s = '0' + s[2:]
    elif len(s) == 10 and s.startswith('9'):
        s = '0' + s
    return s


def find_col_idx(headers, keywords):
    """یافتن ایندکس ستون بر اساس کلمات کلیدی"""
    for i, h in enumerate(headers):
        h_lower = str(h).strip().lower()
        for kw in keywords:
            if kw.lower() in h_lower:
                return i
    return None


class Command(BaseCommand):
    help = 'فاز دوم ایمپورت تاریخی: ایجاد متقاضیان و درخواست‌های همکاری از فایل اکسل'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='مسیر فایل اکسل'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default=None,
            help='نام شیت (پیش‌فرض: جستجوی خودکار شیت ثبت‌نام)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='فقط تحلیل کن، هیچ رکوردی در دیتابیس ایجاد نکن'
        )

    def handle(self, *args, **options):
        excel_path = options['excel_file']
        dry_run = options['dry_run']
        sheet_name_arg = options.get('sheet')

        self.stdout.write('')
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(self.style.HTTP_INFO('  ایمپورت تاریخی متقاضیان و درخواست‌ها — فاز دوم'))
        if dry_run:
            self.stdout.write(self.style.WARNING('  [DRY-RUN] هیچ تغییری در دیتابیس ایجاد نخواهد شد'))
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write('')

        # --- بارگذاری اکسل ---
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'فایل یافت نشد: {excel_path}')
        except Exception as e:
            raise CommandError(f'خطا در بارگذاری فایل: {e}')

        sheet_names = wb.sheetnames
        self.stdout.write(f'شیت‌های موجود: {", ".join(sheet_names)}')

        # --- انتخاب شیت ---
        main_sheet = None
        if sheet_name_arg:
            if sheet_name_arg not in sheet_names:
                raise CommandError(f"شیت '{sheet_name_arg}' یافت نشد.")
            main_sheet = sheet_name_arg
        else:
            for name in sheet_names:
                if 'ثبت_نام' in name or 'ثبت نام' in name or 'candidate' in name.lower() or 'signup' in name.lower():
                    main_sheet = name
                    break
            if not main_sheet:
                main_sheet = sheet_names[0]
                self.stdout.write(self.style.WARNING(
                    f"شیت ثبت‌نام پیدا نشد. از اولین شیت استفاده می‌شود: '{main_sheet}'"
                ))

        self.stdout.write(f'شیت انتخاب شده: {self.style.SUCCESS(main_sheet)}')
        self.stdout.write('')

        ws = wb[main_sheet]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise CommandError('شیت انتخاب شده خالی است.')

        headers = [str(h).strip() if h is not None else '' for h in all_rows[0]]
        data_rows = [r for r in all_rows[1:] if any(c is not None for c in r)]

        self.stdout.write(f'تعداد ستون‌ها: {len(headers)}')
        self.stdout.write(f'تعداد ردیف‌های داده: {len(data_rows)}')
        self.stdout.write('')

        # --- تشخیص ستون‌ها ---
        col_map = {}
        col_map['last_name'] = find_col_idx(headers, ['نام خانوادگی', 'family', 'last_name', 'شهرت'])
        col_map['phone_number'] = find_col_idx(headers, ['شماره همراه', 'همراه', 'mobile', 'تلفن همراه', 'phone', 'تلفن'])
        col_map['national_id'] = find_col_idx(headers, ['nationcode', 'national', 'کد ملی', 'کدملی'])
        col_map['exam_code'] = find_col_idx(headers, ['examcode', 'شماره آزمون', 'کد آزمون', 'کد فرصت', 'کد شغل'])
        col_map['result'] = find_col_idx(headers, ['result', 'نتیجه', 'وضعیت'])

        # فیلد نام (شامل کلمه نام ولی فاقد کلمه خانوادگی)
        first_name_idx = None
        for i, h in enumerate(headers):
            h_lower = str(h).strip().lower()
            if 'نام' in h_lower and 'خانوادگی' not in h_lower:
                first_name_idx = i
                break
        if first_name_idx is None:
            first_name_idx = find_col_idx(headers, ['first_name', 'name', 'نام'])
        col_map['first_name'] = first_name_idx

        self.stdout.write('نگاشت ستون‌ها:')
        field_labels = {
            'exam_code': 'شماره آزمون (کد شغل)',
            'national_id': 'کد ملی متقاضی',
            'first_name': 'نام',
            'last_name': 'نام خانوادگی',
            'phone_number': 'شماره همراه',
            'result': 'وضعیت نتیجه ثبت‌نام',
        }
        for field, label in field_labels.items():
            idx = col_map.get(field)
            col_name = f'"{headers[idx]}"' if idx is not None else '— یافت نشد'
            symbol = '✓' if idx is not None else '✗'
            style = self.style.SUCCESS if idx is not None else self.style.WARNING
            self.stdout.write(f'  {symbol} {label}: {style(col_name)}')

        if any(col_map[f] is None for f in ['exam_code', 'national_id', 'first_name', 'last_name', 'phone_number']):
            raise CommandError('برخی از ستون‌های حیاتی متقاضیان (کد ملی، نام، نام خانوادگی، همراه یا کد آزمون) یافت نشد.')

        self.stdout.write('')

        # --- اجرای ایمپورت ---
        self._run_import(
            data_rows=data_rows,
            col_map=col_map,
            dry_run=dry_run
        )

    @transaction.atomic
    def _run_import(self, data_rows, col_map, dry_run):
        from apps.candidates.models import Candidate, JobApplication
        from apps.jobs.models import JobOpportunity

        stats = {
            'candidates_created': 0,
            'candidates_existing': 0,
            'applications_created': 0,
            'applications_skipped': 0,
            'warnings': [],
            'errors': [],
        }

        # Cache برای JobOpportunity ها جهت سرعت بیشتر
        job_cache = {}

        def get_cell(row, field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row_num, row in enumerate(data_rows, start=2):
            raw_exam_code = get_cell(row, 'exam_code')
            raw_national_id = get_cell(row, 'national_id')
            raw_first_name = get_cell(row, 'first_name')
            raw_last_name = get_cell(row, 'last_name')
            raw_phone_number = get_cell(row, 'phone_number')
            raw_result = get_cell(row, 'result')

            exam_code = normalize_digits(raw_exam_code).strip() if raw_exam_code is not None else ''
            national_id = normalize_digits(raw_national_id).zfill(10) if raw_national_id is not None else ''
            first_name = str(raw_first_name).strip() if raw_first_name is not None else ''
            last_name = str(raw_last_name).strip() if raw_last_name is not None else ''
            phone_number = clean_phone_number(raw_phone_number)
            result = str(raw_result).strip() if raw_result is not None else ''

            # اعتبارسنجی‌های پایه
            if not national_id or len(national_id) != 10 or not national_id.isdigit():
                stats['warnings'].append(f'ردیف {row_num}: کد ملی نامعتبر است "{raw_national_id}" — نادیده گرفته شد')
                continue
            if not exam_code:
                stats['warnings'].append(f'ردیف {row_num}: شماره آزمون خالی است — نادیده گرفته شد')
                continue
            if not first_name or not last_name:
                stats['warnings'].append(f'ردیف {row_num} (کد ملی: {national_id}): نام یا نام خانوادگی خالی است — نادیده گرفته شد')
                continue

            # پیدا کردن فرصت شغلی
            if exam_code not in job_cache:
                job = JobOpportunity.objects.filter(code=exam_code, is_deleted=False).first()
                job_cache[exam_code] = job
            job = job_cache[exam_code]

            if not job:
                stats['warnings'].append(f'ردیف {row_num}: فرصت شغلی با کد "{exam_code}" در دیتابیس یافت نشد — نادیده گرفته شد')
                continue

            # ایجاد یا یافتن Candidate
            candidate = None
            candidate_created = False
            try:
                candidate = Candidate.objects.filter(national_id=national_id, is_deleted=False).first()
                if not candidate:
                    # بررسی سافت دیلیت شده
                    candidate = Candidate.all_objects.filter(national_id=national_id).first()
                    if candidate:
                        candidate.is_deleted = False
                        candidate.deleted_at = None
                        candidate.first_name = first_name
                        candidate.last_name = last_name
                        candidate.phone_number = phone_number
                        candidate.save()
                    else:
                        if not dry_run:
                            candidate = Candidate.objects.create(
                                national_id=national_id,
                                first_name=first_name,
                                last_name=last_name,
                                phone_number=phone_number,
                                email=''  # فیلد ایمیل خالی
                            )
                        candidate_created = True
            except Exception as e:
                stats['errors'].append(f'ردیف {row_num}: خطا در ایجاد متقاضی {national_id} — {e}')
                continue

            if candidate_created:
                stats['candidates_created'] += 1
            else:
                stats['candidates_existing'] += 1

            # تعیین وضعیت درخواست همکاری
            # مجاز -> IN_PROGRESS
            # غیر مجاز -> REJECTED
            app_status = JobApplication.STATUS_IN_PROGRESS
            if 'غیر' in result or 'ineligible' in result.lower():
                app_status = JobApplication.STATUS_REJECTED

            # بررسی و ایجاد JobApplication
            if not dry_run:
                # بررسی وجود درخواست (کلید یونیک زوج job و candidate)
                existing_app = JobApplication.objects.filter(job=job, candidate=candidate, is_deleted=False).exists()
                if existing_app:
                    stats['applications_skipped'] += 1
                    continue

                # بررسی درخواست‌های سافت‌دیلیت شده
                app = JobApplication.all_objects.filter(job=job, candidate=candidate).first()
                if app:
                    app.is_deleted = False
                    app.deleted_at = None
                    app.status = app_status
                    app._bypass_stage_creation = True  # دور زدن ساخت خودکار استیج استیت‌ها
                    app.save()
                else:
                    app = JobApplication(
                        job=job,
                        candidate=candidate,
                        status=app_status
                    )
                    app._bypass_stage_creation = True  # دور زدن ساخت خودکار استیج استیت‌ها
                    app.save()

                stats['applications_created'] += 1
            else:
                # در حالت dry-run فرض می‌کنیم درخواست ایجاد می‌شود اگر از قبل موجود نباشد
                # (با توجه به اینکه کاندیدای جدید یا قدیم وجود دارد)
                # در صورت عدم اجرای کوئری در درای‌ران، فرض می‌کنیم درخواست‌ها ساخته می‌شوند
                if candidate:
                    existing_app = JobApplication.objects.filter(job=job, candidate=candidate, is_deleted=False).exists()
                    if existing_app:
                        stats['applications_skipped'] += 1
                    else:
                        stats['applications_created'] += 1
                else:
                    stats['applications_created'] += 1

        # در صورت dry-run تراکنش را rollback می‌کنیم
        if dry_run:
            transaction.set_rollback(True)

        # --- گزارش نهایی ---
        self.stdout.write('')
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(self.style.HTTP_INFO('  گزارش نهایی ایمپورت متقاضیان'))
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(f'  Candidate های جدید ایجاد شده : {self.style.SUCCESS(str(stats["candidates_created"]))}')
        self.stdout.write(f'  Candidate های موجود (بازنشانی)  : {stats["candidates_existing"]}')
        self.stdout.write(f'  JobApplication های ایجاد شده  : {self.style.SUCCESS(str(stats["applications_created"]))}')
        self.stdout.write(f'  JobApplication های تکراری (رد) : {stats["applications_skipped"]}')
        self.stdout.write(f'  هشدارها                      : {self.style.WARNING(str(len(stats["warnings"])))}')
        self.stdout.write(f'  خطاهای بحرانی                : {self.style.ERROR(str(len(stats["errors"])))}')

        if stats['warnings']:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING('هشدارها:'))
            for w in stats['warnings'][:50]:  # نمایش حداکثر ۵۰ هشدار اول
                self.stdout.write(f'  ⚠  {w}')
            if len(stats['warnings']) > 50:
                self.stdout.write(self.style.WARNING(f'  ... و {len(stats["warnings"])-50} هشدار دیگر.'))

        if stats['errors']:
            self.stdout.write('')
            self.stdout.write(self.style.ERROR('خطاهای بحرانی:'))
            for e in stats['errors']:
                self.stdout.write(f'  ✗  {e}')

        self.stdout.write('')
        if dry_run:
            self.stdout.write(self.style.WARNING('[DRY-RUN] هیچ تغییری اعمال نشد. برای اجرای واقعی --dry-run را حذف کنید.'))
        else:
            self.stdout.write(self.style.SUCCESS('ایمپورت متقاضیان با موفقیت به پایان رسید.'))
        self.stdout.write('')
