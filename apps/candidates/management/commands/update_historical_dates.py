"""
Management Command: update_historical_dates
==========================================
فاز ۴ ایمپورت تاریخی — بروزرسانی تاریخ‌های مراحل و فرصت‌های شغلی بر اساس تاریخ‌های شمسی شیت وضعیت

استفاده:
    python manage.py update_historical_dates <path_to_excel>
                       [--sheet وضعیت]
                       [--dry-run]
"""

import re
import sys
import datetime
import openpyxl
import jdatetime
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


def normalize_digits(s):
    """تبدیل ارقام فارسی/عربی به انگلیسی"""
    if s is None:
        return ''
    s = str(s).strip()
    for fa, en in zip('۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩', '01234567890123456789'):
        s = s.replace(fa, en)
    return s


def find_shamsi_col_idx(headers, keywords):
    """یافتن ایندکس ستون بر اساس کلمات کلیدی، با اولویت ستون‌های غیرمیلادی"""
    for i, h in enumerate(headers):
        h_lower = str(h).strip().lower()
        if 'میلادی' in h_lower:
            continue
        # Exclude numeric/count statistics columns
        if any(bad in h_lower for bad in ['تعداد', 'روز', 'نفرات', 'حاضرین', 'نامه', 'چرخه']):
            continue
        for kw in keywords:
            if kw.lower() in h_lower:
                return i
    return None


def parse_shamsi_date_str(val):
    """تبدیل تاریخ شمسی (رشته یا عدد) به میلادی. در صورت شکست None برمی‌گرداند."""
    if val is None:
        return None
    
    # اگر خود شیء تاریخ باشد
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val

    val_str = normalize_digits(val).strip()
    if not val_str or val_str.lower() in ('none', 'null', '-', '', '*'):
        return None

    # فرمت YYYYMMDD (۸ رقم بدون جداکننده)
    if len(val_str) == 8 and val_str.isdigit():
        y = int(val_str[0:4])
        m = int(val_str[4:6])
        d = int(val_str[6:8])
        try:
            return jdatetime.date(y, m, d).togregorian()
        except Exception:
            return None

    # فرمت YYYY/MM/DD یا YYYY-MM-DD
    parts = [int(p) for p in re.findall(r'\d+', val_str)]
    if len(parts) == 3:
        try:
            y, mo, d = parts
            if y < 100:
                y += 1400
            return jdatetime.date(y, mo, d).togregorian()
        except Exception:
            return None
            
    return None


STATUS_KEYWORD_MAP = [
    (['لغو', 'کنسل', 'cancel', 'لغو درخواست'],                                'CANCELLED'),
    (['توقف', 'متوقف'],                                                         'CANCELLED'),
    (['پایان فرآیند', 'اتمام', 'بسته', 'closed', 'تمام'],                     'CLOSED'),
    (['ثبت نمرات کانون', 'کانون', 'assessment'],                               'ASSESSMENT'),
    (['هماهنگی مصاحبه', 'مصاحبه', 'interview'],                               'INTERVIEW'),
    (['ثبت نمرات آزمون مهارتی', 'آزمون مهارتی', 'ارزیابی مهارتی', 'مهارتی'], 'SKILL_TEST'),
    (['هماهنگی آزمون کتبی', 'ثبت نمرات کتبی', 'آزمون کتبی', 'کتبی'],        'EXAM'),
    (['غربالگری', 'غربال', 'screening'],                                        'SCREENING'),
    (['پیش نویس آگهی', 'منتشر', 'ثبت‌نام', 'ثبت نام', 'published'],          'PUBLISHED'),
    (['برنامه‌ریزی', 'برنامه ریزی', 'planning'],                               'PLANNING'),
    (['انتخاب نهایی', 'final'],                                                 'FINAL_SELECTION'),
    (['خاتمه ثبت'],                                                             'REGISTRATION_CLOSED'),
    (['دریافت', 'در انتظار', 'ثبت درخواست'],                                   'RECEIVED'),
]


def map_status(raw_val):
    if not raw_val:
        return 'RECEIVED'
    raw_lower = str(raw_val).strip().lower()
    for keywords, code in STATUS_KEYWORD_MAP:
        for kw in keywords:
            if kw.lower() in raw_lower:
                return code
    return 'RECEIVED'



class Command(BaseCommand):
    help = 'فاز ۴ ایمپورت تاریخی: بروزرسانی تاریخ‌های مراحل و فرصت‌های شغلی بر اساس تاریخ‌های شمسی شیت وضعیت'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='مسیر فایل اکسل'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default=None,
            help='نام شیت (پیش‌فرض: جستجوی خودکار شیت وضعیت)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='فقط تحلیل کن، هیچ رکوردی در دیتابیس تغییر نخواهد کرد'
        )

    def handle(self, *args, **options):
        excel_path = options['excel_file']
        dry_run = options['dry_run']
        sheet_name_arg = options.get('sheet')

        self.stdout.write('')
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(self.style.HTTP_INFO('  بروزرسانی تاریخ‌های شمسی مراحل و مشاغل — فاز ۴'))
        if dry_run:
            self.stdout.write(self.style.WARNING('  [DRY-RUN] هیچ تغییری در دیتابیس ایجاد نخواهد شد'))
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write('')

        # --- بارگذاری اکسل ---
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'فایل یافت نشد: {excel_path}')
        except Exception as e:
            raise CommandError(f'خطا در بارگذاری فایل: {e}')

        sheet_names = wb.sheetnames
        self.stdout.write(f'شیت‌های موجود: {", ".join(sheet_names)}')

        # --- انتخاب شیت ---
        main_sheet = None
        if sheet_name_arg:
            if sheet_name_arg not in sheet_names:
                raise CommandError(f"شیت '{sheet_name_arg}' یافت نشد.")
            main_sheet = sheet_name_arg
        else:
            for name in sheet_names:
                if 'وضعیت' in name or 'status' in name.lower() or 'فرصت' in name:
                    main_sheet = name
                    break
            if not main_sheet:
                raise CommandError("شیت حاوی اطلاعات 'وضعیت' یافت نشد.")

        self.stdout.write(f'شیت انتخاب شده: {self.style.SUCCESS(main_sheet)}')
        self.stdout.write('')

        ws = wb[main_sheet]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise CommandError('شیت انتخاب شده خالی است.')

        headers = [str(h).strip() if h is not None else '' for h in all_rows[0]]
        data_rows = [r for r in all_rows[1:] if any(c is not None for c in r)]

        self.stdout.write(f'تعداد ستون‌ها: {len(headers)}')
        self.stdout.write(f'تعداد ردیف‌های داده: {len(data_rows)}')
        self.stdout.write('')

        # --- تشخیص ستون‌ها ---
        col_map = {}
        col_map['code'] = find_shamsi_col_idx(headers, ['کد', 'کد فرصت', 'کد شغل'])
        col_map['start_reg'] = find_shamsi_col_idx(headers, ['شروع ثبت‌نام', 'شروع ثبت نام'])
        col_map['end_reg'] = find_shamsi_col_idx(headers, ['پایان ثبت‌نام', 'پایان ثبت نام'])
        col_map['screening'] = find_shamsi_col_idx(headers, ['پایان غربالگری'])
        col_map['exam'] = find_shamsi_col_idx(headers, ['آزمون کتبی'])
        col_map['skill'] = find_shamsi_col_idx(headers, ['آزمون مهارتی'])
        col_map['interview'] = find_shamsi_col_idx(headers, ['مصاحبه'])
        col_map['assessment'] = find_shamsi_col_idx(headers, ['معرفی به کانون'])
        col_map['final_selection'] = find_shamsi_col_idx(headers, ['اعلام نتیجه نهایی'])
        col_map['current_status'] = find_shamsi_col_idx(headers, ['آخرین مرحله', 'وضعیت فعلی', 'وضعیت'])

        self.stdout.write('نگاشت ستون‌های تاریخ شمسی:')
        field_labels = {
            'code': 'کد فرصت شغلی',
            'start_reg': 'شروع ثبت نام',
            'end_reg': 'پایان ثبت نام (پایان فرصت شغلی)',
            'screening': 'تاریخ پایان غربالگری',
            'exam': 'تاریخ آزمون کتبی',
            'skill': 'تاریخ آزمون مهارتی',
            'interview': 'تاریخ مصاحبه',
            'assessment': 'تاریخ معرفی به کانون',
            'final_selection': 'تاریخ اعلام نتیجه نهایی',
            'current_status': 'آخرین مرحله (وضعیت)',
        }
        for field, label in field_labels.items():
            idx = col_map.get(field)
            col_name = f'"{headers[idx]}"' if idx is not None else '— یافت نشد (نادیده گرفته می‌شود)'
            symbol = '✓' if idx is not None else '?'
            style = self.style.SUCCESS if idx is not None else self.style.WARNING
            self.stdout.write(f'  {symbol} {label}: {style(col_name)}')

        if col_map.get('code') is None:
            raise CommandError('ستون حیاتی کد فرصت شغلی یافت نشد.')

        self.stdout.write('')

        # --- اجرای بروزرسانی ---
        self._run_update(
            data_rows=data_rows,
            col_map=col_map,
            dry_run=dry_run
        )

    @transaction.atomic
    def _run_update(self, data_rows, col_map, dry_run):
        from apps.jobs.models import JobOpportunity
        from apps.recruitment_planning.models import JobRecruitmentPlan, JobStagePlan
        from apps.candidates.models import ApplicationStageState

        stats = {
            'processed': 0,
            'jobs_updated': 0,
            'stage_plans_updated': 0,
            'candidate_states_updated': 0,
            'warnings': [],
            'errors': [],
        }

        def get_cell(row, field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row_num, row in enumerate(data_rows, start=2):
            raw_code = get_cell(row, 'code')
            code = normalize_digits(raw_code).strip() if raw_code is not None else ''

            if not code:
                stats['warnings'].append(f'ردیف {row_num}: کد فرصت شغلی خالی است — نادیده گرفته شد')
                continue

            # یافتن فرصت شغلی
            job = JobOpportunity.objects.filter(code=code, is_deleted=False).first()
            if not job:
                stats['warnings'].append(f'ردیف {row_num}: فرصت شغلی با کد "{code}" در دیتابیس یافت نشد — رد شد')
                continue

            stats['processed'] += 1

            # استخراج تاریخ‌ها
            start_date = parse_shamsi_date_str(get_cell(row, 'start_reg'))
            end_date = parse_shamsi_date_str(get_cell(row, 'end_reg'))
            
            # استخراج تاریخ‌های مراحل
            stage_dates = {
                'SCREENING': parse_shamsi_date_str(get_cell(row, 'screening')),
                'EXAM': parse_shamsi_date_str(get_cell(row, 'exam')),
                'SKILL_TEST': parse_shamsi_date_str(get_cell(row, 'skill')),
                'INTERVIEW': parse_shamsi_date_str(get_cell(row, 'interview')),
                'ASSESSMENT': parse_shamsi_date_str(get_cell(row, 'assessment')),
            }

            # تعیین آخرین وضعیت بر اساس ستون «آخرین مرحله»
            new_status = None
            
            # بررسی اینکه آیا فرصت شغلی پذیرفته‌شده نهایی دارد یا خیر
            has_accepted = job.applications.filter(status='SELECTED', is_deleted=False).exists()
            
            if has_accepted:
                new_status = JobOpportunity.STATUS_CLOSED
            else:
                raw_status = get_cell(row, 'current_status')
                new_status = map_status(raw_status)

            # ۱. بروزرسانی تاریخ‌ها و وضعیت JobOpportunity
            job_changed = False
            if start_date and job.start_date != start_date:
                job.start_date = start_date
                job_changed = True
            if end_date and job.end_date != end_date:
                job.end_date = end_date
                job_changed = True
            if new_status and job.status != new_status:
                job.status = new_status
                job_changed = True

            if job_changed:
                if not dry_run:
                    job.save(update_fields=['start_date', 'end_date', 'status'])
                stats['jobs_updated'] += 1

            # ۲. بروزرسانی JobRecruitmentPlan
            plan = JobRecruitmentPlan.objects.filter(job=job, is_deleted=False).first()
            if not plan and not dry_run:
                # ایجاد پلان در صورت عدم وجود
                plan = JobRecruitmentPlan.objects.create(
                    job=job,
                    start_date=start_date or datetime.date.today(),
                    predicted_end_date=end_date or datetime.date.today(),
                    status='COMPLETED' if job.status in ('CLOSED', 'CANCELLED') else 'ACTIVE'
                )

            if plan and start_date and plan.start_date != start_date:
                if not dry_run:
                    plan.start_date = start_date
                    plan.save(update_fields=['start_date'])

            # ۳. بروزرسانی تاریخ مراحل JobStagePlan و ApplicationStageState
            all_valid_dates = []
            for stage_type, s_date in stage_dates.items():
                if not s_date:
                    continue
                all_valid_dates.append(s_date)

                # پیدا کردن مرحله ارزیابی شغل
                stage = job.stages.filter(stage_type=stage_type, is_deleted=False).first()
                if not stage:
                    continue

                # بروزرسانی یا ایجاد JobStagePlan
                if plan:
                    if not dry_run:
                        stage_plan, created = JobStagePlan.objects.update_or_create(
                            plan=plan,
                            stage=stage,
                            defaults={
                                'stage_type': stage_type,
                                'planned_start_date': s_date,
                                'planned_end_date': s_date,
                                'sla_days': 1,
                            }
                        )
                        if created or stage_plan.planned_start_date != s_date:
                            stats['stage_plans_updated'] += 1
                    else:
                        stats['stage_plans_updated'] += 1

                # بروزرسانی تاریخ ارزیابی در وضعیت‌های مراحل کاندیداها
                states_to_update = ApplicationStageState.objects.filter(
                    application__job=job,
                    stage=stage,
                    is_deleted=False
                )
                if states_to_update.exists():
                    cnt = states_to_update.count()
                    if not dry_run:
                        states_to_update.update(evaluation_date=s_date)
                    stats['candidate_states_updated'] += cnt

            # ۴. بروزرسانی predicted_end_date پلان جذب بر اساس آخرین تاریخ مراحل
            if plan and all_valid_dates:
                max_date = max(all_valid_dates)
                if plan.predicted_end_date != max_date:
                    if not dry_run:
                        plan.predicted_end_date = max_date
                        plan.save(update_fields=['predicted_end_date'])

        # در صورت dry-run تراکنش را rollback می‌کنیم
        if dry_run:
            transaction.set_rollback(True)

        # --- گزارش نهایی ---
        self.stdout.write('')
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(self.style.HTTP_INFO('  گزارش نهایی بروزرسانی تاریخ‌های شمسی'))
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(f'  تعداد ردیف‌های پردازش شده         : {stats["processed"]}')
        self.stdout.write(f'  تعداد فرصت‌های شغلی بروزرسانی شده  : {self.style.SUCCESS(str(stats["jobs_updated"]))}')
        self.stdout.write(f'  تعداد برنامه‌های مرحله بروزرسانی شده: {self.style.SUCCESS(str(stats["stage_plans_updated"]))}')
        self.stdout.write(f'  تعداد تاریخ کاندیداها بروزرسانی شده: {self.style.SUCCESS(str(stats["candidate_states_updated"]))}')
        self.stdout.write(f'  هشدارها                          : {self.style.WARNING(str(len(stats["warnings"])))}')
        self.stdout.write(f'  خطاهای سیستم                     : {self.style.ERROR(str(len(stats["errors"])))}')

        if stats['warnings']:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING('برخی هشدارها (حداکثر ۵۰ مورد):'))
            for w in stats['warnings'][:50]:
                self.stdout.write(f'  ⚠  {w}')
            if len(stats['warnings']) > 50:
                self.stdout.write(self.style.WARNING(f'  ... و {len(stats["warnings"])-50} هشدار دیگر.'))

        if stats['errors']:
            self.stdout.write('')
            self.stdout.write(self.style.ERROR('خطاهای سیستم:'))
            for e in stats['errors']:
                self.stdout.write(f'  ✗  {e}')

        self.stdout.write('')
        if dry_run:
            self.stdout.write(self.style.WARNING('[DRY-RUN] هیچ تغییری اعمال نشد. برای اجرای واقعی --dry-run را حذف کنید.'))
        else:
            self.stdout.write(self.style.SUCCESS('بروزرسانی تاریخ‌ها با موفقیت به پایان رسید.'))
        self.stdout.write('')
