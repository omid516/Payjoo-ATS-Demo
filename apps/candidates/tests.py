from django.test import TestCase
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.urls import reverse
from datetime import date
import jdatetime

from apps.accounts.models import UserProfile
from apps.jobs.models import WorkflowTemplate, WorkflowStageTemplate, JobOpportunity, JobOpportunityStage
from apps.candidates.models import Candidate, CandidateEducation, CandidateExperience, JobApplication, ApplicationStageState, CandidateSkill, CandidateLanguage
from apps.candidates.forms import CandidateForm, CandidateEducationFormSet, CandidateExperienceFormSet, JobApplicationForm

class CandidateModuleTests(TestCase):
    def setUp(self):
        # Create a recruiter user
        self.recruiter = User.objects.create_user(username='recruiter_user', password='password123')
        self.recruiter.profile.role = UserProfile.ROLE_RECRUITMENT_SPECIALIST
        self.recruiter.profile.save()

        # Create a WorkflowTemplate with default stages
        self.workflow = WorkflowTemplate.objects.create(
            name='فرآیند استخدام استاندارد',
            description='فرآیند شامل آزمون کتبی و مصاحبه فنی'
        )
        self.stage_template_1 = WorkflowStageTemplate.objects.create(
            workflow=self.workflow,
            name='آزمون کتبی',
            default_weight=40,
            sequence=1
        )
        self.stage_template_2 = WorkflowStageTemplate.objects.create(
            workflow=self.workflow,
            name='مصاحبه فنی',
            default_weight=60,
            sequence=2
        )

        # Create a JobOpportunity
        self.job = JobOpportunity.objects.create(
            request_number='REQ-1405-001',
            title='کارشناس هوش مصنوعی',
            code='AI-01',
            department='فناوری اطلاعات',
            assigned_recruiter=self.recruiter,
            workflow=self.workflow,
            status=JobOpportunity.STATUS_PUBLISHED,
            description='کارشناس هوش مصنوعی و یادگیری ماشین'
        )

    def test_candidate_creation_and_relations(self):
        """تست ایجاد موفقیت‌آمیز متقاضی همراه با سوابق تحصیلی و کاری"""
        candidate = Candidate.objects.create(
            first_name='امید',
            last_name='صالحی',
            email='omid@example.com',
            phone_number='09123456789',
            national_id='0012345678'
        )
        edu = CandidateEducation.objects.create(
            candidate=candidate,
            degree='MASTER',
            major='هوش مصنوعی',
            university='دانشگاه صنعتی شریف',
            gpa=18.5,
            graduation_year=1402
        )
        exp = CandidateExperience.objects.create(
            candidate=candidate,
            company='شرکت داده‌پرداز',
            job_title='توسعه‌دهنده پایتون',
            start_date=date(2023, 1, 1),
            end_date=date(2024, 1, 1),
            description='توسعه وب جنگو'
        )

        self.assertEqual(Candidate.objects.count(), 1)
        self.assertEqual(candidate.education.count(), 1)
        self.assertEqual(candidate.experience.count(), 1)
        self.assertEqual(edu.graduation_year, 1402)
        self.assertEqual(exp.job_title, 'توسعه‌دهنده پایتون')

    def test_stage_dates_and_intervals(self):
        """تست ویژگی‌های فاصله زمانی و تاریخ انجام ارزیابی بین مراحل"""
        candidate = Candidate.objects.create(
            first_name='فرید',
            last_name='فرهمند',
            email='farid@example.com',
            phone_number='09129999999',
            national_id='9999999999'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        
        # Verify stage states are created
        states = list(app.stage_states.all().order_by('stage__sequence'))
        self.assertEqual(len(states), 2)
        
        state1, state2 = states[0], states[1]
        
        # 1. Initially (PENDING), interval should be None
        self.assertIsNone(state1.prev_stage_state)
        self.assertIsNone(state2.days_since_prev_stage)
        
        # 2. Complete both stages with distinct dates
        state1.status = ApplicationStageState.STATUS_COMPLETED
        state1.evaluation_date = date(2026, 6, 1)
        state1.save()
        
        state2.status = ApplicationStageState.STATUS_COMPLETED
        state2.evaluation_date = date(2026, 6, 6) # 5 days later
        state2.save()
        
        # Refresh and verify
        state2.refresh_from_db()
        self.assertEqual(state2.prev_stage_state, state1)
        self.assertEqual(state2.days_since_prev_stage, 5)

    def test_national_id_validation_form(self):
        """تست اعتبارسنجی یکتا بودن و فرمت ۱۰ رقمی کد ملی متقاضی در فرم"""
        # Create an existing candidate
        Candidate.objects.create(
            first_name='علی',
            last_name='کریمی',
            email='ali@example.com',
            phone_number='09121111111',
            national_id='1111111111'
        )

        # 1. Invalid length/format
        form_invalid = CandidateForm(data={
            'first_name': 'رضا',
            'last_name': 'احمدی',
            'email': 'reza@example.com',
            'phone_number': '09122222222',
            'national_id': '123'  # Less than 10 digits
        })
        self.assertFalse(form_invalid.is_valid())
        self.assertIn('national_id', form_invalid.errors)

        # 2. Duplicate national ID
        form_duplicate = CandidateForm(data={
            'first_name': 'رضا',
            'last_name': 'احمدی',
            'email': 'reza@example.com',
            'phone_number': '09122222222',
            'national_id': '1111111111'  # Duplicate
        })
        self.assertFalse(form_duplicate.is_valid())
        self.assertIn('national_id', form_duplicate.errors)

        # 3. Valid national ID
        form_valid = CandidateForm(data={
            'first_name': 'رضا',
            'last_name': 'احمدی',
            'email': 'reza@example.com',
            'phone_number': '09122222222',
            'national_id': '2222222222'
        })
        self.assertTrue(form_valid.is_valid())

    def test_auto_creation_of_stage_states(self):
        """تست ایجاد خودکار مراحل ارزیابی در درخواست همکاری متقاضی بر اساس مراحل فعال شغل"""
        candidate = Candidate.objects.create(
            first_name='سارا',
            last_name='کمالی',
            email='sara@example.com',
            phone_number='09123333333',
            national_id='3333333333'
        )
        
        # Verify the Job has 2 stages copied from the workflow
        self.assertEqual(self.job.stages.count(), 2)

        # Create JobApplication
        application = JobApplication.objects.create(
            job=self.job,
            candidate=candidate
        )

        # ApplicationStageState objects should be created automatically for the 2 stages of the job
        stage_states = ApplicationStageState.objects.filter(application=application)
        self.assertEqual(stage_states.count(), 2)
        
        # Ensure status is PENDING and score is 0.0 initially
        for state in stage_states:
            self.assertEqual(state.status, ApplicationStageState.STATUS_PENDING)
            self.assertEqual(state.score, 0.0)

        # Verify current stage is set to the first stage (sequence=1)
        first_stage = self.job.stages.filter(is_deleted=False).order_by('sequence').first()
        self.assertEqual(application.current_stage, first_stage)

    def test_duplicate_application_prevention(self):
        """تست جلوگیری از ثبت درخواست تکراری برای یک متقاضی و یک فرصت شغلی خاص"""
        candidate = Candidate.objects.create(
            first_name='نرگس',
            last_name='حسینی',
            email='narges@example.com',
            phone_number='09124444444',
            national_id='4444444444'
        )

        # First Application: Valid
        JobApplication.objects.create(
            job=self.job,
            candidate=candidate
        )

        # Second Application (Direct create): should raise exception/validation error due to unique_together
        from django.db import transaction, IntegrityError
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                JobApplication.objects.create(
                    job=self.job,
                    candidate=candidate
                )

        # Form test for prevention
        form = JobApplicationForm(data={'job': self.job.id})
        # Simulate view's validation check
        is_duplicate = JobApplication.objects.filter(job=self.job, candidate=candidate, is_deleted=False).exists()
        self.assertTrue(is_duplicate)

    def test_soft_delete_candidate_and_cascades(self):
        """تست حذف نرم متقاضی و بررسی حذف نرم سوابق تحصیلی، کاری و درخواست‌های او"""
        candidate = Candidate.objects.create(
            first_name='مهرداد',
            last_name='رضایی',
            email='mehrdad@example.com',
            phone_number='09125555555',
            national_id='5555555555'
        )
        edu = CandidateEducation.objects.create(
            candidate=candidate,
            degree='BACHELOR',
            major='فیزیک',
            university='دانشگاه تهران',
            gpa=15.0,
            graduation_year=1398
        )
        exp = CandidateExperience.objects.create(
            candidate=candidate,
            company='مدارس انرژی اتمی',
            job_title='دبیر فیزیک',
            start_date=date(2020, 9, 1),
            description='تدریس فیزیک کنکور'
        )
        app = JobApplication.objects.create(
            job=self.job,
            candidate=candidate
        )

        candidate_pk = candidate.pk
        edu_pk = edu.pk
        exp_pk = exp.pk
        app_pk = app.pk

        # Simulate Cascade Delete from Delete View
        candidate.delete()
        candidate.education.all().delete()
        candidate.experience.all().delete()
        candidate.applications.all().delete()

        # Check soft delete statuses
        self.assertTrue(Candidate.all_objects.get(pk=candidate_pk).is_deleted)
        self.assertTrue(CandidateEducation.all_objects.get(pk=edu_pk).is_deleted)
        self.assertTrue(CandidateExperience.all_objects.get(pk=exp_pk).is_deleted)
        self.assertTrue(JobApplication.all_objects.get(pk=app_pk).is_deleted)

        # Check default manager excludes them
        self.assertFalse(Candidate.objects.filter(pk=candidate_pk).exists())
        self.assertFalse(CandidateEducation.objects.filter(pk=edu_pk).exists())
        self.assertFalse(CandidateExperience.objects.filter(pk=exp_pk).exists())
        self.assertFalse(JobApplication.objects.filter(pk=app_pk).exists())

    def test_experience_form_jalali_date_conversion(self):
        """تست تبدیل خودکار تاریخ شمسی سوابق کاری در فرم به میلادی جهت ثبت در دیتابیس"""
        form_data = {
            'company': 'سازمان فناوری اطلاعات',
            'job_title': 'کارشناس شبکه',
            'start_date': '1400/01/15',  # Shamsi
            'end_date': '1401/12/29',    # Shamsi
            'description': 'پشتیبانی شبکه',
        }
        form = CandidateExperienceFormSet(data={
            'experience-TOTAL_FORMS': '1',
            'experience-INITIAL_FORMS': '0',
            'experience-MIN_NUM_FORMS': '0',
            'experience-MAX_NUM_FORMS': '1000',
            'experience-0-company': form_data['company'],
            'experience-0-job_title': form_data['job_title'],
            'experience-0-start_date': form_data['start_date'],
            'experience-0-end_date': form_data['end_date'],
            'experience-0-description': form_data['description'],
        })
        self.assertTrue(form.is_valid(), form.errors)
        
        # Save experience with a dummy candidate
        candidate = Candidate.objects.create(
            first_name='حمید',
            last_name='عباسی',
            email='hamid@example.com',
            phone_number='09126666666',
            national_id='6666666666'
        )
        form.instance = candidate
        instances = form.save()
        
        self.assertEqual(len(instances), 1)
        exp = instances[0]
        # 1400/01/15 is 2021-04-04 gregorian
        self.assertEqual(exp.start_date, date(2021, 4, 4))
        # 1401/12/29 is 2023-03-20 gregorian
        self.assertEqual(exp.end_date, date(2023, 3, 20))

    def test_personnel_number_validation_and_uniqueness(self):
        """تست اعتبارسنجی و یکتا بودن شماره پرسنلی در فرم کاندیدا"""
        # Create a candidate with a personnel number
        Candidate.objects.create(
            first_name='مهدی',
            last_name='صالحی',
            email='mehdi@example.com',
            phone_number='09121112222',
            national_id='1111222233',
            personnel_number='P-1002'
        )

        # Attempt to create another candidate with the same personnel number
        form = CandidateForm(data={
            'first_name': 'حسین',
            'last_name': 'علوی',
            'email': 'hossein@example.com',
            'phone_number': '09122223333',
            'national_id': '2222333344',
            'personnel_number': 'P-1002'  # Duplicate
        })
        self.assertFalse(form.is_valid())
        self.assertIn('personnel_number', form.errors)

        # Form with different personnel number should be valid
        form_valid = CandidateForm(data={
            'first_name': 'حسین',
            'last_name': 'علوی',
            'email': 'hossein@example.com',
            'phone_number': '09122223333',
            'national_id': '2222333344',
            'personnel_number': 'P-1003'
        })
        self.assertTrue(form_valid.is_valid(), form_valid.errors)

    def test_weighted_final_score_calculation(self):
        """تست محاسبه خودکار امتیاز نهایی وزنی درخواست همکاری پس از ثبت نمره برای هر مرحله"""
        candidate = Candidate.objects.create(
            first_name='نیلوفر',
            last_name='رحیمی',
            email='niloofar@example.com',
            phone_number='09127778888',
            national_id='7777888899'
        )
        # Create JobApplication (which auto-creates ApplicationStageState records for the 2 stages)
        app = JobApplication.objects.create(
            job=self.job,
            candidate=candidate
        )
        
        # Verify initial final score is 0.0
        self.assertEqual(app.final_score, 0.0)

        # Retrieve the automatically created stage states
        states = app.stage_states.all().order_by('stage__sequence')
        self.assertEqual(states.count(), 2)

        # Update first stage (Weight 40%) score to 80
        state1 = states[0]
        state1.score = 80.0
        state1.status = ApplicationStageState.STATUS_COMPLETED
        state1.save()

        # Reload application from db and assert final_score is 80 * 0.40 = 32.0
        app.refresh_from_db()
        self.assertEqual(app.final_score, 32.0)

        # Update second stage (Weight 60%) score to 90
        state2 = states[1]
        state2.score = 90.0
        state2.status = ApplicationStageState.STATUS_COMPLETED
        state2.save()

        # Reload application and assert final_score is 80 * 0.40 + 90 * 0.60 = 86.0
        app.refresh_from_db()
        self.assertEqual(app.final_score, 86.0)

    def test_careers_list_public_access(self):
        """تست دسترسی عمومی به لیست مشاغل منتشر شده و دسته‌بندی دپارتمانی"""
        # Create a non-published job to verify it's excluded
        JobOpportunity.objects.create(
            request_number='REQ-1405-PLAN',
            title='کارشناس برنامه‌ریزی',
            code='PLAN-01',
            department='برنامه‌ریزی',
            status=JobOpportunity.STATUS_PLANNING
        )

        response = self.client.get(reverse('careers_list'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'کارشناس هوش مصنوعی')
        self.assertNotContains(response, 'کارشناس برنامه‌ریزی')

    def test_careers_apply_and_candidate_creation(self):
        """تست ثبت‌نام آنلاین متقاضی جدید برای فرصت شغلی"""
        url = reverse('careers_apply', kwargs={'pk': self.job.id})
        post_data = {
            'first_name': 'رضا',
            'last_name': 'احمدی',
            'email': 'reza.ah@example.com',
            'phone_number': '09129999999',
            'national_id': '9999999999',
            'personnel_number': ''
        }
        
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302) # Redirect to success

        # Verify candidate and application are created
        candidate = Candidate.objects.get(national_id='9999999999')
        self.assertEqual(candidate.first_name, 'رضا')
        self.assertTrue(JobApplication.objects.filter(job=self.job, candidate=candidate).exists())

    def test_careers_apply_existing_candidate(self):
        """تست اینکه ثبت‌نام مجدد متقاضی موجود، کاندیدای جدید ایجاد نمی‌کند بلکه اطلاعات را به‌روز و درخواست ثبت می‌کند"""
        # Pre-create candidate
        candidate = Candidate.objects.create(
            first_name='زهرا',
            last_name='کریمی',
            email='zahra.k@example.com',
            phone_number='09128888888',
            national_id='8888888888'
        )

        url = reverse('careers_apply', kwargs={'pk': self.job.id})
        post_data = {
            'first_name': 'زهرا',
            'last_name': 'کریمی',
            'email': 'zahra.new@example.com', # Updated email
            'phone_number': '09128888888',
            'national_id': '8888888888',
            'personnel_number': 'P-9900' # Added personnel number
        }

        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)

        # Assert candidate count remains same but details are updated
        self.assertEqual(Candidate.objects.filter(national_id='8888888888').count(), 1)
        candidate.refresh_from_db()
        self.assertEqual(candidate.email, 'zahra.new@example.com')
        self.assertEqual(candidate.personnel_number, 'P-9900')
        
        # Verify job application is linked
        self.assertTrue(JobApplication.objects.filter(job=self.job, candidate=candidate).exists())

    def test_careers_track_success_and_failure(self):
        """تست سیستم پیگیری لاگ درخواست‌های متقاضی با کدملی و شماره تماس"""
        # Pre-create candidate and application
        candidate = Candidate.objects.create(
            first_name='حامد',
            last_name='نوری',
            email='hamed@example.com',
            phone_number='09125555555',
            national_id='5555566666'
        )
        JobApplication.objects.create(job=self.job, candidate=candidate)

        url = reverse('careers_track')

        # 1. Invalid tracking data
        response_fail = self.client.post(url, {'national_id': '5555566666', 'phone_number': '09120000000'})
        self.assertEqual(response_fail.status_code, 200)
        self.assertContains(response_fail, 'یافت نشد')

        # 2. Valid tracking data
        response_success = self.client.post(url, {'national_id': '5555566666', 'phone_number': '09125555555'})
        self.assertEqual(response_success.status_code, 200)
        self.assertContains(response_success, 'حامد نوری')
        self.assertContains(response_success, 'کارشناس هوش مصنوعی')

    def test_candidate_self_registration(self):
        """تست ثبت‌نام خودکار متقاضی"""
        url = reverse('candidate_signup')
        post_data = {
            'first_name': 'کامران',
            'last_name': 'راد',
            'email': 'kamran@example.com',
            'phone_number': '09121112233',
            'national_id': '1111222233',
            'personnel_number': '',
            'password': 'password123',
            'password_confirm': 'password123',
            
            # Formset management parameters
            'edu-TOTAL_FORMS': '0',
            'edu-INITIAL_FORMS': '0',
            'edu-MIN_NUM_FORMS': '0',
            'edu-MAX_NUM_FORMS': '1000',
            
            'exp-TOTAL_FORMS': '0',
            'exp-INITIAL_FORMS': '0',
            'exp-MIN_NUM_FORMS': '0',
            'exp-MAX_NUM_FORMS': '1000',
            
            'lang-TOTAL_FORMS': '0',
            'lang-INITIAL_FORMS': '0',
            'lang-MIN_NUM_FORMS': '0',
            'lang-MAX_NUM_FORMS': '1000',
            
            'skill-TOTAL_FORMS': '0',
            'skill-INITIAL_FORMS': '0',
            'skill-MIN_NUM_FORMS': '0',
            'skill-MAX_NUM_FORMS': '1000',
            
            'cert-TOTAL_FORMS': '0',
            'cert-INITIAL_FORMS': '0',
            'cert-MIN_NUM_FORMS': '0',
            'cert-MAX_NUM_FORMS': '1000',
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302) # Redirects to dashboard
        
        # Verify Django user created
        user_exists = User.objects.filter(username='1111222233').exists()
        self.assertTrue(user_exists)
        user = User.objects.get(username='1111222233')
        
        # Verify UserProfile role is CANDIDATE
        self.assertEqual(user.profile.role, UserProfile.ROLE_CANDIDATE)
        
        # Verify Candidate profile linked
        candidate = Candidate.objects.get(national_id='1111222233')
        self.assertEqual(candidate.user, user)
        self.assertEqual(candidate.first_name, 'کامران')

    def test_candidate_self_registration_restores_soft_deleted(self):
        """تست بازیابی کاندیدای حذف نرم شده در هنگام ثبت‌نام مجدد"""
        # Pre-create a soft-deleted candidate
        deleted_candidate = Candidate.objects.create(
            first_name='حمید',
            last_name='رضایی',
            email='hamid.r@example.com',
            phone_number='09127776666',
            national_id='7777666655',
            personnel_number='P-88220'
        )
        deleted_candidate.delete()  # soft delete it
        
        self.assertTrue(Candidate.all_objects.filter(national_id='7777666655').first().is_deleted)

        url = reverse('candidate_signup')
        post_data = {
            'first_name': 'حمید',
            'last_name': 'رضایی',
            'email': 'hamid.r@example.com',
            'phone_number': '09127776666',
            'national_id': '7777666655',
            'personnel_number': 'P-88220',
            'password': 'password123',
            'password_confirm': 'password123',
            'edu-TOTAL_FORMS': '0',
            'edu-INITIAL_FORMS': '0',
            'edu-MIN_NUM_FORMS': '0',
            'edu-MAX_NUM_FORMS': '1000',
            'exp-TOTAL_FORMS': '0',
            'exp-INITIAL_FORMS': '0',
            'exp-MIN_NUM_FORMS': '0',
            'exp-MAX_NUM_FORMS': '1000',
            'lang-TOTAL_FORMS': '0',
            'lang-INITIAL_FORMS': '0',
            'lang-MIN_NUM_FORMS': '0',
            'lang-MAX_NUM_FORMS': '1000',
            'skill-TOTAL_FORMS': '0',
            'skill-INITIAL_FORMS': '0',
            'skill-MIN_NUM_FORMS': '0',
            'skill-MAX_NUM_FORMS': '1000',
            'cert-TOTAL_FORMS': '0',
            'cert-INITIAL_FORMS': '0',
            'cert-MIN_NUM_FORMS': '0',
            'cert-MAX_NUM_FORMS': '1000',
        }
        
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)
        
        # Verify candidate is restored
        candidate = Candidate.objects.get(national_id='7777666655')
        self.assertFalse(candidate.is_deleted)
        self.assertEqual(candidate.personnel_number, 'P-88220')

    def test_reapply_restores_soft_deleted_application(self):
        """تست احیای درخواست همکاری حذف نرم شده در هنگام ثبت‌نام/درخواست مجدد"""
        candidate = Candidate.objects.create(
            first_name='سارا',
            last_name='احمدی',
            email='sara.a@example.com',
            phone_number='09121234567',
            national_id='1234512345'
        )
        # Create application and soft delete it
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        app.delete()
        
        self.assertTrue(JobApplication.all_objects.filter(job=self.job, candidate=candidate).first().is_deleted)

        # Re-apply via direct apply view
        candidate_user = User.objects.create_user(username='1234512345', password='password123')
        candidate_user.profile.role = UserProfile.ROLE_CANDIDATE
        candidate_user.profile.save()
        candidate.user = candidate_user
        candidate.save()

        self.client.login(username='1234512345', password='password123')
        url = reverse('candidate_apply_direct', kwargs={'pk': self.job.id})
        response = self.client.post(url)
        self.assertEqual(response.status_code, 302)

        # Verify application is restored and active
        restored_app = JobApplication.objects.get(job=self.job, candidate=candidate)
        self.assertFalse(restored_app.is_deleted)

    def test_candidate_dashboard_access_control(self):
        """تست کنترل دسترسی تفکیک‌شده داشبورد متقاضی و کارشناسان"""
        # Create candidate user
        candidate_user = User.objects.create_user(username='2222333344', password='password123')
        candidate_user.profile.role = UserProfile.ROLE_CANDIDATE
        candidate_user.profile.save()
        
        # Create recruiter user
        recruiter_user = User.objects.create_user(username='recruiter_test', password='password123')
        recruiter_user.profile.role = UserProfile.ROLE_RECRUITMENT_SPECIALIST
        recruiter_user.profile.save()

        # 1. Candidate tries to access recruiter dashboard -> should redirect
        self.client.login(username='2222333344', password='password123')
        response = self.client.get(reverse('dashboard'))
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('candidate_dashboard'))
        
        # 2. Candidate accesses candidate dashboard -> 200
        response = self.client.get(reverse('candidate_dashboard'))
        self.assertEqual(response.status_code, 200)
        self.client.logout()

        # 3. Recruiter tries to access candidate dashboard -> should redirect to recruiter dashboard
        self.client.login(username='recruiter_test', password='password123')
        response = self.client.get(reverse('candidate_dashboard'))
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('dashboard'))

    def test_candidate_dashboard_apply(self):
        """تست ثبت درخواست با یک کلیک از داخل پنل متقاضی"""
        candidate_user = User.objects.create_user(username='3333444455', password='password123')
        candidate_user.profile.role = UserProfile.ROLE_CANDIDATE
        candidate_user.profile.save()
        
        candidate = Candidate.objects.create(
            user=candidate_user,
            first_name='پیمان',
            last_name='احمدی',
            email='peyman@example.com',
            phone_number='09124445555',
            national_id='3333444455'
        )

        self.client.login(username='3333444455', password='password123')
        url = reverse('candidate_apply_direct', kwargs={'pk': self.job.id})
        
        response = self.client.post(url)
        self.assertEqual(response.status_code, 302) # Redirects to dashboard
        
        # Verify JobApplication created
        self.assertTrue(JobApplication.objects.filter(job=self.job, candidate=candidate).exists())

    def test_candidate_profile_attributes_update(self):
        """تست ثبت زبان‌ها، مهارت‌ها و دوره‌ها/مدرک‌ها از طریق فرم‌ست‌ها در پنل متقاضی"""
        candidate_user = User.objects.create_user(username='4444555566', password='password123')
        candidate_user.profile.role = UserProfile.ROLE_CANDIDATE
        candidate_user.profile.save()
        
        candidate = Candidate.objects.create(
            user=candidate_user,
            first_name='احسان',
            last_name='مرادی',
            email='ehsan@example.com',
            phone_number='09125556666',
            national_id='4444555566'
        )
        self.client.login(username='4444555566', password='password123')

        # 1. Test language formset submit
        lang_url = reverse('candidate_language_update')
        lang_data = {
            'lang-TOTAL_FORMS': '1',
            'lang-INITIAL_FORMS': '0',
            'lang-MIN_NUM_FORMS': '0',
            'lang-MAX_NUM_FORMS': '1000',
            'lang-0-name': 'انگلیسی',
            'lang-0-proficiency': 'ADVANCED',
        }
        response = self.client.post(lang_url, lang_data)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(candidate.languages.count(), 1)
        self.assertEqual(candidate.languages.first().name, 'انگلیسی')

        # 2. Test skill formset submit
        skill_url = reverse('candidate_skill_update')
        skill_data = {
            'skill-TOTAL_FORMS': '1',
            'skill-INITIAL_FORMS': '0',
            'skill-MIN_NUM_FORMS': '0',
            'skill-MAX_NUM_FORMS': '1000',
            'skill-0-name': 'Python',
            'skill-0-level': 'EXPERT',
        }
        response = self.client.post(skill_url, skill_data)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(candidate.skills.count(), 1)
        self.assertEqual(candidate.skills.first().name, 'Python')

        # 3. Test certificate formset submit & Jalali conversion
        cert_url = reverse('candidate_certificate_update')
        cert_data = {
            'cert-TOTAL_FORMS': '1',
            'cert-INITIAL_FORMS': '0',
            'cert-MIN_NUM_FORMS': '0',
            'cert-MAX_NUM_FORMS': '1000',
            'cert-0-name': 'دوره تفکر الگوریتمی',
            'cert-0-issuer': 'رهنمون',
            'cert-0-issue_date': '1400/01/15',
            'cert-0-expiration_date': '1402/12/29',
        }
        response = self.client.post(cert_url, cert_data)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(candidate.certificates.count(), 1)
        
        cert = candidate.certificates.first()
        self.assertEqual(cert.name, 'دوره تفکر الگوریتمی')
        self.assertEqual(cert.issue_date, date(2021, 4, 4))
        self.assertEqual(cert.expiration_date, date(2024, 3, 19))  # 1402/12/29 is 2024-03-19

    def test_one_click_apply_authenticated(self):
        """تست دکمه ثبت درخواست مستقیم/یک کلیکی برای متقاضی وارد شده در صفحه عمومی جزئیات آگهی"""
        candidate_user = User.objects.create_user(username='5555666677', password='password123')
        candidate_user.profile.role = UserProfile.ROLE_CANDIDATE
        candidate_user.profile.save()
        
        candidate = Candidate.objects.create(
            user=candidate_user,
            first_name='مهسا',
            last_name='کریمی',
            email='mahsa@example.com',
            phone_number='09126667777',
            national_id='5555666677'
        )

        url = reverse('careers_apply', kwargs={'pk': self.job.id})

        # 1. Anonymous user sees registration form
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'ثبت‌نام و ارسال رزومه')
        self.assertNotContains(response, 'ارسال درخواست همکاری سریع')

        # 2. Authenticated user sees One-click Apply button
        self.client.login(username='5555666677', password='password123')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'ارسال درخواست همکاری سریع')
        self.assertNotContains(response, 'ثبت‌نام و ارسال رزومه')
        
        # Try direct apply via CareersDetailAndApplyView's candidate_apply_direct
        apply_url = reverse('candidate_apply_direct', kwargs={'pk': self.job.id})
        response = self.client.post(apply_url)
        self.assertEqual(response.status_code, 302)
        
        # Verify application registered
        self.assertTrue(JobApplication.objects.filter(job=self.job, candidate=candidate).exists())

    def test_stage_progression_gating(self):
        """تست کنترل پیشرفت مراحل، قفل بودن مراحل بعدی و باز شدن آن‌ها پس از تایید مرحله قبل"""
        candidate = Candidate.objects.create(
            first_name='زهرا',
            last_name='موسوی',
            email='zahra@example.com',
            phone_number='09129999999',
            national_id='9999999999'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        stages = list(self.job.stages.filter(is_deleted=False).order_by('sequence'))
        
        state1 = app.stage_states.get(stage=stages[0])
        state2 = app.stage_states.get(stage=stages[1])
        
        # Stage 1 should be accessible, Stage 2 should not be accessible
        self.assertTrue(state1.is_accessible)
        self.assertFalse(state2.is_accessible)
        
        # Attempting to edit/update Stage 2 via API should fail
        self.client.login(username='recruiter_user', password='password123')
        edit_url = reverse('edit_stage_state', kwargs={'pk': state2.pk})
        response = self.client.get(edit_url)
        self.assertEqual(response.status_code, 400)
        
        # Complete Stage 1
        state1.status = ApplicationStageState.STATUS_COMPLETED
        state1.score = 85.0
        state1.save()
        
        # Verify candidate automatically advanced to Stage 2
        app.refresh_from_db()
        self.assertEqual(app.current_stage, stages[1])
        
        # Stage 2 should now be accessible
        state2.refresh_from_db()
        self.assertTrue(state2.is_accessible)

    def test_score_entry_view(self):
        """تست صفحه ورود نمرات آزمون و فیلتر کردن متقاضیان فعال بر اساس فرصت شغلی و فیلترهای پیشرفته"""
        candidate1 = Candidate.objects.create(
            first_name='فاطمه',
            last_name='تقوی',
            email='fatemeh.t@example.com',
            phone_number='09122222222',
            national_id='2222222222'
        )
        app1 = JobApplication.objects.create(job=self.job, candidate=candidate1)
        stages = list(self.job.stages.filter(is_deleted=False).order_by('sequence'))
        state1 = app1.stage_states.get(stage=stages[0])

        candidate2 = Candidate.objects.create(
            first_name='محمد',
            last_name='امیری',
            email='mohammad.a@example.com',
            phone_number='09123333333',
            national_id='3333333333'
        )
        app2 = JobApplication.objects.create(job=self.job, candidate=candidate2)
        state2 = app2.stage_states.get(stage=stages[0])
        # Mark state2 as COMPLETED
        state2.status = ApplicationStageState.STATUS_COMPLETED
        state2.score = 90.0
        state2.save()

        self.client.login(username='recruiter_user', password='password123')
        url = reverse('candidate_score_entry')
        
        # 1. Access page without job_id
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'فرصت شغلی انتخاب نشده است')
        
        # 2. Access page with job_id and stage_id (default status is PENDING, so c1 is visible, c2 is hidden)
        response = self.client.get(f"{url}?job_id={self.job.id}&stage_id={stages[0].id}")
        self.assertEqual(response.status_code, 200)
        self.assertIn(state1, response.context['pending_states'])
        self.assertNotIn(state2, response.context['pending_states'])

        # 3. Filter by COMPLETED_FAILED status (c1 hidden, c2 visible)
        response = self.client.get(f"{url}?job_id={self.job.id}&stage_id={stages[0].id}&eval_status=COMPLETED_FAILED")
        self.assertEqual(response.status_code, 200)
        self.assertNotIn(state1, response.context['pending_states'])
        self.assertIn(state2, response.context['pending_states'])

        # 4. Filter by ALL status (both visible)
        response = self.client.get(f"{url}?job_id={self.job.id}&stage_id={stages[0].id}&eval_status=ALL")
        self.assertEqual(response.status_code, 200)
        self.assertIn(state1, response.context['pending_states'])
        self.assertIn(state2, response.context['pending_states'])

        # 5. Search query filtering
        response = self.client.get(f"{url}?job_id={self.job.id}&stage_id={stages[0].id}&eval_status=ALL&q=تقوی")
        self.assertEqual(response.status_code, 200)
        self.assertIn(state1, response.context['pending_states'])
        self.assertNotIn(state2, response.context['pending_states'])

    def test_bulk_score_submission(self):
        """تست ثبت فله‌ای نمرات و وضعیت‌های ارزیابی متقاضیان"""
        self.client.login(username='recruiter_user', password='password123')
        candidate = Candidate.objects.create(
            first_name='سارا',
            last_name='احمدی',
            email='sara@example.com',
            phone_number='09121111111',
            national_id='1111111111'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        state = app.stage_states.filter(stage=self.job.stages.first()).first()
        
        post_data = {
            'job_id': self.job.id,
            f'score_{state.id}': '92.5',
            f'status_{state.id}': 'COMPLETED',
            f'notes_{state.id}': 'عالی در آزمون کتبی',
        }
        url = reverse('candidate_score_entry')
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200)
        
        # Verify changes saved
        state.refresh_from_db()
        self.assertEqual(state.score, 92.5)
        self.assertEqual(state.status, 'COMPLETED')
        self.assertEqual(state.notes, 'عالی در آزمون کتبی')

    def test_interviewer_weighted_scoring(self):
        """تست اختصاص مصاحبه‌گران و ارزیابی با محاسبه میانگین وزنی"""
        from apps.jobs.models import JobStageInterviewer
        from apps.candidates.models import InterviewerScore
        
        # Create second interviewer
        interviewer_a = User.objects.create_user(username='int_a', password='password123')
        interviewer_a.profile.role = UserProfile.ROLE_INTERVIEWER
        interviewer_a.profile.save()
        
        interviewer_b = User.objects.create_user(username='int_b', password='password123')
        interviewer_b.profile.role = UserProfile.ROLE_INTERVIEWER
        interviewer_b.profile.save()
        
        stage1 = self.job.stages.first()
        
        # Assign interviewers to stage 1: A (weight 60%), B (weight 40%)
        JobStageInterviewer.objects.create(job=self.job, stage=stage1, user=interviewer_a, weight=60)
        JobStageInterviewer.objects.create(job=self.job, stage=stage1, user=interviewer_b, weight=40)
        
        candidate = Candidate.objects.create(
            first_name='رضا',
            last_name='کریمی',
            email='reza@example.com',
            phone_number='09122222222',
            national_id='2222222222'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        state = app.stage_states.filter(stage=stage1).first()
        
        # Submit interviewer A score: 80
        InterviewerScore.objects.create(stage_state=state, interviewer=interviewer_a, score=80.0, status='COMPLETED')
        state.refresh_from_db()
        # Since only one has submitted, the average weight is just A's score (80.0)
        self.assertEqual(state.score, 80.0)
        
        # Submit interviewer B score: 90
        InterviewerScore.objects.create(stage_state=state, interviewer=interviewer_b, score=90.0, status='COMPLETED')
        state.refresh_from_db()
        # Weighted average: (80 * 60 + 90 * 40) / 100 = 84.0
        self.assertEqual(state.score, 84.0)
        # Since both completed, status is COMPLETED
        self.assertEqual(state.status, 'COMPLETED')

    def test_passing_score_and_discrepancy_alert(self):
        """تست حد نصاب قبولی و هشدار اختلاف فاحش نمرات مصاحبه‌گران"""
        from apps.jobs.models import JobStageInterviewer
        from apps.candidates.models import InterviewerScore

        interviewer_a = User.objects.create_user(username='int_c', password='password123')
        interviewer_b = User.objects.create_user(username='int_d', password='password123')
        
        stage1 = self.job.stages.first()
        stage1.passing_score = 75.0
        stage1.save()
        
        JobStageInterviewer.objects.create(job=self.job, stage=stage1, user=interviewer_a, weight=50)
        JobStageInterviewer.objects.create(job=self.job, stage=stage1, user=interviewer_b, weight=50)
        
        candidate = Candidate.objects.create(
            first_name='مهدی', last_name='کریمی', email='mehdi@example.com', phone_number='09121234567', national_id='7777777777'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        state = app.stage_states.filter(stage=stage1).first()
        
        # Scenario 1: Total score 70.0 (below 75.0 passing_score) -> FAILED status
        iscore_a = InterviewerScore.objects.create(stage_state=state, interviewer=interviewer_a, score=70.0, status='COMPLETED')
        iscore_b = InterviewerScore.objects.create(stage_state=state, interviewer=interviewer_b, score=70.0, status='COMPLETED')
        state.refresh_from_db()
        self.assertEqual(state.score, 70.0)
        self.assertEqual(state.status, 'FAILED')
        self.assertFalse(state.score_discrepancy_alert)
        
        # Scenario 2: High discrepancy (90 and 60 -> diff 30 >= 20) -> Alert = True
        iscore_a.score = 90.0
        iscore_a.save()
        iscore_b.score = 60.0
        iscore_b.save()
        state.refresh_from_db()
        # Average: 75.0 (>= 75.0 passing_score) -> COMPLETED status
        self.assertEqual(state.score, 75.0)
        self.assertEqual(state.status, 'COMPLETED')
        self.assertTrue(state.score_discrepancy_alert)

    def test_assessment_center_weighted_scores(self):
        """تست اختصاص شایستگی‌ها به کانون ارزیابی و محاسبه خودکار میانگین وزنی"""
        from apps.jobs.models import AssessmentCompetency
        from apps.candidates.models import InterviewerScore, AssessorCompetencyScore

        # Create assessor user
        assessor = User.objects.create_user(username='assessor_user', password='password123')
        assessor.profile.role = UserProfile.ROLE_EXTERNAL_ASSESSOR
        assessor.profile.save()

        stage1 = self.job.stages.first()
        stage1.name = 'کانون ارزیابی'
        stage1.save()
        
        # Configure competencies on stage 1
        comp_a = AssessmentCompetency.objects.create(stage=stage1, name='کار تیمی', weight=60)
        comp_b = AssessmentCompetency.objects.create(stage=stage1, name='حل مسئله', weight=40)

        # Assign assessor to stage
        from apps.jobs.models import JobStageInterviewer
        JobStageInterviewer.objects.create(job=self.job, stage=stage1, user=assessor, weight=100)

        candidate = Candidate.objects.create(
            first_name='فاطمه', last_name='حسینی', email='fatemeh@example.com', phone_number='09123334444', national_id='8888888888'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        state = app.stage_states.filter(stage=stage1).first()

        self.client.login(username='assessor_user', password='password123')
        
        # Get assessment sheet
        response = self.client.get(reverse('candidate_assessment_sheet', kwargs={'pk': state.id}))
        self.assertEqual(response.status_code, 200)

        # Post competency scores
        post_data = {
            f'comp_score_{comp_a.id}': '80.0',
            f'comp_notes_{comp_a.id}': 'کار تیمی عالی',
            f'comp_score_{comp_b.id}': '90.0',
            f'comp_notes_{comp_b.id}': 'حل مسئله خوب',
            'notes': 'ارزیابی کلی عالی',
        }
        
        response = self.client.post(reverse('candidate_assessment_sheet', kwargs={'pk': state.id}), post_data)
        self.assertEqual(response.status_code, 200)

        # Retrieve InterviewerScore and AssessorCompetencyScore from DB
        iscore = InterviewerScore.objects.filter(stage_state=state, interviewer=assessor).first()
        self.assertIsNotNone(iscore)
        # Check competency scores
        cs_a = AssessorCompetencyScore.objects.get(interviewer_score=iscore, competency=comp_a)
        cs_b = AssessorCompetencyScore.objects.get(interviewer_score=iscore, competency=comp_b)
        self.assertEqual(cs_a.score, 80.0)
        self.assertEqual(cs_b.score, 90.0)

        # Check weighted average: (80 * 60 + 90 * 40) / 100 = 84.0
        self.assertEqual(iscore.score, 84.0)

        # Test reports list view
        report_list_url = reverse('assessment_center_report')
        response = self.client.get(report_list_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'فاطمه حسینی')
        self.assertContains(response, 'کانون ارزیابی')

        # Test detailed report view
        detail_report_url = reverse('assessment_center_detail_report', kwargs={'pk': state.id})
        response = self.client.get(detail_report_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'کارنامه نهایی کانون ارزیابی')
        self.assertContains(response, 'فاطمه حسینی')
        self.assertContains(response, '84')

    def test_readd_deleted_competency(self):
        """تست بازگردانی و ثبت مجدد شایستگی حذف شده بدون بروز خطای یکتایی"""
        from apps.jobs.models import AssessmentCompetency
        
        stage = self.job.stages.first()
        comp = AssessmentCompetency.objects.create(stage=stage, name='تفکر سیستمی', weight=50)
        
        # Soft delete it
        comp.delete()
        self.assertTrue(AssessmentCompetency.all_objects.filter(pk=comp.pk).first().is_deleted)
        
        # Try to add it back via the manage view POST
        self.client.login(username='recruiter_user', password='password123')
        url = reverse('manage_stage_interviewers', kwargs={'stage_id': stage.id})
        post_data = {
            'action': 'add_competency',
            'competency_name': 'تفکر سیستمی',
            'competency_weight': '60'
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200)
        
        # Verify it is active again with updated weight
        comp.refresh_from_db()
        self.assertFalse(comp.is_deleted)
        self.assertEqual(comp.weight, 60)

    def test_final_ranking_dashboard(self):
        """تست داشبورد رتبه‌بندی نهایی و مرتب‌سازی متقاضیان"""
        candidate1 = Candidate.objects.create(first_name='علی', last_name='امینی', email='ali@example.com', phone_number='09121111111', national_id='1111111111')
        candidate2 = Candidate.objects.create(first_name='رضا', last_name='کریمی', email='reza@example.com', phone_number='09122222222', national_id='2222222222')
        app1 = JobApplication.objects.create(job=self.job, candidate=candidate1, final_score=90.0)
        app2 = JobApplication.objects.create(job=self.job, candidate=candidate2, final_score=75.5)

        self.client.login(username='recruiter_user', password='password123')
        url = reverse('job_final_ranking', kwargs={'pk': self.job.id})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        
        # Verify ordering in template context
        applications = response.context['applications']
        self.assertEqual(list(applications), [app1, app2])
        self.assertContains(response, 'علی امینی')
        self.assertContains(response, 'رضا کریمی')

    def test_bulk_status_updates(self):
        """تست تغییر وضعیت گروهی متقاضیان"""
        candidate1 = Candidate.objects.create(first_name='مریم', last_name='راد', email='maryam@example.com', phone_number='09123333333', national_id='3333333333')
        candidate2 = Candidate.objects.create(first_name='سارا', last_name='احمدی', email='sara@example.com', phone_number='09124444444', national_id='4444444444')
        app1 = JobApplication.objects.create(job=self.job, candidate=candidate1, status=JobApplication.STATUS_IN_PROGRESS)
        app2 = JobApplication.objects.create(job=self.job, candidate=candidate2, status=JobApplication.STATUS_IN_PROGRESS)

        self.client.login(username='recruiter_user', password='password123')
        url = reverse('bulk_update_application_status', kwargs={'job_id': self.job.id})
        
        # Select both applications and update to SELECTED
        response = self.client.post(url, {
            'application_ids': [app1.id, app2.id],
            'status_action': 'SELECTED'
        })
        self.assertEqual(response.status_code, 302)
        
        app1.refresh_from_db()
        app2.refresh_from_db()
        self.assertEqual(app1.status, JobApplication.STATUS_SELECTED)
        self.assertEqual(app2.status, JobApplication.STATUS_SELECTED)

    def test_bulk_advance_stage(self):
        """تست ارتقای گروهی مرحله متقاضیان"""
        candidate1 = Candidate.objects.create(first_name='امید', last_name='نوری', email='omid@example.com', phone_number='09125555555', national_id='5555555555')
        candidate2 = Candidate.objects.create(first_name='پیمان', last_name='حسنی', email='peyman@example.com', phone_number='09126666666', national_id='6666666666')
        app1 = JobApplication.objects.create(job=self.job, candidate=candidate1, status=JobApplication.STATUS_IN_PROGRESS)
        app2 = JobApplication.objects.create(job=self.job, candidate=candidate2, status=JobApplication.STATUS_IN_PROGRESS)
        
        stages = list(self.job.stages.filter(is_deleted=False).order_by('sequence'))
        self.assertEqual(app1.current_stage, stages[0])
        self.assertEqual(app2.current_stage, stages[0])
        
        # Complete Stage 1 states
        state1_1 = app1.stage_states.get(stage=stages[0])
        state1_1.status = ApplicationStageState.STATUS_COMPLETED
        state1_1.save()
        
        state2_1 = app2.stage_states.get(stage=stages[0])
        state2_1.status = ApplicationStageState.STATUS_COMPLETED
        state2_1.save()
        
        self.client.login(username='recruiter_user', password='password123')
        url = reverse('bulk_advance_stage', kwargs={'job_id': self.job.id})
        
        # Perform bulk advance
        response = self.client.post(url, {
            'application_ids': [app1.id, app2.id]
        })
        self.assertEqual(response.status_code, 302)
        
        app1.refresh_from_db()
        app2.refresh_from_db()
        self.assertEqual(app1.current_stage, stages[1])
        self.assertEqual(app2.current_stage, stages[1])

    def test_talent_bank_advanced_filtering(self):
        """تست فیلترهای پیشرفته بانک استعداد"""
        c1 = Candidate.objects.create(first_name='احسان', last_name='راد', email='ehsan@example.com', phone_number='09127777777', national_id='7777777777')
        CandidateEducation.objects.create(candidate=c1, degree='MASTER', major='مهندسی کامپیوتر', university='شریف', gpa=17.5, graduation_year=1400)
        from datetime import date, timedelta
        CandidateExperience.objects.create(candidate=c1, company='رایا', job_title='برنامه‌نویس', start_date=date.today() - timedelta(days=1200), end_date=date.today())
        CandidateSkill.objects.create(candidate=c1, name='Python', level='EXPERT')

        c2 = Candidate.objects.create(first_name='مهسا', last_name='کریمی', email='mahsa@example.com', phone_number='09128888888', national_id='8888888888')
        CandidateEducation.objects.create(candidate=c2, degree='BACHELOR', major='ریاضیات', university='تهران', gpa=16.0, graduation_year=1401)

        self.client.login(username='recruiter_user', password='password123')
        url = reverse('candidate_list')

        # 1. Filter by degree
        response = self.client.get(url, {'degree': 'MASTER'})
        self.assertEqual(response.status_code, 200)
        self.assertIn(c1, response.context['candidates'])
        self.assertNotIn(c2, response.context['candidates'])

        # 2. Filter by major
        response = self.client.get(url, {'major': 'ریاضی'})
        self.assertEqual(response.status_code, 200)
        self.assertNotIn(c1, response.context['candidates'])
        self.assertIn(c2, response.context['candidates'])

        # 3. Filter by skill
        response = self.client.get(url, {'skill': 'python'})
        self.assertEqual(response.status_code, 200)
        self.assertIn(c1, response.context['candidates'])
        self.assertNotIn(c2, response.context['candidates'])

        # 4. Filter by experience
        response = self.client.get(url, {'min_experience': '2'})
        self.assertEqual(response.status_code, 200)
        self.assertIn(c1, response.context['candidates'])
        self.assertNotIn(c2, response.context['candidates'])

    def test_assign_candidate_to_job(self):
        """تست انتساب مستقیم متقاضی به فرصت شغلی فعال"""
        candidate = Candidate.objects.create(first_name='نوید', last_name='احمدی', email='navid@example.com', phone_number='09120000000', national_id='0000000000')
        self.client.login(username='recruiter_user', password='password123')
        url = reverse('assign_candidate_job', kwargs={'candidate_id': candidate.id})

        # Perform assignment
        response = self.client.post(url, {'job_id': self.job.id})
        self.assertEqual(response.status_code, 302)

        # Verify application created
        self.assertTrue(JobApplication.objects.filter(job=self.job, candidate=candidate, is_deleted=False).exists())

        # Attempt duplicate assignment
        response = self.client.post(url, {'job_id': self.job.id})
        self.assertEqual(response.status_code, 400)

    def test_job_status_updates_on_selected_candidate(self):
        """تست اینکه با پذیرش نهایی متقاضی، وضعیت فرصت شغلی به اتمام یافته (CLOSED) تغییر کند"""
        candidate = Candidate.objects.create(
            first_name='علی', last_name='رضایی', email='ali@example.com',
            phone_number='09123334455', national_id='1112223334'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        
        # Change candidate status to SELECTED
        self.client.login(username='recruiter_user', password='password123')
        url = reverse('update_application_status', kwargs={'pk': app.id})
        response = self.client.post(url, {'status': 'SELECTED'})
        self.assertEqual(response.status_code, 200)
        
        # Verify job is CLOSED (اتمام یافته)
        self.job.refresh_from_db()
        self.assertEqual(self.job.status, JobOpportunity.STATUS_CLOSED)

    def test_job_status_updates_on_stage_advance(self):
        """تست اینکه با انتقال متقاضی به مرحله بعد، وضعیت فرصت شغلی به صورت خودکار تغییر کند"""
        candidate = Candidate.objects.create(
            first_name='محمد', last_name='کریمی', email='m@example.com',
            phone_number='09124445566', national_id='2223334445'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        
        # Mark first stage as completed
        state = app.stage_states.filter(stage=self.job.stages.order_by('sequence').first()).first()
        state.status = 'COMPLETED'
        state.score = 80.0
        state.save()
        
        # Advance candidate to next stage (مصاحبه فنی) using bulk advance
        self.client.login(username='recruiter_user', password='password123')
        url = reverse('bulk_advance_stage', kwargs={'job_id': self.job.id})
        response = self.client.post(url, {'application_ids': [app.id]})
        self.assertEqual(response.status_code, 302)
        
        # Verify job status changed to STATUS_INTERVIEW (since next stage is 'مصاحبه فنی')
        self.job.refresh_from_db()
        self.assertEqual(self.job.status, JobOpportunity.STATUS_INTERVIEW)

    def test_registration_deadline_enforced(self):
        """تست اینکه پس از اتمام مهلت ثبت‌نام، امکان ثبت‌نام مجدد در فرصت شغلی نباشد"""
        from datetime import date, timedelta
        # Update job end_date to yesterday
        self.job.end_date = date.today() - timedelta(days=1)
        self.job.save()
        
        # 1. Check CareersDetailAndApplyView redirects/renders closed warning
        url_apply = reverse('careers_apply', kwargs={'pk': self.job.id})
        response = self.client.get(url_apply)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context.get('registration_closed'))
        
        # 2. Check direct apply fails
        candidate = Candidate.objects.create(
            first_name='حسین', last_name='نوری', email='h@example.com',
            phone_number='09125556677', national_id='3334445556'
        )
        candidate_user = User.objects.create_user(username='cand_user_test', password='password123')
        candidate_user.profile.role = UserProfile.ROLE_CANDIDATE
        candidate_user.profile.save()
        candidate.user = candidate_user
        candidate.save()
        
        self.client.login(username='cand_user_test', password='password123')
        url_direct = reverse('candidate_apply_direct', kwargs={'pk': self.job.id})
        response = self.client.post(url_direct)
        self.assertEqual(response.status_code, 400) # Bad Request/Expired

    def test_closed_jobs_excluded_from_score_entry(self):
        """تست اینکه فرصت‌های شغلی اتمام‌یافته در صفحه ثبت نمرات نمایش داده نشوند"""
        self.job.status = JobOpportunity.STATUS_CLOSED
        self.job.save()
        
        self.client.login(username='recruiter_user', password='password123')
        url = reverse('candidate_score_entry')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        
        # Ensure CLOSED job is NOT in the jobs query of context
        self.assertNotIn(self.job, response.context['jobs'])

    def test_talent_bank_advanced_filters_and_recommendations(self):
        """تست فیلترهای نمرات پیشرفته و سیستم پیشنهاددهی هوشمند بر اساس انطباق شغل"""
        from apps.candidates.models import CandidateSkill, CandidateExperience, ApplicationStageState
        
        # ۱. ایجاد متقاضی اول (با مهارت پایتون و نمره آزمون ۸۵)
        cand1 = Candidate.objects.create(
            first_name='مهدی', last_name='رضایی', email='m.rezaei@example.com',
            phone_number='09121111111', national_id='1111111111'
        )
        CandidateSkill.objects.create(candidate=cand1, name='Python', level='EXPERT')
        CandidateExperience.objects.create(candidate=cand1, company='Company A', job_title='Python Developer', start_date='2020-01-01')
        
        # ثبت درخواست و اختصاص نمره آزمون کتبی به کاندیدای اول
        app1 = JobApplication.objects.create(job=self.job, candidate=cand1)
        exam_stage = self.job.stages.order_by('sequence').first()
        state1 = app1.stage_states.filter(stage=exam_stage).first()
        state1.score = 85.0
        state1.status = 'COMPLETED'
        state1.save()
        
        # ۲. ایجاد متقاضی دوم (با مهارت جاوا و نمره مصاحبه ۹۰)
        cand2 = Candidate.objects.create(
            first_name='سارا', last_name='احمدی', email='sara@example.com',
            phone_number='09122222222', national_id='2222222222'
        )
        CandidateSkill.objects.create(candidate=cand2, name='Java', level='EXPERT')
        
        app2 = JobApplication.objects.create(job=self.job, candidate=cand2)
        # فرض می‌کنیم مرحله دوم مصاحبه فنی است
        interview_stage = self.job.stages.order_by('sequence')[1]
        state2 = app2.stage_states.filter(stage=interview_stage).first()
        state2.score = 90.0
        state2.status = 'COMPLETED'
        state2.save()

        # ورود به پنل استخدام‌کننده و درخواست صفحه بانک استعدادها
        self.client.login(username='recruiter_user', password='password123')
        url = reverse('candidate_list')

        # الف) تست فیلتر حداقل نمره آزمون کتبی (min_exam_score=80)
        response = self.client.get(url, {'min_exam_score': '80'})
        self.assertEqual(response.status_code, 200)
        candidates = response.context['candidates']
        self.assertIn(cand1, candidates)
        self.assertNotIn(cand2, candidates)

        # ب) تست فیلتر حداقل نمره مصاحبه حضوری (min_interview_score=85)
        response = self.client.get(url, {'min_interview_score': '85'})
        self.assertEqual(response.status_code, 200)
        candidates = response.context['candidates']
        self.assertIn(cand2, candidates)
        self.assertNotIn(cand1, candidates)

        # ج) تست پیشنهاددهی هوشمند متقاضیان مناسب برای شغل پایتون
        # ایجاد فرصت شغلی جدید با عنوان "Python Engineer"
        python_job = JobOpportunity.objects.create(
            request_number='REQ-999', title='Python Engineer', code='PY-999',
            department='فنی', recruitment_type='EXTERNAL', workflow=self.workflow
        )
        response = self.client.get(url, {'similar_to_job': python_job.id})
        self.assertEqual(response.status_code, 200)
        candidates = response.context['candidates']
        
        # متقاضی اول به دلیل داشتن مهارت و سابقه کار مرتبط با کلمه Python باید انطباق مثبت داشته و در لیست قرار گیرد
        self.assertIn(cand1, candidates)
        # متقاضی دوم هیچ کلمه مشترکی ندارد و نباید منطبق تشخیص داده شود
        self.assertNotIn(cand2, candidates)
        
        # بررسی ثبت درصد انطباق
        matched_cand1 = next(c for c in candidates if c.id == cand1.id)
        self.assertGreater(matched_cand1.match_score, 0)
        self.assertIn("مهارت‌ها", matched_cand1.match_reasons)

    def test_candidates_and_scores_export_excel(self):
        """تست خروجی اکسل متقاضیان و نمرات"""
        from apps.candidates.models import Candidate, JobApplication, ApplicationStageState
        
        # 1. Test Export Candidates Excel
        self.client.login(username='recruiter_user', password='password123')
        response = self.client.get(reverse('candidate_export_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # 2. Test Export Score Entry Excel
        stage = self.job.stages.first()
        cand = Candidate.objects.create(
            first_name='زهرا', last_name='امینی', email='zahra@example.com', phone_number='09124445555', national_id='9999999999'
        )
        app = JobApplication.objects.create(job=self.job, candidate=cand)
        state = app.stage_states.filter(stage=stage).first()
        
        url = reverse('score_entry_export_excel')
        response = self.client.get(url, {'job_id': self.job.id, 'stage_id': stage.id})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # 3. Test Export Interviews Excel
        response = self.client.get(reverse('interviews_export_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # 4. Test Export Assessment Center Excel
        response = self.client.get(reverse('assessment_center_export_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # 5. Test Export Job Final Ranking Excel
        response = self.client.get(reverse('job_ranking_export_excel', kwargs={'pk': self.job.id}))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_candidate_excel_import(self):
        """تست ورود اطلاعات متقاضیان از اکسل"""
        import io
        import openpyxl
        from apps.candidates.models import Candidate, JobApplication

        self.client.login(username='recruiter_user', password='password123')

        # 1. Test template download view
        response = self.client.get(reverse('candidate_import_template'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # 2. Build a valid in-memory Excel file to import
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["نام", "نام خانوادگی", "کد ملی", "ایمیل", "شماره تماس", "شماره پرسنلی (اختیاری)"])
        ws.append(["مریم", "کریمی", "1122334455", "maryam@example.com", "09129998888", "12345"])
        
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        excel_file = io.BytesIO(excel_io.read())
        excel_file.name = 'candidates.xlsx'

        # Import valid file without job assignment
        response = self.client.post(reverse('candidate_import_excel'), {'excel_file': excel_file})
        self.assertEqual(response.status_code, 302)
        
        # Verify candidate created
        cand = Candidate.objects.filter(national_id="1122334455").first()
        self.assertIsNotNone(cand)
        self.assertEqual(cand.first_name, "مریم")
        self.assertEqual(cand.last_name, "کریمی")
        self.assertEqual(cand.personnel_number, "12345")

        # 3. Import with job assignment
        wb_job = openpyxl.Workbook()
        ws_job = wb_job.active
        ws_job.append(["نام", "نام خانوادگی", "کد ملی", "ایمیل", "شماره تماس", "شماره پرسنلی (اختیاری)"])
        ws_job.append(["سعید", "امینی", "5544332211", "saeed@example.com", "09127776666", "54321"])
        
        excel_io_job = io.BytesIO()
        wb_job.save(excel_io_job)
        excel_io_job.seek(0)
        excel_file_job = io.BytesIO(excel_io_job.read())
        excel_file_job.name = 'candidates_job.xlsx'

        response = self.client.post(reverse('candidate_import_excel'), {
            'excel_file': excel_file_job,
            'job_id': self.job.id
        })
        self.assertEqual(response.status_code, 302)
        
        # Verify candidate and application created
        cand_job = Candidate.objects.filter(national_id="5544332211").first()
        self.assertIsNotNone(cand_job)
        self.assertTrue(JobApplication.objects.filter(job=self.job, candidate=cand_job).exists())

        # 4. Import invalid file (invalid national id format)
        wb_invalid = openpyxl.Workbook()
        ws_invalid = wb_invalid.active
        ws_invalid.append(["نام", "نام خانوادگی", "کد ملی", "ایمیل", "شماره تماس", "شماره پرسنلی (اختیاری)"])
        ws_invalid.append(["رضا", "راد", "123", "invalid-email", "09121111111", ""])
        
        excel_io_invalid = io.BytesIO()
        wb_invalid.save(excel_io_invalid)
        excel_io_invalid.seek(0)
        excel_file_invalid = io.BytesIO(excel_io_invalid.read())
        excel_file_invalid.name = 'candidates_invalid.xlsx'

        response = self.client.post(reverse('candidate_import_excel'), {'excel_file': excel_file_invalid})
        self.assertEqual(response.status_code, 302)
        
        # Verify candidate not created
        self.assertFalse(Candidate.objects.filter(first_name="رضا").exists())

    def test_candidate_excel_import_with_education_and_skills(self):
        """تست ورود اطلاعات متقاضیان از اکسل به همراه سوابق تحصیلی و مهارت‌ها"""
        import io
        import openpyxl
        from apps.candidates.models import Candidate, CandidateEducation, CandidateSkill

        self.client.login(username='recruiter_user', password='password123')

        # Build an Excel file with the new columns
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "نام", "نام خانوادگی", "کد ملی", "ایمیل", "شماره تماس", 
            "شماره پرسنلی (اختیاری)", "مدرک تحصیلی (کاردانی/کارشناسی/ارشد/دکتری)", "رشته تحصیلی", "مهارت‌ها (جدا شده با کاما)"
        ])
        ws.append([
            "سارا", "کریمی", "9876543210", "sara@example.com", "09123456789", 
            "99001", "کارشناسی ارشد", "مهندسی کامپیوتر", "Python, Django, SQL، Docker"
        ])
        
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        excel_file = io.BytesIO(excel_io.read())
        excel_file.name = 'candidates_edu_skills.xlsx'

        response = self.client.post(reverse('candidate_import_excel'), {'excel_file': excel_file})
        self.assertEqual(response.status_code, 302)

        # Verify candidate created
        cand = Candidate.objects.filter(national_id="9876543210").first()
        self.assertIsNotNone(cand)
        self.assertEqual(cand.first_name, "سارا")
        self.assertEqual(cand.last_name, "کریمی")
        self.assertEqual(cand.personnel_number, "99001")

        # Verify education created
        edu = cand.education.filter(is_deleted=False).first()
        self.assertIsNotNone(edu)
        self.assertEqual(edu.degree, "MASTER")
        self.assertEqual(edu.major, "مهندسی کامپیوتر")
        self.assertEqual(edu.university, "ثبت نشده")
        self.assertEqual(edu.gpa, 0.0)
        self.assertEqual(edu.graduation_year, 1400)

        # Verify skills created
        skills = list(cand.skills.filter(is_deleted=False).values_list('name', 'level'))
        self.assertEqual(len(skills), 4)
        self.assertIn(('Python', 'INTERMEDIATE'), skills)
        self.assertIn(('Django', 'INTERMEDIATE'), skills)
        self.assertIn(('SQL', 'INTERMEDIATE'), skills)
        self.assertIn(('Docker', 'INTERMEDIATE'), skills)

    def test_export_job_final_ranking_excel(self):
        """تست خروجی اکسل رتبه‌بندی نهایی داوطلبان یک فرصت شغلی"""
        import openpyxl
        import io
        from apps.candidates.models import Candidate, JobApplication, ApplicationStageState

        # Create candidates and applications
        candidate = Candidate.objects.create(
            first_name='نیلوفر',
            last_name='رحیمی',
            email='niloofar@example.com',
            phone_number='09127778888',
            national_id='7777888899',
            personnel_number='11223'
        )
        app = JobApplication.objects.create(
            job=self.job,
            candidate=candidate
        )
        # Complete first stage with score 85
        states = app.stage_states.all().order_by('stage__sequence')
        state1 = states[0]
        state1.score = 85.0
        state1.status = ApplicationStageState.STATUS_COMPLETED
        state1.save()

        self.client.login(username='recruiter_user', password='password123')

        url = reverse('job_ranking_export_excel', kwargs={'pk': self.job.id})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Load Excel content
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Verify headers
        expected_headers = ["رتبه", "نام", "نام خانوادگی", "کد ملی", "شماره پرسنلی", "امتیاز نهایی وزنی"]
        for stage in self.job.stages.filter(is_deleted=False).order_by('sequence'):
            expected_headers.append(stage.name)
        expected_headers.append("وضعیت درخواست")

        # Check headers match
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, expected_headers)

        # Check candidate row data
        row_values = [cell.value for cell in ws[2]]
        # Rank should be 1
        self.assertEqual(row_values[0], 1)
        # First name
        self.assertEqual(row_values[1], "نیلوفر")
        # Last name
        self.assertEqual(row_values[2], "رحیمی")
        # National ID
        self.assertEqual(row_values[3], "7777888899")
        # Personnel Number
        self.assertEqual(row_values[4], "11223")
        # Score on Stage 1 should be 85
        self.assertEqual(row_values[6], 85.0)
        # Score on Stage 2 should be "در حال ارزیابی"
        self.assertEqual(row_values[7], "در حال ارزیابی")

    def test_candidate_list_pagination(self):
        """تست صفحه‌بندی (Pagination) در لیست متقاضیان بانک استعدادها"""
        from apps.candidates.models import Candidate
        
        # Create 12 additional candidates (total 13 with setup, but wait - setUp doesn't create candidate, only job/workflow)
        for i in range(12):
            Candidate.objects.create(
                first_name=f'داوطلب_{i}',
                last_name=f'فامیلی_{i}',
                email=f'candidate_{i}@example.com',
                phone_number=f'091200000{i:02d}',
                national_id=f'{i:010d}'
            )

        self.client.login(username='recruiter_user', password='password123')
        
        # Request page 1
        response = self.client.get(reverse('candidate_list'))
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['is_paginated'])
        self.assertEqual(len(response.context['candidates']), 10) # 10 items per page

        # Request page 2
        response = self.client.get(reverse('candidate_list') + '?page=2')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['candidates']), 2) # remaining 2 items

    def test_score_entry_pagination(self):
        """تست صفحه‌بندی (Pagination) در صفحه ورود نمرات"""
        from apps.candidates.models import Candidate, JobApplication
        
        # Get the first stage of the job
        active_stage = self.job.stages.filter(is_deleted=False).order_by('sequence').first()
        self.assertIsNotNone(active_stage)

        # Create 12 candidates and applications
        for i in range(12):
            candidate = Candidate.objects.create(
                first_name=f'داوطلب_{i}',
                last_name=f'فامیلی_{i}',
                email=f'candidate_{i}@example.com',
                phone_number=f'091200000{i:02d}',
                national_id=f'{i:010d}'
            )
            JobApplication.objects.create(
                job=self.job,
                candidate=candidate
            )

        self.client.login(username='recruiter_user', password='password123')
        
        # Request page 1
        url = reverse('candidate_score_entry') + f'?job_id={self.job.id}&stage_id={active_stage.id}'
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['is_paginated'])
        self.assertEqual(len(response.context['pending_states']), 10)

        # Request page 2
        response = self.client.get(url + '&page=2')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['pending_states']), 2)

    def test_import_score_entry_excel_success(self):
        """تست ورود موفقیت‌آمیز نمرات و وضعیت داوطلبان از فایل اکسل"""
        import io
        import openpyxl
        
        # Setup candidate & application
        candidate = Candidate.objects.create(
            first_name='زهرا', last_name='امینی', email='zahra@example.com',
            phone_number='09124445555', national_id='9999999999'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        stage = self.job.stages.order_by('sequence').first()
        state = app.stage_states.filter(stage=stage).first()
        
        # Build an Excel file in memory
        wb = openpyxl.Workbook()
        ws = wb.active
        # Columns must match expected headers
        ws.append(["شناسه وضعیت", "نام", "نام خانوادگی", "نمره نهایی مرحله", "وضعیت ارزیابی", "توضیحات و یادداشت ارزیاب"])
        ws.append([state.id, "زهرا", "امینی", "92.5", "قبول شده در این مرحله", "مصاحبه عالی بود"])
        
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        excel_file = io.BytesIO(excel_io.read())
        excel_file.name = 'score_import_test.xlsx'
        
        self.client.login(username='recruiter_user', password='password123')
        
        url = reverse('score_entry_import_excel')
        response = self.client.post(url, {
            'job_id': self.job.id,
            'stage_id': stage.id,
            'excel_file': excel_file
        })
        
        self.assertEqual(response.status_code, 302)
        
        # Verify db updates
        state.refresh_from_db()
        self.assertEqual(state.score, 92.5)
        self.assertEqual(state.status, ApplicationStageState.STATUS_COMPLETED)
        self.assertEqual(state.notes, "مصاحبه عالی بود")
        self.assertEqual(state.evaluator, self.recruiter)

    def test_import_score_entry_excel_validation_errors(self):
        """تست خطاهای مختلف در ورود نمرات از اکسل"""
        import io
        import openpyxl
        
        candidate = Candidate.objects.create(
            first_name='علی', last_name='رضایی', email='ali.r@example.com',
            phone_number='09125556666', national_id='8888888888'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        stage1 = self.job.stages.order_by('sequence').first()
        stage2 = self.job.stages.order_by('sequence')[1]
        
        state1 = app.stage_states.filter(stage=stage1).first()
        state2 = app.stage_states.filter(stage=stage2).first()
        
        self.client.login(username='recruiter_user', password='password123')
        url = reverse('score_entry_import_excel')
        
        # 1. Missing parameter
        response = self.client.post(url, {
            'job_id': self.job.id,
            'stage_id': stage1.id
            # missing excel_file
        })
        self.assertEqual(response.status_code, 302)
        
        # 2. Excel with invalid headers (missing critical columns)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["نام", "نام خانوادگی", "نمره نهایی مرحله"]) # missing 'شناسه وضعیت' & 'وضعیت ارزیابی'
        ws.append(["علی", "رضایی", "85"])
        
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        excel_file = io.BytesIO(excel_io.read())
        excel_file.name = 'invalid_headers.xlsx'
        
        response = self.client.post(url, {
            'job_id': self.job.id,
            'stage_id': stage1.id,
            'excel_file': excel_file
        })
        self.assertEqual(response.status_code, 302)
        
        # 3. Excel row with state_id of stage2 (which is locked/not accessible because stage1 is not COMPLETED)
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.append(["شناسه وضعیت", "نمره نهایی مرحله", "وضعیت ارزیابی"])
        ws2.append([state2.id, "90.0", "قبول"])
        
        excel_io2 = io.BytesIO()
        wb2.save(excel_io2)
        excel_io2.seek(0)
        excel_file2 = io.BytesIO(excel_io2.read())
        excel_file2.name = 'locked_stage.xlsx'
        
        response = self.client.post(url, {
            'job_id': self.job.id,
            'stage_id': stage2.id,
            'excel_file': excel_file2
        })
        self.assertEqual(response.status_code, 302)
        
        # Verify state2 remains unchanged
        state2.refresh_from_db()
        self.assertEqual(state2.score, 0.0)
        self.assertEqual(state2.status, ApplicationStageState.STATUS_PENDING)

    def test_score_entry_autosave_on_pagination(self):
        """تست ذخیره‌سازی خودکار نمرات هنگام جابجایی بین صفحات با استفاده از HTMX"""
        # Create 12 candidates to ensure 2 pages
        active_stage = self.job.stages.order_by('sequence').first()
        states_list = []
        for i in range(12):
            candidate = Candidate.objects.create(
                first_name=f'داوطلب_{i}',
                last_name=f'فامیلی_{i}',
                email=f'candidate_{i}@example.com',
                phone_number=f'091200000{i:02d}',
                national_id=f'{i:010d}'
            )
            app = JobApplication.objects.create(job=self.job, candidate=candidate)
            states_list.append(app.stage_states.filter(stage=active_stage).first())
            
        self.client.login(username='recruiter_user', password='password123')
        
        # Target state is one of the candidates on page 1 (which will be saved during page change)
        # The candidates are ordered by last_name, so we can identify a state to update
        state_to_save = states_list[0]
        
        post_data = {
            'job_id': self.job.id,
            'stage_id': active_stage.id,
            'eval_status': 'ALL',
            f'score_{state_to_save.id}': '88.5',
            f'status_{state_to_save.id}': 'COMPLETED',
            f'notes_{state_to_save.id}': 'توضیحات ذخیره خودکار پیجینگ',
        }
        
        # We simulate clicking page 2 link, which posts the current page's scores and requests page 2
        url = reverse('candidate_score_entry') + '?page=2'
        response = self.client.post(url, post_data, HTTP_HX_REQUEST='true')
        
        self.assertEqual(response.status_code, 200)
        # Verify page 2 contains 2 candidates (total 12, 10 per page)
        self.assertEqual(len(response.context['pending_states']), 2)
        self.assertEqual(response.context['page_obj'].number, 2)
        
        # Verify page 1 state's score was saved
        state_to_save.refresh_from_db()
        self.assertEqual(state_to_save.score, 88.5)
        self.assertEqual(state_to_save.status, 'COMPLETED')
        self.assertEqual(state_to_save.notes, 'توضیحات ذخیره خودکار پیجینگ')

    def test_score_entry_failed_prior_filtering(self):
        """تست فیلتر شدن و نمایش متقاضیانی که در مراحل قبلی مردود شده‌اند در لیست ورود نمرات"""
        candidate1 = Candidate.objects.create(
            first_name='علی',
            last_name='تقوی',
            email='ali.t@example.com',
            phone_number='09121111111',
            national_id='1111111111'
        )
        app1 = JobApplication.objects.create(job=self.job, candidate=candidate1)
        stages = list(self.job.stages.filter(is_deleted=False).order_by('sequence'))
        
        # Mark Stage 1 as FAILED for candidate1
        state1_stage1 = app1.stage_states.get(stage=stages[0])
        state1_stage1.status = ApplicationStageState.STATUS_FAILED
        state1_stage1.score = 50.0
        state1_stage1.save()
        
        self.client.login(username='recruiter_user', password='password123')
        url = reverse('candidate_score_entry')
        
        # 1. By default (show_failed_prior is False), candidate1 (who failed stage 1) should NOT be in stage 2's list
        response = self.client.get(f"{url}?job_id={self.job.id}&stage_id={stages[1].id}")
        self.assertEqual(response.status_code, 200)
        self.assertNotIn(app1.stage_states.get(stage=stages[1]), response.context['pending_states'])
        
        # 2. When show_failed_prior is True, candidate1 should be in the list
        response = self.client.get(f"{url}?job_id={self.job.id}&stage_id={stages[1].id}&show_failed_prior=true")
        self.assertEqual(response.status_code, 200)
        state1_stage2 = app1.stage_states.get(stage=stages[1])
        self.assertIn(state1_stage2, response.context['pending_states'])
        
        # Check if the HTML contains the correct label for the failed candidate
        self.assertContains(response, 'مردود (در مراحل قبلی)')

    def test_score_entry_custom_evaluation_date(self):
        """تست اختصاص تاریخ ارزیابی سفارشی و پیش‌فرض تاریخ امروز در صورت عدم ورود تاریخ"""
        self.client.login(username='recruiter_user', password='password123')
        
        candidate = Candidate.objects.create(
            first_name='زهرا',
            last_name='تقوی',
            email='zahra.t@example.com',
            phone_number='09121111112',
            national_id='1111111112'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        stages = list(self.job.stages.filter(is_deleted=False).order_by('sequence'))
        state = app.stage_states.filter(stage=stages[0]).first()
        
        # 1. Submit with custom Jalali date
        post_data = {
            'job_id': self.job.id,
            f'score_{state.id}': '85.0',
            f'status_{state.id}': 'COMPLETED',
            f'notes_{state.id}': 'نمرات با تاریخ سفارشی',
            f'date_{state.id}': '1405/03/16',  # custom Jalali date
        }
        url = reverse('candidate_score_entry')
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200)
        
        state.refresh_from_db()
        self.assertEqual(state.score, 85.0)
        self.assertEqual(state.status, 'COMPLETED')
        # Jalali 1405/03/16 corresponds to Gregorian 2026-06-06
        self.assertEqual(state.evaluation_date, date(2026, 6, 6))

        # 2. Reset status and submit without date -> should default to today's date
        state.status = 'PENDING'
        state.evaluation_date = None
        state.save()
        
        post_data = {
            'job_id': self.job.id,
            f'score_{state.id}': '90.0',
            f'status_{state.id}': 'COMPLETED',
            f'notes_{state.id}': 'نمرات با تاریخ امروز',
            f'date_{state.id}': '',  # empty date
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200)
        
        state.refresh_from_db()
        self.assertEqual(state.evaluation_date, date.today())

    def test_candidate_password_change_view(self):
        """تست عملکرد نمای تغییر رمز عبور متقاضی"""
        candidate_user = User.objects.create_user(
            username='1234567890',
            password='09121112222',
            first_name='علی',
            last_name='احمدی'
        )
        candidate_user.profile.role = UserProfile.ROLE_CANDIDATE
        candidate_user.profile.phone_number = '09121112222'
        candidate_user.profile.save()

        # Login as candidate
        self.client.login(username='1234567890', password='09121112222')
        
        # Access password change page
        url = reverse('candidate_password_change')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

        # POST form to change password
        post_data = {
            'old_password': '09121112222',
            'new_password1': 'newpassword123',
            'new_password2': 'newpassword123',
        }
        response = self.client.post(url, post_data)
        # Should redirect to candidate dashboard (code 302)
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('candidate_dashboard'))

        # Check login with new password
        self.client.logout()
        login_success = self.client.login(username='1234567890', password='newpassword123')
        self.assertTrue(login_success)

    def test_candidate_excel_import_creates_user(self):
        """تست ایجاد خودکار اکانت کاربری با مشخصات صحیح (کد ملی و شماره موبایل) در زمان ایمپورت متقاضی از اکسل"""
        import io
        import openpyxl
        
        self.client.login(username='recruiter_user', password='password123')
        
        # Build in-memory excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["نام", "نام خانوادگی", "کد ملی", "ایمیل", "شماره تماس", "شماره پرسنلی (اختیاری)"])
        # Phone number with 10 digits without leading zero -> should be normalized to 09123334444
        ws.append(["محسن", "رضایی", "9876543210", "mohsen@example.com", "9123334444", "9988"])
        
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        excel_file = io.BytesIO(excel_io.read())
        excel_file.name = 'import_test.xlsx'

        url = reverse('candidate_import_excel')
        response = self.client.post(url, {'excel_file': excel_file})
        self.assertEqual(response.status_code, 302)

        # Check candidate profile in DB
        candidate = Candidate.objects.filter(national_id="9876543210").first()
        self.assertIsNotNone(candidate)
        self.assertEqual(candidate.first_name, "محسن")
        self.assertEqual(candidate.last_name, "رضایی")
        # Assert normalized phone number in profile
        self.assertEqual(candidate.phone_number, "09123334444")
        self.assertIsNotNone(candidate.user)
        
        # Assert Django User is created with national ID as username
        user = candidate.user
        self.assertEqual(user.username, "9876543210")
        self.assertEqual(user.profile.role, UserProfile.ROLE_CANDIDATE)
        self.assertEqual(user.profile.phone_number, "09123334444")
        
        # Assert candidate can login with national_id and normalized phone number (with leading zero) as password
        self.client.logout()
        login_success = self.client.login(username='9876543210', password='09123334444')
        self.assertTrue(login_success)

    def test_conditional_stage_promotion_logic(self):
        """تست منطق ارجاع مشروط متقاضیان و دسترسی به مراحل بعدی"""
        candidate = Candidate.objects.create(
            first_name='حمید',
            last_name='رضایی',
            email='hamid.r@example.com',
            phone_number='09121112233',
            national_id='1111112233'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        stages = list(self.job.stages.filter(is_deleted=False).order_by('sequence'))
        state1 = app.stage_states.get(stage=stages[0])
        state2 = app.stage_states.get(stage=stages[1])

        # 1. By default, stage 2 is not accessible
        self.assertTrue(state1.is_accessible)
        self.assertFalse(state2.is_accessible)
        self.assertFalse(state2.has_failed_prior_stages)

        # 2. Candidate fails stage 1, stage 2 remains inaccessible, has_failed_prior_stages becomes True
        state1.status = ApplicationStageState.STATUS_FAILED
        state1.score = 40.0
        state1.save()
        
        state2.refresh_from_db()
        self.assertFalse(state2.is_accessible)
        self.assertTrue(state2.has_failed_prior_stages)

        # 3. Enable conditional pass on stage 1 -> stage 2 becomes accessible, has_failed_prior_stages becomes False, and candidate's current_stage advances to stage 2
        state1.is_conditional_pass = True
        state1.save()

        state2.refresh_from_db()
        app.refresh_from_db()
        self.assertTrue(state2.is_accessible)
        self.assertFalse(state2.has_failed_prior_stages)
        self.assertEqual(app.current_stage, stages[1])

    def test_conditional_stage_promotion_views(self):
        """تست نماهای به‌روزرسانی و ثبت نمره با قبولی ارفاقی/ارجاع مشروط"""
        self.client.login(username='recruiter_user', password='password123')
        
        candidate = Candidate.objects.create(
            first_name='سارا',
            last_name='کریمی',
            email='sara.k@example.com',
            phone_number='09121112244',
            national_id='1111112244'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        stages = list(self.job.stages.filter(is_deleted=False).order_by('sequence'))
        state1 = app.stage_states.get(stage=stages[0])

        # 1. Test single update view POST
        url = reverse('update_stage_state', kwargs={'pk': state1.id})
        post_data = {
            'score': '45.0',
            'status': 'FAILED',
            'notes': 'رد مشروط تک متقاضی',
            'is_conditional_pass': 'on'
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200)

        state1.refresh_from_db()
        app.refresh_from_db()
        self.assertTrue(state1.is_conditional_pass)
        self.assertEqual(state1.status, 'FAILED')
        self.assertEqual(app.current_stage, stages[1])

        # Reset to check bulk save
        state1.is_conditional_pass = False
        state1.status = 'PENDING'
        state1.save()
        app.current_stage = stages[0]
        app.save()

        # 2. Test bulk score entry POST
        bulk_url = reverse('candidate_score_entry')
        bulk_post_data = {
            'job_id': self.job.id,
            f'score_{state1.id}': '40.0',
            f'status_{state1.id}': 'FAILED',
            f'notes_{state1.id}': 'رد مشروط گروهی',
            f'is_conditional_pass_{state1.id}': 'on',
        }
        response = self.client.post(bulk_url, bulk_post_data)
        self.assertEqual(response.status_code, 200)

        state1.refresh_from_db()
        app.refresh_from_db()
        self.assertTrue(state1.is_conditional_pass)
        self.assertEqual(state1.status, 'FAILED')
        self.assertEqual(app.current_stage, stages[1])

    def test_score_bypass_locks_regular_post(self):
        """تست ثبت نمره در مرحله قفل شده با استفاده از قابلیت ثبت آزاد (بای‌پس)"""
        self.client.login(username='recruiter_user', password='password123')
        
        # Create a candidate and application
        candidate = Candidate.objects.create(
            first_name='بای‌پس',
            last_name='تست',
            email='bypass@example.com',
            phone_number='09121112255',
            national_id='1111112255'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        stages = list(self.job.stages.filter(is_deleted=False).order_by('sequence'))
        
        # Stage 1 and Stage 2
        state1 = app.stage_states.get(stage=stages[0])
        state2 = app.stage_states.get(stage=stages[1])
        
        # Stage 2 is currently locked (state1 is not completed, and state2.is_accessible is False)
        self.assertFalse(state2.is_accessible)
        
        # 1. Post score to stage 2 WITHOUT bypass_locks -> should NOT save
        bulk_url = reverse('candidate_score_entry')
        post_data_no_bypass = {
            'job_id': self.job.id,
            f'score_{state2.id}': '85.0',
            f'status_{state2.id}': 'COMPLETED',
            f'notes_{state2.id}': 'تست بدون بای‌پس',
        }
        response = self.client.post(bulk_url, post_data_no_bypass)
        state2.refresh_from_db()
        self.assertNotEqual(state2.score, 85.0)
        
        # 2. Post score to stage 2 WITH bypass_locks -> should save successfully
        post_data_with_bypass = post_data_no_bypass.copy()
        post_data_with_bypass['bypass_locks'] = 'true'
        response = self.client.post(bulk_url, post_data_with_bypass)
        self.assertEqual(response.status_code, 200)
        
        state2.refresh_from_db()
        self.assertEqual(state2.score, 85.0)
        self.assertEqual(state2.status, 'COMPLETED')

    def test_score_bypass_locks_excel_import(self):
        """تست بارگذاری نمرات از طریق فایل اکسل برای مرحله قفل شده با و بدون بای‌پس"""
        import io
        import openpyxl
        
        self.client.login(username='recruiter_user', password='password123')
        
        candidate = Candidate.objects.create(
            first_name='اکسل',
            last_name='بای‌پس',
            email='excel.bypass@example.com',
            phone_number='09121112266',
            national_id='1111112266'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        stages = list(self.job.stages.filter(is_deleted=False).order_by('sequence'))
        
        # Stage 2 is locked
        state2 = app.stage_states.get(stage=stages[1])
        self.assertFalse(state2.is_accessible)
        
        # Create an in-memory excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["شناسه وضعیت", "نام متقاضی", "نمره نهایی مرحله", "وضعیت ارزیابی", "توضیحات و یادداشت ارزیاب"])
        ws.append([state2.id, f"{candidate.first_name} {candidate.last_name}", 92.5, "قبول شده در این مرحله", "تست اکسل"])
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'test_scores.xlsx'
        
        # 1. Post excel WITHOUT bypass_locks
        url = reverse('score_entry_import_excel')
        response = self.client.post(url, {
            'job_id': self.job.id,
            'stage_id': stages[1].id,
            'excel_file': excel_file,
            'bypass_locks': 'false'
        })
        state2.refresh_from_db()
        self.assertNotEqual(state2.score, 92.5)
        
        # Reset excel file stream for another read
        excel_file.seek(0)
        
        # 2. Post excel WITH bypass_locks
        response = self.client.post(url, {
            'job_id': self.job.id,
            'stage_id': stages[1].id,
            'excel_file': excel_file,
            'bypass_locks': 'true'
        })
        state2.refresh_from_db()
        self.assertEqual(state2.score, 92.5)
        self.assertEqual(state2.status, 'COMPLETED')

    def test_external_interviewer_scores_panel(self):
        """تست پنل ورود دستی نمرات مصاحبه‌گران بدون یوزر و محاسبه میانگین وزنی"""
        self.client.login(username='recruiter_user', password='password123')
        
        # Create a candidate and application
        candidate = Candidate.objects.create(
            first_name='پوریا',
            last_name='احمدی',
            email='p.ahmadi@example.com',
            phone_number='09121112277',
            national_id='1111112277'
        )
        app = JobApplication.objects.create(job=self.job, candidate=candidate)
        stages = list(self.job.stages.filter(is_deleted=False).order_by('sequence'))
        
        stage1 = stages[0]
        stage1.stage_type = 'INTERVIEW'
        stage1.save()
        
        state1 = app.stage_states.get(stage=stage1)
        
        # 1. GET panel page
        url = reverse('interview_scores_panel', kwargs={'pk': state1.id})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'کاربرگ نمرات مصاحبه‌گران')
        
        # 2. POST panel data with 2 interviewers
        post_data = {
            'interviewer_name[]': ['مصاحبه‌گر اول', 'مصاحبه‌گر دوم'],
            'score[]': ['80.0', '95.0'],
            'weight[]': ['2', '1'],
            'notes[]': ['یادداشت اول', 'یادداشت دوم'],
            'bypass_locks': '1'
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200)
        
        # Verify db recalculation
        state1.refresh_from_db()
        self.assertEqual(state1.score, 85.00)
        self.assertEqual(state1.status, 'COMPLETED')
        
        # Verify external scores records
        ext_scores = state1.external_interviewer_scores.filter(is_deleted=False)
        self.assertEqual(ext_scores.count(), 2)
        score1 = ext_scores.get(interviewer_name='مصاحبه‌گر اول')
        self.assertEqual(score1.score, 80.0)
        self.assertEqual(score1.weight, 2)
        self.assertEqual(score1.notes, 'یادداشت اول')
        
        # 3. Verify GET panel again returns the saved scores
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'مصاحبه‌گر اول')
        self.assertContains(response, 'مصاحبه‌گر دوم')


