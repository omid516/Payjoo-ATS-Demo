from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, TemplateView
from django.views import View
from django.urls import reverse_lazy, reverse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, HttpResponseRedirect
from django.db import transaction, models
from django.core.exceptions import PermissionDenied

from apps.accounts.permissions import RoleRequiredMixin
from apps.accounts.models import UserProfile
from .models import (
    Candidate, CandidateEducation, CandidateExperience, JobApplication,
    CandidateLanguage, CandidateSkill, CandidateCertificate
)
from .forms import (
    CandidateForm, CandidateEducationFormSet, CandidateExperienceFormSet, JobApplicationForm,
    CandidateLanguageFormSet, CandidateSkillFormSet, CandidateCertificateFormSet
)

class CandidateListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    model = Candidate
    template_name = 'candidates/candidate_list.html'
    context_object_name = 'candidates'
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]
    paginate_by = 10

    def extract_keywords(self, text):
        if not text:
            return []
        import re
        # Remove non-word characters
        cleaned = re.sub(r'[^\w\s]', ' ', text)
        words = cleaned.split()
        stop_words = {
            'در', 'و', 'به', 'از', 'با', 'بر', 'تا', 'یا', 'برای', 'کارشناس', 'مدیر', 
            'مسئول', 'دبارتمان', 'بخش', 'سازمان', 'شرکت', 'توسعه', 'توسعه‌دهنده', 
            'ارشد', 'جذب', 'استخدام', 'نفر', 'ها', 'های', 'and', 'the', 'for', 'of', 'in'
        }
        keywords = [w.lower() for w in words if len(w) >= 3 and w.lower() not in stop_words]
        return list(set(keywords))

    def compute_job_match(self, candidate, job_keywords):
        score = 0
        matched_skills = []
        matched_exps = []
        matched_edus = []

        # 1. Skill Match (up to 45 points)
        skill_points = 0
        for skill in candidate.skills.filter(is_deleted=False):
            skill_name_lower = skill.name.lower()
            if any(kw in skill_name_lower for kw in job_keywords):
                matched_skills.append(skill.name)
                skill_points += 15
        score += min(skill_points, 45)

        # 2. Experience Match (up to 40 points)
        exp_points = 0
        for exp in candidate.experience.filter(is_deleted=False):
            exp_title_lower = exp.job_title.lower()
            if any(kw in exp_title_lower for kw in job_keywords):
                matched_exps.append(exp.job_title)
                exp_points += 20
        score += min(exp_points, 40)

        # 3. Education Match (up to 15 points)
        edu_points = 0
        for edu in candidate.education.filter(is_deleted=False):
            edu_major_lower = edu.major.lower()
            if any(kw in edu_major_lower for kw in job_keywords):
                matched_edus.append(edu.major)
                edu_points += 15
        score += min(edu_points, 15)

        reasons = []
        if matched_skills:
            reasons.append(f"مهارت‌ها: {', '.join(matched_skills[:3])}")
        if matched_exps:
            reasons.append(f"سوابق: {', '.join(matched_exps[:2])}")
        if matched_edus:
            reasons.append(f"رشته: {', '.join(matched_edus[:1])}")

        return score, ' | '.join(reasons)

    def get_queryset(self):
        queryset = Candidate.objects.filter(is_deleted=False).prefetch_related(
            'applications__stage_states__stage',
            'skills',
            'experience',
            'education'
        )
        
        # 1. Text Search
        search_query = self.request.GET.get('q')
        if search_query:
            queryset = queryset.filter(
                models.Q(first_name__icontains=search_query) |
                models.Q(last_name__icontains=search_query) |
                models.Q(national_id__icontains=search_query) |
                models.Q(email__icontains=search_query)
            )

        # 2. Education Degree Filter
        degree = self.request.GET.get('degree')
        if degree:
            queryset = queryset.filter(education__degree=degree, education__is_deleted=False).distinct()

        # 3. Major Filter
        major = self.request.GET.get('major')
        if major:
            queryset = queryset.filter(education__major__icontains=major, education__is_deleted=False).distinct()

        # 4. Skill Filter
        skill = self.request.GET.get('skill')
        if skill:
            queryset = queryset.filter(skills__name__icontains=skill, skills__is_deleted=False).distinct()

        # 5. Work Experience Years Filter (Python-side filtering for cross-DB safety)
        min_exp = self.request.GET.get('min_experience')
        if min_exp:
            try:
                min_exp_years = float(min_exp)
                valid_candidate_ids = []
                from datetime import date
                for candidate in queryset.prefetch_related('experience'):
                    total_days = 0
                    for exp in candidate.experience.filter(is_deleted=False):
                        start = exp.start_date
                        end = exp.end_date or date.today()
                        if start:
                            total_days += (end - start).days
                    if (total_days / 365.25) >= min_exp_years:
                        valid_candidate_ids.append(candidate.id)
                queryset = queryset.filter(id__in=valid_candidate_ids)
            except ValueError:
                pass

        # 6. Score Filters
        min_exam = self.request.GET.get('min_exam_score')
        if min_exam:
            try:
                val = float(min_exam)
                exam_kws = ["آزمون", "امتحان", "کتبی", "exam", "test"]
                q_obj = models.Q()
                for kw in exam_kws:
                    q_obj |= models.Q(
                        applications__stage_states__stage__name__icontains=kw,
                        applications__stage_states__score__gte=val,
                        applications__stage_states__is_deleted=False,
                        applications__is_deleted=False
                    )
                queryset = queryset.filter(q_obj).distinct()
            except ValueError:
                pass

        min_interview = self.request.GET.get('min_interview_score')
        if min_interview:
            try:
                val = float(min_interview)
                intv_kws = ["مصاحبه", "interview"]
                q_obj = models.Q()
                for kw in intv_kws:
                    q_obj |= models.Q(
                        applications__stage_states__stage__name__icontains=kw,
                        applications__stage_states__score__gte=val,
                        applications__stage_states__is_deleted=False,
                        applications__is_deleted=False
                    )
                queryset = queryset.filter(q_obj).distinct()
            except ValueError:
                pass

        min_assessment = self.request.GET.get('min_assessment_score')
        if min_assessment:
            try:
                val = float(min_assessment)
                asmt_kws = ["کانون", "ارزیابی", "assessment", "سنتر", "competency"]
                q_obj = models.Q()
                for kw in asmt_kws:
                    q_obj |= models.Q(
                        applications__stage_states__stage__name__icontains=kw,
                        applications__stage_states__score__gte=val,
                        applications__stage_states__is_deleted=False,
                        applications__is_deleted=False
                    )
                queryset = queryset.filter(q_obj).distinct()
            except ValueError:
                pass

        # 7. Applied Job Opportunity Filter
        applied_job_id = self.request.GET.get('applied_job')
        if applied_job_id:
            queryset = queryset.filter(
                applications__job_id=applied_job_id,
                applications__is_deleted=False
            ).distinct()

        # 8. Similarity Match / Suitability Recommendation for a target Job
        similar_to_job_id = self.request.GET.get('similar_to_job')
        if similar_to_job_id:
            similar_job = JobOpportunity.objects.filter(id=similar_to_job_id, is_deleted=False).first()
            if similar_job:
                # Calculate match scores and filter
                job_keywords = self.extract_keywords(similar_job.title) + self.extract_keywords(similar_job.department)
                candidates_list = list(queryset)
                filtered_candidates = []
                for candidate in candidates_list:
                    match_score, matched_reasons = self.compute_job_match(candidate, job_keywords)
                    if match_score > 0:
                        candidate.match_score = match_score
                        candidate.match_reasons = matched_reasons
                        filtered_candidates.append(candidate)
                # Sort by match score descending
                filtered_candidates.sort(key=lambda c: c.match_score, reverse=True)
                return filtered_candidates

        return queryset

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        from datetime import date
        
        # Calculate dashboard analytics
        all_candidates = Candidate.objects.filter(is_deleted=False)
        total_candidates = all_candidates.count()
        data['total_candidates'] = total_candidates
        
        # Higher Education % (Master / PhD)
        higher_edu_count = all_candidates.filter(
            education__degree__in=['MASTER', 'PHD'],
            education__is_deleted=False
        ).distinct().count()
        data['higher_edu_percentage'] = round((higher_edu_count / total_candidates) * 100, 1) if total_candidates > 0 else 0.0
        
        # Average Experience Years
        total_exp_days = 0
        candidates_with_exp = all_candidates.prefetch_related('experience')
        for candidate in candidates_with_exp:
            for exp in candidate.experience.filter(is_deleted=False):
                start = exp.start_date
                end = exp.end_date or date.today()
                if start:
                    total_exp_days += (end - start).days
        data['average_experience_years'] = round((total_exp_days / 365.25) / total_candidates, 1) if total_candidates > 0 else 0.0

        # Pass active job opportunities for assignment dropdown (fixed query)
        data['active_jobs'] = JobOpportunity.objects.filter(is_deleted=False).exclude(status__in=['CLOSED', 'CANCELLED', 'SUSPENDED']).order_by('-created_at')
        
        # Pass all jobs for search/filters
        data['all_jobs'] = JobOpportunity.objects.filter(is_deleted=False).order_by('-created_at')

        # Preserve active filters in context
        data['selected_degree'] = self.request.GET.get('degree', '')
        data['selected_major'] = self.request.GET.get('major', '')
        data['selected_skill'] = self.request.GET.get('skill', '')
        data['selected_min_experience'] = self.request.GET.get('min_experience', '')
        data['selected_min_exam_score'] = self.request.GET.get('min_exam_score', '')
        data['selected_min_interview_score'] = self.request.GET.get('min_interview_score', '')
        data['selected_min_assessment_score'] = self.request.GET.get('min_assessment_score', '')
        data['selected_applied_job'] = self.request.GET.get('applied_job', '')
        data['selected_similar_to_job'] = self.request.GET.get('similar_to_job', '')
        
        return data


class AssignCandidateToJobView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def post(self, request, candidate_id):
        candidate = get_object_or_404(Candidate, pk=candidate_id, is_deleted=False)
        job_id = request.POST.get('job_id')
        if not job_id:
            return HttpResponse("شناسه فرصت شغلی الزامی است.", status=400)
            
        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
        
        # Check if already exists
        exists = JobApplication.objects.filter(job=job, candidate=candidate, is_deleted=False).exists()
        if exists:
            return HttpResponse("این متقاضی قبلاً برای این فرصت شغلی ثبت‌نام شده است.", status=400)
            
        with transaction.atomic():
            JobApplication.objects.create(job=job, candidate=candidate)
            
        view_param = request.GET.get('view')
        redirect_url = reverse('candidate_list')
        if view_param:
            redirect_url += f'?view={view_param}'

        if request.headers.get('HX-Request'):
            response = HttpResponse("موفقیت‌آمیز")
            response['HX-Redirect'] = redirect_url
            return response
            
        return redirect(redirect_url)


class CandidateDetailView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    model = Candidate
    template_name = 'candidates/candidate_detail.html'
    context_object_name = 'candidate'
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        # دریافت سوابق تحصیلی و کاری فعال
        data['education_list'] = self.object.education.filter(is_deleted=False)
        data['experience_list'] = self.object.experience.filter(is_deleted=False)
        data['language_list'] = self.object.languages.filter(is_deleted=False)
        data['skill_list'] = self.object.skills.filter(is_deleted=False)
        data['certificate_list'] = self.object.certificates.filter(is_deleted=False)
        # دریافت لیست درخواست‌ها به همراه مراحلشان
        data['applications'] = self.object.applications.filter(is_deleted=False).select_related('job', 'current_stage')
        return data


class CandidateCreateView(LoginRequiredMixin, RoleRequiredMixin, CreateView):
    model = Candidate
    form_class = CandidateForm
    template_name = 'candidates/candidate_form.html'
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get_success_url(self):
        view_param = self.request.GET.get('view')
        if view_param:
            return reverse('candidate_list') + f'?view={view_param}'
        return reverse('candidate_list')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['education_formset'] = CandidateEducationFormSet(self.request.POST)
            data['experience_formset'] = CandidateExperienceFormSet(self.request.POST)
        else:
            data['education_formset'] = CandidateEducationFormSet()
            data['experience_formset'] = CandidateExperienceFormSet()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        education_formset = context['education_formset']
        experience_formset = context['experience_formset']
        
        if education_formset.is_valid() and experience_formset.is_valid():
            with transaction.atomic():
                self.object = form.save()
                
                education_formset.instance = self.object
                education_formset.save()
                
                experience_formset.instance = self.object
                experience_formset.save()
                
            return redirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))


class CandidateUpdateView(LoginRequiredMixin, RoleRequiredMixin, UpdateView):
    model = Candidate
    form_class = CandidateForm
    template_name = 'candidates/candidate_form.html'
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['education_formset'] = CandidateEducationFormSet(self.request.POST, instance=self.object)
            data['experience_formset'] = CandidateExperienceFormSet(self.request.POST, instance=self.object)
        else:
            data['education_formset'] = CandidateEducationFormSet(
                instance=self.object,
                queryset=CandidateEducation.objects.filter(is_deleted=False).order_by('-graduation_year')
            )
            data['experience_formset'] = CandidateExperienceFormSet(
                instance=self.object,
                queryset=CandidateExperience.objects.filter(is_deleted=False).order_by('-start_date')
            )
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        education_formset = context['education_formset']
        experience_formset = context['experience_formset']
        
        if education_formset.is_valid() and experience_formset.is_valid():
            with transaction.atomic():
                self.object = form.save()
                
                education_formset.instance = self.object
                education_formset.save()
                
                experience_formset.instance = self.object
                experience_formset.save()
                
            return redirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))

    def get_success_url(self):
        return reverse('candidate_detail', kwargs={'pk': self.object.pk})


class CandidateDeleteView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def delete(self, request, pk):
        candidate = get_object_or_404(Candidate, pk=pk)
        
        with transaction.atomic():
            # حذف نرم متقاضی
            candidate.delete()
            # حذف نرم درخواست‌ها و سوابق وی
            candidate.education.all().delete()
            candidate.experience.all().delete()
            candidate.applications.all().delete()
            candidate.languages.all().delete()
            candidate.skills.all().delete()
            candidate.certificates.all().delete()
            
        return HttpResponse("")


class JobApplicationCreateView(LoginRequiredMixin, RoleRequiredMixin, CreateView):
    model = JobApplication
    form_class = JobApplicationForm
    template_name = 'candidates/application_form.html'
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        data['candidate'] = get_object_or_404(Candidate, pk=self.kwargs.get('candidate_id'))
        return data

    def form_valid(self, form):
        candidate = get_object_or_404(Candidate, pk=self.kwargs.get('candidate_id'))
        form.instance.candidate = candidate
        
        # جلوگیری از درخواست تکراری برای یک شغل
        if JobApplication.objects.filter(job=form.instance.job, candidate=candidate, is_deleted=False).exists():
            form.add_error('job', 'این متقاضی قبلاً برای این فرصت شغلی ثبت‌نام شده است.')
            return self.form_invalid(form)
            
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('candidate_detail', kwargs={'pk': self.kwargs.get('candidate_id')})


from apps.jobs.models import JobOpportunity, JobOpportunityStage
from .models import ApplicationStageState

class JobOpportunityPipelineView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    model = JobOpportunity
    template_name = 'jobs/job_pipeline.html'
    context_object_name = 'job'
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        stages = self.object.stages.filter(is_deleted=False).prefetch_related('interviewers__user').order_by('sequence')
        data['stages'] = stages
        data['table_stages'] = stages.filter(models.Q(weight__gt=0) | models.Q(stage_type='SCREENING')).order_by('sequence')
        
        stages_list = list(stages)
        first_stage = stages_list[0] if stages_list and ('غربالگری' in stages_list[0].name or stages_list[0].stage_type == 'SCREENING') else None
        
        apps_qs = self.object.applications.filter(is_deleted=False).select_related(
            'candidate', 'job__recruitment_plan'
        ).prefetch_related(
            'stage_states__stage', 'job__recruitment_plan__stage_plans'
        )
        
        def get_app_priority(app):
            eff_status = app.effective_status
            if eff_status == 'SELECTED':
                return 0
            elif eff_status in ['IN_PROGRESS', 'RESERVE']:
                return 1
            elif eff_status == 'REJECTED':
                if first_stage:
                    for state in app.stage_states.all():
                        if state.stage_id == first_stage.id:
                            if state.status == 'FAILED':
                                return 3
                            break
                return 2
            return 4

        apps_list = list(apps_qs)
        for app in apps_list:
            for state in app.stage_states.all():
                state.application = app
        apps_list.sort(key=lambda app: (get_app_priority(app), -app.final_score, -app.id))
        
        data['applications'] = apps_list
        data['status_choices'] = JobApplication.STATUS_CHOICES
        
        # Fetch stage plans
        plan = getattr(self.object, 'recruitment_plan', None)
        stage_plans = {}
        if plan and not plan.is_deleted:
            for sp in plan.stage_plans.filter(is_deleted=False):
                stage_plans[sp.stage_id] = sp
        data['stage_plans'] = stage_plans

        # Calculate gaps (days) between consecutive stages based on planned dates
        stage_gaps = {}
        stages_list = list(stages)
        for i in range(len(stages_list) - 1):
            curr_stage = stages_list[i]
            next_stage = stages_list[i+1]
            sp_curr = stage_plans.get(curr_stage.id)
            sp_next = stage_plans.get(next_stage.id)
            if sp_curr and sp_next:
                gap = (sp_next.planned_start_date - sp_curr.planned_end_date).days
                stage_gaps[curr_stage.id] = gap
        data['stage_gaps'] = stage_gaps

        return data


class ToggleStageCompletionView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def post(self, request, stage_id):
        stage = get_object_or_404(JobOpportunityStage, pk=stage_id, is_deleted=False)
        status = request.GET.get('status')
        if status == 'completed':
            stage.is_manually_completed = True
        elif status == 'pending':
            stage.is_manually_completed = False
        elif status == 'auto':
            stage.is_manually_completed = None
        stage.save()
        
        response = HttpResponse()
        response['HX-Refresh'] = 'true'
        return response


class EditApplicationStageStateView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get(self, request, pk):
        state = get_object_or_404(ApplicationStageState, pk=pk)
        if not state.is_accessible and not (request.GET.get('bypass_locks') == 'true'):
            return HttpResponse("مرحله ارزیابی برای این متقاضی هنوز قابل دسترسی نیست.", status=400)
        return render(request, 'candidates/partials/stage_state_edit.html', {'state': state})


class UpdateApplicationStageStateView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def post(self, request, pk):
        state = get_object_or_404(ApplicationStageState, pk=pk)
        
        score_val = request.POST.get('score', '0')
        status_val = request.POST.get('status', ApplicationStageState.STATUS_PENDING)
        notes_val = request.POST.get('notes', '')
        is_conditional_pass_val = request.POST.get('is_conditional_pass') == 'true' or request.POST.get('is_conditional_pass') == 'on'
        date_val = request.POST.get('date', '').strip()
        time_val = request.POST.get('time', '10:00').strip()

        try:
            state.score = float(score_val)
        except ValueError:
            state.score = 0.0

        state.status = status_val
        state.notes = notes_val
        state.is_conditional_pass = is_conditional_pass_val
        state.evaluator = request.user

        if date_val:
            state.evaluation_date = parse_jalali_date(date_val)
        else:
            state.evaluation_date = None

        state.evaluation_time = time_val or "10:00"
        state._bypass_stage_score_calculation = True
        state.is_manually_edited = True
        state.save()

        source = request.POST.get('source')
        if source == 'score_entry':
            # Get the new active stage state for the application if it advanced
            app = state.application
            new_active_state = ApplicationStageState.objects.filter(
                application=app,
                stage=app.current_stage,
                is_deleted=False
            ).first()
            return render(request, 'candidates/partials/score_entry_row.html', {
                'state': new_active_state or state,
                'just_saved': True,
                'saved_stage_name': state.stage.name,
                'saved_status': state.get_status_display(),
                'saved_score': state.score,
            })

        # Render read-only view cell along with out-of-band update for final score
        return render(request, 'candidates/partials/stage_state_view.html', {
            'state': state,
            'application': state.application,
            'oob_update': True
        })


class ViewApplicationStageStateView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get(self, request, pk):
        state = get_object_or_404(ApplicationStageState, pk=pk)
        return render(request, 'candidates/partials/stage_state_view.html', {'state': state})


class UpdateApplicationStatusView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def post(self, request, pk):
        application = get_object_or_404(JobApplication, pk=pk)
        status_val = request.POST.get('status')
        if status_val in dict(JobApplication.STATUS_CHOICES):
            application.status = status_val
            application.save(update_fields=['status'])
            application.job.update_status()
            
        badge_class = 'bg-warning text-warning'
        border_class = 'border-warning'
        if application.status == JobApplication.STATUS_SELECTED:
            badge_class = 'bg-success text-success'
            border_class = 'border-success'
        elif application.status == JobApplication.STATUS_RESERVE:
            badge_class = 'bg-info text-info'
            border_class = 'border-info'
        elif application.status == JobApplication.STATUS_REJECTED:
            badge_class = 'bg-danger text-danger'
            border_class = 'border-danger'

        html = f'<span class="badge {badge_class} bg-opacity-10 border {border_class} border-opacity-25 px-2 py-1.5 rounded text-xs">{application.get_status_display()}</span>'
        return HttpResponse(html)


class CareersListView(ListView):
    model = JobOpportunity
    template_name = 'candidates/public_job_list.html'
    context_object_name = 'jobs'

    def get_queryset(self):
        from datetime import date
        from django.db.models import Q
        return JobOpportunity.objects.filter(
            status=JobOpportunity.STATUS_PUBLISHED,
            is_deleted=False
        ).filter(
            Q(end_date__isnull=True) | Q(end_date__gte=date.today())
        ).prefetch_related('stages')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        # Group by department
        jobs = self.get_queryset()
        grouped_jobs = {}
        for job in jobs:
            dept = job.department or 'سایر دپارتمان‌ها'
            if dept not in grouped_jobs:
                grouped_jobs[dept] = []
            grouped_jobs[dept].append(job)
        data['grouped_jobs'] = grouped_jobs
        return data


class CareersDetailAndApplyView(View):
    def get(self, request, pk):
        from datetime import date
        job = get_object_or_404(JobOpportunity, pk=pk, status=JobOpportunity.STATUS_PUBLISHED, is_deleted=False)
        if job.end_date and job.end_date < date.today():
            return render(request, 'candidates/public_job_detail.html', {
                'job': job,
                'registration_closed': True,
            })
        form = CandidateForm()
        has_applied = False
        candidate = None
        if request.user.is_authenticated and hasattr(request.user, 'profile') and request.user.profile.role == UserProfile.ROLE_CANDIDATE:
            candidate = getattr(request.user, 'candidate_profile', None)
            if candidate:
                has_applied = JobApplication.objects.filter(job=job, candidate=candidate, is_deleted=False).exists()
        return render(request, 'candidates/public_job_detail.html', {
            'job': job,
            'form': form,
            'has_applied': has_applied,
            'candidate': candidate,
        })

    def post(self, request, pk):
        from datetime import date
        job = get_object_or_404(JobOpportunity, pk=pk, status=JobOpportunity.STATUS_PUBLISHED, is_deleted=False)
        if job.end_date and job.end_date < date.today():
            return render(request, 'candidates/public_job_detail.html', {
                'job': job,
                'registration_closed': True,
            })
        
        national_id = request.POST.get('national_id', '').strip()

        candidate = None
        if national_id:
            candidate = Candidate.objects.filter(national_id=national_id, is_deleted=False).first()

        if candidate:
            form = CandidateForm(request.POST, request.FILES, instance=candidate)
        else:
            form = CandidateForm(request.POST, request.FILES)

        if form.is_valid():
            with transaction.atomic():
                candidate = form.save()

                # Check if this candidate already applied for this job
                existing_app = JobApplication.all_objects.filter(job=job, candidate=candidate).first()
                if existing_app:
                    if not existing_app.is_deleted:
                        form.add_error(None, 'شما قبلاً برای این فرصت شغلی ثبت‌نام کرده‌اید.')
                        return render(request, 'candidates/public_job_detail.html', {
                            'job': job,
                            'form': form
                        })
                    else:
                        # Restore soft-deleted application
                        existing_app.is_deleted = False
                        existing_app.deleted_at = None
                        existing_app.status = JobApplication.STATUS_IN_PROGRESS
                        existing_app.save()
                else:
                    # Create JobApplication
                    JobApplication.objects.create(
                        job=job,
                        candidate=candidate,
                        status=JobApplication.STATUS_IN_PROGRESS
                    )
            
            return redirect('careers_success')
        else:
            return render(request, 'candidates/public_job_detail.html', {
                'job': job,
                'form': form
            })


class CareersTrackView(View):
    def get(self, request):
        return render(request, 'candidates/public_track.html')

    def post(self, request):
        national_id = request.POST.get('national_id', '').strip()
        phone_number = request.POST.get('phone_number', '').strip()

        candidate = Candidate.objects.filter(
            national_id=national_id,
            phone_number=phone_number,
            is_deleted=False
        ).prefetch_related('applications__job', 'applications__current_stage').first()

        if not candidate:
            return render(request, 'candidates/public_track.html', {
                'error': 'متقاضی با مشخصات فوق در سیستم یافت نشد. لطفاً اطلاعات وارد شده را مجدداً بررسی نمایید.',
                'national_id': national_id,
                'phone_number': phone_number
            })

        applications = candidate.applications.filter(is_deleted=False)
        return render(request, 'candidates/public_track.html', {
            'candidate': candidate,
            'applications': applications,
            'national_id': national_id,
            'phone_number': phone_number
        })


class CareersSuccessView(TemplateView):
    template_name = 'candidates/public_success.html'


from django.contrib.auth import login
from django.contrib.auth.models import User
from .forms import CandidateSignUpForm

class CandidateSignUpView(View):
    def get(self, request):
        if request.user.is_authenticated:
            return redirect('candidate_dashboard')
        form = CandidateSignUpForm()
        education_formset = CandidateEducationFormSet(prefix='edu')
        experience_formset = CandidateExperienceFormSet(prefix='exp')
        language_formset = CandidateLanguageFormSet(prefix='lang')
        skill_formset = CandidateSkillFormSet(prefix='skill')
        certificate_formset = CandidateCertificateFormSet(prefix='cert')
        
        return render(request, 'candidates/signup.html', {
            'form': form,
            'education_formset': education_formset,
            'experience_formset': experience_formset,
            'language_formset': language_formset,
            'skill_formset': skill_formset,
            'certificate_formset': certificate_formset,
        })

    def post(self, request):
        if request.user.is_authenticated:
            return redirect('candidate_dashboard')
        form = CandidateSignUpForm(request.POST, request.FILES)
        education_formset = CandidateEducationFormSet(request.POST, prefix='edu')
        experience_formset = CandidateExperienceFormSet(request.POST, prefix='exp')
        language_formset = CandidateLanguageFormSet(request.POST, prefix='lang')
        skill_formset = CandidateSkillFormSet(request.POST, prefix='skill')
        certificate_formset = CandidateCertificateFormSet(request.POST, prefix='cert')

        if (form.is_valid() and education_formset.is_valid() and 
            experience_formset.is_valid() and language_formset.is_valid() and 
            skill_formset.is_valid() and certificate_formset.is_valid()):
            
            with transaction.atomic():
                first_name = form.cleaned_data['first_name']
                last_name = form.cleaned_data['last_name']
                email = form.cleaned_data['email']
                phone_number = form.cleaned_data['phone_number']
                national_id = form.cleaned_data['national_id']
                personnel_number = form.cleaned_data['personnel_number']
                password = form.cleaned_data['password']

                # Create Django User
                user = User.objects.create_user(
                    username=national_id,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name
                )
                
                # Update UserProfile role to CANDIDATE
                profile = user.profile
                profile.role = UserProfile.ROLE_CANDIDATE
                profile.phone_number = phone_number
                profile.save()

                # Link or create Candidate record
                candidate = Candidate.all_objects.filter(national_id=national_id).first()
                if candidate:
                    candidate.is_deleted = False
                    candidate.deleted_at = None
                    candidate.user = user
                    candidate.first_name = first_name
                    candidate.last_name = last_name
                    candidate.email = email
                    candidate.phone_number = phone_number
                    candidate.personnel_number = personnel_number or None
                    if 'resume' in request.FILES:
                        candidate.resume = request.FILES['resume']
                    candidate.save()
                else:
                    candidate = Candidate.objects.create(
                        user=user,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        phone_number=phone_number,
                        national_id=national_id,
                        personnel_number=personnel_number or None,
                        resume=request.FILES.get('resume')
                    )

                # Save formsets linked to candidate
                education_formset.instance = candidate
                education_formset.save()
                
                experience_formset.instance = candidate
                experience_formset.save()
                
                language_formset.instance = candidate
                language_formset.save()
                
                skill_formset.instance = candidate
                skill_formset.save()
                
                certificate_formset.instance = candidate
                certificate_formset.save()
            
            # Log the user in
            login(request, user)
            return redirect('candidate_dashboard')
        
        return render(request, 'candidates/signup.html', {
            'form': form,
            'education_formset': education_formset,
            'experience_formset': experience_formset,
            'language_formset': language_formset,
            'skill_formset': skill_formset,
            'certificate_formset': certificate_formset,
        })


class CandidateDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'candidates/dashboard.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not hasattr(request.user, 'profile') or request.user.profile.role != UserProfile.ROLE_CANDIDATE:
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        # Find or create linked Candidate profile
        candidate, created = Candidate.objects.get_or_create(
            user=self.request.user,
            defaults={
                'first_name': self.request.user.first_name,
                'last_name': self.request.user.last_name,
                'email': self.request.user.email,
                'phone_number': self.request.user.profile.phone_number,
                'national_id': self.request.user.username,
            }
        )
        data['candidate'] = candidate
        
        # Load published jobs grouped by department
        jobs = JobOpportunity.objects.filter(
            status=JobOpportunity.STATUS_PUBLISHED,
            is_deleted=False
        ).prefetch_related('stages')
        
        grouped_jobs = {}
        for job in jobs:
            dept = job.department or 'سایر دپارتمان‌ها'
            if dept not in grouped_jobs:
                grouped_jobs[dept] = []
            grouped_jobs[dept].append(job)
        data['grouped_jobs'] = grouped_jobs
        
        # Candidate's applications
        data['applications'] = candidate.applications.filter(is_deleted=False).select_related('job', 'current_stage').prefetch_related('stage_states__stage')
        
        # Already applied job ids
        data['applied_job_ids'] = list(candidate.applications.filter(is_deleted=False).values_list('job_id', flat=True))
        
        # Profile & CV forms
        data['profile_form'] = CandidateForm(instance=candidate)
        data['education_formset'] = CandidateEducationFormSet(instance=candidate, prefix='edu')
        data['experience_formset'] = CandidateExperienceFormSet(instance=candidate, prefix='exp')
        data['language_formset'] = CandidateLanguageFormSet(instance=candidate, prefix='lang')
        data['skill_formset'] = CandidateSkillFormSet(instance=candidate, prefix='skill')
        data['certificate_formset'] = CandidateCertificateFormSet(instance=candidate, prefix='cert')
        
        return data


class CandidateProfileUpdateView(LoginRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not hasattr(request.user, 'profile') or request.user.profile.role != UserProfile.ROLE_CANDIDATE:
            return HttpResponse(status=403)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        candidate = get_object_or_404(Candidate, user=request.user, is_deleted=False)
        form = CandidateForm(request.POST, request.FILES, instance=candidate)
        if form.is_valid():
            form.save()
            return redirect('candidate_dashboard')
        
        dashboard_view = CandidateDashboardView()
        dashboard_view.request = request
        context = dashboard_view.get_context_data()
        context['profile_form'] = form
        context['active_tab'] = 'profile'
        return render(request, 'candidates/dashboard.html', context)


class CandidateEducationUpdateView(LoginRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not hasattr(request.user, 'profile') or request.user.profile.role != UserProfile.ROLE_CANDIDATE:
            return HttpResponse(status=403)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        candidate = get_object_or_404(Candidate, user=request.user, is_deleted=False)
        formset = CandidateEducationFormSet(request.POST, instance=candidate, prefix='edu')
        if formset.is_valid():
            formset.save()
            return redirect('candidate_dashboard')
        
        dashboard_view = CandidateDashboardView()
        dashboard_view.request = request
        context = dashboard_view.get_context_data()
        context['education_formset'] = formset
        context['active_tab'] = 'profile'
        return render(request, 'candidates/dashboard.html', context)


class CandidateExperienceUpdateView(LoginRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not hasattr(request.user, 'profile') or request.user.profile.role != UserProfile.ROLE_CANDIDATE:
            return HttpResponse(status=403)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        candidate = get_object_or_404(Candidate, user=request.user, is_deleted=False)
        formset = CandidateExperienceFormSet(request.POST, instance=candidate, prefix='exp')
        if formset.is_valid():
            formset.save()
            return redirect('candidate_dashboard')
        
        dashboard_view = CandidateDashboardView()
        dashboard_view.request = request
        context = dashboard_view.get_context_data()
        context['experience_formset'] = formset
        context['active_tab'] = 'profile'
        return render(request, 'candidates/dashboard.html', context)


class CandidateApplyDirectView(LoginRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not hasattr(request.user, 'profile') or request.user.profile.role != UserProfile.ROLE_CANDIDATE:
            return HttpResponse(status=403)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, pk):
        from datetime import date
        job = get_object_or_404(JobOpportunity, pk=pk, status=JobOpportunity.STATUS_PUBLISHED, is_deleted=False)
        if job.end_date and job.end_date < date.today():
            return HttpResponse("مهلت ثبت‌نام در این فرصت شغلی به پایان رسیده است.", status=400)
            
        candidate = get_object_or_404(Candidate, user=request.user, is_deleted=False)
        
        # Check if already applied
        existing_app = JobApplication.all_objects.filter(job=job, candidate=candidate).first()
        if existing_app:
            if existing_app.is_deleted:
                existing_app.is_deleted = False
                existing_app.deleted_at = None
                existing_app.status = JobApplication.STATUS_IN_PROGRESS
                existing_app.save()
        else:
            JobApplication.objects.create(
                job=job,
                candidate=candidate,
                status=JobApplication.STATUS_IN_PROGRESS
            )
            
        return redirect('candidate_dashboard')


class CandidateLanguageUpdateView(LoginRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not hasattr(request.user, 'profile') or request.user.profile.role != UserProfile.ROLE_CANDIDATE:
            return HttpResponse(status=403)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        candidate = get_object_or_404(Candidate, user=request.user, is_deleted=False)
        formset = CandidateLanguageFormSet(request.POST, instance=candidate, prefix='lang')
        if formset.is_valid():
            formset.save()
            return redirect('candidate_dashboard')
        
        dashboard_view = CandidateDashboardView()
        dashboard_view.request = request
        context = dashboard_view.get_context_data()
        context['language_formset'] = formset
        context['active_tab'] = 'profile'
        return render(request, 'candidates/dashboard.html', context)


class CandidateSkillUpdateView(LoginRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not hasattr(request.user, 'profile') or request.user.profile.role != UserProfile.ROLE_CANDIDATE:
            return HttpResponse(status=403)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        candidate = get_object_or_404(Candidate, user=request.user, is_deleted=False)
        formset = CandidateSkillFormSet(request.POST, instance=candidate, prefix='skill')
        if formset.is_valid():
            formset.save()
            return redirect('candidate_dashboard')
        
        dashboard_view = CandidateDashboardView()
        dashboard_view.request = request
        context = dashboard_view.get_context_data()
        context['skill_formset'] = formset
        context['active_tab'] = 'profile'
        return render(request, 'candidates/dashboard.html', context)


class CandidateCertificateUpdateView(LoginRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not hasattr(request.user, 'profile') or request.user.profile.role != UserProfile.ROLE_CANDIDATE:
            return HttpResponse(status=403)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request):
        candidate = get_object_or_404(Candidate, user=request.user, is_deleted=False)
        formset = CandidateCertificateFormSet(request.POST, instance=candidate, prefix='cert')
        if formset.is_valid():
            formset.save()
            return redirect('candidate_dashboard')
        
        dashboard_view = CandidateDashboardView()
        dashboard_view.request = request
        context = dashboard_view.get_context_data()
        context['certificate_formset'] = formset
        context['active_tab'] = 'profile'
        return render(request, 'candidates/dashboard.html', context)
def parse_jalali_date(val):
    if not val:
        return None
    try:
        import jdatetime
        val_str = str(val).strip()
        for fa, en in zip('۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩', '01234567890123456789'):
            val_str = val_str.replace(fa, en)
        parts = [int(p) for p in val_str.split('/')]
        if len(parts) == 3:
            return jdatetime.date(parts[0], parts[1], parts[2]).togregorian()
    except Exception:
        pass
    return None


class ScoreEntryListView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_SPECIALIST, UserProfile.ROLE_RECRUITMENT_DIRECTOR]

    def get(self, request):
        from django.core.paginator import Paginator
        from django.db.models import Q as DQ
        job_q = request.GET.get('job_q', '').strip()
        jobs_qs = JobOpportunity.objects.filter(is_deleted=False).exclude(
            status__in=[JobOpportunity.STATUS_CLOSED, JobOpportunity.STATUS_CANCELLED, JobOpportunity.STATUS_SUSPENDED]
        )
        if job_q:
            jobs_qs = jobs_qs.filter(
                DQ(title__icontains=job_q) |
                DQ(code__icontains=job_q) |
                DQ(request_number__icontains=job_q)
            )
        
        job_id = request.GET.get('job_id')
        selected_job = None
        if job_id:
            selected_job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)

        jobs = list(jobs_qs.order_by('code'))
        if selected_job and selected_job not in jobs:
            jobs.insert(0, selected_job)

        is_job_closed_or_cancelled = False
        if selected_job:
            is_job_closed_or_cancelled = selected_job.status in [JobOpportunity.STATUS_CLOSED, JobOpportunity.STATUS_CANCELLED, JobOpportunity.STATUS_SUSPENDED]

        stage_id = request.GET.get('stage_id')
        selected_stage = None
        q = request.GET.get('q')
        eval_status = request.GET.get('eval_status', 'PENDING')
        show_failed_prior_val = request.GET.get('show_failed_prior')
        show_failed_prior = show_failed_prior_val in ['true', 'on', '1']
        bypass_locks_val = request.GET.get('bypass_locks') or request.POST.get('bypass_locks')
        bypass_locks = bypass_locks_val in ['true', 'on', '1']

        stages = []
        applications = []
        page_obj = None
        paginator = None
        is_paginated = False
        pending_states = []

        if selected_job:
            stages = list(selected_job.stages.filter(is_deleted=False).order_by('sequence'))
            
            if stage_id:
                from django.db.models import Exists, OuterRef
                selected_stage = get_object_or_404(JobOpportunityStage, pk=stage_id, is_deleted=False)
                app_statuses = [JobApplication.STATUS_IN_PROGRESS]
                if show_failed_prior or bypass_locks:
                    app_statuses.append(JobApplication.STATUS_REJECTED)
                pending_states_qs = ApplicationStageState.objects.filter(
                    application__job=selected_job,
                    application__status__in=app_statuses,
                    application__is_deleted=False,
                    stage=selected_stage,
                    is_deleted=False
                ).select_related('application__candidate', 'application__job', 'stage', 'evaluator')
                
                prior_failed_subquery = ApplicationStageState.objects.filter(
                    application=OuterRef('application'),
                    stage__sequence__lt=OuterRef('stage__sequence'),
                    status=ApplicationStageState.STATUS_FAILED,
                    is_conditional_pass=False,
                    is_deleted=False
                )
                pending_states_qs = pending_states_qs.annotate(
                    has_failed_prior=Exists(prior_failed_subquery)
                )
                
                if not show_failed_prior:
                    pending_states_qs = pending_states_qs.filter(has_failed_prior=False)
                
                if eval_status == 'PENDING':
                    pending_states_qs = pending_states_qs.filter(status=ApplicationStageState.STATUS_PENDING)
                elif eval_status == 'COMPLETED_FAILED':
                    pending_states_qs = pending_states_qs.filter(status__in=[ApplicationStageState.STATUS_COMPLETED, ApplicationStageState.STATUS_FAILED])
                
                if q:
                    pending_states_qs = pending_states_qs.filter(
                        models.Q(application__candidate__first_name__icontains=q) |
                        models.Q(application__candidate__last_name__icontains=q) |
                        models.Q(application__candidate__national_id__icontains=q)
                    )
                    
                ordered_states = pending_states_qs.order_by('application__candidate__last_name')
                paginator = Paginator(ordered_states, 10)
                page_number = request.GET.get('page')
                page_obj = paginator.get_page(page_number)
                pending_states = list(page_obj)
                is_paginated = page_obj.has_other_pages()
                
                applications = [state.application for state in pending_states]
            else:
                app_statuses = [JobApplication.STATUS_IN_PROGRESS]
                if show_failed_prior or bypass_locks:
                    app_statuses.append(JobApplication.STATUS_REJECTED)
                    
                applications_qs = JobApplication.objects.filter(
                    job=selected_job,
                    status__in=app_statuses,
                    is_deleted=False
                ).select_related('candidate').prefetch_related('stage_states__stage')
                
                if q:
                    applications_qs = applications_qs.filter(
                        models.Q(candidate__first_name__icontains=q) |
                        models.Q(candidate__last_name__icontains=q) |
                        models.Q(candidate__national_id__icontains=q)
                    )
                
                if eval_status == 'PENDING':
                    applications_qs = applications_qs.filter(status=JobApplication.STATUS_IN_PROGRESS)
                elif eval_status == 'COMPLETED_FAILED':
                    applications_qs = applications_qs.filter(status__in=[JobApplication.STATUS_SELECTED, JobApplication.STATUS_RESERVE, JobApplication.STATUS_REJECTED])

                applications_qs = applications_qs.order_by('candidate__last_name')
                paginator = Paginator(applications_qs, 10)
                page_number = request.GET.get('page')
                page_obj = paginator.get_page(page_number)
                is_paginated = page_obj.has_other_pages()
                
                applications = list(page_obj)
                pending_states = applications
            
            for app in applications:
                state_map = {state.stage_id: state for state in app.stage_states.filter(is_deleted=False)}
                app.stage_cells = []
                for stage in stages:
                    state = state_map.get(stage.id)
                    if not state:
                        state = ApplicationStageState.objects.create(
                            application=app,
                            stage=stage,
                            status=ApplicationStageState.STATUS_PENDING,
                            score=0.0
                        )
                    app.stage_cells.append(state)

        if request.headers.get('HX-Request'):
            return render(request, 'candidates/partials/score_entry_list.html', {
                'selected_job': selected_job,
                'stages': stages,
                'applications': applications,
                'pending_states': pending_states,
                'selected_q': q,
                'selected_eval_status': eval_status,
                'show_failed_prior': show_failed_prior,
                'bypass_locks': bypass_locks,
                'page_obj': page_obj,
                'is_paginated': is_paginated,
                'paginator': paginator,
                'is_job_closed_or_cancelled': is_job_closed_or_cancelled,
            })

        return render(request, 'candidates/score_entry.html', {
            'jobs': jobs,
            'job_q': job_q,
            'selected_job': selected_job,
            'stages': stages,
            'applications': applications,
            'pending_states': pending_states,
            'selected_q': q,
            'selected_eval_status': eval_status,
            'show_failed_prior': show_failed_prior,
            'bypass_locks': bypass_locks,
            'page_obj': page_obj,
            'is_paginated': is_paginated,
            'paginator': paginator,
            'is_job_closed_or_cancelled': is_job_closed_or_cancelled,
        })

    def post(self, request):
        from django.core.paginator import Paginator
        job_id = request.POST.get('job_id')
        stage_id = request.POST.get('stage_id')
        q = request.POST.get('q')
        eval_status = request.POST.get('eval_status', 'PENDING')
        show_failed_prior_val = request.POST.get('show_failed_prior') or request.GET.get('show_failed_prior')
        show_failed_prior = show_failed_prior_val in ['true', 'on', '1']
        bypass_locks_val = request.POST.get('bypass_locks') or request.GET.get('bypass_locks')
        bypass_locks = bypass_locks_val in ['true', 'on', '1']
        
        selected_job = None
        stages = []
        applications = []
        page_obj = None
        paginator = None
        is_paginated = False
        pending_states = []

        jobs_qs = JobOpportunity.objects.filter(is_deleted=False).exclude(
            status__in=[JobOpportunity.STATUS_CLOSED, JobOpportunity.STATUS_CANCELLED, JobOpportunity.STATUS_SUSPENDED]
        )

        if job_id:
            selected_job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
            stages = list(selected_job.stages.filter(is_deleted=False).order_by('sequence'))
            is_job_closed_or_cancelled = selected_job.status in [JobOpportunity.STATUS_CLOSED, JobOpportunity.STATUS_CANCELLED, JobOpportunity.STATUS_SUSPENDED]

            jobs = list(jobs_qs.order_by('-created_at'))
            if selected_job and selected_job not in jobs:
                jobs.insert(0, selected_job)

            if not is_job_closed_or_cancelled:
                state_ids = set()
                for key in request.POST.keys():
                    if key.startswith('score_') or key.startswith('status_') or key.startswith('notes_') or key.startswith('date_') or key.startswith('is_conditional_pass_'):
                        try:
                            state_ids.add(int(key.split('_')[-1]))
                        except ValueError:
                            pass

                app_ids = set()
                for key in request.POST.keys():
                    if key.startswith('app_status_'):
                        try:
                            app_ids.add(int(key.split('_')[-1]))
                        except ValueError:
                            pass

                with transaction.atomic():
                    for aid in app_ids:
                        app = JobApplication.objects.filter(pk=aid, is_deleted=False).first()
                        if app:
                            status_val = request.POST.get(f'app_status_{aid}')
                            if status_val in [JobApplication.STATUS_IN_PROGRESS, JobApplication.STATUS_SELECTED, JobApplication.STATUS_RESERVE, JobApplication.STATUS_REJECTED]:
                                if app.status != status_val:
                                    app.status = status_val
                                    app.save(update_fields=['status'])

                    for sid in state_ids:
                        state = ApplicationStageState.objects.filter(pk=sid, is_deleted=False).first()
                        if state and (bypass_locks or (state.is_accessible and not state.stage.is_completed)):
                            status_val = request.POST.get(f'status_{sid}', ApplicationStageState.STATUS_PENDING)
                            notes_val = request.POST.get(f'notes_{sid}', '')
                            date_val = request.POST.get(f'date_{sid}', '').strip()
                            is_conditional_pass_val = request.POST.get(f'is_conditional_pass_{sid}') in ['true', 'on']
                            
                            if state.stage.stage_type != 'SCREENING':
                                score_val = request.POST.get(f'score_{sid}', '0')
                                try:
                                    score_float = float(score_val)
                                except ValueError:
                                    score_float = 0.0
                                
                                if state.stage.competencies.exists():
                                    if abs(state.score - score_float) > 0.01:
                                        state.score = score_float
                                        state.is_manually_edited = True
                                else:
                                    state.score = score_float
                                    state.is_manually_edited = True
                            elif state.stage.stage_type == 'SCREENING':
                                state.score = 0.0
                                
                            state.status = status_val
                            state.notes = notes_val
                            state.is_conditional_pass = is_conditional_pass_val
                            state.evaluator = request.user
                            
                            if date_val:
                                state.evaluation_date = parse_jalali_date(date_val)
                            else:
                                state.evaluation_date = None
                                
                            state._bypass_stage_score_calculation = True
                            state.is_manually_edited = True
                            state.save()

                if selected_job.has_all_evaluations_completed:
                    if selected_job.status not in [JobOpportunity.STATUS_CLOSED, JobOpportunity.STATUS_CANCELLED, JobOpportunity.STATUS_SUSPENDED]:
                        selected_job.status = JobOpportunity.STATUS_FINAL_SELECTION
                        selected_job.save(update_fields=['status'])
                    
                    from django.contrib import messages
                    messages.success(request, "ارزیابی تمامی مراحل داوطلبان با موفقیت به پایان رسید. اکنون می‌توانید نفرات برتر را انتخاب و نهایی کنید.")
                    
                    if request.headers.get('HX-Request'):
                        response = HttpResponse()
                        response['HX-Redirect'] = reverse('job_pipeline', kwargs={'pk': selected_job.id})
                        return response
                    return redirect('job_pipeline', pk=selected_job.id)

            if stage_id:
                from django.db.models import Exists, OuterRef
                selected_stage = get_object_or_404(JobOpportunityStage, pk=stage_id, is_deleted=False)
                app_statuses = [JobApplication.STATUS_IN_PROGRESS]
                if show_failed_prior or bypass_locks:
                    app_statuses.append(JobApplication.STATUS_REJECTED)
                pending_states_qs = ApplicationStageState.objects.filter(
                    application__job=selected_job,
                    application__status__in=app_statuses,
                    application__is_deleted=False,
                    stage=selected_stage,
                    is_deleted=False
                ).select_related('application__candidate', 'application__job', 'stage', 'evaluator')
                
                prior_failed_subquery = ApplicationStageState.objects.filter(
                    application=OuterRef('application'),
                    stage__sequence__lt=OuterRef('stage__sequence'),
                    status=ApplicationStageState.STATUS_FAILED,
                    is_conditional_pass=False,
                    is_deleted=False
                )
                pending_states_qs = pending_states_qs.annotate(
                    has_failed_prior=Exists(prior_failed_subquery)
                )
                
                if not show_failed_prior:
                    pending_states_qs = pending_states_qs.filter(has_failed_prior=False)
                
                if eval_status == 'PENDING':
                    pending_states_qs = pending_states_qs.filter(status=ApplicationStageState.STATUS_PENDING)
                elif eval_status == 'COMPLETED_FAILED':
                    pending_states_qs = pending_states_qs.filter(status__in=[ApplicationStageState.STATUS_COMPLETED, ApplicationStageState.STATUS_FAILED])
                
                if q:
                    pending_states_qs = pending_states_qs.filter(
                        models.Q(application__candidate__first_name__icontains=q) |
                        models.Q(application__candidate__last_name__icontains=q) |
                        models.Q(application__candidate__national_id__icontains=q)
                    )
                    
                ordered_states = pending_states_qs.order_by('application__candidate__last_name')
                paginator = Paginator(ordered_states, 10)
                page_number = request.GET.get('page') or request.POST.get('page')
                page_obj = paginator.get_page(page_number)
                pending_states = list(page_obj)
                is_paginated = page_obj.has_other_pages()
                
                applications = [state.application for state in pending_states]
            else:
                app_statuses = [JobApplication.STATUS_IN_PROGRESS]
                if show_failed_prior or bypass_locks:
                    app_statuses.append(JobApplication.STATUS_REJECTED)
                    
                applications_qs = JobApplication.objects.filter(
                    job=selected_job,
                    status__in=app_statuses,
                    is_deleted=False
                ).select_related('candidate').prefetch_related('stage_states__stage')
                
                if q:
                    applications_qs = applications_qs.filter(
                        models.Q(candidate__first_name__icontains=q) |
                        models.Q(candidate__last_name__icontains=q) |
                        models.Q(candidate__national_id__icontains=q)
                    )
                
                if eval_status == 'PENDING':
                    applications_qs = applications_qs.filter(status=JobApplication.STATUS_IN_PROGRESS)
                elif eval_status == 'COMPLETED_FAILED':
                    applications_qs = applications_qs.filter(status__in=[JobApplication.STATUS_SELECTED, JobApplication.STATUS_RESERVE, JobApplication.STATUS_REJECTED])

                applications_qs = applications_qs.order_by('candidate__last_name')
                paginator = Paginator(applications_qs, 10)
                page_number = request.GET.get('page') or request.POST.get('page')
                page_obj = paginator.get_page(page_number)
                is_paginated = page_obj.has_other_pages()
                
                applications = list(page_obj)
                pending_states = applications
                
            for app in applications:
                state_map = {state.stage_id: state for state in app.stage_states.filter(is_deleted=False)}
                app.stage_cells = []
                for stage in stages:
                    state = state_map.get(stage.id)
                    if not state:
                        state = ApplicationStageState.objects.create(
                            application=app,
                            stage=stage,
                            status=ApplicationStageState.STATUS_PENDING,
                            score=0.0
                        )
                    app.stage_cells.append(state)

        if request.headers.get('HX-Request'):
            return render(request, 'candidates/partials/score_entry_list.html', {
                'selected_job': selected_job,
                'stages': stages,
                'applications': applications,
                'pending_states': pending_states,
                'bulk_success': not is_job_closed_or_cancelled and job_id is not None,
                'selected_q': q,
                'selected_eval_status': eval_status,
                'show_failed_prior': show_failed_prior,
                'bypass_locks': bypass_locks,
                'page_obj': page_obj,
                'is_paginated': is_paginated,
                'paginator': paginator,
                'is_job_closed_or_cancelled': is_job_closed_or_cancelled,
            })

        return render(request, 'candidates/score_entry.html', {
            'jobs': jobs,
            'selected_job': selected_job,
            'stages': stages,
            'applications': applications,
            'pending_states': pending_states,
            'bulk_success': not is_job_closed_or_cancelled and job_id is not None,
            'selected_q': q,
            'selected_eval_status': eval_status,
            'show_failed_prior': show_failed_prior,
            'bypass_locks': bypass_locks,
            'page_obj': page_obj,
            'is_paginated': is_paginated,
            'paginator': paginator,
            'is_job_closed_or_cancelled': is_job_closed_or_cancelled,
        })


class ImportScoreEntryExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_SPECIALIST, UserProfile.ROLE_RECRUITMENT_DIRECTOR]

    def post(self, request):
        import openpyxl
        from django.contrib import messages
        from django.shortcuts import get_object_or_404, redirect
        from django.db import transaction
        from apps.jobs.models import JobOpportunity, JobOpportunityStage
        from apps.candidates.models import ApplicationStageState

        job_id = request.POST.get('job_id')
        stage_id = request.POST.get('stage_id')
        excel_file = request.FILES.get('excel_file')
        bypass_locks_val = request.POST.get('bypass_locks')
        bypass_locks = bypass_locks_val in ['true', 'on', '1']

        if not job_id or not stage_id or not excel_file:
            messages.error(request, "اطلاعات فرصت شغلی، مرحله یا فایل ارسالی نامعتبر است.")
            return redirect('candidate_score_entry')

        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
        stage = get_object_or_404(JobOpportunityStage, pk=stage_id, is_deleted=False)

        redirect_url = f"{reverse('candidate_score_entry')}?job_id={job.id}&stage_id={stage.id}"
        if bypass_locks:
            redirect_url += "&bypass_locks=1"

        if job.status in [JobOpportunity.STATUS_CLOSED, JobOpportunity.STATUS_CANCELLED, JobOpportunity.STATUS_SUSPENDED]:
            messages.error(request, f"خطا: امکان ثبت نمرات از طریق اکسل برای فرصت شغلی در وضعیت '{job.get_status_display()}' وجود ندارد.")
            return redirect(redirect_url)

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            messages.error(request, f"خطا در خواندن فایل اکسل: {str(e)}")
            return redirect(redirect_url)

        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            messages.error(request, "فایل اکسل خالی است یا فاقد داده‌های معتبر می‌باشد.")
            return redirect(redirect_url)

        headers = rows[0]
        data_rows = rows[1:]

        # Map headers
        header_map = {name: index for index, name in enumerate(headers)}
        
        state_id_idx = header_map.get("شناسه وضعیت")
        score_idx = header_map.get("نمره نهایی مرحله")
        status_idx = header_map.get("وضعیت ارزیابی")
        
        # Robust check for comments/notes column
        notes_idx = None
        for key in ["توضیحات و یادداشت ارزیاب", "توضیحات ارزیاب", "توضیحات", "توضیح", "یادداشت", "notes", "comment", "description"]:
            for h_name in header_map:
                if h_name and key.strip().lower() == str(h_name).strip().lower():
                    notes_idx = header_map[h_name]
                    break
            if notes_idx is not None:
                break

        # Check for evaluator column
        evaluator_idx = None
        for key in ["ارزیاب", "نام ارزیاب", "evaluator", "assessor"]:
            for h_name in header_map:
                if h_name and key.strip().lower() == str(h_name).strip().lower():
                    evaluator_idx = header_map[h_name]
                    break
            if evaluator_idx is not None:
                break

        # Check for evaluation date column
        eval_date_idx = None
        for key in ["تاریخ ارزیابی", "تاریخ", "date", "evaluation_date"]:
            for h_name in header_map:
                if h_name and key.strip().lower() == str(h_name).strip().lower():
                    eval_date_idx = header_map[h_name]
                    break
            if eval_date_idx is not None:
                break

        if state_id_idx is None or score_idx is None or status_idx is None:
            messages.error(request, "ستون‌های حیاتی 'شناسه وضعیت'، 'نمره نهایی مرحله' یا 'وضعیت ارزیابی' یافت نشدند.")
            return redirect(redirect_url)

        from .models import JobDefaultInterviewer, ExternalInterviewerScore
        default_interviewers = list(JobDefaultInterviewer.objects.filter(job=job, is_deleted=False).order_by('id'))
        interviewer_col_map = {}
        if stage.stage_type == 'INTERVIEW' and default_interviewers:
            for iv in default_interviewers:
                col_name = f"نمره {iv.interviewer_name}"
                for header_name, col_idx in header_map.items():
                    if header_name and (col_name.strip() == header_name.strip() or iv.interviewer_name.strip() in header_name):
                        interviewer_col_map[iv.interviewer_name] = col_idx
                        break

        success_count = 0
        error_count = 0
        errors = []

        # Map Persian status string to DB status choices
        status_mapping = {
            "در انتظار ارزیابی": 'PENDING',
            "در انتظار": 'PENDING',
            "pending": 'PENDING',
            "قبول شده در این مرحله": 'COMPLETED',
            "قبول": 'COMPLETED',
            "completed": 'COMPLETED',
            "مردود شده در این مرحله": 'FAILED',
            "مردود": 'FAILED',
            "failed": 'FAILED',
        }

        try:
            with transaction.atomic():
                for idx, row in enumerate(data_rows, start=2):
                    # Skip completely empty rows
                    if not any(row):
                        continue

                    state_id = row[state_id_idx] if state_id_idx < len(row) else None
                    score_val = row[score_idx] if score_idx < len(row) else None
                    status_str = row[status_idx] if status_idx < len(row) else None
                    
                    notes_val = ""
                    if notes_idx is not None and notes_idx < len(row):
                        notes_val = row[notes_idx]
                        if notes_val is None:
                            notes_val = ""

                    eval_date = None
                    has_eval_date = False
                    if eval_date_idx is not None and eval_date_idx < len(row):
                        eval_date_val = row[eval_date_idx]
                        has_eval_date = True
                        if eval_date_val:
                            from apps.historical_import.utils import parse_date_safely
                            eval_date = parse_date_safely(eval_date_val)

                    evaluator = None
                    if evaluator_idx is not None and evaluator_idx < len(row):
                        evaluator_val = row[evaluator_idx]
                        if evaluator_val:
                            evaluator_str = str(evaluator_val).strip()
                            from django.contrib.auth.models import User
                            from django.db.models import Value
                            from django.db.models.functions import Concat
                            evaluator = User.objects.annotate(
                                full_name=Concat('first_name', Value(' '), 'last_name')
                            ).filter(
                                models.Q(username__iexact=evaluator_str) |
                                models.Q(email__iexact=evaluator_str) |
                                models.Q(full_name__iexact=evaluator_str)
                            ).first()

                    if not evaluator:
                        evaluator = request.user

                    if not state_id:
                        errors.append(f"ردیف {idx}: شناسه وضعیت خالی است.")
                        error_count += 1
                        continue

                    # Fetch and verify stage state
                    state = ApplicationStageState.objects.filter(
                        id=state_id,
                        application__job=job,
                        stage=stage,
                        is_deleted=False
                    ).first()

                    if not state:
                        errors.append(f"ردیف {idx}: شناسه وضعیت {state_id} برای این شغل و مرحله معتبر نیست.")
                        error_count += 1
                        continue

                    if not (state.is_accessible or bypass_locks):
                        errors.append(f"ردیف {idx} ({state.application.candidate}): این مرحله قفل یا در دسترس نیست.")
                        error_count += 1
                        continue

                    # Parse status
                    db_status = 'PENDING'
                    if status_str:
                        clean_status = str(status_str).strip().lower()
                        db_status = status_mapping.get(clean_status, 'PENDING')

                    # Process interviewer scores
                    has_interviewer_scores = False
                    any_interviewer_changes = False
                    if stage.stage_type == 'INTERVIEW' and default_interviewers:
                        for iv in default_interviewers:
                            col_idx = interviewer_col_map.get(iv.interviewer_name)
                            if col_idx is not None and col_idx < len(row):
                                raw_val = row[col_idx]
                                score_val_iv = _normalize_number(raw_val)
                                if score_val_iv is not None:
                                    score_val_iv = max(0.0, min(100.0, score_val_iv))
                                    has_interviewer_scores = True
                                    
                                    # upsert
                                    es_obj = ExternalInterviewerScore.objects.filter(
                                        stage_state=state,
                                        interviewer_name=iv.interviewer_name,
                                        is_deleted=False
                                    ).first()
                                    if es_obj:
                                        if es_obj.score != score_val_iv or es_obj.weight != iv.weight:
                                            es_obj.score = score_val_iv
                                            es_obj.weight = iv.weight
                                            es_obj.save(update_fields=['score', 'weight'])
                                            any_interviewer_changes = True
                                    else:
                                        ExternalInterviewerScore.objects.create(
                                            stage_state=state,
                                            interviewer_name=iv.interviewer_name,
                                            score=score_val_iv,
                                            weight=iv.weight
                                        )
                                        any_interviewer_changes = True
                                else:
                                    es_obj = ExternalInterviewerScore.objects.filter(
                                        stage_state=state,
                                        interviewer_name=iv.interviewer_name,
                                        is_deleted=False
                                    ).first()
                                    if es_obj:
                                        es_obj.is_deleted = True
                                        es_obj.save()
                                        any_interviewer_changes = True

                    if has_interviewer_scores or any_interviewer_changes:
                        state.is_manually_edited = False
                        state.status = db_status
                        if notes_idx is not None:
                            state.notes = str(notes_val).strip()
                        if has_eval_date:
                            state.evaluation_date = eval_date
                        state.evaluator = evaluator
                        state.save()
                    else:
                        # Fallback to manual score
                        try:
                            score = float(score_val) if score_val is not None else 0.0
                        except (ValueError, TypeError):
                            score = 0.0
                        state.score = score
                        state.status = db_status
                        if notes_idx is not None:
                            state.notes = str(notes_val).strip()
                        if has_eval_date:
                            state.evaluation_date = eval_date
                        state.evaluator = evaluator
                        state.is_manually_edited = True
                        state._bypass_stage_score_calculation = True
                        state.save()

                    success_count += 1
        except Exception as ex:
            messages.error(request, f"خطا در حین تراکنش بروزرسانی نمرات: {str(ex)}")
            return redirect(redirect_url)

        if success_count > 0:
            messages.success(request, f"نمرات تعداد {success_count} متقاضی با موفقیت از فایل اکسل بروزرسانی شد.")
        if error_count > 0:
            messages.error(request, f"بروزرسانی تعداد {error_count} ردیف با خطا مواجه شد:<br>" + "<br>".join(errors[:10]))

        return redirect(redirect_url)


class ManageStageInterviewersView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_SPECIALIST, UserProfile.ROLE_RECRUITMENT_DIRECTOR]

    def get(self, request, stage_id):
        from apps.jobs.models import JobOpportunityStage, JobStageInterviewer
        stage = get_object_or_404(JobOpportunityStage, pk=stage_id, is_deleted=False)
        interviewers = stage.interviewers.filter(is_deleted=False).select_related('user')
        competencies = stage.competencies.filter(is_deleted=False)
        
        from django.contrib.auth.models import User
        available_users = User.objects.filter(
            profile__role__in=[
                UserProfile.ROLE_ADMIN,
                UserProfile.ROLE_RECRUITMENT_SPECIALIST,
                UserProfile.ROLE_RECRUITMENT_DIRECTOR,
                UserProfile.ROLE_INTERVIEWER,
                UserProfile.ROLE_EXTERNAL_ASSESSOR
            ],
            is_active=True
        ).order_by('first_name', 'username')

        return render(request, 'candidates/partials/stage_interviewers_manage.html', {
            'stage': stage,
            'interviewers': interviewers,
            'available_users': available_users,
            'competencies': competencies,
        })

    def post(self, request, stage_id):
        from apps.jobs.models import JobOpportunityStage, JobStageInterviewer, AssessmentCompetency
        stage = get_object_or_404(JobOpportunityStage, pk=stage_id, is_deleted=False)
        
        passing_score_val = request.POST.get('passing_score')
        if passing_score_val:
            try:
                stage.passing_score = float(passing_score_val)
                stage.save()
                for state in stage.candidate_states.filter(is_deleted=False):
                    state.save()
            except ValueError:
                pass
                
        action = request.POST.get('action')

        if action == 'delete':
            assignment_id = request.POST.get('assignment_id')
            assignment = get_object_or_404(JobStageInterviewer, pk=assignment_id, stage=stage)
            assignment.delete()
        elif action == 'add':
            user_id = request.POST.get('user_id')
            weight_val = request.POST.get('weight', '100')
            group_name = request.POST.get('group_name', '')
            
            try:
                weight = int(weight_val)
            except ValueError:
                weight = 100

            from django.contrib.auth.models import User
            user = get_object_or_404(User, pk=user_id)
            
            assignment = JobStageInterviewer.all_objects.filter(stage=stage, user=user).first()
            if assignment:
                assignment.is_deleted = False
                assignment.deleted_at = None
                assignment.weight = weight
                assignment.group_name = group_name
                assignment.save()
            else:
                JobStageInterviewer.objects.create(
                    job=stage.job,
                    stage=stage,
                    user=user,
                    weight=weight,
                    group_name=group_name
                )
        elif action == 'add_competency':
            comp_name = request.POST.get('competency_name')
            comp_weight_val = request.POST.get('competency_weight', '100')
            try:
                comp_weight = int(comp_weight_val)
            except ValueError:
                comp_weight = 100
            
            if comp_name:
                comp = AssessmentCompetency.all_objects.filter(stage=stage, name=comp_name).first()
                if comp:
                    comp.is_deleted = False
                    comp.deleted_at = None
                    comp.weight = comp_weight
                    comp.save()
                else:
                    AssessmentCompetency.objects.create(
                        stage=stage,
                        name=comp_name,
                        weight=comp_weight
                    )
        elif action == 'delete_competency':
            comp_id = request.POST.get('competency_id')
            AssessmentCompetency.objects.filter(pk=comp_id, stage=stage).delete()

        interviewers = stage.interviewers.filter(is_deleted=False).select_related('user')
        competencies = stage.competencies.filter(is_deleted=False)
        from django.contrib.auth.models import User
        available_users = User.objects.filter(
            profile__role__in=[
                UserProfile.ROLE_ADMIN,
                UserProfile.ROLE_RECRUITMENT_SPECIALIST,
                UserProfile.ROLE_RECRUITMENT_DIRECTOR,
                UserProfile.ROLE_INTERVIEWER,
                UserProfile.ROLE_EXTERNAL_ASSESSOR
            ],
            is_active=True
        ).order_by('first_name', 'username')

        return render(request, 'candidates/partials/stage_interviewers_manage.html', {
            'stage': stage,
            'interviewers': interviewers,
            'available_users': available_users,
            'competencies': competencies,
        })


class InterviewsPanelView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [
        UserProfile.ROLE_ADMIN,
        UserProfile.ROLE_RECRUITMENT_SPECIALIST,
        UserProfile.ROLE_RECRUITMENT_DIRECTOR,
        UserProfile.ROLE_INTERVIEWER,
        UserProfile.ROLE_EXTERNAL_ASSESSOR
    ]

    def get(self, request):
        from apps.jobs.models import JobStageInterviewer
        from django.db.models import F
        from apps.accounts.permissions import check_stage_access

        user_profile = request.user.profile

        # 1. Staff / Admin can view all active stage states
        if user_profile.is_recruitment_staff or user_profile.role == UserProfile.ROLE_ADMIN:
            active_stage_states = ApplicationStageState.objects.filter(
                application__status=JobApplication.STATUS_IN_PROGRESS,
                application__is_deleted=False,
                stage=F('application__current_stage'),
                is_deleted=False
            ).select_related('application__candidate', 'application__job', 'stage').order_by('application__candidate__last_name')
        else:
            # 2. Non-staff (Interviewers / Assessors) can only see stage states where they are explicitly assigned
            assigned_stage_ids = JobStageInterviewer.objects.filter(
                user=request.user, is_deleted=False
            ).values_list('stage_id', flat=True)

            active_stage_states = ApplicationStageState.objects.filter(
                application__status=JobApplication.STATUS_IN_PROGRESS,
                application__is_deleted=False,
                stage_id__in=assigned_stage_ids,
                stage=F('application__current_stage'),
                is_deleted=False
            ).select_related('application__candidate', 'application__job', 'stage').order_by('application__candidate__last_name')

            # 3. Apply role-based stage type check
            filtered_states = []
            for state in active_stage_states:
                if check_stage_access(request.user, state.stage):
                    filtered_states.append(state)
            active_stage_states = filtered_states

        for state in active_stage_states:
            my_score = state.interviewer_scores.filter(interviewer=request.user, is_deleted=False).first()
            state.my_score_obj = my_score

        return render(request, 'candidates/interviews.html', {
            'stage_states': active_stage_states,
        })


class SubmitInterviewerScoreView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [
        UserProfile.ROLE_ADMIN,
        UserProfile.ROLE_RECRUITMENT_SPECIALIST,
        UserProfile.ROLE_RECRUITMENT_DIRECTOR,
        UserProfile.ROLE_INTERVIEWER,
        UserProfile.ROLE_EXTERNAL_ASSESSOR
    ]

    def post(self, request, pk):
        from .models import InterviewerScore
        from apps.accounts.permissions import check_stage_access
        
        state = get_object_or_404(ApplicationStageState, pk=pk, is_deleted=False)
        
        # Enforce role-based stage type check
        if not check_stage_access(request.user, state.stage):
            raise PermissionDenied("شما دسترسی لازم برای ثبت نمره در این مرحله را ندارید.")
            
        # Enforce explicit interviewer assignment check for non-staff/admins
        if not (request.user.profile.is_recruitment_staff or request.user.profile.role == UserProfile.ROLE_ADMIN):
            from apps.jobs.models import JobStageInterviewer
            is_assigned = JobStageInterviewer.objects.filter(
                stage=state.stage,
                user=request.user,
                is_deleted=False
            ).exists()
            if not is_assigned:
                raise PermissionDenied("شما به عنوان مصاحبه‌گر به این مرحله تخصیص داده نشده‌اید.")
        
        score_val = request.POST.get('score', '0')
        status_val = 'COMPLETED'
        notes_val = request.POST.get('notes', '')

        try:
            score_float = float(score_val)
        except ValueError:
            score_float = 0.0

        with transaction.atomic():
            iscore, created = InterviewerScore.objects.get_or_create(
                stage_state=state,
                interviewer=request.user,
                defaults={'score': score_float, 'status': status_val, 'notes': notes_val}
            )
            if not created:
                iscore.score = score_float
                iscore.status = status_val
                iscore.notes = notes_val
                iscore.save()

        return render(request, 'candidates/partials/interviewer_score_row_saved.html', {
            'state': state,
            'iscore': iscore,
        })


class JobOpportunityDescriptionView(LoginRequiredMixin, View):
    def get(self, request, pk):
        from apps.jobs.models import JobOpportunity
        job = get_object_or_404(JobOpportunity, pk=pk, is_deleted=False)
        return render(request, 'candidates/partials/job_description_modal.html', {'job': job})


class AssessmentCenterSheetView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [
        UserProfile.ROLE_ADMIN,
        UserProfile.ROLE_RECRUITMENT_SPECIALIST,
        UserProfile.ROLE_RECRUITMENT_DIRECTOR,
        UserProfile.ROLE_INTERVIEWER,
        UserProfile.ROLE_EXTERNAL_ASSESSOR
    ]

    def get(self, request, pk):
        from .models import InterviewerScore, AssessorCompetencyScore
        from apps.accounts.permissions import check_stage_access
        
        state = get_object_or_404(ApplicationStageState, pk=pk, is_deleted=False)
        
        # Enforce role-based stage type check
        if not check_stage_access(request.user, state.stage):
            raise PermissionDenied("شما دسترسی لازم برای مشاهده این برگ ارزیابی را ندارید.")
            
        # Enforce explicit interviewer assignment check for non-staff/admins
        if not (request.user.profile.is_recruitment_staff or request.user.profile.role == UserProfile.ROLE_ADMIN):
            from apps.jobs.models import JobStageInterviewer
            is_assigned = JobStageInterviewer.objects.filter(
                stage=state.stage,
                user=request.user,
                is_deleted=False
            ).exists()
            if not is_assigned:
                raise PermissionDenied("شما به عنوان ارزیاب به این مرحله تخصیص داده نشده‌اید.")

        competencies = state.stage.competencies.filter(is_deleted=False)
        
        iscore = state.interviewer_scores.filter(interviewer=request.user, is_deleted=False).first()
        comp_scores = {}
        if iscore:
            for cs in iscore.competency_scores.filter(is_deleted=False):
                comp_scores[cs.competency_id] = cs
                
        comp_data = []
        for c in competencies:
            comp_data.append({
                'competency': c,
                'score_obj': comp_scores.get(c.id)
            })

        bypass_locks = request.GET.get('bypass_locks') == '1'

        return render(request, 'candidates/partials/assessment_center_sheet.html', {
            'state': state,
            'comp_data': comp_data,
            'iscore': iscore,
            'source': request.GET.get('source'),
            'bypass_locks': bypass_locks,
        })

    def post(self, request, pk):
        from .models import InterviewerScore, AssessorCompetencyScore
        from apps.accounts.permissions import check_stage_access
        
        state = get_object_or_404(ApplicationStageState, pk=pk, is_deleted=False)
        bypass_locks = request.POST.get('bypass_locks') == '1'
        
        # Enforce role-based stage type check
        if not check_stage_access(request.user, state.stage):
            raise PermissionDenied("شما دسترسی لازم برای ثبت نمره در این مرحله را ندارید.")
            
        # Enforce explicit interviewer assignment check for non-staff/admins
        if not (request.user.profile.is_recruitment_staff or request.user.profile.role == UserProfile.ROLE_ADMIN):
            from apps.jobs.models import JobStageInterviewer
            is_assigned = JobStageInterviewer.objects.filter(
                stage=state.stage,
                user=request.user,
                is_deleted=False
            ).exists()
            if not is_assigned:
                raise PermissionDenied("شما به عنوان ارزیاب به این مرحله تخصیص داده نشده‌اید.")

        competencies = state.stage.competencies.filter(is_deleted=False)

        notes_val = request.POST.get('notes', '')
        source = request.POST.get('source')

        with transaction.atomic():
            iscore, created = InterviewerScore.objects.get_or_create(
                stage_state=state,
                interviewer=request.user,
                defaults={'score': 0.0, 'status': 'COMPLETED', 'notes': notes_val}
            )
            iscore.status = 'COMPLETED'
            iscore.notes = notes_val
            iscore.save()

            for c in competencies:
                score_val = request.POST.get(f'comp_score_{c.id}', '0')
                comp_notes = request.POST.get(f'comp_notes_{c.id}', '')
                try:
                    score_float = float(score_val)
                except ValueError:
                    score_float = 0.0

                cs_obj, cs_created = AssessorCompetencyScore.objects.get_or_create(
                    interviewer_score=iscore,
                    competency=c,
                    defaults={'score': score_float, 'notes': comp_notes}
                )
                if not cs_created:
                    cs_obj.score = score_float
                    cs_obj.notes = comp_notes
                    cs_obj.save()

            iscore.save()
            state.is_manually_edited = False
            state.save()

        if source == 'score_entry':
            stages = list(state.application.job.stages.filter(is_deleted=False).order_by('sequence'))
            state_map = {s.stage_id: s for s in state.application.stage_states.filter(is_deleted=False)}
            state.application.stage_cells = []
            for stg in stages:
                s_cell = state_map.get(stg.id)
                if not s_cell:
                    s_cell = ApplicationStageState.objects.create(
                        application=state.application,
                        stage=stg,
                        status=ApplicationStageState.STATUS_PENDING,
                        score=0.0
                    )
                state.application.stage_cells.append(s_cell)
            return render(request, 'candidates/partials/score_entry_row.html', {
                'app': state.application,
                'stages': stages,
                'bypass_locks': bypass_locks,
            })

        return render(request, 'candidates/partials/interviewer_score_row_saved.html', {
            'state': state,
            'iscore': iscore,
        })


class InterviewScoresPanelView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [
        UserProfile.ROLE_ADMIN,
        UserProfile.ROLE_RECRUITMENT_SPECIALIST,
        UserProfile.ROLE_RECRUITMENT_DIRECTOR
    ]

    def get(self, request, pk):
        from .models import ExternalInterviewerScore, JobDefaultInterviewer, ExternalInterviewerCompetencyScore
        state = get_object_or_404(ApplicationStageState, pk=pk, is_deleted=False)
        
        if state.stage.stage_type != 'INTERVIEW':
            raise PermissionDenied("این پنل فقط برای مراحل مصاحبه در دسترس است.")
            
        external_scores = list(state.external_interviewer_scores.filter(is_deleted=False))
        bypass_locks = request.GET.get('bypass_locks') == '1'
        
        competencies = list(state.stage.competencies.filter(is_deleted=False).order_by('id'))

        # Map competencies to existing scores
        for score in external_scores:
            existing_comp_scores = {
                cs.competency_id: cs.score 
                for cs in score.competency_scores.filter(is_deleted=False)
            }
            score.comp_cells = [
                {
                    'competency': comp,
                    'score': existing_comp_scores.get(comp.id, 0.0)
                }
                for comp in competencies
            ]

        # If no score exists, load job default interviewers to pre-populate form
        default_interviewers = []
        if not external_scores:
            default_interviewers = list(
                JobDefaultInterviewer.objects.filter(
                    job=state.application.job,
                    is_deleted=False
                ).order_by('id')
            )
            for iv in default_interviewers:
                iv.comp_cells = [
                    {
                        'competency': comp,
                        'score': 0.0
                    }
                    for comp in competencies
                ]
        
        return render(request, 'candidates/partials/interview_scores_panel.html', {
            'state': state,
            'external_scores': external_scores,
            'bypass_locks': bypass_locks,
            'default_interviewers': default_interviewers,
            'competencies': competencies,
        })

    def post(self, request, pk):
        from .models import ExternalInterviewerScore, ExternalInterviewerCompetencyScore
        from django.db import transaction
        
        state = get_object_or_404(ApplicationStageState, pk=pk, is_deleted=False)
        bypass_locks = request.POST.get('bypass_locks') == '1'
        
        if state.stage.stage_type != 'INTERVIEW':
            raise PermissionDenied("این پنل فقط برای مراحل مصاحبه در دسترس است.")

        competencies = list(state.stage.competencies.filter(is_deleted=False).order_by('id'))
        names = request.POST.getlist('interviewer_name[]')
        scores = request.POST.getlist('score[]')
        weights = request.POST.getlist('weight[]')
        notes = request.POST.getlist('notes[]')

        # Since we are returning score_entry_row.html, we need stages and application context
        from apps.jobs.models import JobOpportunityStage
        stages = list(state.application.job.stages.filter(is_deleted=False).order_by('sequence'))
        
        # Ensure all states exist to prevent UI crashes when rendering the row
        state_map = {s.stage_id: s for s in state.application.stage_states.filter(is_deleted=False)}
        state.application.stage_cells = []
        for stg in stages:
            s_cell = state_map.get(stg.id)
            if not s_cell:
                s_cell = ApplicationStageState.objects.create(
                    application=state.application,
                    stage=stg,
                    status=ApplicationStageState.STATUS_PENDING,
                    score=0.0
                )
            state.application.stage_cells.append(s_cell)

        with transaction.atomic():
            # Delete old interviewer scores (soft-delete)
            old_scores = state.external_interviewer_scores.filter(is_deleted=False)
            for old in old_scores:
                old.competency_scores.filter(is_deleted=False).update(is_deleted=True)
                old.is_deleted = True
                old.save()
            
            for i in range(len(names)):
                name = names[i].strip()
                if not name:
                    continue
                
                try:
                    weight_val = int(weights[i])
                    if weight_val < 1:
                        weight_val = 100
                except (ValueError, IndexError):
                    weight_val = 100
                    
                notes_val = notes[i].strip() if i < len(notes) else ""
                
                # Fetch dynamically submitted competency scores for this row index
                row_comp_scores = []
                total_c_weight = 0
                weighted_c_sum = 0.0
                
                for comp in competencies:
                    key = f"comp_score_{i}_{comp.id}"
                    raw_score = request.POST.get(key)
                    if raw_score is not None:
                        try:
                            c_score = float(raw_score)
                        except ValueError:
                            c_score = 0.0
                        row_comp_scores.append((comp, c_score))
                        weighted_c_sum += c_score * comp.weight
                        total_c_weight += comp.weight
                
                if row_comp_scores:
                    final_score_val = round(weighted_c_sum / total_c_weight, 2) if total_c_weight > 0 else 0.0
                else:
                    try:
                        final_score_val = float(scores[i])
                    except (ValueError, IndexError):
                        final_score_val = 0.0
                
                ext_score = ExternalInterviewerScore.objects.create(
                    stage_state=state,
                    interviewer_name=name,
                    score=final_score_val,
                    weight=weight_val,
                    notes=notes_val
                )
                
                for comp, c_score in row_comp_scores:
                    ExternalInterviewerCompetencyScore.objects.create(
                        external_interviewer_score=ext_score,
                        competency=comp,
                        score=c_score
                    )
            
            # Recalculate average and save state
            state.is_manually_edited = False
            state.save()

        return render(request, 'candidates/partials/score_entry_row.html', {
            'app': state.application,
            'stages': stages,
            'bypass_locks': bypass_locks,
        })


class AssessmentCenterReportView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_SPECIALIST, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_EXTERNAL_ASSESSOR]

    def get(self, request):
        states = ApplicationStageState.objects.filter(
            stage__competencies__isnull=False,
            is_deleted=False
        ).exclude(status=ApplicationStageState.STATUS_PENDING).select_related(
            'application__candidate', 'application__job', 'stage', 'evaluator'
        ).prefetch_related('interviewer_scores__competency_scores__competency').distinct().order_by('-updated_at')

        return render(request, 'candidates/assessment_report_list.html', {
            'states': states
        })


class AssessmentCenterDetailReportView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_SPECIALIST, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_EXTERNAL_ASSESSOR]

    def get(self, request, pk):
        state = get_object_or_404(ApplicationStageState, pk=pk, is_deleted=False)
        competencies = state.stage.competencies.filter(is_deleted=False)
        
        iscores = state.interviewer_scores.filter(is_deleted=False).exclude(status='PENDING').prefetch_related(
            'competency_scores__competency', 'interviewer'
        )

        comp_rows = []
        for comp in competencies:
            scores_detail = []
            for iscore in iscores:
                cs = iscore.competency_scores.filter(competency=comp, is_deleted=False).first()
                scores_detail.append({
                    'assessor': iscore.interviewer.get_full_name() or iscore.interviewer.username,
                    'score': cs.score if cs else None,
                    'notes': cs.notes if cs else ''
                })
            comp_rows.append({
                'competency': comp,
                'scores': scores_detail
            })

        return render(request, 'candidates/assessment_detail_report.html', {
            'state': state,
            'iscores': iscores,
            'comp_rows': comp_rows,
            'competencies': competencies,
        })


class JobOpportunityFinalRankingView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    model = JobOpportunity
    template_name = 'jobs/job_final_ranking.html'
    context_object_name = 'job'
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_SPECIALIST, UserProfile.ROLE_RECRUITMENT_DIRECTOR]

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        apps = self.object.applications.filter(is_deleted=False).select_related('candidate')
        data['applications'] = apps.order_by('-final_score')
        data['stages'] = self.object.stages.filter(is_deleted=False).exclude(weight=0).order_by('sequence')
        data['all_stages'] = self.object.stages.filter(is_deleted=False).order_by('sequence')
        data['selected_count'] = apps.filter(status=JobApplication.STATUS_SELECTED).count()
        data['reserve_count'] = apps.filter(status=JobApplication.STATUS_RESERVE).count()
        data['rejected_count'] = apps.filter(status=JobApplication.STATUS_REJECTED).count()
        return data


class BulkUpdateApplicationStatusView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_SPECIALIST, UserProfile.ROLE_RECRUITMENT_DIRECTOR]

    def post(self, request, job_id):
        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
        application_ids = request.POST.getlist('application_ids')
        status_action = request.POST.get('status_action')

        if status_action in dict(JobApplication.STATUS_CHOICES):
            with transaction.atomic():
                JobApplication.objects.filter(pk__in=application_ids, job=job, is_deleted=False).update(status=status_action)
                job.update_status()

        return redirect('job_final_ranking', pk=job_id)


class BulkAdvanceStageView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_SPECIALIST, UserProfile.ROLE_RECRUITMENT_DIRECTOR]

    def post(self, request, job_id):
        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
        application_ids = request.POST.getlist('application_ids')

        with transaction.atomic():
            for app_id in application_ids:
                app = JobApplication.objects.filter(pk=app_id, job=job, status=JobApplication.STATUS_IN_PROGRESS, is_deleted=False).first()
                if app and app.current_stage:
                    current_state = app.stage_states.filter(stage=app.current_stage, is_deleted=False).first()
                    if current_state and (current_state.status == ApplicationStageState.STATUS_COMPLETED or current_state.is_conditional_pass):
                        next_stage = job.stages.filter(
                            is_deleted=False, 
                            sequence__gt=app.current_stage.sequence
                        ).order_by('sequence').first()
                        
                        if next_stage:
                            app.current_stage = next_stage
                            app.save(update_fields=['current_stage'])
            job.update_status()

        return redirect('job_pipeline', pk=job_id)


class ExportCandidatesExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get(self, request):
        from apps.candidates.views import CandidateListView
        view = CandidateListView()
        view.request = request
        view.args = ()
        view.kwargs = {}
        queryset = view.get_queryset()
        
        headers = [
            "شناسه", "نام", "نام خانوادگی", "کد ملی", "ایمیل", "تلفن", 
            "آخرین مدرک تحصیلی", "رشته تحصیلی", "دانشگاه", "سوابق شغلی", "مهارت‌ها", "فرصت‌های شغلی متقاضی"
        ]
        
        rows = []
        for candidate in queryset:
            edu = candidate.education.filter(is_deleted=False).order_by('-graduation_year').first()
            degree_display = edu.get_degree_display() if edu else ""
            major_display = edu.major if edu else ""
            univ_display = edu.university if edu else ""
            
            exps = [f"{exp.job_title} ({exp.company})" for exp in candidate.experience.filter(is_deleted=False)]
            exps_str = " | ".join(exps)
            
            skills = [s.name for s in candidate.skills.filter(is_deleted=False)]
            skills_str = ", ".join(skills)
            
            jobs = [app.job.title for app in candidate.applications.filter(is_deleted=False)]
            jobs_str = ", ".join(jobs)
            
            rows.append([
                candidate.id,
                candidate.first_name,
                candidate.last_name,
                candidate.national_id,
                candidate.email,
                candidate.phone_number,
                degree_display,
                major_display,
                univ_display,
                exps_str,
                skills_str,
                jobs_str
            ])
            
        from apps.core.utils import export_to_excel_response
        return export_to_excel_response("candidates_report.xlsx", headers, rows)


class ExportScoreEntryExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_SPECIALIST, UserProfile.ROLE_RECRUITMENT_DIRECTOR]

    def get(self, request):
        from apps.jobs.models import JobOpportunity, JobOpportunityStage
        from apps.candidates.models import ApplicationStageState, JobApplication
        
        job_id = request.GET.get('job_id')
        stage_id = request.GET.get('stage_id')
        q = request.GET.get('q')
        eval_status = request.GET.get('eval_status', 'PENDING')
        show_failed_prior_val = request.GET.get('show_failed_prior')
        show_failed_prior = show_failed_prior_val in ['true', 'on', '1']
        bypass_locks_val = request.GET.get('bypass_locks')
        bypass_locks = bypass_locks_val in ['true', 'on', '1']
        
        if not job_id or not stage_id:
            return HttpResponse("شناسه فرصت شغلی و مرحله الزامی است.", status=400)
            
        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
        stage = get_object_or_404(JobOpportunityStage, pk=stage_id, is_deleted=False)
        
        app_statuses = [JobApplication.STATUS_IN_PROGRESS]
        if show_failed_prior or bypass_locks:
            app_statuses.append(JobApplication.STATUS_REJECTED)
            
        # Ensure all active applications have a stage state record for this stage before exporting
        active_apps = JobApplication.objects.filter(
            job=job,
            status__in=app_statuses,
            is_deleted=False
        )
        existing_app_ids = set(ApplicationStageState.objects.filter(
            application__job=job,
            stage=stage,
            is_deleted=False
        ).values_list('application_id', flat=True))
        
        to_create = []
        for app in active_apps:
            if app.id not in existing_app_ids:
                to_create.append(ApplicationStageState(
                    application=app,
                    stage=stage,
                    status=ApplicationStageState.STATUS_PENDING,
                    score=0.0
                ))
        if to_create:
            ApplicationStageState.objects.bulk_create(to_create)

        pending_states_qs = ApplicationStageState.objects.filter(
            application__job=job,
            application__status__in=app_statuses,
            application__is_deleted=False,
            stage=stage,
            is_deleted=False
        ).select_related('application__candidate', 'application__job', 'stage')
        
        from django.db.models import Exists, OuterRef
        # Annotate if there is a failed prior stage state
        prior_failed_subquery = ApplicationStageState.objects.filter(
            application=OuterRef('application'),
            stage__sequence__lt=OuterRef('stage__sequence'),
            status=ApplicationStageState.STATUS_FAILED,
            is_conditional_pass=False,
            is_deleted=False
        )
        pending_states_qs = pending_states_qs.annotate(
            has_failed_prior=Exists(prior_failed_subquery)
        )
        
        # Filter out prior failed stages unless specifically requested
        if not show_failed_prior:
            pending_states_qs = pending_states_qs.filter(has_failed_prior=False)

        if eval_status == 'PENDING':
            pending_states_qs = pending_states_qs.filter(status=ApplicationStageState.STATUS_PENDING)
        elif eval_status == 'COMPLETED_FAILED':
            pending_states_qs = pending_states_qs.filter(status__in=[ApplicationStageState.STATUS_COMPLETED, ApplicationStageState.STATUS_FAILED])
            
        if q:
            pending_states_qs = pending_states_qs.filter(
                models.Q(application__candidate__first_name__icontains=q) |
                models.Q(application__candidate__last_name__icontains=q) |
                models.Q(application__candidate__national_id__icontains=q)
            )
            
        states = pending_states_qs.order_by('application__candidate__last_name')
        
        headers = [
            "شناسه وضعیت", "نام متقاضی", "نام خانوادگی", "کد ملی", "عنوان شغل", "مرحله", 
            "نمره نهایی مرحله", "وضعیت ارزیابی", "توضیحات و یادداشت ارزیاب", "تاریخ ارزیابی", "ارزیاب", "آخرین تغییر"
        ]
        
        from .models import JobDefaultInterviewer
        default_interviewers = list(JobDefaultInterviewer.objects.filter(job=job, is_deleted=False).order_by('id'))
        if stage.stage_type == 'INTERVIEW' and default_interviewers:
            for iv in default_interviewers:
                headers.append(f"نمره {iv.interviewer_name}")
        
        from apps.core.templatetags.jalali_tags import to_jalali
        rows = []
        for state in states:
            evaluator_name = state.evaluator.get_full_name() if state.evaluator else str(state.evaluator or '')
            row_data = [
                state.id,
                state.application.candidate.first_name,
                state.application.candidate.last_name,
                state.application.candidate.national_id,
                job.title,
                stage.name,
                state.score if state.score is not None else "",
                state.get_status_display(),
                state.notes if state.notes else "",
                to_jalali(state.evaluation_date),
                evaluator_name,
                to_jalali(state.updated_at)
            ]
            if stage.stage_type == 'INTERVIEW' and default_interviewers:
                existing = {
                    es.interviewer_name: es.score
                    for es in state.external_interviewer_scores.filter(is_deleted=False)
                }
                for iv in default_interviewers:
                    row_data.append(existing.get(iv.interviewer_name, ""))
            rows.append(row_data)
            
        from apps.core.utils import export_to_excel_response
        filename = f"score_entry_{job.id}_{stage.id}.xlsx"
        return export_to_excel_response(filename, headers, rows)


class ExportInterviewsExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [
        UserProfile.ROLE_ADMIN,
        UserProfile.ROLE_RECRUITMENT_SPECIALIST,
        UserProfile.ROLE_RECRUITMENT_DIRECTOR,
        UserProfile.ROLE_INTERVIEWER,
        UserProfile.ROLE_EXTERNAL_ASSESSOR
    ]

    def get(self, request):
        from apps.jobs.models import JobStageInterviewer
        from django.db.models import F
        from apps.accounts.permissions import check_stage_access
        
        user_profile = request.user.profile
        
        if user_profile.is_recruitment_staff or user_profile.role == UserProfile.ROLE_ADMIN:
            active_stage_states = ApplicationStageState.objects.filter(
                application__status=JobApplication.STATUS_IN_PROGRESS,
                application__is_deleted=False,
                stage=F('application__current_stage'),
                is_deleted=False
            ).select_related('application__candidate', 'application__job', 'stage').order_by('application__candidate__last_name')
        else:
            assigned_stage_ids = JobStageInterviewer.objects.filter(
                user=request.user, is_deleted=False
            ).values_list('stage_id', flat=True)

            active_stage_states = ApplicationStageState.objects.filter(
                application__status=JobApplication.STATUS_IN_PROGRESS,
                application__is_deleted=False,
                stage_id__in=assigned_stage_ids,
                stage=F('application__current_stage'),
                is_deleted=False
            ).select_related('application__candidate', 'application__job', 'stage').order_by('application__candidate__last_name')

            filtered_states = []
            for state in active_stage_states:
                if check_stage_access(request.user, state.stage):
                    filtered_states.append(state)
            active_stage_states = filtered_states

        from apps.core.templatetags.jalali_tags import to_jalali
        headers = [
            "شناسه", "نام متقاضی", "نام خانوادگی", "عنوان شغل", "مرحله جاری", "نمره شما", "آخرین تغییر"
        ]
        
        rows = []
        for state in active_stage_states:
            my_score_obj = state.interviewer_scores.filter(interviewer=request.user, is_deleted=False).first()
            my_score = my_score_obj.score if my_score_obj else ""
            
            rows.append([
                state.id,
                state.application.candidate.first_name,
                state.application.candidate.last_name,
                state.application.job.title,
                state.stage.name,
                my_score,
                to_jalali(state.updated_at)
            ])
            
        from apps.core.utils import export_to_excel_response
        return export_to_excel_response("my_interviews.xlsx", headers, rows)


class ExportAssessmentCenterExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_SPECIALIST, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_EXTERNAL_ASSESSOR]

    def get(self, request):
        states = ApplicationStageState.objects.filter(
            stage__competencies__isnull=False,
            is_deleted=False
        ).exclude(status=ApplicationStageState.STATUS_PENDING).select_related(
            'application__candidate', 'application__job', 'stage'
        ).distinct().order_by('-updated_at')
        
        headers = [
            "شناسه وضعیت", "نام متقاضی", "نام خانوادگی", "کد ملی", "فرصت شغلی", "مرحله کانون", 
            "میانگین وزنی کانون", "حد نصاب قبولی", "وضعیت نتیجه", "آخرین به‌روزرسانی"
        ]
        
        rows = []
        for state in states:
            passing_score = state.stage.passing_score or 60.0
            score_val = state.score if state.score is not None else 0.0
            status_display = "قبول" if score_val >= passing_score else "مردود"
            
            rows.append([
                state.id,
                state.application.candidate.first_name,
                state.application.candidate.last_name,
                state.application.candidate.national_id,
                state.application.job.title,
                state.stage.name,
                state.score if state.score is not None else "",
                passing_score,
                status_display,
                state.updated_at.strftime('%Y-%m-%d %H:%M') if state.updated_at else ""
            ])
            
        from apps.core.utils import export_to_excel_response
        return export_to_excel_response("assessment_center_report.xlsx", headers, rows)


class ExportJobFinalRankingExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_SPECIALIST, UserProfile.ROLE_RECRUITMENT_DIRECTOR]

    def get(self, request, pk):
        from apps.jobs.models import JobOpportunity
        from apps.candidates.models import JobApplication, ApplicationStageState
        job = get_object_or_404(JobOpportunity, pk=pk, is_deleted=False)
        apps = job.applications.filter(is_deleted=False).select_related('candidate').order_by('-final_score')
        stages = job.stages.filter(is_deleted=False).exclude(weight=0).order_by('sequence')

        headers = ["رتبه", "نام", "نام خانوادگی", "کد ملی", "شماره پرسنلی", "امتیاز نهایی وزنی"]
        for stage in stages:
            headers.append(stage.name)
        headers.append("وضعیت درخواست")

        rows = []
        for idx, app in enumerate(apps, start=1):
            candidate = app.candidate
            row = [
                idx,
                candidate.first_name,
                candidate.last_name,
                candidate.national_id,
                candidate.personnel_number or "",
                app.final_score,
            ]
            for stage in stages:
                state = app.stage_states.filter(stage=stage, is_deleted=False).first()
                if state:
                    if state.status == 'COMPLETED':
                        row.append(state.score)
                    else:
                        row.append("در حال ارزیابی")
                else:
                    row.append("-")
            
            status_map = {
                'IN_PROGRESS': 'در حال بررسی / ارزیابی',
                'SELECTED': 'پذیرفته شده نهایی',
                'RESERVE': 'ذخیره',
                'REJECTED': 'رد شده',
            }
            row.append(status_map.get(app.status, app.status))
            rows.append(row)

        from apps.core.utils import export_to_excel_response
        filename = f"job_ranking_{job.id}.xlsx"
        return export_to_excel_response(filename, headers, rows)


class DownloadCandidateTemplateView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get(self, request):
        headers = [
            "نام", "نام خانوادگی", "کد ملی", "ایمیل", "شماره تماس", 
            "شماره پرسنلی (اختیاری)", "مدرک تحصیلی (کاردانی/کارشناسی/ارشد/دکتری)", "رشته تحصیلی", "مهارت‌ها (جدا شده با کاما)"
        ]
        sample_row = [
            "علی", "احمدی", "0012345678", "ali@example.com", "09121112222", 
            "98001", "کارشناسی ارشد", "هوش مصنوعی", "Python, Machine Learning, Git"
        ]
        
        from apps.core.utils import export_to_excel_response
        return export_to_excel_response("candidate_import_template.xlsx", headers, [sample_row])


class ImportCandidatesView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def post(self, request):
        from django.contrib import messages
        import openpyxl
        import re
        from django.contrib.auth.models import User
        from apps.accounts.models import UserProfile
        from apps.candidates.models import Candidate, JobApplication, CandidateEducation, CandidateSkill
        from apps.jobs.models import JobOpportunity

        view_param = request.GET.get('view')
        redirect_url = reverse('candidate_list')
        if view_param:
            redirect_url += f'?view={view_param}'

        excel_file = request.FILES.get('excel_file')
        job_id = request.POST.get('job_id')

        if not excel_file:
            messages.error(request, "لطفاً یک فایل اکسل انتخاب کنید.")
            return redirect(redirect_url)

        # Read Excel using openpyxl
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            messages.error(request, f"خطا در خواندن فایل اکسل: {str(e)}")
            return redirect(redirect_url)

        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            messages.error(request, "فایل اکسل خالی است یا فاقد داده‌های معتبر می‌باشد.")
            return redirect(redirect_url)

        # First row is headers
        headers = rows[0]
        data_rows = rows[1:]

        # Check headers compatibility (basic check)
        expected_headers = ["نام", "نام خانوادگی", "کد ملی", "ایمیل", "شماره تماس"]
        for header in expected_headers:
            if header not in headers:
                messages.error(request, f"ستون حیاتی '{header}' در سربرگ‌های فایل اکسل یافت نشد. لطفاً از فایل نمونه استفاده کنید.")
                return redirect(redirect_url)

        # Map headers to indices
        header_map = {name: index for index, name in enumerate(headers)}
        
        success_count = 0
        errors = []

        # Target job (if any)
        target_job = None
        if job_id:
            target_job = JobOpportunity.objects.filter(id=job_id, is_deleted=False).first()

        for idx, row in enumerate(data_rows, start=2):
            # Skip completely empty rows
            if not any(row):
                continue

            first_name = row[header_map.get("نام")]
            last_name = row[header_map.get("نام خانوادگی")]
            national_id = row[header_map.get("کد ملی")]
            email = row[header_map.get("ایمیل")]
            phone_number = row[header_map.get("شماره تماس")]
            
            p_num_idx = header_map.get("شماره پرسنلی (اختیاری)")
            personnel_number = str(row[p_num_idx]).strip() if p_num_idx is not None and row[p_num_idx] is not None else None
            if not personnel_number:
                p_num_idx_alt = header_map.get("شماره پرسنلی")
                personnel_number = str(row[p_num_idx_alt]).strip() if p_num_idx_alt is not None and row[p_num_idx_alt] is not None else None

            # Optional new fields: degree, major, skills
            degree_idx = header_map.get("مدرک تحصیلی (کاردانی/کارشناسی/ارشد/دکتری)")
            if degree_idx is None:
                degree_idx = header_map.get("مدرک تحصیلی")
            if degree_idx is None:
                degree_idx = header_map.get("مقطع تحصیلی")
            degree_str = str(row[degree_idx]).strip() if degree_idx is not None and row[degree_idx] is not None else None

            major_idx = header_map.get("رشته تحصیلی")
            if major_idx is None:
                major_idx = header_map.get("رشته")
            major_str = str(row[major_idx]).strip() if major_idx is not None and row[major_idx] is not None else None

            skills_idx = header_map.get("مهارت‌ها (جدا شده با کاما)")
            if skills_idx is None:
                skills_idx = header_map.get("مهارت‌ها")
            if skills_idx is None:
                skills_idx = header_map.get("مهارت")
            skills_str = str(row[skills_idx]).strip() if skills_idx is not None and row[skills_idx] is not None else None

            # Basic Validation
            if not first_name or not last_name or not national_id or not email or not phone_number:
                errors.append(f"ردیف {idx}: تمامی فیلدهای نام، نام خانوادگی، کد ملی، ایمیل و شماره تماس اجباری هستند.")
                continue

            first_name = str(first_name).strip()
            last_name = str(last_name).strip()
            national_id = str(national_id).strip()
            email = str(email).strip()
            
            # Clean and normalize phone number (e.g. 9123456789 -> 09123456789)
            phone_str = str(phone_number).strip()
            if phone_str.endswith('.0'):
                phone_str = phone_str[:-2]
            phone_str = phone_str.replace(" ", "")
            if phone_str and not phone_str.startswith('0') and len(phone_str) == 10:
                phone_str = '0' + phone_str
            phone_number = phone_str

            # Validate National ID (10 digits)
            if not re.match(r'^\d{10}$', national_id):
                errors.append(f"ردیف {idx} ({first_name} {last_name}): کد ملی باید دقیقاً ۱۰ رقم عددی باشد.")
                continue

            # Validate Email format
            if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
                errors.append(f"ردیف {idx} ({first_name} {last_name}): فرمت ایمیل نامعتبر است.")
                continue

            # Determine if candidate already exists (including soft-deleted)
            existing_candidate = Candidate.all_objects.filter(national_id=national_id).first()

            # Check duplicate personnel number (if provided)
            if personnel_number:
                p_num_taken = Candidate.objects.filter(personnel_number=personnel_number, is_deleted=False)
                if existing_candidate:
                    p_num_taken = p_num_taken.exclude(pk=existing_candidate.pk)
                if p_num_taken.exists():
                    errors.append(f"ردیف {idx} ({first_name} {last_name}): متقاضی با شماره پرسنلی {personnel_number} از قبل وجود دارد.")
                    continue

            # Map Persian degree string to DB choices keys
            degree_choice = None
            if degree_str:
                val = degree_str.lower()
                if "ارشد" in val or "master" in val or "فوق لیسانس" in val:
                    degree_choice = 'MASTER'
                elif "دکتری" in val or "phd" in val or "دکترا" in val:
                    degree_choice = 'PHD'
                elif "کاردانی" in val or "associate" in val or "فوق دیپلم" in val:
                    degree_choice = 'ASSOCIATE'
                elif "کارشناسی" in val or "bachelor" in val or "لیسانس" in val:
                    degree_choice = 'BACHELOR'

            # Create or update candidate
            try:
                from django.db import transaction
                with transaction.atomic():
                    if existing_candidate:
                        # Restore if soft-deleted
                        if existing_candidate.is_deleted:
                            existing_candidate.is_deleted = False
                            existing_candidate.deleted_at = None
                        
                        # Update candidate fields
                        existing_candidate.first_name = first_name or existing_candidate.first_name
                        existing_candidate.last_name = last_name or existing_candidate.last_name
                        existing_candidate.email = email or existing_candidate.email
                        existing_candidate.phone_number = phone_number or existing_candidate.phone_number
                        if personnel_number:
                            existing_candidate.personnel_number = personnel_number
                        existing_candidate.save()
                        
                        # Update associated User and Profile
                        user = existing_candidate.user
                        if user:
                            user.first_name = first_name or user.first_name
                            user.last_name = last_name or user.last_name
                            user.email = email or user.email
                            user.save()
                            
                            profile = user.profile
                            profile.role = UserProfile.ROLE_CANDIDATE
                            profile.phone_number = phone_number or profile.phone_number
                            profile.save()
                            
                        candidate = existing_candidate
                    else:
                        # Create new Django User
                        user = User.objects.filter(username=national_id).first()
                        if not user:
                            user = User.objects.create_user(
                                username=national_id,
                                email=email,
                                password=phone_number,
                                first_name=first_name,
                                last_name=last_name
                            )
                        
                        # Update UserProfile role to CANDIDATE
                        profile = user.profile
                        profile.role = UserProfile.ROLE_CANDIDATE
                        profile.phone_number = phone_number
                        profile.save()

                        # Create Candidate
                        candidate = Candidate.objects.create(
                            user=user,
                            first_name=first_name,
                            last_name=last_name,
                            national_id=national_id,
                            email=email,
                            phone_number=phone_number,
                            personnel_number=personnel_number if personnel_number else None
                        )
                    
                    # Create education record if degree and major are specified
                    if degree_choice and major_str:
                        # Only create education if it doesn't already exist for this candidate
                        edu_exists = CandidateEducation.objects.filter(
                            candidate=candidate,
                            degree=degree_choice,
                            major=major_str,
                            is_deleted=False
                        ).exists()
                        if not edu_exists:
                            CandidateEducation.objects.create(
                                candidate=candidate,
                                degree=degree_choice,
                                major=major_str,
                                university="ثبت نشده",
                                gpa=0.0,
                                graduation_year=1400
                            )

                    # Create skill records if specified
                    if skills_str:
                        skills_list = [s.strip() for s in re.split(r'[,،]', skills_str) if s.strip()]
                        for s_name in skills_list:
                            skill_exists = CandidateSkill.objects.filter(
                                candidate=candidate,
                                name=s_name,
                                is_deleted=False
                            ).exists()
                            if not skill_exists:
                                CandidateSkill.objects.create(
                                    candidate=candidate,
                                    name=s_name,
                                    level='INTERMEDIATE'
                                )

                    # If target job is specified, create or restore JobApplication
                    if target_job:
                        application = JobApplication.all_objects.filter(job=target_job, candidate=candidate).first()
                        if application:
                            if application.is_deleted:
                                application.is_deleted = False
                                application.deleted_at = None
                                application.status = JobApplication.STATUS_IN_PROGRESS
                                application.save()
                        else:
                            JobApplication.objects.create(
                                job=target_job,
                                candidate=candidate,
                                status=JobApplication.STATUS_IN_PROGRESS
                            )
                success_count += 1
            except Exception as ex:
                errors.append(f"ردیف {idx} ({first_name} {last_name}): خطا در ثبت داده‌ها: {str(ex)}")

        # Build final feedback message
        if success_count > 0:
            msg = f"تعداد {success_count} متقاضی جدید با موفقیت از فایل اکسل وارد شدند."
            if target_job:
                msg += f" و به فرصت شغلی '{target_job.title}' انتساب یافتند."
            messages.success(request, msg)

        if errors:
            err_msg = "برخی ردیف‌ها با خطا مواجه شدند:<br>" + "<br>".join(errors)
            messages.error(request, err_msg)

        return redirect(redirect_url)


from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth.forms import PasswordChangeForm

class CandidatePasswordChangeView(LoginRequiredMixin, PasswordChangeView):
    form_class = PasswordChangeForm
    template_name = 'candidates/password_change.html'

    def get_success_url(self):
        user = self.request.user
        if hasattr(user, 'profile') and user.profile.role == UserProfile.ROLE_CANDIDATE:
            return reverse('candidate_dashboard')
        return reverse('dashboard')

    def form_valid(self, form):
        from django.contrib import messages
        messages.success(self.request, "رمز عبور شما با موفقیت تغییر یافت.")
        return super().form_valid(form)


class CandidateResumePrintView(LoginRequiredMixin, RoleRequiredMixin, View):
    """نمایش رزومه چاپ‌آماده بر اساس اطلاعات موجود در سامانه"""
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get(self, request, pk):
        candidate = get_object_or_404(Candidate, pk=pk, is_deleted=False)
        context = {
            'candidate': candidate,
            'education_list': candidate.education.filter(is_deleted=False).order_by('-graduation_year'),
            'experience_list': candidate.experience.filter(is_deleted=False).order_by('-start_date'),
            'language_list': candidate.languages.filter(is_deleted=False),
            'skill_list': candidate.skills.filter(is_deleted=False),
            'certificate_list': candidate.certificates.filter(is_deleted=False).order_by('-issue_date'),
            'applications': candidate.applications.filter(is_deleted=False).select_related('job', 'current_stage'),
        }
        return render(request, 'candidates/candidate_resume_print.html', context)


class CandidateTranscriptPrintView(LoginRequiredMixin, RoleRequiredMixin, View):
    """کارنامه آزمون / ارزیابی یک درخواست شغلی خاص"""
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get(self, request, pk):
        application = get_object_or_404(
            JobApplication,
            pk=pk,
            is_deleted=False
        )
        stage_states = application.stage_states.filter(
            is_deleted=False
        ).select_related('stage', 'evaluator').order_by('stage__sequence')
        context = {
            'application': application,
            'candidate': application.candidate,
            'job': application.job,
            'stage_states': stage_states,
        }
        return render(request, 'candidates/candidate_transcript_print.html', context)


# ─────────────────────────────────────────────────────────────────────────────
# Component 1: Stage Rollback — برآورد مرحله
# ─────────────────────────────────────────────────────────────────────────────

class StageRollbackView(LoginRequiredMixin, RoleRequiredMixin, View):
    """
    برگرداندن متقاضی به مرحله قبل (یا هر مرحله مشخص‌شده) جهت ورود مجدد نمرات.
    وضعیت مرحله به PENDING برمی‌گردد؛ نمرات خارجی قبلی پاک می‌شوند.
    """
    allowed_roles = [
        UserProfile.ROLE_ADMIN,
        UserProfile.ROLE_RECRUITMENT_SPECIALIST,
        UserProfile.ROLE_RECRUITMENT_DIRECTOR,
    ]

    def post(self, request, app_id, stage_id):
        from .models import ExternalInterviewerScore

        app = get_object_or_404(
            JobApplication, pk=app_id, is_deleted=False
        )
        target_stage = get_object_or_404(
            JobOpportunityStage, pk=stage_id, job=app.job, is_deleted=False
        )
        stage_state = get_object_or_404(
            ApplicationStageState,
            application=app,
            stage=target_stage,
            is_deleted=False
        )

        bypass_locks = request.POST.get('bypass_locks') == '1' or request.POST.get('bypass_locks') == 'on'

        with transaction.atomic():
            # 1. برگرداندن مرحله فعلی متقاضی
            app.current_stage = target_stage
            app.save(update_fields=['current_stage'])

            # 2. اگه وضعیت نهایی درخواست رد شده بود، به in_progress برگردان
            if app.status == JobApplication.STATUS_REJECTED:
                app.status = JobApplication.STATUS_IN_PROGRESS
                app.save(update_fields=['status'])

            # 3. ریست وضعیت مرحله به PENDING و پاکسازی نمره
            stage_state.status = ApplicationStageState.STATUS_PENDING
            stage_state.score = 0.0
            stage_state.is_conditional_pass = False
            stage_state.save(update_fields=['status', 'score', 'is_conditional_pass'])

            # 4. حذف نمرات خارجی قبلی این مرحله (soft-delete)
            ExternalInterviewerScore.objects.filter(
                stage_state=stage_state,
                is_deleted=False
            ).update(is_deleted=True)

        stages = list(app.job.stages.filter(is_deleted=False).order_by('sequence'))
        state_map = {s.stage_id: s for s in app.stage_states.filter(is_deleted=False)}
        app.stage_cells = []
        for stg in stages:
            s_cell = state_map.get(stg.id)
            if not s_cell:
                s_cell = ApplicationStageState.objects.create(
                    application=app,
                    stage=stg,
                    status=ApplicationStageState.STATUS_PENDING,
                    score=0.0
                )
            app.stage_cells.append(s_cell)
        return render(request, 'candidates/partials/score_entry_row.html', {
            'app': app,
            'stages': stages,
            'bypass_locks': bypass_locks,
        })


# ─────────────────────────────────────────────────────────────────────────────
# Component 2: مصاحبه‌گران پیش‌فرض فرصت شغلی
# ─────────────────────────────────────────────────────────────────────────────

class ManageJobDefaultInterviewersView(LoginRequiredMixin, RoleRequiredMixin, View):
    """
    HTMX view: لیست مصاحبه‌گران پیش‌فرض یک فرصت شغلی را مدیریت می‌کند.
    GET  → قالب مودال با فرم و جدول.
    POST → ذخیره مصاحبه‌گران پیش‌فرض (حذف قبلی‌ها و ایجاد مجدد).
    """
    allowed_roles = [
        UserProfile.ROLE_ADMIN,
        UserProfile.ROLE_RECRUITMENT_SPECIALIST,
        UserProfile.ROLE_RECRUITMENT_DIRECTOR,
    ]

    def get(self, request, job_id):
        from .models import JobDefaultInterviewer
        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
        interviewers = JobDefaultInterviewer.objects.filter(
            job=job, is_deleted=False
        ).order_by('id')
        return render(request, 'candidates/partials/manage_default_interviewers.html', {
            'job': job,
            'interviewers': interviewers,
        })

    def post(self, request, job_id):
        from .models import JobDefaultInterviewer
        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)

        names = request.POST.getlist('interviewer_name[]')
        weights = request.POST.getlist('weight[]')

        with transaction.atomic():
            # نرم‌حذف همه مصاحبه‌گران قبلی
            JobDefaultInterviewer.objects.filter(job=job, is_deleted=False).update(is_deleted=True)

            for i, name in enumerate(names):
                name = name.strip()
                if not name:
                    continue
                try:
                    weight_val = int(weights[i])
                    if weight_val < 1:
                        weight_val = 100
                except (ValueError, IndexError):
                    weight_val = 100

                JobDefaultInterviewer.objects.create(
                    job=job,
                    interviewer_name=name,
                    weight=weight_val
                )

        interviewers = JobDefaultInterviewer.objects.filter(
            job=job, is_deleted=False
        ).order_by('id')
        return render(request, 'candidates/partials/manage_default_interviewers.html', {
            'job': job,
            'interviewers': interviewers,
            'saved': True,
        })

    def delete(self, request, job_id):
        """حذف یک مصاحبه‌گر منفرد (با job_id و interviewer pk در query string)"""
        from .models import JobDefaultInterviewer
        interviewer_pk = request.GET.get('pk')
        if interviewer_pk:
            JobDefaultInterviewer.objects.filter(
                pk=interviewer_pk, job_id=job_id, is_deleted=False
            ).update(is_deleted=True)
        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
        interviewers = JobDefaultInterviewer.objects.filter(
            job=job, is_deleted=False
        ).order_by('id')
        return render(request, 'candidates/partials/manage_default_interviewers.html', {
            'job': job,
            'interviewers': interviewers,
        })


# ─────────────────────────────────────────────────────────────────────────────
# Component 3: ورود سریع نمرات گروهی مصاحبه‌گران — Bulk Interview Scores
# ─────────────────────────────────────────────────────────────────────────────

class BulkInterviewScoresView(LoginRequiredMixin, RoleRequiredMixin, View):
    """
    GET  → جدول ماتریسی (متقاضی × مصاحبه‌گران پیش‌فرض) برای ورود سریع نمرات.
    POST → ذخیره همه نمرات به‌عنوان ExternalInterviewerScore و محاسبه مجدد score هر state.
    """
    allowed_roles = [
        UserProfile.ROLE_ADMIN,
        UserProfile.ROLE_RECRUITMENT_SPECIALIST,
        UserProfile.ROLE_RECRUITMENT_DIRECTOR,
    ]

    def _get_stage_states(self, job, stage):
        from django.db.models import Exists, OuterRef
        prior_failed_subquery = ApplicationStageState.objects.filter(
            application=OuterRef('application'),
            stage__sequence__lt=OuterRef('stage__sequence'),
            status=ApplicationStageState.STATUS_FAILED,
            is_conditional_pass=False,
            is_deleted=False
        )
        return ApplicationStageState.objects.filter(
            application__job=job,
            application__status=JobApplication.STATUS_IN_PROGRESS,
            application__is_deleted=False,
            stage=stage,
            is_deleted=False,
        ).annotate(
            has_failed_prior=Exists(prior_failed_subquery)
        ).filter(has_failed_prior=False).select_related(
            'application__candidate'
        ).order_by('application__candidate__last_name')

    def get(self, request):
        from .models import JobDefaultInterviewer
        job_id = request.GET.get('job_id')
        stage_id = request.GET.get('stage_id')

        if not job_id or not stage_id:
            return render(request, 'candidates/partials/bulk_interview_scores_panel.html', {
                'error': 'ابتدا یک فرصت شغلی و مرحله انتخاب کنید.'
            })

        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
        stage = get_object_or_404(JobOpportunityStage, pk=stage_id, job=job, is_deleted=False)

        if stage.stage_type != 'INTERVIEW':
            return render(request, 'candidates/partials/bulk_interview_scores_panel.html', {
                'error': 'ورود سریع نمرات فقط برای مراحل مصاحبه در دسترس است.'
            })

        default_interviewers = JobDefaultInterviewer.objects.filter(
            job=job, is_deleted=False
        ).order_by('id')

        stage_states = self._get_stage_states(job, stage)

        # برای هر state، نمرات خارجی موجود را بارگذاری کن
        for state in stage_states:
            existing = {
                es.interviewer_name: es
                for es in state.external_interviewer_scores.filter(is_deleted=False)
            }
            state.interviewer_cells = [
                {
                    'interviewer': iv,
                    'existing_score': existing.get(iv.interviewer_name),
                }
                for iv in default_interviewers
            ]

        return render(request, 'candidates/partials/bulk_interview_scores_panel.html', {
            'job': job,
            'stage': stage,
            'default_interviewers': default_interviewers,
            'stage_states': stage_states,
        })

    def post(self, request):
        from .models import JobDefaultInterviewer, ExternalInterviewerScore
        job_id = request.POST.get('job_id')
        stage_id = request.POST.get('stage_id')

        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
        stage = get_object_or_404(JobOpportunityStage, pk=stage_id, job=job, is_deleted=False)
        default_interviewers = JobDefaultInterviewer.objects.filter(
            job=job, is_deleted=False
        ).order_by('id')
        stage_states = self._get_stage_states(job, stage)

        with transaction.atomic():
            for state in stage_states:
                has_score = False
                any_changes = False
                for iv in default_interviewers:
                    field_key = f'score_{state.pk}_{iv.pk}'
                    raw = request.POST.get(field_key, '').strip()
                    if raw == '':
                        # Delete existing score if it was cleared
                        es_obj = ExternalInterviewerScore.objects.filter(
                            stage_state=state,
                            interviewer_name=iv.interviewer_name,
                            is_deleted=False
                        ).first()
                        if es_obj:
                            es_obj.is_deleted = True
                            es_obj.save()
                            any_changes = True
                        continue
                    try:
                        score_val = float(raw)
                    except ValueError:
                        continue

                    has_score = True
                    # upsert: به‌روزرسانی یا ایجاد
                    es_obj = ExternalInterviewerScore.objects.filter(
                        stage_state=state,
                        interviewer_name=iv.interviewer_name,
                        is_deleted=False
                    ).first()
                    if es_obj:
                        if es_obj.score != score_val or es_obj.weight != iv.weight:
                            es_obj.score = score_val
                            es_obj.weight = iv.weight
                            es_obj.save(update_fields=['score', 'weight'])
                            any_changes = True
                    else:
                        ExternalInterviewerScore.objects.create(
                            stage_state=state,
                            interviewer_name=iv.interviewer_name,
                            score=score_val,
                            weight=iv.weight
                        )
                        any_changes = True

                if any_changes or has_score:
                    if state.external_interviewer_scores.filter(is_deleted=False).exists():
                        state.is_manually_edited = False
                        state.save()
                    else:
                        state.score = 0.0
                        state.status = ApplicationStageState.STATUS_PENDING
                        state.is_manually_edited = False
                        state.save()

        response = render(request, 'candidates/partials/bulk_interview_scores_panel.html', {
            'job': job,
            'stage': stage,
            'default_interviewers': default_interviewers,
            'stage_states': self._get_stage_states(job, stage),
            'saved': True,
            '_reload_list': True,
        })
        response['HX-Trigger'] = 'refresh-score-entry-list'
        return response


# ─────────────────────────────────────────────────────────────────────────────
# Excel Import/Export for Interview Scores
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_number(raw):
    """تبدیل اعداد فارسی/عربی به لاتین و تبدیل به float"""
    if raw is None:
        return None
    raw = str(raw).strip()
    # جدول تبدیل: ۰-۹ (فارسی U+06F0-U+06F9) و ٠-٩ (عربی U+0660-U+0669)
    for persian, arabic, western in zip(
        '۰۱۲۳۴۵۶۷۸۹', '٠١٢٣٤٥٦٧٨٩', '0123456789'
    ):
        raw = raw.replace(persian, western).replace(arabic, western)
    # پاکسازی کاما فارسی/عربی
    raw = raw.replace('٫', '.').replace('،', '').replace(',', '')
    try:
        return float(raw)
    except ValueError:
        return None


class DownloadInterviewScoresTemplateView(LoginRequiredMixin, RoleRequiredMixin, View):
    """
    دانلود قالب اکسل برای ورود نمرات مصاحبه‌گران.
    ستون‌ها: کد ملی | نام و نام خانوادگی | [مصاحبه‌گر ۱] | [مصاحبه‌گر ۲] | ...
    سطرها: متقاضیان فعال مرحله انتخاب‌شده.
    """
    allowed_roles = [
        UserProfile.ROLE_ADMIN,
        UserProfile.ROLE_RECRUITMENT_SPECIALIST,
        UserProfile.ROLE_RECRUITMENT_DIRECTOR,
    ]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from .models import JobDefaultInterviewer

        job_id = request.GET.get('job_id')
        stage_id = request.GET.get('stage_id')

        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
        stage = get_object_or_404(JobOpportunityStage, pk=stage_id, job=job, is_deleted=False)

        default_interviewers = list(
            JobDefaultInterviewer.objects.filter(job=job, is_deleted=False).order_by('id')
        )

        # واکشی متقاضیان فعال (مشابه BulkInterviewScoresView)
        from django.db.models import Exists, OuterRef
        prior_failed_subquery = ApplicationStageState.objects.filter(
            application=OuterRef('application'),
            stage__sequence__lt=OuterRef('stage__sequence'),
            status=ApplicationStageState.STATUS_FAILED,
            is_conditional_pass=False,
            is_deleted=False
        )
        stage_states = ApplicationStageState.objects.filter(
            application__job=job,
            application__status=JobApplication.STATUS_IN_PROGRESS,
            application__is_deleted=False,
            stage=stage,
            is_deleted=False,
        ).annotate(
            has_failed_prior=Exists(prior_failed_subquery)
        ).filter(has_failed_prior=False).select_related(
            'application__candidate'
        ).order_by('application__candidate__last_name')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "نمرات مصاحبه"
        ws.sheet_view.rightToLeft = True

        # استایل هدر
        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        interviewer_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right = Alignment(horizontal='right', vertical='center')
        thin = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1'),
        )

        competencies = list(stage.competencies.filter(is_deleted=False).order_by('id'))

        # سطر هدر — ردیف ۲
        headers = ['کد ملی', 'نام و نام خانوادگی', 'وضعیت فعلی']
        
        # سطر اول — توضیح
        total_cols = 3
        if competencies:
            for iv in default_interviewers:
                for comp in competencies:
                    headers.append(f"{iv.interviewer_name} - {comp.name}")
                    total_cols += 1
        else:
            for iv in default_interviewers:
                headers.append(f'{iv.interviewer_name}\n(وزن: {iv.weight}%)')
                total_cols += 1

        headers.append('توضیحات ارزیاب')
        total_cols += 1

        ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
        desc_cell = ws['A1']
        desc_cell.value = f'قالب ورود نمرات مصاحبه | فرصت شغلی: {job.title} | مرحله: {stage.name} | ستون‌های آبی را ویرایش نکنید. فقط نمرات (۰-۱۰۰) وارد کنید.'
        desc_cell.font = Font(name='Calibri', bold=True, color='1E3A5F', size=10)
        desc_cell.alignment = right
        ws.row_dimensions[1].height = 28

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            if col_idx <= 3:
                cell.fill = header_fill
            elif col_idx == total_cols:
                cell.fill = interviewer_fill
            else:
                cell.fill = interviewer_fill
            cell.alignment = center
            cell.border = thin

        ws.row_dimensions[2].height = 36

        # سطرهای داده
        status_map = {
            ApplicationStageState.STATUS_PENDING: 'در انتظار',
            ApplicationStageState.STATUS_COMPLETED: 'قبول',
            ApplicationStageState.STATUS_FAILED: 'مردود',
        }
        row_fill_even = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        row_fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        for row_idx, state in enumerate(stage_states, start=3):
            candidate = state.application.candidate
            row_fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd

            data = [
                candidate.national_id,
                f'{candidate.first_name} {candidate.last_name}',
                status_map.get(state.status, state.status),
            ]

            if competencies:
                for iv in default_interviewers:
                    es = state.external_interviewer_scores.filter(
                        interviewer_name=iv.interviewer_name, 
                        is_deleted=False
                    ).first()
                    existing_comp_scores = {}
                    if es:
                        existing_comp_scores = {
                            cs.competency_id: cs.score 
                            for cs in es.competency_scores.filter(is_deleted=False)
                        }
                    for comp in competencies:
                        data.append(existing_comp_scores.get(comp.id, ''))
            else:
                existing = {
                    es.interviewer_name: es.score
                    for es in state.external_interviewer_scores.filter(is_deleted=False)
                }
                for iv in default_interviewers:
                    data.append(existing.get(iv.interviewer_name, ''))

            data.append(state.notes if state.notes else '')

            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin
                cell.fill = row_fill
                if col_idx <= 3:
                    cell.alignment = right
                    cell.font = Font(name='Calibri', size=10)
                    if col_idx == 1:  # کد ملی — قفل
                        cell.font = Font(name='Calibri', size=10, color='374151')
                elif col_idx == total_cols:
                    cell.alignment = right
                    cell.font = Font(name='Calibri', size=10)
                else:
                    cell.alignment = center
                    cell.number_format = '0.00'

        # عرض ستون‌ها
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 12
        for i in range(total_cols - 4):
            col_letter = get_column_letter(4 + i)
            ws.column_dimensions[col_letter].width = 15
        ws.column_dimensions[get_column_letter(total_cols)].width = 25

        # یادداشت پایین
        note_row = len(list(stage_states)) + 3
        ws.merge_cells(f'A{note_row}:{get_column_letter(total_cols)}{note_row}')
        note_cell = ws[f'A{note_row}']
        note_cell.value = '📌 راهنما: فقط ستون‌های آبی رنگ (نمرات) را ویرایش کنید. کد ملی را تغییر ندهید. اعداد فارسی نیز قابل قبول است.'
        note_cell.font = Font(name='Calibri', italic=True, color='6B7280', size=9)
        note_cell.alignment = right

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="interview_scores_{job.code}_{stage.name}.xlsx"'
        wb.save(response)
        return response


class ImportInterviewScoresExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    """
    آپلود فایل اکسل نمرات مصاحبه‌گران و ذخیره ExternalInterviewerScore ها.
    - ستون اول: کد ملی (برای تطابق)
    - ستون‌های ۴ به بعد: نمرات مصاحبه‌گران (مطابق قالب دانلود شده)
    - اعداد فارسی/عربی به‌صورت خودکار تبدیل می‌شوند
    """
    allowed_roles = [
        UserProfile.ROLE_ADMIN,
        UserProfile.ROLE_RECRUITMENT_SPECIALIST,
        UserProfile.ROLE_RECRUITMENT_DIRECTOR,
    ]

    def post(self, request):
        import openpyxl
        from .models import JobDefaultInterviewer, ExternalInterviewerScore, ExternalInterviewerCompetencyScore
        from django.db import transaction

        job_id = request.POST.get('job_id')
        stage_id = request.POST.get('stage_id')
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            return render(request, 'candidates/partials/bulk_interview_scores_panel.html', {
                'error': 'فایل اکسل انتخاب نشده است.'
            })

        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
        stage = get_object_or_404(JobOpportunityStage, pk=stage_id, job=job, is_deleted=False)
        default_interviewers = list(
            JobDefaultInterviewer.objects.filter(job=job, is_deleted=False).order_by('id')
        )

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
        except Exception as e:
            return render(request, 'candidates/partials/bulk_interview_scores_panel.html', {
                'error': f'خطا در باز کردن فایل اکسل: {e}'
            })

        # خواندن هدر (ردیف ۲ در قالب ما) و شناسایی ستون‌های مصاحبه‌گران
        header_row = None
        for row in ws.iter_rows(min_row=2, max_row=3):
            if row[0].value and 'کد ملی' in str(row[0].value):
                header_row = [cell.value for cell in row]
                header_row_idx = row[0].row
                break

        if header_row is None:
            return render(request, 'candidates/partials/bulk_interview_scores_panel.html', {
                'error': 'فرمت فایل اکسل معتبر نیست. لطفاً از قالب دانلود شده استفاده کنید.'
            })

        # نگاشت: نام مصاحبه‌گر → شماره ستون (0-indexed)
        interviewer_col_map = {}
        for iv in default_interviewers:
            for col_idx, header in enumerate(header_row):
                if header and iv.interviewer_name in str(header):
                    interviewer_col_map[iv.interviewer_name] = col_idx
                    break

        # نگاشت ستون توضیحات ارزیاب
        notes_col_idx = None
        for col_idx, header in enumerate(header_row):
            if header and any(key in str(header).lower() for key in ["توضیحات ارزیاب", "توضیحات", "توضیح", "یادداشت", "notes", "comment", "description"]):
                notes_col_idx = col_idx
                break

        saved_count = 0
        error_rows = []

        with transaction.atomic():
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if not row[0]:  # ردیف خالی
                    continue
                national_id = str(row[0]).strip()
                # تبدیل اعداد فارسی در کد ملی هم
                national_id = str(_normalize_number(national_id) or national_id).replace('.0', '')

                # پیدا کردن متقاضی
                from apps.candidates.models import Candidate as CandidateModel
                candidate = CandidateModel.objects.filter(
                    national_id=national_id, is_deleted=False
                ).first()
                if not candidate:
                    error_rows.append(f'کد ملی پیدا نشد: {national_id}')
                    continue

                state = ApplicationStageState.objects.filter(
                    application__job=job,
                    application__candidate=candidate,
                    stage=stage,
                    is_deleted=False,
                ).first()
                if not state:
                    error_rows.append(f'رکورد مرحله برای {national_id} پیدا نشد')
                    continue

                has_score = False
                for iv in default_interviewers:
                    col_idx = interviewer_col_map.get(iv.interviewer_name)
                    if col_idx is None or col_idx >= len(row):
                        continue
                    raw_val = row[col_idx]
                    score_val = _normalize_number(raw_val)
                    if score_val is None:
                        continue
                    # clamp
                    score_val = max(0.0, min(100.0, score_val))
                    has_score = True

                    es_obj = ExternalInterviewerScore.objects.filter(
                        stage_state=state,
                        interviewer_name=iv.interviewer_name,
                        is_deleted=False
                    ).first()
                    if es_obj:
                        es_obj.score = score_val
                        es_obj.weight = iv.weight
                        es_obj.save(update_fields=['score', 'weight'])
                    else:
                        ExternalInterviewerScore.objects.create(
                            stage_state=state,
                            interviewer_name=iv.interviewer_name,
                            score=score_val,
                            weight=iv.weight
                        )

                # Import stage state notes if present
                notes_val = None
                if notes_col_idx is not None and notes_col_idx < len(row):
                    notes_val = row[notes_col_idx]
                    if notes_val is not None:
                        state.notes = str(notes_val).strip()

                if has_score or (notes_col_idx is not None and notes_val is not None):
                    state.is_manually_edited = False
                    state.save()
                    saved_count += 1

        # بازگشت جدول به‌روز شده
        from django.db.models import Exists, OuterRef
        prior_failed_subquery = ApplicationStageState.objects.filter(
            application=OuterRef('application'),
            stage__sequence__lt=OuterRef('stage__sequence'),
            status=ApplicationStageState.STATUS_FAILED,
            is_conditional_pass=False,
            is_deleted=False
        )
        stage_states = ApplicationStageState.objects.filter(
            application__job=job,
            application__status=JobApplication.STATUS_IN_PROGRESS,
            application__is_deleted=False,
            stage=stage,
            is_deleted=False,
        ).annotate(
            has_failed_prior=Exists(prior_failed_subquery)
        ).filter(has_failed_prior=False).select_related(
            'application__candidate'
        ).order_by('application__candidate__last_name')

        for state in stage_states:
            existing = {
                es.interviewer_name: es
                for es in state.external_interviewer_scores.filter(is_deleted=False)
            }
            state.interviewer_cells = [
                {'interviewer': iv, 'existing_score': existing.get(iv.interviewer_name)}
                for iv in default_interviewers
            ]

        response = render(request, 'candidates/partials/bulk_interview_scores_panel.html', {
            'job': job,
            'stage': stage,
            'default_interviewers': default_interviewers,
            'stage_states': stage_states,
            'saved': True,
            'import_summary': {
                'saved_count': saved_count,
                'error_rows': error_rows,
            },
        })
        response['HX-Trigger'] = 'refresh-score-entry-list'
        return response


class JobOpportunityReportView(LoginRequiredMixin, RoleRequiredMixin, View):
    """
    نمایش شناسنامه جامع فرصت شغلی (گزارش آماری و فرآیند جذب متقاضیان)
    امکان چاپ به عنوان PDF
    """
    allowed_roles = [
        UserProfile.ROLE_ADMIN,
        UserProfile.ROLE_RECRUITMENT_DIRECTOR,
        UserProfile.ROLE_RECRUITMENT_SPECIALIST,
        UserProfile.ROLE_JOB_CLASSIFICATION_USER,
        UserProfile.ROLE_DEPARTMENT_USER,
        UserProfile.ROLE_READ_ONLY_AUDITOR,
    ]

    def get(self, request, pk):
        from django.db.models import Count, Q, Min, Max, Avg, Exists, OuterRef
        from django.utils import timezone
        from apps.jobs.models import JobOpportunity, JobOpportunityStage
        from apps.candidates.models import JobApplication, ApplicationStageState
        
        job = get_object_or_404(JobOpportunity, pk=pk, is_deleted=False)
        stages = list(job.stages.filter(is_deleted=False).order_by('sequence'))
        applications = job.applications.filter(is_deleted=False)
        
        # ۱. اطلاعات و آمارهای ثبت‌نامی و وضعیت‌ها
        total_registered = applications.count()
        status_counts = applications.aggregate(
            selected=Count('id', filter=Q(status='SELECTED')),
            reserve=Count('id', filter=Q(status='RESERVE')),
            inprogress=Count('id', filter=Q(status='IN_PROGRESS')),
            rejected=Count('id', filter=Q(status='REJECTED')),
        )
        
        # محاسبه روزهای سپری‌شده (مدت زمان جذب)
        start_date = job.start_date or job.created_at.date()
        if job.status in [JobOpportunity.STATUS_CLOSED, JobOpportunity.STATUS_CANCELLED, JobOpportunity.STATUS_SUSPENDED] and job.end_date:
            end_date = job.end_date
        elif job.status in [JobOpportunity.STATUS_CLOSED, JobOpportunity.STATUS_CANCELLED, JobOpportunity.STATUS_SUSPENDED]:
            end_date = job.updated_at.date()
        else:
            end_date = timezone.now().date()
        
        duration_days = (end_date - start_date).days
        if duration_days < 0:
            duration_days = 0

        # ۲. محاسبه جزئیات و آمار تفکیکی مراحل
        stage_details = []
        for stage in stages:
            # بررسی برنامه‌ریزی جذب (JobStagePlan)
            stage_plan = None
            if hasattr(job, 'recruitment_plan') and job.recruitment_plan and not job.recruitment_plan.is_deleted:
                stage_plan = stage.planning_states.filter(plan=job.recruitment_plan, is_deleted=False).first()
            
            # پیدا کردن متقاضیانی که به این مرحله رسیده‌اند (یعنی در مراحل قبل رد نشده‌اند)
            prior_failed_subquery = ApplicationStageState.objects.filter(
                application=OuterRef('application'),
                stage__sequence__lt=stage.sequence,
                status=ApplicationStageState.STATUS_FAILED,
                is_conditional_pass=False,
                is_deleted=False
            )
            stage_states = ApplicationStageState.objects.filter(
                application__job=job,
                application__is_deleted=False,
                stage=stage,
                is_deleted=False
            ).annotate(
                has_failed_prior=Exists(prior_failed_subquery)
            ).filter(has_failed_prior=False)
            
            total_entered = stage_states.count()
            passed = stage_states.filter(status=ApplicationStageState.STATUS_COMPLETED).count()
            failed = stage_states.filter(status=ApplicationStageState.STATUS_FAILED).count()
            pending = stage_states.filter(status=ApplicationStageState.STATUS_PENDING).count()
            
            # تاریخ واقعی اولین و آخرین ارزیابی
            evaluated_states = stage_states.filter(
                status__in=[ApplicationStageState.STATUS_COMPLETED, ApplicationStageState.STATUS_FAILED]
            )
            eval_dates = evaluated_states.aggregate(min_date=Min('evaluation_date'), max_date=Max('evaluation_date'))
            actual_start = eval_dates['min_date']
            actual_end = eval_dates['max_date']
            
            # Fallback to stage plan dates if evaluation_date is not set on state objects
            if not actual_start and stage_plan:
                actual_start = stage_plan.planned_start_date
            if not actual_end and stage_plan:
                actual_end = stage_plan.planned_end_date
            
            # نمره‌ها
            scores = evaluated_states.aggregate(
                min_score=Min('score'),
                max_score=Max('score'),
                avg_score=Avg('score')
            )
            
            stage_details.append({
                'stage': stage,
                'plan': stage_plan,
                'total_entered': total_entered,
                'passed': passed,
                'failed': failed,
                'pending': pending,
                'actual_start': actual_start,
                'actual_end': actual_end,
                'min_score': scores['min_score'],
                'max_score': scores['max_score'],
                'avg_score': scores['avg_score'],
            })

        # ۳. متقاضیان پذیرفته‌شده نهایی
        selected_applications = applications.filter(status='SELECTED').select_related('candidate').order_by('-final_score')
        for app in selected_applications:
            # سوابق تحصیلی بالاترین مقطع
            edu = app.candidate.education.filter(is_deleted=False).order_by('-graduation_year').first()
            app.highest_education = edu
            # آخرین سابقه کاری
            exp = app.candidate.experience.filter(is_deleted=False).order_by('-start_date').first()
            app.latest_experience = exp

        context = {
            'job': job,
            'total_registered': total_registered,
            'status_counts': status_counts,
            'duration_days': duration_days,
            'start_date': start_date,
            'end_date': end_date,
            'stage_details': stage_details,
            'selected_applications': selected_applications,
            'print_date': timezone.now().date(),
        }
        return render(request, 'jobs/job_report.html', context)


class DataIntegrityDashboardView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    template_name = 'candidates/data_integrity.html'
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.candidates.integrity_engine import IntegrityScanner
        from apps.core.models import AuditLog
        
        force_refresh = self.request.GET.get('refresh') == '1'
        scan_results = IntegrityScanner.run_scan(force=force_refresh)
        
        # Fetch recent log entries for candidate integrity models
        recent_actions = AuditLog.objects.filter(
            model_name__in=['jobapplication', 'applicationstagestate', 'applicationstagestate_bulk']
        ).select_related('user').order_by('-timestamp')[:10]
        
        context.update({
            'scan_results': scan_results,
            'recent_actions': recent_actions,
        })
        return context


class ResolveDiscrepancyView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def post(self, request):
        from apps.candidates.integrity_engine import IntegrityScanner
        from apps.core.models import AuditLog
        
        issue_code = request.POST.get('issue_code')
        entity_id = request.POST.get('entity_id')
        action_key = request.POST.get('action_key')
        choice_val = request.POST.get('choice')
        
        success, msg = IntegrityScanner.resolve_issue(
            issue_code=issue_code,
            entity_id=entity_id,
            action_key=action_key,
            choice_val=choice_val,
            user=request.user
        )
        
        if request.headers.get('HX-Request'):
            results = IntegrityScanner.run_scan(force=False)
            response_content = f"""
            <div class="alert alert-success border-success border-opacity-25 bg-success bg-opacity-10 text-success text-xs py-2 px-3 rounded d-flex align-items-center gap-1.5 mb-0" style="display: inline-flex !important;">
                <span>✓ {msg}</span>
            </div>
            <span id="integrity-total-count" hx-swap-oob="true">{results['total_count']}</span>
            <span id="integrity-clean-pct" hx-swap-oob="true">{results['clean_percentage']}%</span>
            """
            
            # Fetch recent actions to update OOB
            recent_actions = AuditLog.objects.filter(
                model_name__in=['jobapplication', 'applicationstagestate', 'applicationstagestate_bulk']
            ).select_related('user').order_by('-timestamp')[:10]
            
            log_rows_html = ""
            for log in recent_actions:
                user_name = log.user.get_full_name() if log.user and log.user.get_full_name() else (log.user.username if log.user else 'سیستم')
                log_rows_html += f"""
                <tr id="action-log-row-{log.id}">
                    <td class="ps-3 text-xs">{log.timestamp.strftime('%Y/%m/%d %H:%M')}</td>
                    <td class="text-xs">{user_name}</td>
                    <td class="text-xs font-semibold">{log.get_action_type_display()}</td>
                    <td class="text-xs">{log.model_name} (ID: {log.object_id})</td>
                    <td class="pe-3 text-end">
                        <button class="btn btn-outline-danger btn-xxs" hx-post="/candidates/data-integrity/undo/{log.id}/" hx-target="#action-log-row-{log.id}" hx-swap="outerHTML">بازگردانی (Undo)</button>
                    </td>
                </tr>
                """
            
            response_content += f"""
            <tbody id="integrity-actions-body" hx-swap-oob="true">
                {log_rows_html}
            </tbody>
            """
            return HttpResponse(response_content)
            
        return redirect('data_integrity_dashboard')


class UndoIntegrityActionView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN]

    def post(self, request, log_id):
        from apps.candidates.integrity_engine import undo_audit_log_action, IntegrityScanner
        success, msg = undo_audit_log_action(log_id)
        
        if request.headers.get('HX-Request'):
            if success:
                results = IntegrityScanner.run_scan(force=True)
                response_content = f"""
                <tr class="animate__animated animate__fadeOut" style="animation-duration: 0.5s;">
                    <td colspan="5" class="text-center py-2 bg-danger bg-opacity-10 text-danger text-xs font-semibold">
                        {msg}
                    </td>
                </tr>
                <span id="integrity-total-count" hx-swap-oob="true">{results['total_count']}</span>
                <span id="integrity-clean-pct" hx-swap-oob="true">{results['clean_percentage']}%</span>
                """
                return HttpResponse(response_content)
            else:
                return HttpResponse(f"<td colspan='5' class='text-center py-2 text-danger text-xs'>{msg}</td>", status=400)
                
        return redirect('data_integrity_dashboard')


class SelectedCandidatesListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    model = JobApplication
    template_name = 'candidates/selected_candidates_list.html'
    context_object_name = 'applications'
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]
    paginate_by = 15

    def get_queryset(self):
        from django.db.models import Q
        queryset = JobApplication.objects.filter(
            status=JobApplication.STATUS_SELECTED,
            is_deleted=False
        ).select_related('candidate', 'job').order_by(
            models.F('admission_date').desc(nulls_last=True),
            '-updated_at'
        )

        # 1. Search Query
        q = self.request.GET.get('q', '').strip()
        if q:
            queryset = queryset.filter(
                Q(candidate__first_name__icontains=q) |
                Q(candidate__last_name__icontains=q) |
                Q(candidate__national_id__icontains=q) |
                Q(job__title__icontains=q) |
                Q(job__code__icontains=q)
            )

        # 2. Job Opportunity Filter
        job_id = self.request.GET.get('job_opportunity', '').strip()
        if job_id:
            queryset = queryset.filter(job_id=job_id)

        # 3. Department Filter
        dept = self.request.GET.get('department', '').strip()
        if dept:
            queryset = queryset.filter(job__department=dept)

        return queryset

    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'excel':
            from apps.core.utils import export_to_excel_response
            from apps.core.templatetags.jalali_tags import to_jalali
            queryset = self.get_queryset()
            
            headers = [
                "نام و نام خانوادگی",
                "کد ملی",
                "شماره تماس",
                "ایمیل",
                "کد فرصت شغلی",
                "عنوان فرصت شغلی",
                "دپارتمان / بخش",
                "واحد سازمانی",
                "امتیاز کل ارزیابی",
                "تاریخ پذیرش نهایی"
            ]
            
            rows = []
            for app in queryset:
                rows.append([
                    f"{app.candidate.first_name} {app.candidate.last_name}",
                    app.candidate.national_id,
                    app.candidate.phone_number,
                    app.candidate.email,
                    app.job.code,
                    app.job.title,
                    app.job.department or "-",
                    app.job.unit or "-",
                    f"{app.final_score}%" if app.final_score is not None else "-",
                    to_jalali(app.admission_date) if app.admission_date else "-"
                ])
                
            return export_to_excel_response("لیست_پذیرفته_شدگان_نهایی.xlsx", headers, rows)
            
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.jobs.models import JobOpportunity
        
        context['q'] = self.request.GET.get('q', '').strip()
        context['selected_job_id'] = self.request.GET.get('job_opportunity', '').strip()
        context['selected_dept'] = self.request.GET.get('department', '').strip()
        
        # Unique active jobs and departments for filters
        context['jobs'] = JobOpportunity.objects.filter(is_deleted=False).order_by('title')
        context['departments'] = JobOpportunity.objects.filter(is_deleted=False).exclude(
            department=''
        ).exclude(department=None).values_list('department', flat=True).distinct().order_by('department')
        
        # Calculate selected candidates jobs and host departments count
        context['recruited_jobs_count'] = JobOpportunity.objects.filter(
            applications__status=JobApplication.STATUS_SELECTED,
            applications__is_deleted=False,
            is_deleted=False
        ).distinct().count()
        
        context['recruited_depts_count'] = JobOpportunity.objects.filter(
            applications__status=JobApplication.STATUS_SELECTED,
            applications__is_deleted=False,
            is_deleted=False
        ).exclude(
            department=''
        ).exclude(
            department=None
        ).values('department').distinct().count()
        
        return context


class EditAdmissionDateView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get(self, request, pk):
        app = get_object_or_404(JobApplication, pk=pk, is_deleted=False)
        return render(request, 'candidates/partials/admission_date_edit.html', {'app': app})


class UpdateAdmissionDateView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def post(self, request, pk):
        app = get_object_or_404(JobApplication, pk=pk, is_deleted=False)
        date_val = request.POST.get('admission_date', '').strip()
        
        if date_val:
            app.admission_date = parse_jalali_date(date_val)
        else:
            app.admission_date = None
            
        app.save(update_fields=['admission_date'])
        return render(request, 'candidates/partials/admission_date_view.html', {'app': app})

    def get(self, request, pk):
        # Fallback to standard view if GET is called (e.g. for cancel/discard action)
        app = get_object_or_404(JobApplication, pk=pk, is_deleted=False)
        return render(request, 'candidates/partials/admission_date_view.html', {'app': app})


class CandidatesByStageListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    model = JobApplication
    template_name = 'candidates/candidates_by_stage_list.html'
    context_object_name = 'applications'
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]
    paginate_by = 15

    def get_queryset(self):
        from django.db.models import Q
        queryset = JobApplication.objects.filter(
            status=JobApplication.STATUS_IN_PROGRESS,
            is_deleted=False
        ).exclude(job__status__in=['CLOSED', 'CANCELLED', 'SUSPENDED']).select_related('candidate', 'job', 'current_stage').order_by('-updated_at')

        # 1. Search Query
        q = self.request.GET.get('q', '').strip()
        if q:
            queryset = queryset.filter(
                Q(candidate__first_name__icontains=q) |
                Q(candidate__last_name__icontains=q) |
                Q(candidate__national_id__icontains=q) |
                Q(job__title__icontains=q) |
                Q(job__code__icontains=q)
            )

        # 2. Job Opportunity Filter
        job_id = self.request.GET.get('job_opportunity', '').strip()
        if job_id:
            queryset = queryset.filter(job_id=job_id)

        # 3. Department Filter
        dept = self.request.GET.get('department', '').strip()
        if dept:
            queryset = queryset.filter(job__department=dept)

        # 4. Stage Type Filter
        stage_type = self.request.GET.get('stage_type', '').strip()
        if stage_type:
            queryset = queryset.filter(job__status=stage_type)

        return queryset

    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'excel':
            from apps.core.utils import export_to_excel_response
            from apps.core.templatetags.jalali_tags import to_jalali
            queryset = self.get_queryset()
            
            headers = [
                "نام و نام خانوادگی",
                "کد ملی",
                "شماره تماس",
                "ایمیل",
                "کد فرصت شغلی",
                "عنوان فرصت شغلی",
                "دپارتمان / بخش",
                "مرحله ارزیابی جاری",
                "امتیاز کل ارزیابی",
                "آخرین فعالیت"
            ]
            
            rows = []
            for app in queryset:
                rows.append([
                    f"{app.candidate.first_name} {app.candidate.last_name}",
                    app.candidate.national_id,
                    app.candidate.phone_number,
                    app.candidate.email,
                    app.job.code,
                    app.job.title,
                    app.job.department or "-",
                    app.current_stage.name if app.current_stage else "شروع ارزیابی",
                    f"{app.final_score}%" if app.final_score is not None else "-",
                    to_jalali(app.updated_at) if app.updated_at else "-"
                ])
                
            return export_to_excel_response("متقاضیان_به_تفکیک_مراحل_ارزیابی.xlsx", headers, rows)
            
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.jobs.models import JobOpportunity
        from apps.jobs.models import STAGE_TYPE_CHOICES
        
        context['q'] = self.request.GET.get('q', '').strip()
        context['selected_job_id'] = self.request.GET.get('job_opportunity', '').strip()
        context['selected_dept'] = self.request.GET.get('department', '').strip()
        context['selected_stage_type'] = self.request.GET.get('stage_type', '').strip()
        
        # Unique active jobs, departments, and stage types
        context['jobs'] = JobOpportunity.objects.filter(is_deleted=False).order_by('title')
        context['departments'] = JobOpportunity.objects.filter(is_deleted=False).exclude(
            department=''
        ).exclude(department=None).values_list('department', flat=True).distinct().order_by('department')
        context['stage_types'] = STAGE_TYPE_CHOICES
        
        # Calculate stats for the summary cards
        active_apps = JobApplication.objects.filter(
            status=JobApplication.STATUS_IN_PROGRESS, 
            is_deleted=False
        ).exclude(job__status__in=['CLOSED', 'CANCELLED', 'SUSPENDED'])
        context['total_in_progress'] = active_apps.count()
        
        from django.db.models import Q
        context['screening_count'] = active_apps.filter(job__status='SCREENING').count()
        context['exam_count'] = active_apps.filter(job__status='EXAM').count()
        context['skill_count'] = active_apps.filter(job__status='SKILL_TEST').count()
        context['interview_count'] = active_apps.filter(job__status='INTERVIEW').count()
        context['assessment_count'] = active_apps.filter(job__status='ASSESSMENT').count()
        
        return context


class BulkSendNotificationModalView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_SPECIALIST, UserProfile.ROLE_RECRUITMENT_DIRECTOR]

    def get(self, request, job_id):
        from apps.accounts.models import SMSTemplate
        from apps.jobs.models import JobOpportunity
        from apps.candidates.models import JobApplication

        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
        app_ids_str = request.GET.get('app_ids', '')
        if not app_ids_str:
            return HttpResponse("متقاضیی انتخاب نشده است.", status=400)
            
        app_ids = [int(x) for x in app_ids_str.split(',') if x.isdigit()]
        applications = JobApplication.objects.filter(
            pk__in=app_ids, 
            job=job, 
            status=JobApplication.STATUS_IN_PROGRESS, 
            is_deleted=False
        ).prefetch_related('candidate', 'stage_states__stage')
        
        if not applications.exists():
            return HttpResponse("هیچ متقاضی واجد شرایطی یافت نشد.", status=400)
            
        templates = SMSTemplate.objects.filter(is_deleted=False).order_by('name')
        
        # Prepare previews or details for each
        preview_candidates = []
        for app in applications:
            state = app.stage_states.filter(stage=app.current_stage, is_deleted=False).first()
            preview_candidates.append({
                'app': app,
                'candidate': app.candidate,
                'stage': app.current_stage,
                'state': state
            })
            
        context = {
            'job': job,
            'preview_candidates': preview_candidates,
            'templates': templates,
            'app_ids_str': app_ids_str
        }
        return render(request, 'candidates/partials/bulk_send_notification_modal.html', context)


class BulkSendNotificationView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_SPECIALIST, UserProfile.ROLE_RECRUITMENT_DIRECTOR]

    def post(self, request, job_id):
        from apps.accounts.models import SMSTemplate
        from apps.jobs.models import JobOpportunity, OrganizationSetting
        from apps.candidates.models import JobApplication
        from apps.candidates.signals import send_dynamic_email, send_gateway_sms, render_notification_template
        from apps.accounts.views import render_template_text
        from django.contrib import messages

        job = get_object_or_404(JobOpportunity, pk=job_id, is_deleted=False)
        app_ids_str = request.POST.get('app_ids', '')
        if not app_ids_str:
            messages.error(request, "هیچ متقاضی انتخاب نشده است.")
            return redirect('job_pipeline', pk=job_id)

        app_ids = [int(x) for x in app_ids_str.split(',') if x.isdigit()]
        applications = JobApplication.objects.filter(
            pk__in=app_ids, 
            job=job, 
            status=JobApplication.STATUS_IN_PROGRESS, 
            is_deleted=False
        )

        org_setting = OrganizationSetting.get_active_setting()
        if not org_setting:
            messages.error(request, "تنظیمات سازمان یافت نشد.")
            return redirect('job_pipeline', pk=job_id)

        notification_type = request.POST.get('notification_type', 'SMART')
        send_sms = request.POST.get('send_sms') == 'on'
        send_email = request.POST.get('send_email') == 'on'
        custom_body = request.POST.get('custom_body', '').strip()

        if not send_sms and not send_email:
            messages.error(request, "لطفاً حداقل یک کانال ارتباطی (پیامک یا ایمیل) را انتخاب کنید.")
            return redirect('job_pipeline', pk=job_id)

        sent_count = 0
        for app in applications:
            state = app.stage_states.filter(stage=app.current_stage, is_deleted=False).first()
            if not state:
                continue

            import jdatetime
            date_str = ""
            if state.evaluation_date:
                jd = jdatetime.date.fromgregorian(date=state.evaluation_date)
                date_str = jd.strftime('%Y/%m/%d')
            time_str = state.evaluation_time or "10:00"

            if notification_type == 'SMART':
                stage_name_lower = state.stage.name.lower()
                is_exam = (state.stage.stage_type in ['EXAM', 'SKILL_TEST']) or any(kw in stage_name_lower for kw in ['آزمون', 'کتبی', 'مهارتی', 'سنجش', 'تخصصی', 'عمومی', 'عملکردی'])
                is_interview = (state.stage.stage_type in ['INTERVIEW', 'ASSESSMENT']) or any(kw in stage_name_lower for kw in ['مصاحبه', 'ارزیابی', 'کانون', 'گفتگو', 'حضوری', 'شایستگی'])

                if is_exam:
                    if send_email and org_setting.exam_email_enabled:
                        subject = render_notification_template(org_setting.exam_email_subject, app.candidate, job, stage_name=state.stage.name, date=date_str, time=time_str)
                        body = render_notification_template(org_setting.exam_email_body, app.candidate, job, stage_name=state.stage.name, date=date_str, time=time_str)
                        send_dynamic_email(org_setting, app.candidate.email, subject, body)
                    if send_sms and org_setting.exam_sms_enabled:
                        body = render_notification_template(org_setting.exam_sms_body, app.candidate, job, stage_name=state.stage.name, date=date_str, time=time_str)
                        send_gateway_sms(org_setting, app.candidate.phone_number, body)
                    sent_count += 1
                elif is_interview:
                    if send_email and org_setting.interview_email_enabled:
                        subject = render_notification_template(org_setting.interview_email_subject, app.candidate, job, stage_name=state.stage.name, date=date_str, time=time_str)
                        body = render_notification_template(org_setting.interview_email_body, app.candidate, job, stage_name=state.stage.name, date=date_str, time=time_str)
                        send_dynamic_email(org_setting, app.candidate.email, subject, body)
                    if send_sms and org_setting.interview_sms_enabled:
                        body = render_notification_template(org_setting.interview_sms_body, app.candidate, job, stage_name=state.stage.name, date=date_str, time=time_str)
                        send_gateway_sms(org_setting, app.candidate.phone_number, body)
                    sent_count += 1
            else:
                # Custom Template or Manual Body
                if notification_type.isdigit():
                    tmpl = SMSTemplate.objects.filter(pk=int(notification_type), is_deleted=False).first()
                    text = render_template_text(tmpl.body, app.candidate, job, state.stage, state, app) if tmpl else ""
                else:
                    text = render_template_text(custom_body, app.candidate, job, state.stage, state, app)

                if text:
                    if send_sms:
                        send_gateway_sms(org_setting, app.candidate.phone_number, text)
                    if send_email:
                        subject = f"اطلاعیه جدید از {org_setting.name}"
                        body = f"""<div dir="rtl" style="font-family: Tahoma, Arial, sans-serif; text-align: right; padding: 20px; line-height: 1.6;">
                            <p>{text}</p>
                        </div>"""
                        send_dynamic_email(org_setting, app.candidate.email, subject, body)
                    sent_count += 1

        messages.success(request, f"اعلانات با موفقیت برای {sent_count} متقاضی ارسال شد.")
        return redirect('job_pipeline', pk=job_id)


