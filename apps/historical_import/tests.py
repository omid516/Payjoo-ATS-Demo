import datetime
import jdatetime
from django.test import TestCase
from django.urls import reverse
from django.contrib.auth.models import User
from apps.accounts.models import UserProfile
from apps.jobs.models import JobOpportunity, WorkflowTemplate, WorkflowStageTemplate, JobOpportunityStage
from apps.candidates.models import Candidate, JobApplication, ApplicationStageState
from apps.historical_import.models import ImportSession, StagingJobOpportunity, StagingCandidate, ImportSessionLog
from apps.historical_import.utils import (
    normalize_persian_digits,
    parse_date_safely,
    validate_national_id,
    execute_final_import
)

class UtilsTestCase(TestCase):
    def test_normalize_persian_digits(self):
        self.assertEqual(normalize_persian_digits("۱۲۳۴۵۶۷۸۹۰"), "1234567890")
        self.assertEqual(normalize_persian_digits("١٢٣٤٥٦٧٨٩٠"), "1234567890")
        self.assertEqual(normalize_persian_digits("abc 123"), "abc 123")
        self.assertEqual(normalize_persian_digits(None), "")

    def test_parse_date_safely(self):
        # Jalali string format
        self.assertEqual(parse_date_safely("1402/05/12"), datetime.date(2023, 8, 3))
        self.assertEqual(parse_date_safely("۱۴۰۱-۱۲-۰۵"), datetime.date(2023, 2, 24))
        
        # Jalali YYYYMMDD format
        self.assertEqual(parse_date_safely("14040716"), datetime.date(2025, 10, 8))
        self.assertEqual(parse_date_safely("14040716.0"), datetime.date(2025, 10, 8))
        
        # Already Python date or datetime
        d = datetime.date(2022, 1, 1)
        dt = datetime.datetime(2022, 1, 1, 12, 0)
        self.assertEqual(parse_date_safely(d), d)
        self.assertEqual(parse_date_safely(dt), d)
        
        # Excel float dates (approximate epoch matching)
        self.assertEqual(parse_date_safely("44197"), datetime.date(2021, 1, 1))

        # Invalid formats
        self.assertIsNone(parse_date_safely("invalid-date"))
        self.assertIsNone(parse_date_safely(""))
        self.assertIsNone(parse_date_safely(None))

    def test_validate_national_id(self):
        # Valid ID (using a mathematically valid national ID)
        self.assertTrue(validate_national_id("7731689956"))
        
        # Invalid IDs
        self.assertFalse(validate_national_id("1111111111")) # repeating
        self.assertFalse(validate_national_id("123456")) # short
        self.assertFalse(validate_national_id("abcdefghij")) # non-numeric
        self.assertFalse(validate_national_id(""))
        self.assertFalse(validate_national_id(None))


class ExecuteImportTestCase(TestCase):
    def setUp(self):
        # Create user
        self.user = User.objects.create_user(username='admin_test', password='password')
        self.profile, created = UserProfile.objects.get_or_create(user=self.user)
        self.profile.role = UserProfile.ROLE_ADMIN
        self.profile.save()
        
        # Create standard workflow template
        self.wf = WorkflowTemplate.objects.create(name="آزمون کتبی و مصاحبه")
        self.stage1 = WorkflowStageTemplate.objects.create(
            workflow=self.wf, name="آزمون کتبی", sequence=1, stage_type="EXAM", default_weight=40
        )
        self.stage2 = WorkflowStageTemplate.objects.create(
            workflow=self.wf, name="مصاحبه فنی", sequence=2, stage_type="INTERVIEW", default_weight=60
        )
        
        # Create import session
        self.session = ImportSession.objects.create(
            excel_file="historical_imports/test.xlsx",
            created_by=self.user,
            status="PREVIEWED",
            mapping_config={
                "main_sheet": "وضعیت",
                "workflow_mappings": {
                    "کتبی + مصاحبه": str(self.wf.id)
                }
            }
        )

    def test_execute_final_import_new_job(self):
        # Create staging job
        sj = StagingJobOpportunity.objects.create(
            import_session=self.session,
            row_index=2,
            job_code="JOB-TEST-100",
            title="برنامه‌نویس پایتون",
            department="فناوری اطلاعات",
            headcount="3",
            status="آزمون کتبی",
            start_date_str="1402/01/15",
            workflow_pattern="کتبی + مصاحبه"
        )
        
        # Create staging candidate
        sc = StagingCandidate.objects.create(
            import_session=self.session,
            sheet_name="آزمون کتبی",
            row_index=2,
            job_code="JOB-TEST-100",
            national_id="7731689956",
            first_name="امید",
            last_name="صالحی",
            phone_number="09123456789",
            email="omid@example.com",
            score="85.5",
            evaluation_date_str="1402/02/10",
            stage_type="EXAM"
        )

        execute_final_import(self.session, "UPDATE", self.user)
        
        # Verify job opportunity created
        job = JobOpportunity.objects.filter(code="JOB-TEST-100", is_deleted=False).first()
        self.assertIsNotNone(job)
        self.assertEqual(job.title, "برنامه‌نویس پایتون")
        self.assertEqual(job.headcount, 3)
        self.assertEqual(job.workflow, self.wf)
        self.assertEqual(job.start_date, datetime.date(2023, 4, 4))
        
        # Verify job stage counts
        stages = job.stages.filter(is_deleted=False)
        self.assertEqual(stages.count(), 2)
        
        # Verify candidate created
        candidate = Candidate.objects.filter(national_id="7731689956", is_deleted=False).first()
        self.assertIsNotNone(candidate)
        self.assertEqual(candidate.first_name, "امید")
        self.assertEqual(candidate.last_name, "صالحی")
        
        # Verify application created
        app = JobApplication.objects.filter(job=job, candidate=candidate, is_deleted=False).first()
        self.assertIsNotNone(app)
        
        # Verify stage state created and marked complete
        state = app.stage_states.filter(stage__stage_type="EXAM", is_deleted=False).first()
        self.assertIsNotNone(state)
        self.assertEqual(state.status, ApplicationStageState.STATUS_COMPLETED)
        self.assertEqual(state.score, 85.5)
        self.assertEqual(state.evaluation_date, datetime.date(2023, 4, 30))

    def test_execute_final_import_prior_stages_completion(self):
        # If candidate is in INTERVIEW, the prior stage (EXAM) should be marked COMPLETED automatically
        sj = StagingJobOpportunity.objects.create(
            import_session=self.session,
            row_index=2,
            job_code="JOB-TEST-200",
            title="کارشناس شبکه",
            department="فناوری اطلاعات",
            headcount="1",
            status="مصاحبه حضوری",
            start_date_str="1402/01/15",
            workflow_pattern="کتبی + مصاحبه"
        )
        
        sc = StagingCandidate.objects.create(
            import_session=self.session,
            sheet_name="مصاحبه",
            row_index=2,
            job_code="JOB-TEST-200",
            national_id="7731689956",
            first_name="امید",
            last_name="صالحی",
            phone_number="09123456789",
            email="omid@example.com",
            score="90",
            evaluation_date_str="1402/03/15",
            stage_type="INTERVIEW"
        )

        execute_final_import(self.session, "UPDATE", self.user)
        
        job = JobOpportunity.objects.filter(code="JOB-TEST-200").first()
        candidate = Candidate.objects.filter(national_id="7731689956").first()
        app = JobApplication.objects.filter(job=job, candidate=candidate).first()
        
        # Interview stage state should be completed with score 90
        intv_state = app.stage_states.filter(stage__stage_type="INTERVIEW").first()
        self.assertEqual(intv_state.status, ApplicationStageState.STATUS_COMPLETED)
        self.assertEqual(intv_state.score, 90.0)
        
        # Exam stage state (prior stage) should be completed automatically with cutoff score
        exam_state = app.stage_states.filter(stage__stage_type="EXAM").first()
        self.assertEqual(exam_state.status, ApplicationStageState.STATUS_COMPLETED)
        self.assertEqual(exam_state.score, 60.0) # default passing score of stage

    def test_execute_final_import_conflict_skip(self):
        # Create pre-existing JobOpportunity
        existing_job = JobOpportunity.objects.create(
            code="JOB-EXISTS-1",
            request_number="JOB-EXISTS-1",
            title="فرصت شغلی قدیمی",
            department="فروش",
            headcount=5,
            workflow=self.wf,
            status=JobOpportunity.STATUS_RECEIVED
        )
        
        sj = StagingJobOpportunity.objects.create(
            import_session=self.session,
            row_index=2,
            job_code="JOB-EXISTS-1",
            title="فرصت شغلی بروزرسانی شده", # different title
            department="پشتیبانی",
            headcount="10",
            status="آزمون کتبی",
            start_date_str="1402/01/15",
            workflow_pattern="کتبی + مصاحبه"
        )

        execute_final_import(self.session, "SKIP", self.user)
        
        # Refresh from DB
        existing_job.refresh_from_db()
        self.assertEqual(existing_job.title, "فرصت شغلی قدیمی") # should not have changed
        self.assertEqual(existing_job.headcount, 5)

    def test_execute_final_import_conflict_update(self):
        existing_job = JobOpportunity.objects.create(
            code="JOB-EXISTS-2",
            request_number="JOB-EXISTS-2",
            title="فرصت شغلی قدیمی",
            department="فروش",
            headcount=5,
            workflow=self.wf,
            status=JobOpportunity.STATUS_RECEIVED
        )
        
        sj = StagingJobOpportunity.objects.create(
            import_session=self.session,
            row_index=2,
            job_code="JOB-EXISTS-2",
            title="فرصت شغلی بروزرسانی شده",
            department="پشتیبانی",
            headcount="10",
            status="آزمون کتبی",
            start_date_str="1402/01/15",
            workflow_pattern="کتبی + مصاحبه"
        )

        execute_final_import(self.session, "UPDATE", self.user)
        
        existing_job.refresh_from_db()
        self.assertEqual(existing_job.title, "فرصت شغلی بروزرسانی شده") # title updated!
        self.assertEqual(existing_job.department, "پشتیبانی")
        self.assertEqual(existing_job.headcount, 10)


class FixedTemplateImportTestCase(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='admin_test_fixed', password='password')
        self.profile, created = UserProfile.objects.get_or_create(user=self.user)
        self.profile.role = UserProfile.ROLE_ADMIN
        self.profile.save()
        
    def test_download_template(self):
        self.client.login(username='admin_test_fixed', password='password')
        url = reverse('download_fixed_template')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertTrue(len(b"".join(response.streaming_content)) > 0)
        
    def test_fixed_template_import_valid(self):
        self.client.login(username='admin_test_fixed', password='password')
        url = reverse('fixed_template_import')
        
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "جدول وضعیت"
        ws1.append(["Row", "کد", "آخرین مرحله", "مسئول", "واحد متقاضی", "عنوان پست", "رده شغلی", "تعداد مورد نیاز", "جلسه اولیه", "شروع ثبت‌نام", "پایان ثبت‌نام", "تعداد ثبت‌نام", "پایان غربالگری", "تعداد واجد شرایط", "کتبی", "آزمون کتبی", "تعداد حاضرین آزمون کتبی", "اعلام نتایج کتبی", "تعداد نفرات خروجی کتبی", "مهارتی", "آزمون مهارتی", "تعداد حاضرین آزمون مهارتی", "تعداد دعوت به مصاحبه", "مصاحبه", "تعداد حاضرین مصاحبه", "معرفی به کانون", "تعداد معرفی به کانون", "اعلام نتیجه کانون", "اعلام نتیجه نهایی", "شماره نامه اعلام نتیجه نهایی", "کد پست", "شماره نامه درخواست", "توضیحات", "مسیر پیشنهادی (عنوان)", "Column1", "Column2", "Column3"])
        ws1.append([1, "JOB-FIX-1", "کتبی", "امید", "فناوری", "مدیر پروژه", "کارشناس", 2, None, "1405/03/01", None, None, None, None, "آزمون کتبی", "1405/03/10", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, "کتبی", None, None, None])
        
        ws2 = wb.create_sheet("ثبت نام")
        ws2.append(["Row", "ExamCode", "NationCode", "EXP", "Key", "نام", "نام خانوادگی", "شماره همراه"])
        ws2.append([1, "JOB-FIX-1", "0011223344", None, None, "حمید", "علوی", "09121112233"])
        
        ws3 = wb.create_sheet("کتبی")
        ws3.append(["Row", "ExamCode", "NationCode", "ScoreW", "Result1", "Key"])
        ws3.append([1, "JOB-FIX-1", "0011223344", 85, "قبول", "key-1"])
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = "test_fixed.xlsx"
        
        response = self.client.post(url, {'excel_file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'درون‌ریزی موفقیت‌آمیز اطلاعات')
        
        # Verify db entries
        job = JobOpportunity.objects.filter(code="JOB-FIX-1").first()
        self.assertIsNotNone(job)
        self.assertEqual(job.title, "مدیر پروژه")
        self.assertEqual(job.unit, "فناوری")
        self.assertEqual(job.job_category, "کارشناس")
        
        candidate = Candidate.objects.filter(national_id="0011223344").first()
        self.assertIsNotNone(candidate)
        self.assertEqual(candidate.first_name, "حمید")
        
        app = JobApplication.objects.filter(job=job, candidate=candidate).first()
        self.assertIsNotNone(app)
        
        state = app.stage_states.filter(stage__stage_type="EXAM").first()
        self.assertIsNotNone(state)
        self.assertEqual(state.status, ApplicationStageState.STATUS_COMPLETED)
        self.assertEqual(state.score, 85.0)

    def test_recruitment_specialist_access_allowed(self):
        user = User.objects.create_user(username='specialist_test', password='password')
        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.role = UserProfile.ROLE_RECRUITMENT_SPECIALIST
        profile.save()
        
        self.client.login(username='specialist_test', password='password')
        url = reverse('import_upload')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)

    def test_interviewer_access_denied(self):
        user = User.objects.create_user(username='interviewer_test', password='password')
        profile, _ = UserProfile.objects.get_or_create(user=user)
        profile.role = UserProfile.ROLE_INTERVIEWER
        profile.save()
        
        self.client.login(username='interviewer_test', password='password')
        url = reverse('import_upload')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 403)

    def test_fixed_template_import_with_soft_deleted(self):
        # Create a job, candidate, application, and stage state, and soft-delete them
        job = JobOpportunity.objects.create(
            code="JOB-FIX-1",
            request_number="JOB-FIX-1",
            title="مدیر پروژه",
            department="فناوری",
            headcount=2,
            status=JobOpportunity.STATUS_RECEIVED
        )
        
        wf = WorkflowTemplate.objects.create(name="الگوی خودکار - کتبی")
        WorkflowStageTemplate.objects.create(
            workflow=wf, name="کتبی", sequence=1, stage_type="EXAM", default_weight=100
        )
        job.workflow = wf
        job.save() # this will copy stages
        
        stage = job.stages.filter(stage_type="EXAM").first()
        self.assertIsNotNone(stage)

        candidate = Candidate.objects.create(
            national_id="0011223344",
            first_name="حمید",
            last_name="علوی",
            phone_number="09121112233"
        )
        
        app = JobApplication.objects.create(
            job=job,
            candidate=candidate,
            status=JobApplication.STATUS_IN_PROGRESS
        )
        
        state = ApplicationStageState.objects.filter(
            application=app,
            stage=stage
        ).first()
        self.assertIsNotNone(state)
        
        # Verify they exist
        self.assertEqual(JobOpportunity.objects.filter(code="JOB-FIX-1").count(), 1)
        self.assertEqual(Candidate.objects.filter(national_id="0011223344").count(), 1)
        self.assertEqual(JobApplication.objects.filter(job=job, candidate=candidate).count(), 1)
        self.assertEqual(ApplicationStageState.objects.filter(application=app, stage=stage).count(), 1)
        
        # Soft-delete them
        state.delete()
        app.delete()
        candidate.delete()
        job.delete()
        
        # Verify soft-deleted
        self.assertEqual(JobOpportunity.objects.filter(code="JOB-FIX-1").count(), 0)
        self.assertEqual(Candidate.objects.filter(national_id="0011223344").count(), 0)
        self.assertEqual(JobApplication.objects.filter(job=job, candidate=candidate).count(), 0)
        self.assertEqual(ApplicationStageState.objects.filter(application=app, stage=stage).count(), 0)
        
        # Prepare Excel import data
        self.client.login(username='admin_test_fixed', password='password')
        url = reverse('fixed_template_import')
        
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "جدول وضعیت"
        ws1.append(["Row", "کد", "آخرین مرحله", "مسئول", "واحد متقاضی", "عنوان پست", "رده شغلی", "تعداد مورد نیاز", "جلسه اولیه", "شروع ثبت‌نام", "پایان ثبت‌نام", "تعداد ثبت‌نام", "پایان غربالگری", "تعداد واجد شرایط", "کتبی", "آزمون کتبی", "تعداد حاضرین آزمون کتبی", "اعلام نتایج کتبی", "تعداد نفرات خروجی کتبی", "مهارتی", "آزمون مهارتی", "تعداد حاضرین آزمون مهارتی", "تعداد دعوت به مصاحبه", "مصاحبه", "تعداد حاضرین مصاحبه", "معرفی به کانون", "تعداد معرفی به کانون", "اعلام نتیجه کانون", "اعلام نتیجه نهایی", "شماره نامه اعلام نتیجه نهایی", "کد پست", "شماره نامه درخواست", "توضیحات", "مسیر پیشنهادی (عنوان)", "Column1", "Column2", "Column3"])
        ws1.append([1, "JOB-FIX-1", "کتبی", "امید", "فناوری", "مدیر پروژه", "کارشناس", 2, None, "1405/03/01", None, None, None, None, "آزمون کتبی", "1405/03/10", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, "کتبی", None, None, None])
        
        ws2 = wb.create_sheet("ثبت نام")
        ws2.append(["Row", "ExamCode", "NationCode", "EXP", "Key", "نام", "نام خانوادگی", "شماره همراه"])
        ws2.append([1, "JOB-FIX-1", "0011223344", None, None, "حمید", "علوی", "09121112233"])
        
        ws3 = wb.create_sheet("کتبی")
        ws3.append(["Row", "ExamCode", "NationCode", "ScoreW", "Result1", "Key"])
        ws3.append([1, "JOB-FIX-1", "0011223344", 85, "قبول", "key-1"])
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = "test_fixed.xlsx"
        
        # Run import
        response = self.client.post(url, {'excel_file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'درون‌ریزی موفقیت‌آمیز اطلاعات')
        
        # Verify they are restored
        self.assertEqual(JobOpportunity.objects.filter(code="JOB-FIX-1").count(), 1)
        self.assertEqual(Candidate.objects.filter(national_id="0011223344").count(), 1)
        self.assertEqual(JobApplication.objects.filter(job=job, candidate=candidate).count(), 1)
        self.assertEqual(ApplicationStageState.objects.filter(application=app, stage=stage).count(), 1)

    def test_fixed_template_import_screening(self):
        self.client.login(username='admin_test_fixed', password='password')
        url = reverse('fixed_template_import')
        
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "جدول وضعیت"
        ws1.append(["Row", "کد", "آخرین مرحله", "مسئول", "واحد متقاضی", "عنوان پست", "رده شغلی", "تعداد مورد نیاز", "جلسه اولیه", "شروع ثبت‌نام", "پایان ثبت‌نام", "تعداد ثبت‌نام", "پایان غربالگری", "تعداد واجد شرایط", "کتبی", "آزمون کتبی", "تعداد حاضرین آزمون کتبی", "اعلام نتایج کتبی", "تعداد نفرات خروجی کتبی", "مهارتی", "آزمون مهارتی", "تعداد حاضرین آزمون مهارتی", "تعداد دعوت به مصاحبه", "مصاحبه", "تعداد حاضرین مصاحبه", "معرفی به کانون", "تعداد معرفی به کانون", "اعلام نتیجه کانون", "اعلام نتیجه نهایی", "شماره نامه اعلام نتیجه نهایی", "کد پست", "شماره نامه درخواست", "توضیحات", "مسیر پیشنهادی (عنوان)", "Column1", "Column2", "Column3"])
        ws1.append([1, "JOB-SCR-EXAM", "کتبی", "مسئول ۱", "فناوری", "برنامه‌نویس", "کارشناس", 5, None, "1405/03/01", None, None, None, None, "آزمون کتبی", "1405/03/10", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, "غربالگری + کتبی", None, None, None])
        ws1.append([2, "JOB-SCR-PUB", "ثبت نام", "مسئول ۱", "فناوری", "طراح", "کارشناس", 2, None, "1405/03/01", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, "غربالگری + کتبی", None, None, None])
        
        ws2 = wb.create_sheet("ثبت نام")
        ws2.append(["Row", "ExamCode", "NationCode", "EXP", "Key", "نام", "نام خانوادگی", "شماره همراه"])
        ws2.append([1, "JOB-SCR-EXAM", "1111111111", None, None, "کاندیدا", "یک", "09121111111"])
        ws2.append([2, "JOB-SCR-EXAM", "2222222222", None, None, "کاندیدا", "دو", "09122222222"])
        ws2.append([3, "JOB-SCR-EXAM", "3333333333", None, None, "کاندیدا", "سه", "09123333333"])
        ws2.append([4, "JOB-SCR-EXAM", "4444444444", None, None, "کاندیدا", "چهار", "09124444444"])
        ws2.append([5, "JOB-SCR-PUB", "5555555555", None, None, "کاندیدا", "پنج", "09125555555"])
        
        ws3 = wb.create_sheet("غربالگری")
        ws3.append(["ExamCode", "NationCode", "Result", "Description"])
        ws3.append(["JOB-SCR-EXAM", "1111111111", "مردود", "عدم انطباق مدرک"])
        ws3.append(["JOB-SCR-EXAM", "2222222222", "قبول", "تایید اولیه"])
        
        ws4 = wb.create_sheet("کتبی")
        ws4.append(["Row", "ExamCode", "NationCode", "ScoreW", "Result1", "Key"])
        ws4.append([1, "JOB-SCR-EXAM", "4444444444", 75.0, "قبول", "k-4"])
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = "test_screening.xlsx"
        
        response = self.client.post(url, {'excel_file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'درون‌ریزی موفقیت‌آمیز اطلاعات')
        
        # Verify Candidate 1: explicit FAILED
        job_exam = JobOpportunity.objects.get(code="JOB-SCR-EXAM")
        app_1 = JobApplication.objects.get(job=job_exam, candidate__national_id="1111111111")
        scr_stage = job_exam.stages.get(stage_type="SCREENING")
        state_1 = app_1.stage_states.get(stage=scr_stage)
        self.assertEqual(state_1.status, ApplicationStageState.STATUS_FAILED)
        self.assertEqual(state_1.notes, "عدم انطباق مدرک")
        
        # Verify Candidate 2: explicit COMPLETED
        app_2 = JobApplication.objects.get(job=job_exam, candidate__national_id="2222222222")
        state_2 = app_2.stage_states.get(stage=scr_stage)
        self.assertEqual(state_2.status, ApplicationStageState.STATUS_COMPLETED)
        
        # Verify Candidate 3: empty screening, job is in EXAM stage, no score in EXAM stage -> fallback FAILED
        app_3 = JobApplication.objects.get(job=job_exam, candidate__national_id="3333333333")
        state_3 = app_3.stage_states.get(stage=scr_stage)
        self.assertEqual(state_3.status, ApplicationStageState.STATUS_FAILED)
        
        # Verify Candidate 4: empty screening, job is in EXAM stage, has score in EXAM stage -> fallback COMPLETED
        app_4 = JobApplication.objects.get(job=job_exam, candidate__national_id="4444444444")
        state_4 = app_4.stage_states.get(stage=scr_stage)
        self.assertEqual(state_4.status, ApplicationStageState.STATUS_COMPLETED)
        
        # Verify Candidate 5: empty screening, job is not in evaluation stage -> fallback COMPLETED
        job_pub = JobOpportunity.objects.get(code="JOB-SCR-PUB")
        app_5 = JobApplication.objects.get(job=job_pub, candidate__national_id="5555555555")
        scr_stage_pub = job_pub.stages.get(stage_type="SCREENING")
        state_5 = app_5.stage_states.get(stage=scr_stage_pub)
        self.assertEqual(state_5.status, ApplicationStageState.STATUS_COMPLETED)

    def test_reset_database_view(self):
        # Create test data
        job = JobOpportunity.objects.create(
            code="JOB-RESET-TEST",
            request_number="JOB-RESET-TEST",
            title="شغل تستی ریست",
            department="فروش",
            headcount=1
        )
        candidate = Candidate.objects.create(
            national_id="9999888877",
            first_name="داوطلب",
            last_name="ریست"
        )
        
        self.client.login(username='admin_test_fixed', password='password')
        url = reverse('reset_database')
        
        # Test GET request
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'historical_import/reset_db.html')
        
        # Test POST with invalid confirmation text
        response = self.client.post(url, {'confirm_text': 'INVALID'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'متن تاییدیه به درستی وارد نشده است')
        self.assertEqual(JobOpportunity.all_objects.filter(code="JOB-RESET-TEST").count(), 1)
        self.assertEqual(Candidate.all_objects.filter(national_id="9999888877").count(), 1)
        
        # Test POST with valid confirmation text
        response = self.client.post(url, {'confirm_text': 'RESET DATABASE'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'دیتابیس با موفقیت پاکسازی شد')
        
        # Verify both active and soft-deleted are hard-deleted
        self.assertEqual(JobOpportunity.all_objects.filter(code="JOB-RESET-TEST").count(), 0)
        self.assertEqual(Candidate.all_objects.filter(national_id="9999888877").count(), 0)

    def test_fixed_template_import_with_passed_under_cutoff(self):
        self.client.login(username='admin_test_fixed', password='password')
        url = reverse('fixed_template_import')
        
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "جدول وضعیت"
        ws1.append(["Row", "کد", "آخرین مرحله", "مسئول", "واحد متقاضی", "عنوان پست", "رده شغلی", "تعداد مورد نیاز", "جلسه اولیه", "شروع ثبت‌نام", "پایان ثبت‌نام", "تعداد ثبت‌نام", "پایان غربالگری", "تعداد واجد شرایط", "کتبی", "آزمون کتبی", "تعداد حاضرین آزمون کتبی", "اعلام نتایج کتبی", "تعداد نفرات خروجی کتبی", "مهارتی", "آزمون مهارتی", "تعداد حاضرین آزمون مهارتی", "تعداد دعوت به مصاحبه", "مصاحبه", "تعداد حاضرین مصاحبه", "معرفی به کانون", "تعداد معرفی به کانون", "اعلام نتیجه کانون", "اعلام نتیجه نهایی", "شماره نامه اعلام نتیجه نهایی", "کد پست", "شماره نامه درخواست", "توضیحات", "مسیر پیشنهادی (عنوان)", "Column1", "Column2", "Column3"])
        ws1.append([1, "JOB-FIX-UNDER", "کتبی", "امید", "فناوری", "مدیر پروژه", "کارشناس", 2, None, "1405/03/01", None, None, None, None, "آزمون کتبی", "1405/03/10", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, "کتبی", None, None, None])
        
        ws2 = wb.create_sheet("ثبت نام")
        ws2.append(["Row", "ExamCode", "NationCode", "EXP", "Key", "نام", "نام خانوادگی", "شماره همراه"])
        ws2.append([1, "JOB-FIX-UNDER", "0011223344", None, None, "حمید", "علوی", "09121112233"])
        
        ws3 = wb.create_sheet("کتبی")
        ws3.append(["Row", "ExamCode", "NationCode", "ScoreW", "Result1", "Key"])
        ws3.append([1, "JOB-FIX-UNDER", "0011223344", 45.0, "مجاز", "key-1"])
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = "test_fixed_under.xlsx"
        
        response = self.client.post(url, {'excel_file': excel_file})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'درون‌ریزی موفقیت‌آمیز اطلاعات')
        
        # Verify db entries
        job = JobOpportunity.objects.filter(code="JOB-FIX-UNDER").first()
        self.assertIsNotNone(job)
        
        candidate = Candidate.objects.filter(national_id="0011223344").first()
        self.assertIsNotNone(candidate)
        
        app = JobApplication.objects.filter(job=job, candidate=candidate).first()
        self.assertIsNotNone(app)
        
        state = app.stage_states.filter(stage__stage_type="EXAM").first()
        self.assertIsNotNone(state)
        self.assertEqual(state.status, ApplicationStageState.STATUS_COMPLETED)
        self.assertEqual(state.score, 45.0)



