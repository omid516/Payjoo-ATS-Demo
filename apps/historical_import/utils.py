import re
import datetime
import jdatetime
import openpyxl
from django.db import transaction
from django.core.exceptions import ValidationError
from django.contrib.auth.models import User
from .models import ImportSession, StagingJobOpportunity, StagingCandidate, ImportSessionLog

def normalize_persian_digits(s):
    """
    تبدیل ارقام فارسی و عربی به انگلیسی
    """
    if s is None:
        return ""
    s = str(s).strip()
    persian_to_english = {
        '۰': '0', '۱': '1', '۲': '2', '۳': '3', '۴': '4',
        '۵': '5', '۶': '6', '۷': '7', '۸': '8', '۹': '9',
        '٠': '0', '١': '1', '٢': '2', '٣': '3', '٤': '4',
        '٥': '5', '٦': '6', '٧': '7', '٨': '8', '٩': '9'
    }
    for k, v in persian_to_english.items():
        s = s.replace(k, v)
    return s


def parse_date_safely(val):
    """
    تبدیل تاریخ دریافتی به تاریخ میلادی استاندارد.
    پشتیبانی از انواع فرمت‌های تاریخ شمسی (متنی) و اشیاء datetime/date پایتون.
    """
    if val is None:
        return None

    # اگر از قبل شیء تاریخ بود
    if isinstance(val, (datetime.date, datetime.datetime)):
        if isinstance(val, datetime.datetime):
            return val.date()
        return val

    # تبدیل به رشته و نرمال‌سازی اعداد
    date_str = normalize_persian_digits(val).strip()
    if date_str.endswith(".0"):
        date_str = date_str[:-2]
    if not date_str or date_str.lower() in ['none', 'null', '']:
        return None

    # بررسی اگر تاریخ شمسی عددی ۸ رقمی باشد
    if date_str.isdigit() and len(date_str) == 8:
        year = int(date_str[:4])
        month = int(date_str[4:6])
        day = int(date_str[6:8])
        try:
            return jdatetime.date(year, month, day).togregorian()
        except ValueError:
            return None

    # بررسی اگر شبیه به تاریخ اکسل عددی باشد
    if date_str.isdigit() and len(date_str) in [4, 5]:
        try:
            # تبدیل عدد سریال اکسل به تاریخ میلادی
            excel_epoch = datetime.date(1899, 12, 30)
            return excel_epoch + datetime.timedelta(days=int(date_str))
        except Exception:
            pass

    # الگوهای تاریخ شمسی: e.g. 1402/05/12 or 1401-12-05
    match = re.match(r'^(\d{4})[/-](\d{1,2})[/-](\d{1,2})(?:\s+.*)?$', date_str)
    if match:
        year, month, day = map(int, match.groups())
        # تصحیح سال‌های دو رقمی
        if year < 100:
            year += 1300 if year > 50 else 1400
        try:
            return jdatetime.date(year, month, day).togregorian()
        except ValueError:
            return None

    # حالت‌های متفرقه: فقط سال و ماه یا تاریخ ناقص
    match_partial = re.match(r'^(\d{4})[/-](\d{1,2})$', date_str)
    if match_partial:
        year, month = map(int, match_partial.groups())
        try:
            return jdatetime.date(year, month, 1).togregorian()
        except ValueError:
            return None

    return None


def validate_national_id(national_id):
    """
    بررسی معتبر بودن کد ملی ایران (۱۰ رقم و بررسی رقم کنترلی)
    """
    if not national_id:
        return False
    nid = normalize_persian_digits(national_id).zfill(10)
    if not nid.isdigit() or len(nid) != 10:
        return False
    
    # الگوهای تکراری مانند 1111111111 نامعتبر هستند
    if re.match(r'^(\d)\1{9}$', nid):
        return False

    check = int(nid[9])
    s = sum(int(nid[i]) * (10 - i) for i in range(9))
    r = s % 11
    return (r < 2 and check == r) or (r >= 2 and check == 11 - r)


def analyze_excel_structure(import_session, main_sheet_name=None, workflow_col_name=None):
    """
    تحلیل ساختار فایل اکسل بارگذاری شده جهت استخراج نام شیت‌ها، ردیف‌ها و الگوهای استخدام
    """
    try:
        wb = openpyxl.load_workbook(import_session.excel_file.path, read_only=True)
        sheet_names = wb.sheetnames
        if not sheet_names:
            raise ValidationError("فایل اکسل فاقد هرگونه شیت است.")

        # انتخاب شیت اصلی به صورت هوشمند
        if not main_sheet_name:
            # جستجو برای شیت "وضعیت" یا موارد مشابه
            for name in sheet_names:
                if 'وضعیت' in name or 'status' in name.lower() or 'فرصت' in name:
                    main_sheet_name = name
                    break
            if not main_sheet_name:
                main_sheet_name = sheet_names[-1] # به عنوان زاپاس آخرین شیت

        if main_sheet_name not in sheet_names:
            main_sheet_name = sheet_names[-1]

        # خواندن شیت اصلی برای استخراج هدرها و الگوها
        ws = wb[main_sheet_name]
        headers = []
        rows_count = 0
        workflow_patterns = set()

        # خواندن سطر اول برای هدرها
        rows_iterator = ws.iter_rows(values_only=True)
        try:
            first_row = next(rows_iterator)
            headers = [str(cell).strip() for cell in first_row if cell is not None]
        except StopIteration:
            headers = []

        # تشخیص هوشمند ستون الگوی استخدام
        if not workflow_col_name and headers:
            for h in headers:
                if any(kw in h for kw in ['مسیر پیشنهادی', 'الگوی استخدام', 'الگو', 'فرآیند', 'workflow', 'pattern']):
                    workflow_col_name = h
                    break

        # یافتن ایندکس ستون الگو
        wf_col_idx = None
        if workflow_col_name in headers:
            wf_col_idx = headers.index(workflow_col_name)

        # اگر تا الان ستون الگو پیدا نشده، نمونه داده‌ها را اسکن کنیم تا ستونی که حاوی کاراکتر + است را پیدا کنیم
        # این کار همزمان با شمارش ردیف‌ها انجام می‌شود
        scan_wf_by_values = (wf_col_idx is None)
        detected_wf_col_idx = None

        rows_count = 0
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1:
                continue
            if any(cell is not None for cell in row):
                rows_count += 1
                
                # اسکن هوشمند ستون الگو از روی مقادیر در صورت نیاز
                if scan_wf_by_values and detected_wf_col_idx is None:
                    for col_idx, cell in enumerate(row):
                        if cell and isinstance(cell, str) and '+' in cell:
                            if any(k in cell for k in ['کتبی', 'مصاحبه', 'کانون', 'مهارتی', 'ارزیابی', 'غربال']):
                                detected_wf_col_idx = col_idx
                                if col_idx < len(headers):
                                    workflow_col_name = headers[col_idx]
                                break
                    if detected_wf_col_idx is not None:
                        wf_col_idx = detected_wf_col_idx

                # استخراج الگو
                if wf_col_idx is not None and wf_col_idx < len(row):
                    pattern = row[wf_col_idx]
                    if pattern:
                        workflow_patterns.add(str(pattern).strip())

        # اگر هنوز هم پیدا نشده، آخرین ستون را در نظر بگیریم
        if wf_col_idx is None and headers:
            wf_col_idx = len(headers) - 1
            workflow_col_name = headers[-1]
            for row in ws.iter_rows(values_only=True):
                if len(row) > wf_col_idx:
                    pattern = row[wf_col_idx]
                    if pattern:
                        workflow_patterns.add(str(pattern).strip())

        # تحلیل بقیه شیت‌ها (مراحل)
        stage_sheets = []
        for name in sheet_names:
            if name == main_sheet_name:
                continue
            stage_ws = wb[name]
            stage_headers = []
            stage_rows = 0
            for s_idx, s_row in enumerate(stage_ws.iter_rows(values_only=True), start=1):
                if s_idx == 1:
                    stage_headers = [str(cell).strip() for cell in s_row if cell is not None]
                else:
                    if any(cell is not None for cell in s_row):
                        stage_rows += 1
            
            stage_sheets.append({
                'name': name,
                'rows_count': stage_rows,
                'headers': stage_headers
            })

        summary = {
            'main_sheet': main_sheet_name,
            'main_headers': headers,
            'main_rows_count': rows_count,
            'workflow_col': workflow_col_name,
            'workflow_patterns': sorted(list(workflow_patterns)),
            'stage_sheets': stage_sheets,
            'all_sheets': sheet_names
        }
        
        import_session.summary_data = summary
        import_session.status = 'ANALYZED'
        import_session.save(update_fields=['summary_data', 'status'])
        
        ImportSessionLog.objects.create(
            import_session=import_session,
            level='INFO',
            message=f"ساختار فایل با موفقیت تحلیل شد. شیت اصلی '{main_sheet_name}'، {rows_count} فرصت شغلی و {len(stage_sheets)} شیت دیگر یافت شد."
        )
        return summary
    except Exception as e:
        import_session.status = 'FAILED'
        import_session.save(update_fields=['status'])
        ImportSessionLog.objects.create(
            import_session=import_session,
            level='ERROR',
            message=f"خطا در تحلیل ساختار فایل: {str(e)}"
        )
        raise e


@transaction.atomic
def parse_and_stage_data(import_session, mapping_config):
    """
    استخراج داده‌های خام از اکسل بر اساس پیکربندی نگاشت و ثبت در جداول موقت (Staging Tables) همراه با اعتبارسنجی
    """
    # پاکسازی داده‌های قبلی نشست در صورت تلاش مجدد
    import_session.staging_jobs.all().delete()
    import_session.staging_candidates.all().delete()
    import_session.logs.all().delete()

    import_session.mapping_config = mapping_config
    import_session.status = 'MAPPED'
    import_session.save(update_fields=['mapping_config', 'status'])

    wb = openpyxl.load_workbook(import_session.excel_file.path, data_only=True)
    
    # ۱. پردازش شیت اصلی (فرصت‌های شغلی)
    main_sheet_name = mapping_config.get('main_sheet')
    if main_sheet_name not in wb.sheetnames:
        raise ValidationError(f"شیت اصلی '{main_sheet_name}' یافت نشد.")
    
    main_ws = wb[main_sheet_name]
    job_fields = mapping_config.get('job_fields', {})
    
    # یافتن ایندکس ستون‌ها بر اساس هدرها
    main_rows = list(main_ws.iter_rows(values_only=True))
    if not main_rows:
        raise ValidationError("شیت اصلی فاقد اطلاعات است.")
    
    main_headers = [str(h).strip() for h in main_rows[0]]
    
    def get_col_idx(header_name):
        if not header_name:
            return None
        try:
            return main_headers.index(str(header_name).strip())
        except ValueError:
            return None

    col_indices = {field: get_col_idx(col_name) for field, col_name in job_fields.items()}

    # جمع‌آوری کدهای شغل برای اعتبارسنجی ارجاع کاندیداها
    job_codes_in_excel = set()
    
    from apps.jobs.models import JobOpportunity

    for idx, row in enumerate(main_rows[1:], start=2):
        if not any(cell is not None for cell in row):
            continue

        raw_data = {main_headers[i]: row[i] for i in range(min(len(row), len(main_headers)))}
        
        job_code = normalize_persian_digits(row[col_indices['job_code']]).strip() if col_indices.get('job_code') is not None else None
        title = str(row[col_indices['title']]).strip() if col_indices.get('title') is not None else None
        department = str(row[col_indices['department']]).strip() if col_indices.get('department') is not None else ""
        headcount_raw = row[col_indices['headcount']] if col_indices.get('headcount') is not None else "1"
        status_raw = str(row[col_indices['status']]).strip() if col_indices.get('status') is not None else ""
        start_date_raw = row[col_indices['start_date']] if col_indices.get('start_date') is not None else None
        workflow_raw = str(row[col_indices['workflow_pattern']]).strip() if col_indices.get('workflow_pattern') is not None else ""

        errors = []
        warnings = []

        if not job_code:
            errors.append("کد فرصت شغلی نمی‌تواند خالی باشد.")
        else:
            job_codes_in_excel.add(job_code)

        if not title:
            errors.append("عنوان فرصت شغلی نمی‌تواند خالی باشد.")

        # اعتبارسنجی تعداد ظرفیت
        headcount = 1
        if headcount_raw:
            try:
                headcount = int(normalize_persian_digits(headcount_raw))
            except ValueError:
                warnings.append(f"فرمت ظرفیت استخدام '{headcount_raw}' نامعتبر است. مقدار پیش‌فرض ۱ در نظر گرفته شد.")
        
        # اعتبارسنجی و تبدیل تاریخ شروع
        start_date_parsed = parse_date_safely(start_date_raw)
        if start_date_raw and not start_date_parsed:
            warnings.append(f"امکان تبدیل تاریخ شروع '{start_date_raw}' به میلادی وجود نداشت.")
        
        # تشخیص تعارض با داده‌های دیتابیس فعلی
        final_job = None
        if job_code:
            existing_job = JobOpportunity.objects.filter(code=job_code, is_deleted=False).first()
            if existing_job:
                final_job = existing_job
                warnings.append(f"فرصت شغلی با کد '{job_code}' از قبل در سیستم وجود دارد (احتمال تعارض).")

        staging_job = StagingJobOpportunity.objects.create(
            import_session=import_session,
            row_index=idx,
            job_code=job_code,
            title=title,
            department=department,
            headcount=str(headcount),
            status=status_raw,
            start_date_str=str(start_date_raw) if start_date_raw else "",
            workflow_pattern=workflow_raw,
            raw_data=raw_data,
            is_valid=(len(errors) == 0),
            validation_errors=errors,
            validation_warnings=warnings,
            final_job=final_job
        )

        if errors:
            ImportSessionLog.objects.create(
                import_session=import_session,
                level='ERROR',
                sheet_name=main_sheet_name,
                row_index=idx,
                message=f"خطا در ردیف {idx}: " + " | ".join(errors)
            )
        elif warnings:
            ImportSessionLog.objects.create(
                import_session=import_session,
                level='WARNING',
                sheet_name=main_sheet_name,
                row_index=idx,
                message=f"هشدار در ردیف {idx}: " + " | ".join(warnings)
            )

    # ۲. پردازش شیت‌های مراحل ارزیابی (متقاضیان)
    stage_sheet_mappings = mapping_config.get('stage_sheet_mappings', {})
    from apps.candidates.models import Candidate

    for sheet_name, sheet_config in stage_sheet_mappings.items():
        if sheet_name not in wb.sheetnames:
            ImportSessionLog.objects.create(
                import_session=import_session,
                level='WARNING',
                message=f"شیت ارزیابی '{sheet_name}' در اکسل یافت نشد و نادیده گرفته شد."
            )
            continue
        
        ws = wb[sheet_name]
        stage_type = sheet_config.get('stage_type')
        c_fields = sheet_config.get('candidate_fields', {})
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        
        headers = [str(h).strip() for h in rows[0]]
        
        def get_sheet_col_idx(header_name):
            if not header_name:
                return None
            try:
                return headers.index(str(header_name).strip())
            except ValueError:
                return None

        c_indices = {field: get_sheet_col_idx(col_name) for field, col_name in c_fields.items()}

        for idx, row in enumerate(rows[1:], start=2):
            if not any(cell is not None for cell in row):
                continue
            
            raw_data = {headers[i]: row[i] for i in range(min(len(row), len(headers)))}
            
            job_code = normalize_persian_digits(row[c_indices['job_code']]).strip() if c_indices.get('job_code') is not None else None
            national_id = normalize_persian_digits(row[c_indices['national_id']]).strip() if c_indices.get('national_id') is not None else None
            first_name = str(row[c_indices['first_name']]).strip() if c_indices.get('first_name') is not None else ""
            last_name = str(row[c_indices['last_name']]).strip() if c_indices.get('last_name') is not None else ""
            phone_number = normalize_persian_digits(row[c_indices['phone_number']]).strip() if c_indices.get('phone_number') is not None else ""
            email = str(row[c_indices['email']]).strip() if c_indices.get('email') is not None else ""
            score_raw = row[c_indices['score']] if c_indices.get('score') is not None else "0"
            date_raw = row[c_indices['date']] if c_indices.get('date') is not None else None

            errors = []
            warnings = []

            # اعتبارسنجی کد ملی
            if not national_id:
                errors.append("کد ملی متقاضی نمی‌تواند خالی باشد.")
            elif not validate_national_id(national_id):
                warnings.append(f"قالب کد ملی '{national_id}' نامعتبر یا غیر استاندارد است.")

            if not first_name or not last_name:
                errors.append("نام و نام خانوادگی متقاضی الزامی است.")

            if not job_code:
                errors.append("کد فرصت شغلی برای متقاضی الزامی است.")
            elif job_code not in job_codes_in_excel:
                # بررسی تعارض کد شغل با کدهای شیت اصلی
                errors.append(f"فرصت شغلی با کد '{job_code}' در شیت اصلی وضعیت یافت نشد.")

            # بررسی امتیاز
            score = "0.0"
            if score_raw:
                try:
                    score = str(float(normalize_persian_digits(score_raw)))
                except ValueError:
                    warnings.append(f"امتیاز مرحله '{score_raw}' عددی نیست. مقدار ۰ ثبت شد.")

            # اعتبارسنجی تاریخ ارزیابی
            date_parsed = parse_date_safely(date_raw)
            if date_raw and not date_parsed:
                warnings.append(f"امکان تبدیل تاریخ ارزیابی '{date_raw}' به میلادی وجود نداشت.")

            # بررسی تطابق کد ملی متقاضی در پایگاه داده
            final_candidate = None
            if national_id:
                existing_candidate = Candidate.objects.filter(national_id=national_id, is_deleted=False).first()
                if existing_candidate:
                    final_candidate = existing_candidate
                    warnings.append(f"متقاضی با کد ملی '{national_id}' از قبل در سیستم ثبت شده است.")

            staging_candidate = StagingCandidate.objects.create(
                import_session=import_session,
                sheet_name=sheet_name,
                row_index=idx,
                job_code=job_code,
                national_id=national_id,
                first_name=first_name,
                last_name=last_name,
                phone_number=phone_number,
                email=email,
                score=score,
                evaluation_date_str=str(date_raw) if date_raw else "",
                stage_type=stage_type,
                raw_data=raw_data,
                is_valid=(len(errors) == 0),
                validation_errors=errors,
                validation_warnings=warnings,
                final_candidate=final_candidate
            )

            if errors:
                ImportSessionLog.objects.create(
                    import_session=import_session,
                    level='ERROR',
                    sheet_name=sheet_name,
                    row_index=idx,
                    message=f"خطا در ردیف {idx}: " + " | ".join(errors)
                )
            elif warnings:
                ImportSessionLog.objects.create(
                    import_session=import_session,
                    level='WARNING',
                    sheet_name=sheet_name,
                    row_index=idx,
                    message=f"هشدار در ردیف {idx}: " + " | ".join(warnings)
                )

    # به‌روزرسانی وضعیت نشست
    import_session.status = 'PREVIEWED'
    import_session.save(update_fields=['status'])
    
    ImportSessionLog.objects.create(
        import_session=import_session,
        level='INFO',
        message="فرآیند استخراج داده‌ها و اعتبارسنجی با موفقیت به پایان رسید. پیش‌نمایش آماده بازدید است."
    )


@transaction.atomic
def execute_final_import(import_session, conflict_strategy, user):
    """
    اجرای تراکنشی درون‌ریزی نهایی داده‌ها از جداول موقت به دیتابیس اصلی
    """
    from apps.jobs.models import JobOpportunity, JobOpportunityStage, WorkflowTemplate
    from apps.candidates.models import Candidate, JobApplication, ApplicationStageState
    
    logs = []
    def log_info(msg, sheet=None, row=None):
        ImportSessionLog.objects.create(import_session=import_session, level='INFO', message=msg, sheet_name=sheet, row_index=row)
    def log_warn(msg, sheet=None, row=None):
        ImportSessionLog.objects.create(import_session=import_session, level='WARNING', message=msg, sheet_name=sheet, row_index=row)
    def log_err(msg, sheet=None, row=None):
        ImportSessionLog.objects.create(import_session=import_session, level='ERROR', message=msg, sheet_name=sheet, row_index=row)

    log_info("شروع فرآیند تراکنشی ایمپورت سوابق تاریخی جذب...")

    mapping_config = import_session.mapping_config
    main_sheet = mapping_config.get('main_sheet')
    workflow_mappings = mapping_config.get('workflow_mappings', {})

    # ۱. پردازش فرصت‌های شغلی موقت
    staging_jobs = import_session.staging_jobs.filter(is_valid=True)
    job_map = {} # نقشه‌برداری کد شغل به شیء JobOpportunity واقعی
    
    for sj in staging_jobs:
        # بررسی وجود فرصت شغلی در سیستم
        existing_job = JobOpportunity.objects.filter(code=sj.job_code, is_deleted=False).first()
        job = None

        if existing_job:
            if conflict_strategy == 'SKIP':
                log_info(f"فرصت شغلی '{sj.title}' ({sj.job_code}) از قبل وجود دارد. طبق تنظیمات، نادیده گرفته شد.", main_sheet, sj.row_index)
                job_map[sj.job_code] = existing_job
                continue
            elif conflict_strategy == 'REPLACE':
                log_warn(f"فرصت شغلی '{sj.title}' ({sj.job_code}) از قبل وجود دارد. در حال حذف سوابق قبلی جهت جایگزینی...", main_sheet, sj.row_index)
                # حذف نرم درخواست‌های قبلی متصل به این شغل
                existing_job.applications.all().delete()
                # حذف نرم خود شغل
                existing_job.delete()
                # ایجاد مجدد
                job = None
            elif conflict_strategy == 'UPDATE':
                log_info(f"فرصت شغلی '{sj.title}' ({sj.job_code}) ویرایش و بروزرسانی شد.", main_sheet, sj.row_index)
                job = existing_job
                job.title = sj.title
                job.department = sj.department
                job.headcount = int(sj.headcount)
                job.save()

        if not job:
            # ایجاد فرصت شغلی جدید
            log_info(f"در حال ایجاد فرصت شغلی جدید: '{sj.title}' ({sj.job_code})", main_sheet, sj.row_index)
            
            # تعیین الگو
            wf_template = None
            wf_mapped_val = workflow_mappings.get(sj.workflow_pattern)
            
            if wf_mapped_val == 'auto_create':
                # ایجاد خودکار الگو بر اساس متن الگوی استخدام
                wf_name = f"الگوی خودکار - {sj.workflow_pattern}"
                wf_template = WorkflowTemplate.objects.filter(name=wf_name, is_deleted=False).first()
                if not wf_template:
                    wf_template = WorkflowTemplate.objects.create(name=wf_name, description="ایجاد شده خودکار توسط سیستم ایمپورت سوابق")
                    # تقسیم الگو به مراحل
                    parts = [p.strip() for p in sj.workflow_pattern.split('+')]
                    for seq, part in enumerate(parts, start=1):
                        # تشخیص نوع مرحله بر اساس کلمات کلیدی
                        s_type = 'OTHER'
                        part_lower = part.lower()
                        if any(kw in part_lower for kw in ["غربال", "screening"]):
                            s_type = 'SCREENING'
                        elif any(kw in part_lower for kw in ["آزمون", "کتبی", "exam", "test"]):
                            s_type = 'EXAM'
                        elif any(kw in part_lower for kw in ["مصاحبه", "interview"]):
                            s_type = 'INTERVIEW'
                        elif any(kw in part_lower for kw in ["کانون", "ارزیابی", "assessment"]):
                            s_type = 'ASSESSMENT'
                        
                        from apps.jobs.models import WorkflowStageTemplate
                        WorkflowStageTemplate.objects.create(
                            workflow=wf_template,
                            name=part,
                            sequence=seq,
                            stage_type=s_type,
                            default_weight=100 // len(parts) # تقسیم مساوی وزن
                        )
            elif wf_mapped_val:
                try:
                    wf_template = WorkflowTemplate.objects.get(id=int(wf_mapped_val), is_deleted=False)
                except (ValueError, WorkflowTemplate.DoesNotExist):
                    pass

            # پارس تاریخ شروع
            start_date = parse_date_safely(sj.start_date_str)

            job = JobOpportunity.objects.create(
                code=sj.job_code,
                request_number=sj.job_code,
                title=sj.title,
                department=sj.department,
                headcount=int(sj.headcount),
                workflow=wf_template,
                start_date=start_date,
                status=JobOpportunity.STATUS_RECEIVED, # وضعیت اولیه، بعداً بروزرسانی می‌شود
                assigned_recruiter=user,
                description=f"فرصت شغلی ایمپورت شده تاریخی از اکسل. الگوی اولیه: {sj.workflow_pattern}"
            )
            # متد JobOpportunity.save به طور خودکار مراحل را از الگو به JobOpportunityStage کپی می‌کند
            
        job_map[sj.job_code] = job
        sj.final_job = job
        sj.save(update_fields=['final_job'])

    # ۲. پردازش متقاضیان موقت
    staging_candidates = import_session.staging_candidates.filter(is_valid=True)
    
    # گروه‌بندی کاندیداها بر اساس کد ملی
    candidates_by_nid = {}
    for sc in staging_candidates:
        if sc.national_id not in candidates_by_nid:
            candidates_by_nid[sc.national_id] = []
        candidates_by_nid[sc.national_id].append(sc)

    for nid, sc_list in candidates_by_nid.items():
        sc_first = sc_list[0]
        
        # دریافت یا ایجاد متقاضی
        candidate = Candidate.objects.filter(national_id=nid, is_deleted=False).first()
        if not candidate:
            log_info(f"در حال ایجاد متقاضی جدید: {sc_first.first_name} {sc_first.last_name} ({nid})", sc_first.sheet_name, sc_first.row_index)
            candidate = Candidate.objects.create(
                national_id=nid,
                first_name=sc_first.first_name,
                last_name=sc_first.last_name,
                phone_number=sc_first.phone_number,
                email=sc_first.email
            )
        else:
            log_info(f"متقاضی '{candidate.first_name} {candidate.last_name}' با کد ملی '{nid}' از قبل وجود دارد. استفاده مجدد.", sc_first.sheet_name, sc_first.row_index)
            if sc_first.phone_number and not candidate.phone_number:
                candidate.phone_number = sc_first.phone_number
            if sc_first.email and not candidate.email:
                candidate.email = sc_first.email
            candidate.save()

        # ثبت درخواست همکاری
        jobs_for_cand = set(sc.job_code for sc in sc_list if sc.job_code in job_map)
        
        for jcode in jobs_for_cand:
            job = job_map[jcode]
            
            application = JobApplication.objects.filter(job=job, candidate=candidate, is_deleted=False).first()
            if not application:
                application = JobApplication.objects.create(
                    job=job,
                    candidate=candidate,
                    status=JobApplication.STATUS_IN_PROGRESS
                )
            
            # ثبت نمرات و وضعیت مراحل
            stages_for_job_and_cand = [sc for sc in sc_list if sc.job_code == jcode]
            
            for sc in stages_for_job_and_cand:
                # یافتن مرحله مربوط به فرصت شغلی
                job_stage = job.stages.filter(stage_type=sc.stage_type, is_deleted=False).first()
                if not job_stage:
                    job_stage = job.stages.filter(name__icontains=sc.sheet_name, is_deleted=False).first()

                if not job_stage:
                    log_warn(f"مرحله ارزیابی متناظر با شیت '{sc.sheet_name}' (نوع: {sc.stage_type}) برای فرصت شغلی '{job.title}' یافت نشد.", sc.sheet_name, sc.row_index)
                    continue

                stage_state, _ = ApplicationStageState.objects.get_or_create(
                    application=application,
                    stage=job_stage
                )
                
                stage_state.status = ApplicationStageState.STATUS_COMPLETED
                stage_state.score = float(sc.score) if sc.score else 0.0
                
                # تعیین تاریخ ارزیابی
                eval_date = parse_date_safely(sc.evaluation_date_str)
                if not eval_date:
                    eval_date = job.start_date
                
                stage_state.evaluation_date = eval_date
                stage_state.evaluator = user
                stage_state.save()

                log_info(f"ثبت موفق نمره {stage_state.score} در مرحله '{job_stage.name}' برای متقاضی '{candidate}'", sc.sheet_name, sc.row_index)

                # پاس کردن خودکار مراحل قبلی
                prior_stages = job.stages.filter(sequence__lt=job_stage.sequence, is_deleted=False)
                for ps in prior_stages:
                    ps_state, _ = ApplicationStageState.objects.get_or_create(
                        application=application,
                        stage=ps
                    )
                    if ps_state.status == ApplicationStageState.STATUS_PENDING:
                        ps_state.status = ApplicationStageState.STATUS_COMPLETED
                        ps_state.score = ps.passing_score
                        ps_state.evaluation_date = eval_date or job.start_date
                        ps_state.evaluator = user
                        ps_state.save()
                        log_info(f"پاس‌کردن خودکار مرحله قبلی '{ps.name}' به علت حضور متقاضی در مرحله بالاتر.", sc.sheet_name, sc.row_index)

            application.save()
            
            for sc in stages_for_job_and_cand:
                sc.final_candidate = candidate
                sc.save(update_fields=['final_candidate'])

    # ۳. بروزرسانی وضعیت نهایی کل فرصت‌های شغلی
    for job in job_map.values():
        job.update_status()

    log_info("فرآیند ایمپورت سوابق تاریخی جذب با موفقیت به پایان رسید.")
    import_session.status = 'COMPLETED'
    import_session.save(update_fields=['status'])


@transaction.atomic
def import_fixed_template_excel(excel_file, user):
    """
    درون‌ریزی فایل اکسل قالب ثابت (سوابق جذب تاریخی)
    پشتیبانی از ۱فایل با شیت‌های: جدول وضعیت، ثبت نام، غربالگری، کتبی، مهارتی، مصاحبه، کانون، قبولی نهایی
    """
    from django.contrib.auth.models import User
    from apps.accounts.models import UserProfile
    from apps.jobs.models import JobOpportunity, JobOpportunityStage, WorkflowTemplate, WorkflowStageTemplate
    from apps.candidates.models import Candidate, JobApplication, ApplicationStageState
    from apps.core.models import AuditLog
    import openpyxl
    import datetime

    stats = {
        'jobs_created': 0,
        'jobs_updated': 0,
        'candidates_created': 0,
        'candidates_updated': 0,
        'stages_populated': 0,
        'applications_processed': 0,
        'absent_count': 0,
        'warnings': [],
    }

    # بارگذاری فایل اکسل
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet_names = wb.sheetnames

    # توابع کمکی محلی
    def clean_excel_code(val):
        if val is None:
            return ""
        val_str = str(val).strip()
        if val_str.endswith(".0"):
            return val_str[:-2]
        return val_str

    def clean_str_number(val):
        if val is None:
            return ""
        s = normalize_persian_digits(val).strip()
        if s.endswith('.0'):
            s = s[:-2]
        s = ''.join(c for c in s if c.isdigit())
        return s

    def map_job_status(status_str):
        if not status_str:
            return JobOpportunity.STATUS_RECEIVED
        s = str(status_str).strip()
        if s == '**':
            return JobOpportunity.STATUS_CLOSED
        if s == '***':
            return JobOpportunity.STATUS_CANCELLED
        if any(kw in s for kw in ["لغو"]):
            return JobOpportunity.STATUS_CANCELLED
        if any(kw in s for kw in ["توقف"]):
            return JobOpportunity.STATUS_SUSPENDED
        if any(kw in s for kw in ["غربالگری"]):
            return JobOpportunity.STATUS_SCREENING
        if any(kw in s for kw in ["کتبی"]):
            return JobOpportunity.STATUS_EXAM
        if any(kw in s for kw in ["مهارتی", "عملی"]):
            return JobOpportunity.STATUS_SKILL_TEST
        if any(kw in s for kw in ["مصاحبه"]):
            return JobOpportunity.STATUS_INTERVIEW
        if any(kw in s for kw in ["کانون"]):
            return JobOpportunity.STATUS_ASSESSMENT
        if any(kw in s for kw in ["پایان", "خاتمه"]):
            return JobOpportunity.STATUS_CLOSED
        return JobOpportunity.STATUS_RECEIVED

    def get_stage_status(result_val, score, passing_score):
        if not result_val:
            return ApplicationStageState.STATUS_COMPLETED if score >= passing_score else ApplicationStageState.STATUS_FAILED
        
        res_str = str(result_val).strip()
        res_str = res_str.replace('ي', 'ی').replace('ك', 'ک').replace('\u200c', '').strip()
        
        if "غایب" in res_str:
            return ApplicationStageState.STATUS_FAILED
        if any(kw in res_str for kw in ["غیرمجاز", "غیر مجاز", "مردود", "رد"]):
            return ApplicationStageState.STATUS_FAILED
        if any(kw in res_str for kw in ["مجاز", "قبول", "تایید"]):
            return ApplicationStageState.STATUS_COMPLETED
            
        return ApplicationStageState.STATUS_COMPLETED if score >= passing_score else ApplicationStageState.STATUS_FAILED

    def parse_score(val):
        if val is None:
            return 0.0
        val_str = normalize_persian_digits(val).strip()
        if val_str in ['*', '-', '?', '.', '#n/a', 'N/A', '', 'None', 'null']:
            return 0.0
        try:
            return float(val_str)
        except ValueError:
            return 0.0

    def get_cell_by_header(row, headers, header_name):
        if header_name in headers:
            return row[headers.index(header_name)]
        return None

    def get_or_create_workflow_template(pattern_name):
        if not pattern_name:
            return None
        name = f"الگوی خودکار - {pattern_name}"
        wf = WorkflowTemplate.all_objects.filter(name=name).first()
        if wf:
            if wf.is_deleted:
                wf.is_deleted = False
                wf.deleted_at = None
                wf.save()
            return wf
            
        wf = WorkflowTemplate.objects.create(name=name, description="ایجاد شده خودکار توسط سیستم ایمپورت سوابق قالب ثابت")
        # Split pattern by +
        parts = [p.strip() for p in pattern_name.split('+') if p.strip()]
        if not any(p in ['غربالگری', 'غربال'] for p in parts):
            parts.insert(0, 'غربالگری')
        
        non_screening_parts = [p for p in parts if p not in ['غربالگری', 'غربال']]
        m = len(non_screening_parts)
        
        def map_type(p):
            p_clean = p.lower()
            if any(kw in p_clean for kw in ['غربالگری', 'غربال']):
                return 'SCREENING'
            elif any(kw in p_clean for kw in ['کتبی', 'آزمون کتبی', 'آزمون', 'exam', 'test']):
                return 'EXAM'
            elif any(kw in p_clean for kw in ['مهارتی', 'آزمون مهارتی', 'عملی', 'skill_test']):
                return 'SKILL_TEST'
            elif any(kw in p_clean for kw in ['مصاحبه', 'interview']):
                return 'INTERVIEW'
            elif any(kw in p_clean for kw in ['کانون', 'ارزیابی', 'assessment', 'سنتر', 'competency', 'کانون ارزیابی']):
                return 'ASSESSMENT'
            else:
                return 'OTHER'
        
        for idx, part in enumerate(parts, start=1):
            st_type = map_type(part)
            if st_type == 'SCREENING':
                weight = 0
            else:
                try:
                    p_idx = non_screening_parts.index(part)
                    if p_idx == m - 1:
                        weight = 100 - (100 // m) * (m - 1)
                    else:
                        weight = 100 // m
                except ValueError:
                    weight = 0
            
            WorkflowStageTemplate.objects.create(
                workflow=wf,
                name=part,
                default_weight=weight,
                sequence=idx,
                stage_type=st_type
            )
        return wf

    job_stage_dates = {}
    processed_app_ids = set()

    # ۱. پردازش جدول وضعیت (Jobs master)
    if 'جدول وضعیت' in sheet_names:
        ws = wb['جدول وضعیت']
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            for row_idx, row in enumerate(rows[1:], start=2):
                if not any(cell is not None for cell in row):
                    continue
                
                job_code = clean_excel_code(get_cell_by_header(row, headers, "کد"))
                if not job_code:
                    continue
                
                title = str(get_cell_by_header(row, headers, "عنوان پست") or "").strip()
                department = str(get_cell_by_header(row, headers, "واحد متقاضی") or "").strip()
                unit = department
                
                raw_category = str(get_cell_by_header(row, headers, "رده شغلی") or "").strip()
                job_category = None
                category_choices = [c[0] for c in JobOpportunity.CATEGORY_CHOICES]
                for choice in category_choices:
                    if choice.replace(' ', '').replace('\u200c', '') == raw_category.replace(' ', '').replace('\u200c', ''):
                        job_category = choice
                        break
                if not job_category and raw_category:
                    if "اپراتور" in raw_category:
                        job_category = "اپراتور - تعمیرکار"
                    elif "مسئول" in raw_category:
                        if "کارشناس" in raw_category:
                            job_category = "کارشناس مسئول"
                        else:
                            job_category = "کاردان مسئول"
                    elif "کارشناس" in raw_category:
                        if "مدیریت" in raw_category:
                            job_category = "کارشناس مدیریت"
                        else:
                            job_category = "کارشناس"
                    elif "کاردان" in raw_category:
                        job_category = "کاردان"
                
                headcount_val = get_cell_by_header(row, headers, "تعداد مورد نیاز")
                try:
                    headcount = int(float(str(headcount_val).strip())) if headcount_val is not None else 1
                except (ValueError, TypeError):
                    headcount = 1
                
                start_date_val = get_cell_by_header(row, headers, "شروع ثبت‌نام")
                start_date = parse_date_safely(start_date_val)
                if not start_date:
                    start_date = datetime.date.today()
                
                raw_status = get_cell_by_header(row, headers, "آخرین مرحله")
                job_status = map_job_status(raw_status)
                
                workflow_pattern = str(get_cell_by_header(row, headers, "مسیر پیشنهادی (عنوان)") or "").strip()
                description = str(get_cell_by_header(row, headers, "توضیحات") or "").strip()
                
                # ثبت تاریخ مراحل جهت استفاده بعدی در شیت‌ها
                job_stage_dates[job_code] = {}
                for st_type, header_key in [('SCREENING', 'پایان غربالگری'), ('EXAM', 'آزمون کتبی'), ('SKILL_TEST', 'آزمون مهارتی'), ('INTERVIEW', 'مصاحبه'), ('ASSESSMENT', 'معرفی به کانون')]:
                    date_val = get_cell_by_header(row, headers, header_key)
                    if date_val:
                        parsed_dt = parse_date_safely(date_val)
                        if parsed_dt:
                            job_stage_dates[job_code][st_type] = parsed_dt
                
                wf_template = None
                if workflow_pattern:
                    wf_template = get_or_create_workflow_template(workflow_pattern)
                
                job = JobOpportunity.all_objects.filter(code=job_code).first()
                if job:
                    if job.is_deleted:
                        job.is_deleted = False
                        job.deleted_at = None
                    job.title = title or job.title
                    job.department = department or job.department
                    job.unit = unit or job.unit
                    job.headcount = headcount
                    if job_category:
                        job.job_category = job_category
                    job.start_date = start_date
                    job.status = job_status
                    if description:
                        job.description = description
                    if not job.workflow:
                        job.workflow = wf_template
                    job.save()
                    stats['jobs_updated'] += 1
                else:
                    job = JobOpportunity.objects.create(
                        code=job_code,
                        request_number=job_code,
                        title=title or f"شغل {job_code}",
                        department=department,
                        unit=unit,
                        headcount=headcount,
                        job_category=job_category or "کارشناس",
                        start_date=start_date,
                        status=job_status,
                        workflow=wf_template,
                        assigned_recruiter=user,
                        source=JobOpportunity.SOURCE_IMPORT,
                        description=description or f"فرصت شغلی ایمپورت شده تاریخی از اکسل. الگو: {workflow_pattern}"
                    )
                    stats['jobs_created'] += 1

    # ۲. پردازش ثبت نام
    if 'ثبت نام' in sheet_names:
        ws = wb['ثبت نام']
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            for row_idx, row in enumerate(rows[1:], start=2):
                if not any(cell is not None for cell in row):
                    continue
                
                job_code = clean_excel_code(get_cell_by_header(row, headers, "ExamCode"))
                national_id = clean_str_number(get_cell_by_header(row, headers, "NationCode"))
                
                if not job_code or not national_id:
                    key_val = str(get_cell_by_header(row, headers, "Key") or "").strip()
                    if key_val and len(key_val) > 10:
                        national_id = clean_str_number(key_val[-10:])
                        job_code = clean_excel_code(key_val[:-10])
                
                if not job_code or not national_id:
                    stats['warnings'].append(f"ردیف ناقص در شیت ثبت نام (ردیف {row_idx})")
                    continue
                
                job = JobOpportunity.objects.filter(code=job_code).first()
                if not job:
                    stats['warnings'].append(f"شغل با کد {job_code} یافت نشد (ردیف {row_idx} شیت ثبت نام)")
                    continue
                
                first_name = str(get_cell_by_header(row, headers, "نام") or "").strip()
                last_name = str(get_cell_by_header(row, headers, "نام خانوادگی") or "").strip()
                phone_number = clean_str_number(get_cell_by_header(row, headers, "شماره همراه"))
                
                if len(national_id) < 10:
                    stats['warnings'].append(f"کد ملی کوتاه است (ردیف {row_idx} شیت ثبت نام): {national_id}")
                
                candidate = Candidate.all_objects.filter(national_id=national_id).first()
                if candidate:
                    if candidate.is_deleted:
                        candidate.is_deleted = False
                        candidate.deleted_at = None
                    candidate.first_name = first_name or candidate.first_name
                    candidate.last_name = last_name or candidate.last_name
                    if phone_number:
                        candidate.phone_number = phone_number
                    candidate.save()
                    stats['candidates_updated'] += 1
                else:
                    username = national_id
                    password_phone = phone_number
                    if password_phone and not password_phone.startswith('0'):
                        password_phone = '0' + password_phone
                    
                    user_obj = User.objects.filter(username=username).first()
                    if not user_obj:
                        user_obj = User.objects.create_user(username=username, password=password_phone or '123456')
                    
                    user_profile, _ = UserProfile.objects.get_or_create(user=user_obj)
                    user_profile.role = UserProfile.ROLE_CANDIDATE
                    if phone_number:
                        user_profile.phone_number = phone_number
                    user_profile.save()
                    
                    candidate = Candidate.objects.create(
                        user=user_obj,
                        first_name=first_name,
                        last_name=last_name,
                        national_id=national_id,
                        phone_number=phone_number
                    )
                    stats['candidates_created'] += 1
                
                application = JobApplication.all_objects.filter(job=job, candidate=candidate).first()
                if application:
                    if application.is_deleted:
                        application.is_deleted = False
                        application.deleted_at = None
                    application.status = JobApplication.STATUS_IN_PROGRESS
                    application._bypass_stage_recalculation = True
                    application._bypass_screening_auto_fail = True
                    application.save()
                    stats['applications_processed'] += 1
                else:
                    application = JobApplication.objects.create(
                        job=job,
                        candidate=candidate,
                        status=JobApplication.STATUS_IN_PROGRESS
                    )
                    stats['applications_processed'] += 1
                
                processed_app_ids.add(application.id)

    # ۳. پردازش شیت‌های مراحل
    stage_sheets_config = {
        'غربالگری': {
            'stage_type': 'SCREENING',
            'score_col': None,
            'result_col': 'Result',
            'notes_cols': ['EXP', 'Description', 'توضیحات'],
        },
        'کتبی': {
            'stage_type': 'EXAM',
            'score_col': 'ScoreW',
            'result_col': 'Result1',
            'notes_cols': ['توضیحات'],
        },
        'مهارتی': {
            'stage_type': 'SKILL_TEST',
            'score_col': 'ScoreS',
            'result_col': 'Result2',
            'notes_cols': ['توضیحات'],
        },
        'مصاحبه': {
            'stage_type': 'INTERVIEW',
            'score_col': 'ScoreI',
            'result_col': 'Result3',
            'notes_cols': ['توضیحات'],
        },
        'کانون': {
            'stage_type': 'ASSESSMENT',
            'score_col': 'ScoreAC',
            'result_col': 'Result4',
            'notes_cols': ['توضیحات'],
        }
    }

    for sheet_name, config in stage_sheets_config.items():
        if sheet_name in sheet_names:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                headers = [str(h).strip() if h is not None else "" for h in rows[0]]
                for row_idx, row in enumerate(rows[1:], start=2):
                    if not any(cell is not None for cell in row):
                        continue
                    
                    job_code = clean_excel_code(get_cell_by_header(row, headers, "ExamCode"))
                    national_id = clean_str_number(get_cell_by_header(row, headers, "NationCode"))
                    
                    if not job_code or not national_id:
                        key_val = str(get_cell_by_header(row, headers, "Key") or "").strip()
                        if key_val and len(key_val) > 10:
                            national_id = clean_str_number(key_val[-10:])
                            job_code = clean_excel_code(key_val[:-10])
                    
                    if not job_code or not national_id:
                        continue
                    
                    job = JobOpportunity.objects.filter(code=job_code).first()
                    candidate = Candidate.objects.filter(national_id=national_id).first()
                    
                    if not job or not candidate:
                        stats['warnings'].append(f"رکورد یتیم در شیت {sheet_name}: کدملی {national_id} برای شغل {job_code} (ردیف {row_idx})")
                        continue
                    
                    application = JobApplication.all_objects.filter(job=job, candidate=candidate).first()
                    if not application:
                        stats['warnings'].append(f"درخواست متناظر یافت نشد برای شیت {sheet_name} ردیف {row_idx}")
                        continue
                    
                    stage = job.stages.filter(stage_type=config['stage_type'], is_deleted=False).first()
                    if not stage:
                        stage = job.stages.filter(name__icontains=sheet_name, is_deleted=False).first()
                    
                    if not stage:
                        stats['warnings'].append(f"مرحله {sheet_name} برای شغل {job_code} تعریف نشده است (ردیف {row_idx})")
                        continue
                    
                    stage_state = ApplicationStageState.all_objects.filter(
                        application=application,
                        stage=stage
                    ).first()
                    if stage_state:
                        if stage_state.is_deleted:
                            stage_state.is_deleted = False
                            stage_state.deleted_at = None
                    else:
                        stage_state = ApplicationStageState(
                            application=application,
                            stage=stage
                        )
                    
                    result_val = get_cell_by_header(row, headers, config['result_col'])
                    score_val = get_cell_by_header(row, headers, config['score_col']) if config['score_col'] else None
                    score = parse_score(score_val)
                    
                    notes_val = ""
                    for n_col in config['notes_cols']:
                        v = get_cell_by_header(row, headers, n_col)
                        if v is not None:
                            notes_val = str(v).strip()
                            break
                    
                    is_absent = False
                    if result_val and "غایب" in str(result_val):
                        is_absent = True
                    
                    if is_absent:
                        stage_state.status = ApplicationStageState.STATUS_FAILED
                        stage_state.score = 0.0
                        stage_state.notes = f"غایب در {stage.name}"
                        stats['absent_count'] += 1
                    else:
                        status = get_stage_status(result_val, score, stage.passing_score)
                        stage_state.status = status
                        stage_state.score = score
                        if notes_val:
                            stage_state.notes = notes_val
                    
                    # تاریخ ارزیابی
                    eval_date = None
                    for h in headers:
                        if "تاریخ" in h or "date" in h.lower():
                            date_val = get_cell_by_header(row, headers, h)
                            eval_date = parse_date_safely(date_val)
                            if eval_date:
                                break
                    
                    if not eval_date:
                        eval_date = job_stage_dates.get(job_code, {}).get(config['stage_type'])
                    if not eval_date:
                        eval_date = job.start_date
                    
                    stage_state.evaluation_date = eval_date or datetime.date.today()
                    stage_state.evaluator = user
                    stage_state._bypass_status_calculation = True
                    stage_state.save()
                    
                    stats['stages_populated'] += 1
                    processed_app_ids.add(application.id)

    # ۴. قبولی نهایی
    if 'قبولی نهایی' in sheet_names:
        ws = wb['قبولی نهایی']
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            for row_idx, row in enumerate(rows[1:], start=2):
                if not any(cell is not None for cell in row):
                    continue
                job_code = clean_excel_code(get_cell_by_header(row, headers, "ExamCode"))
                national_id = clean_str_number(get_cell_by_header(row, headers, "NationCode"))
                if not job_code or not national_id:
                    key_val = str(get_cell_by_header(row, headers, "key") or get_cell_by_header(row, headers, "Key") or "").strip()
                    if key_val and len(key_val) > 10:
                        national_id = clean_str_number(key_val[-10:])
                        job_code = clean_excel_code(key_val[:-10])
                
                if not job_code or not national_id:
                    continue
                
                app = JobApplication.objects.filter(job__code=job_code, candidate__national_id=national_id).first()
                if app:
                    app.status = JobApplication.STATUS_SELECTED
                    app._bypass_stage_recalculation = True
                    app._bypass_screening_auto_fail = True
                    app.save()
                    processed_app_ids.add(app.id)

    # ۵. فرآیند بازمحاسبه پس از ایمپورت
    for app_id in processed_app_ids:
        app = JobApplication.objects.get(id=app_id)
        job = app.job
        stages = list(job.stages.filter(is_deleted=False).order_by('sequence'))
        if not stages:
            continue
        
        # ۱. پاس‌کردن خودکار مراحل قبل
        stage_states = {state.stage_id: state for state in app.stage_states.filter(is_deleted=False)}
        active_states = [state for state in stage_states.values() if state.status != ApplicationStageState.STATUS_PENDING]
        
        if active_states:
            max_active_seq = max(state.stage.sequence for state in active_states)
            for stage in stages:
                if stage.sequence < max_active_seq:
                    ps_state = stage_states.get(stage.id)
                    if not ps_state:
                        ps_state = ApplicationStageState.all_objects.filter(application=app, stage=stage).first()
                        if ps_state:
                            if ps_state.is_deleted:
                                ps_state.is_deleted = False
                                ps_state.deleted_at = None
                        else:
                            ps_state = ApplicationStageState(application=app, stage=stage)
                    if ps_state.status == ApplicationStageState.STATUS_PENDING or not ps_state.pk:
                        ps_state.status = ApplicationStageState.STATUS_COMPLETED
                        ps_state.score = stage.passing_score
                        ps_state.evaluation_date = job.start_date or datetime.date.today()
                        ps_state.evaluator = user
                        ps_state._bypass_status_calculation = True
                        ps_state.save()
                        stage_states[stage.id] = ps_state
                        stats['stages_populated'] += 1
        
        # ۲. منطق غربالگری پیش‌فرض و نام نویسی بدون ارزیابی بعدی
        first_stage = stages[0]
        if first_stage.stage_type == 'SCREENING' or 'غربالگری' in first_stage.name:
            scr_state = stage_states.get(first_stage.id)
            if not scr_state:
                scr_state = ApplicationStageState.all_objects.filter(application=app, stage=first_stage).first()
                if scr_state:
                    if scr_state.is_deleted:
                        scr_state.is_deleted = False
                        scr_state.deleted_at = None
                        scr_state._bypass_status_calculation = True
                        scr_state.save()
                else:
                    scr_state = ApplicationStageState.objects.create(
                        application=app,
                        stage=first_stage,
                        status=ApplicationStageState.STATUS_PENDING,
                        score=0.0
                    )
                stage_states[first_stage.id] = scr_state
            
            if scr_state.status == ApplicationStageState.STATUS_PENDING:
                past_screening_statuses = [
                    JobOpportunity.STATUS_EXAM,
                    JobOpportunity.STATUS_SKILL_TEST,
                    JobOpportunity.STATUS_INTERVIEW,
                    JobOpportunity.STATUS_ASSESSMENT,
                    JobOpportunity.STATUS_FINAL_SELECTION,
                    JobOpportunity.STATUS_CLOSED,
                    JobOpportunity.STATUS_CANCELLED,
                    JobOpportunity.STATUS_SUSPENDED
                ]
                if job.status in past_screening_statuses:
                    has_subsequent_eval = False
                    for state in stage_states.values():
                        if state.stage.sequence > first_stage.sequence:
                            if state.status != ApplicationStageState.STATUS_PENDING or state.score > 0.0:
                                has_subsequent_eval = True
                                break
                    
                    should_fail = True
                    if job.status == JobOpportunity.STATUS_EXAM:
                        other_has_scores = ApplicationStageState.objects.filter(
                            application__job=job,
                            stage__sequence__gt=first_stage.sequence,
                            is_deleted=False
                        ).exclude(
                            score=0.0,
                            status=ApplicationStageState.STATUS_PENDING
                        ).exists()
                        if not other_has_scores:
                            should_fail = False
                    
                    if has_subsequent_eval:
                        scr_state.status = ApplicationStageState.STATUS_COMPLETED
                        scr_state.score = first_stage.passing_score
                        scr_state.evaluation_date = job.start_date or datetime.date.today()
                        scr_state.evaluator = user
                        scr_state._bypass_status_calculation = True
                        scr_state.save()
                    elif should_fail:
                        scr_state.status = ApplicationStageState.STATUS_FAILED
                        scr_state.score = 0.0
                        scr_state.evaluation_date = job.start_date or datetime.date.today()
                        scr_state.evaluator = user
                        scr_state._bypass_status_calculation = True
                        scr_state.save()
                        
                        if app.status != JobApplication.STATUS_REJECTED:
                            app.status = JobApplication.STATUS_REJECTED
                            app._bypass_stage_recalculation = True
                            app._bypass_screening_auto_fail = True
                            app.save()
                else:
                    scr_state.status = ApplicationStageState.STATUS_COMPLETED
                    scr_state.score = first_stage.passing_score
                    scr_state.evaluation_date = job.start_date or datetime.date.today()
                    scr_state.evaluator = user
                    scr_state._bypass_status_calculation = True
                    scr_state.save()

        # ۳. محاسبه نمره کل و وضعیت نهایی درخواست
        has_failed_stage = app.stage_states.filter(status=ApplicationStageState.STATUS_FAILED, is_deleted=False).exists()
        total_weighted_score = 0.0
        states = app.stage_states.filter(is_deleted=False)
        for state in states:
            total_weighted_score += (state.score * state.stage.weight) / 100.0
        
        app.final_score = round(total_weighted_score, 2)
        if app.status != JobApplication.STATUS_SELECTED:
            if has_failed_stage:
                app.status = JobApplication.STATUS_REJECTED
            else:
                app.status = JobApplication.STATUS_IN_PROGRESS
        
        app.recalculate_current_stage(save=False)
        app._bypass_stage_recalculation = True
        app._bypass_screening_auto_fail = True
        app.save()

    # ۶. بروزرسانی وضعیت فرصت‌های شغلی
    processed_job_ids = set(app.job_id for app in JobApplication.objects.filter(id__in=processed_app_ids))
    for job_id in processed_job_ids:
        job = JobOpportunity.objects.get(id=job_id)
        job.update_status()

    # ثبت در لاگ حسابرسی (AuditLog)
    AuditLog.objects.create(
        user=user,
        action_type=AuditLog.ACTION_CREATE,
        model_name="ImportSession",
        object_id=str(getattr(excel_file, 'name', 'fixed_template')),
        changes={
            'action': 'Fixed Template Excel Import',
            'jobs_created': stats['jobs_created'],
            'jobs_updated': stats['jobs_updated'],
            'candidates_created': stats['candidates_created'],
            'candidates_updated': stats['candidates_updated'],
            'applications_processed': stats['applications_processed'],
            'absent_count': stats['absent_count'],
            'stages_populated': stats['stages_populated']
        }
    )

    return stats


