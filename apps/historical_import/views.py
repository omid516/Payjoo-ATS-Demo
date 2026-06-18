from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse
from django.db import connection, transaction
from apps.accounts.permissions import RoleRequiredMixin
from apps.accounts.models import UserProfile
from apps.jobs.models import WorkflowTemplate
from .models import ImportSession, StagingJobOpportunity, StagingCandidate, ImportSessionLog
from .utils import analyze_excel_structure, parse_and_stage_data, execute_final_import

class UploadExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]
    template_name = 'historical_import/upload.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return render(request, self.template_name, {'error': 'لطفاً فایل اکسل را بارگذاری کنید.'})
        
        # بررسی فرمت فایل
        if not (excel_file.name.endswith('.xlsx') or excel_file.name.endswith('.xls')):
            return render(request, self.template_name, {'error': 'فرمت فایل باید Excel (.xlsx or .xls) باشد.'})

        # ایجاد نشست جدید
        import_session = ImportSession.objects.create(
            excel_file=excel_file,
            created_by=request.user,
            status='PENDING'
        )

        try:
            # تحلیل ساختار فایل به صورت همزمان
            analyze_excel_structure(import_session)
            return redirect('import_mapping', session_id=import_session.id)
        except Exception as e:
            return render(request, self.template_name, {'error': f'خطا در تحلیل ساختار فایل اکسل: {str(e)}'})


class MappingView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]
    template_name = 'historical_import/mapping.html'

    def get(self, request, session_id):
        import_session = get_object_or_404(ImportSession, id=session_id)
        if import_session.status not in ['ANALYZED', 'MAPPED', 'PREVIEWED', 'COMPLETED', 'FAILED']:
            return redirect('import_upload')

        # بررسی اینکه آیا کاربر شیت اصلی یا ستون الگو را تغییر داده است
        selected_main_sheet = request.GET.get('main_sheet')
        selected_workflow_col = request.GET.get('workflow_col')

        if selected_main_sheet or selected_workflow_col:
            try:
                # تحلیل مجدد با پارامترهای ارسالی
                analyze_excel_structure(
                    import_session,
                    main_sheet_name=selected_main_sheet,
                    workflow_col_name=selected_workflow_col
                )
            except Exception:
                pass

        workflow_templates = WorkflowTemplate.objects.filter(is_deleted=False)
        context = {
            'session': import_session,
            'summary': import_session.summary_data,
            'workflow_templates': workflow_templates,
        }
        return render(request, self.template_name, context)

    def post(self, request, session_id):
        import_session = get_object_or_404(ImportSession, id=session_id)
        summary = import_session.summary_data
        
        # ۱. استخراج فیلدهای نگاشت فرصت شغلی
        job_fields = {
            'job_code': request.POST.get('job_code_col'),
            'title': request.POST.get('title_col'),
            'department': request.POST.get('department_col'),
            'unit': request.POST.get('unit_col'),
            'job_category': request.POST.get('job_category_col'),
            'headcount': request.POST.get('headcount_col'),
            'status': request.POST.get('status_col'),
            'start_date': request.POST.get('start_date_col'),
            'workflow_pattern': request.POST.get('workflow_pattern_col'),
            'description': request.POST.get('description_col'),
            'referral_letter': request.POST.get('referral_letter_col'),
            'screening_date': request.POST.get('screening_date_col'),
            'exam_date': request.POST.get('exam_date_col'),
            'skill_test_date': request.POST.get('skill_test_date_col'),
            'interview_date': request.POST.get('interview_date_col'),
            'assessment_date': request.POST.get('assessment_date_col'),
        }

        # ۲. استخراج نگاشت‌های الگو
        workflow_mappings = {}
        patterns = summary.get('workflow_patterns', [])
        for pattern in patterns:
            key = f"workflow_map_{pattern}"
            val = request.POST.get(key)
            workflow_mappings[pattern] = val

        # ۳. استخراج نگاشت‌های شیت‌های مراحل
        stage_sheet_mappings = {}
        stage_sheets = summary.get('stage_sheets', [])
        for sheet in stage_sheets:
            s_name = sheet['name']
            stage_type = request.POST.get(f"stage_type_{s_name}")
            
            # اگر کاربر برای این شیت نوع مرحله انتخاب کرده باشد
            if stage_type and stage_type != 'IGNORE':
                stage_sheet_mappings[s_name] = {
                    'stage_type': stage_type,
                    'candidate_fields': {
                        'job_code': request.POST.get(f"c_job_code_col_{s_name}"),
                        'national_id': request.POST.get(f"c_national_id_col_{s_name}"),
                        'first_name': request.POST.get(f"c_first_name_col_{s_name}"),
                        'last_name': request.POST.get(f"c_last_name_col_{s_name}"),
                        'phone_number': request.POST.get(f"c_phone_number_col_{s_name}"),
                        'email': request.POST.get(f"c_email_col_{s_name}"),
                        'score': request.POST.get(f"c_score_col_{s_name}"),
                        'date': request.POST.get(f"c_date_col_{s_name}"),
                    }
                }

        # ساختار پیکربندی نگاشت نهایی
        mapping_config = {
            'main_sheet': summary.get('main_sheet'),
            'job_fields': job_fields,
            'workflow_mappings': workflow_mappings,
            'stage_sheet_mappings': stage_sheet_mappings,
        }

        try:
            # شروع استخراج داده‌ها به جداول موقت و اعتبارسنجی
            parse_and_stage_data(import_session, mapping_config)
            return redirect('import_preview', session_id=import_session.id)
        except Exception as e:
            workflow_templates = WorkflowTemplate.objects.filter(is_deleted=False)
            return render(request, self.template_name, {
                'session': import_session,
                'summary': summary,
                'workflow_templates': workflow_templates,
                'error': f'خطا در تحلیل و نگاشت داده‌ها: {str(e)}'
            })


class PreviewView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]
    template_name = 'historical_import/preview.html'

    def get(self, request, session_id):
        import_session = get_object_or_404(ImportSession, id=session_id)
        if import_session.status != 'PREVIEWED':
            return redirect('import_mapping', session_id=import_session.id)

        # استخراج داده‌های موقت
        staging_jobs = import_session.staging_jobs.all()
        staging_candidates = import_session.staging_candidates.all()
        logs = import_session.logs.all()

        # محاسبه آمار خلاصه
        total_jobs_count = staging_jobs.count()
        valid_jobs_count = staging_jobs.filter(is_valid=True).count()
        invalid_jobs_count = staging_jobs.filter(is_valid=False).count()
        conflict_jobs_count = staging_jobs.filter(final_job__isnull=False).count()

        total_candidates_count = staging_candidates.count()
        valid_candidates_count = staging_candidates.filter(is_valid=True).count()
        invalid_candidates_count = staging_candidates.filter(is_valid=False).count()

        error_logs_count = logs.filter(level='ERROR').count()
        warning_logs_count = logs.filter(level='WARNING').count()

        context = {
            'session': import_session,
            'staging_jobs': staging_jobs,
            'staging_candidates': staging_candidates,
            'logs': logs,
            'stats': {
                'total_jobs': total_jobs_count,
                'valid_jobs': valid_jobs_count,
                'invalid_jobs': invalid_jobs_count,
                'conflict_jobs': conflict_jobs_count,
                'total_candidates': total_candidates_count,
                'valid_candidates': valid_candidates_count,
                'invalid_candidates': invalid_candidates_count,
                'errors': error_logs_count,
                'warnings': warning_logs_count,
            }
        }
        return render(request, self.template_name, context)

    def post(self, request, session_id):
        import_session = get_object_or_404(ImportSession, id=session_id)
        conflict_strategy = request.POST.get('conflict_strategy', 'UPDATE')

        try:
            execute_final_import(import_session, conflict_strategy, request.user)
            return redirect('import_success', session_id=import_session.id)
        except Exception as e:
            staging_jobs = import_session.staging_jobs.all()
            staging_candidates = import_session.staging_candidates.all()
            logs = import_session.logs.all()
            
            ImportSessionLog.objects.create(
                import_session=import_session,
                level='ERROR',
                message=f"خطای بحرانی در زمان اجرای نهایی ایمپورت: {str(e)}. پایگاه داده کاملاً Rollback شد."
            )

            return render(request, self.template_name, {
                'session': import_session,
                'staging_jobs': staging_jobs,
                'staging_candidates': staging_candidates,
                'logs': import_session.logs.all(),
                'error': f'خطای غیرمنتظره در ثبت پایگاه داده: {str(e)}'
            })


class RollbackScreeningView(LoginRequiredMixin, RoleRequiredMixin, View):
    """حذف SCREENING stage های اضافه‌شده توسط ثبت_نام 1 و بازیابی current_stage_id"""
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get(self, request):
        import_date = '2026-06-12'
        with connection.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM jobs_jobopportunitystage WHERE stage_type='SCREENING' AND DATE(created_at)=%s", [import_date])
            screening_stages = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM candidates_applicationstagestate WHERE stage_id IN (SELECT id FROM jobs_jobopportunitystage WHERE stage_type='SCREENING' AND DATE(created_at)=%s)", [import_date])
            screening_states = cur.fetchone()[0]
        return JsonResponse({'screening_stages_to_delete': screening_stages, 'screening_states_to_delete': screening_states, 'note': 'POST to this URL to execute rollback'})

    @transaction.atomic
    def post(self, request):
        import_date = '2026-06-12'
        results = {}
        with connection.cursor() as cur:
            # مرحله ۱: بازیابی current_stage_id برای app هایی که stage دیگری دارند
            cur.execute("""
                UPDATE candidates_jobapplication
                SET current_stage_id = (
                    SELECT ss2.stage_id
                    FROM candidates_applicationstagestate ss2
                    JOIN jobs_jobopportunitystage s2 ON ss2.stage_id = s2.id
                    WHERE ss2.application_id = candidates_jobapplication.id
                      AND s2.stage_type != 'SCREENING'
                    ORDER BY ss2.updated_at DESC
                    LIMIT 1
                )
                WHERE current_stage_id IN (
                    SELECT id FROM jobs_jobopportunitystage
                    WHERE stage_type='SCREENING' AND DATE(created_at)=%s
                )
                AND (
                    SELECT ss2.stage_id
                    FROM candidates_applicationstagestate ss2
                    JOIN jobs_jobopportunitystage s2 ON ss2.stage_id = s2.id
                    WHERE ss2.application_id = candidates_jobapplication.id
                      AND s2.stage_type != 'SCREENING'
                    ORDER BY ss2.updated_at DESC
                    LIMIT 1
                ) IS NOT NULL
            """, [import_date])
            results['current_stage_restored'] = cur.rowcount

            # مرحله ۲: app هایی که هیچ stage دیگری ندارند → NULL
            cur.execute("""
                UPDATE candidates_jobapplication
                SET current_stage_id = NULL
                WHERE current_stage_id IN (
                    SELECT id FROM jobs_jobopportunitystage
                    WHERE stage_type='SCREENING' AND DATE(created_at)=%s
                )
            """, [import_date])
            results['current_stage_nulled'] = cur.rowcount

            # مرحله ۳: حذف jobstageplan های مرتبط
            cur.execute("""
                DELETE FROM recruitment_planning_jobstageplan
                WHERE stage_id IN (
                    SELECT id FROM jobs_jobopportunitystage
                    WHERE stage_type='SCREENING' AND DATE(created_at)=%s
                )
            """, [import_date])
            results['jobstageplans_deleted'] = cur.rowcount

            # مرحله ۴: حذف stage state های SCREENING
            cur.execute("""
                DELETE FROM candidates_applicationstagestate
                WHERE stage_id IN (
                    SELECT id FROM jobs_jobopportunitystage
                    WHERE stage_type='SCREENING' AND DATE(created_at)=%s
                )
            """, [import_date])
            results['stage_states_deleted'] = cur.rowcount

            # مرحله ۵: حذف SCREENING stage ها
            cur.execute("""
                DELETE FROM jobs_jobopportunitystage
                WHERE stage_type='SCREENING' AND DATE(created_at)=%s
            """, [import_date])
            results['screening_stages_deleted'] = cur.rowcount

            # تایید نهایی
            cur.execute("""
                SELECT COUNT(*) FROM jobs_jobopportunitystage
                WHERE stage_type='SCREENING' AND DATE(created_at)=%s
            """, [import_date])
            results['remaining'] = cur.fetchone()[0]

        return JsonResponse({'status': 'success', 'results': results})


class FixSelectedCandidatesView(LoginRequiredMixin, RoleRequiredMixin, View):
    """تنظیم is_conditional_pass=True برای stage state های FAILED متقاضیانی که app_status=SELECTED دارند"""
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get(self, request):
        with connection.cursor() as cur:
            cur.execute("""
                SELECT ss.id, c.first_name, c.last_name, s.name as stage_name, ss.score
                FROM candidates_applicationstagestate ss
                JOIN candidates_jobapplication ja ON ss.application_id = ja.id
                JOIN candidates_candidate c ON ja.candidate_id = c.id
                JOIN jobs_jobopportunitystage s ON ss.stage_id = s.id
                WHERE ja.status = 'SELECTED'
                  AND ss.status = 'FAILED'
                  AND ss.is_conditional_pass = 0
                  AND ss.is_deleted = 0
                ORDER BY c.last_name, s.name
            """)
            rows = [{'id': r[0], 'name': f"{r[1]} {r[2]}", 'stage': r[3], 'score': round(r[4], 2)} for r in cur.fetchall()]
        return JsonResponse({'count': len(rows), 'records': rows, 'note': 'POST to fix'})

    @transaction.atomic
    def post(self, request):
        with connection.cursor() as cur:
            cur.execute("""
                UPDATE candidates_applicationstagestate
                SET is_conditional_pass = 1,
                    updated_at = datetime('now')
                WHERE id IN (
                    SELECT ss.id
                    FROM candidates_applicationstagestate ss
                    JOIN candidates_jobapplication ja ON ss.application_id = ja.id
                    WHERE ja.status = 'SELECTED'
                      AND ss.status = 'FAILED'
                      AND ss.is_conditional_pass = 0
                      AND ss.is_deleted = 0
                )
            """)
            fixed = cur.rowcount
        return JsonResponse({'status': 'success', 'fixed_records': fixed})


class SuccessView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]
    template_name = 'historical_import/success.html'

    def get(self, request, session_id):
        import_session = get_object_or_404(ImportSession, id=session_id)
        logs = import_session.logs.all().order_by('id')
        
        # آمار نهایی
        imported_jobs = import_session.staging_jobs.filter(final_job__isnull=False).count()
        imported_candidates = import_session.staging_candidates.filter(final_candidate__isnull=False).count()

        context = {
            'session': import_session,
            'logs': logs,
            'stats': {
                'imported_jobs': imported_jobs,
                'imported_candidates': imported_candidates
            }
        }
        return render(request, self.template_name, context)


class DownloadFixedTemplateView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]

    def get(self, request):
        import os
        import openpyxl
        from django.conf import settings
        from django.http import FileResponse, HttpResponse

        template_path = os.path.join(settings.BASE_DIR, 'static', 'recruitment_import_template.xlsx')
        
        # اگر فایل الگو وجود ندارد، تلاش برای تولید آن از Sch3.xlsx یا تولید پویای یک فایل مشابه
        if not os.path.exists(template_path):
            original_path = os.path.join(settings.BASE_DIR, 'Sch3.xlsx')
            os.makedirs(os.path.dirname(template_path), exist_ok=True)
            
            if os.path.exists(original_path):
                try:
                    wb = openpyxl.load_workbook(original_path)
                    # حذف شیت مراحل
                    if 'مراحل' in wb.sheetnames:
                        wb.remove(wb['مراحل'])
                    
                    # حذف ستون Result از شیت ثبت نام
                    if 'ثبت نام' in wb.sheetnames:
                        ws_reg = wb['ثبت نام']
                        for col_idx in range(1, ws_reg.max_column + 1):
                            cell = ws_reg.cell(row=1, column=col_idx)
                            if cell.value == 'Result':
                                ws_reg.delete_cols(col_idx)
                                break
                    
                    # اضافه کردن شیت غربالگری بعد از شیت ثبت نام
                    if 'غربالگری' not in wb.sheetnames:
                        reg_idx = wb.sheetnames.index('ثبت نام') if 'ثبت نام' in wb.sheetnames else 0
                        ws_scr = wb.create_sheet(title='غربالگری', index=reg_idx + 1)
                        ws_scr.append(['ExamCode', 'NationCode', 'Result', 'Description'])
                        
                        from openpyxl.styles import Font, PatternFill, Alignment
                        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                        header_alignment = Alignment(horizontal="center", vertical="center")
                        for col_num in range(1, 5):
                            cell = ws_scr.cell(row=1, column=col_num)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            ws_scr.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 22
                    
                    for name in ['جدول وضعیت', 'ثبت نام', 'غربالگری', 'کتبی', 'مهارتی', 'مصاحبه', 'کانون', 'قبولی نهایی']:
                        if name in wb.sheetnames:
                            ws = wb[name]
                            max_r = ws.max_row
                            if max_r > 1:
                                ws.delete_rows(2, max_r)
                    wb.save(template_path)
                except Exception:
                    pass
            
            # اگر Sch3.xlsx وجود نداشت یا خطایی رخ داد، فایل پایه را به صورت پویا تولید می‌کنیم
            if not os.path.exists(template_path):
                wb = openpyxl.Workbook()
                # حذف شیت پیش‌فرض
                default_sheet = wb.active
                wb.remove(default_sheet)
                
                sheets_config = {
                    "defList": ['واحد متقاضی', 'رده شغلی', 'نتیجه', 'توضیحات عدم احراز', 'ردیف', 'مسیر پیشنهادی (عنوان)'],
                    "جدول وضعیت": ['Row', 'کد', 'آخرین مرحله', 'مسئول', 'واحد متقاضی', 'عنوان پست', 'رده شغلی', 'تعداد مورد نیاز', 'جلسه اولیه', 'شروع ثبت‌نام', 'پایان ثبت‌نام', 'تعداد ثبت‌نام', 'پایان غربالگری', 'تعداد واجد شرایط', 'کتبی', 'آزمون کتبی', 'تعداد حاضرین آزمون کتبی', 'اعلام نتایج کتبی', 'تعداد نفرات خروجی کتبی', 'مهارتی', 'آزمون مهارتی', 'تعداد حاضرین آزمون مهارتی', 'تعداد دعوت به مصاحبه', 'مصاحبه', 'تعداد حاضرین مصاحبه', 'معرفی به کانون', 'تعداد معرفی به کانون', 'اعلام نتیجه کانون', 'اعلام نتیجه نهایی', 'شماره نامه اعلام نتیجه نهایی', 'کد پست', 'شماره نامه درخواست', 'توضیحات', 'مسیر پیشنهادی (عنوان)', 'Column1', 'Column2', 'Column3'],
                    "ثبت نام": ['Row', 'ExamCode', 'NationCode', 'EXP', 'Key', 'نام', 'نام خانوادگی', 'شماره همراه'],
                    "غربالگری": ['ExamCode', 'NationCode', 'Result', 'Description'],
                    "کتبی": ['Row', 'ExamCode', 'NationCode', 'ScoreW', 'Result1', 'Key'],
                    "مهارتی": ['row', 'ExamCode', 'NationCode', 'ScoreS', 'CScoreS', 'Result2', 'key'],
                    "مصاحبه": ['Row', 'ExamCode', 'NationCode', 'ScoreI', 'CScoreI', 'Result3', 'key'],
                    "کانون": ['Row', 'ExamCode', 'NationCode', 'ScoreAC', 'CScoreAC', 'Result4', 'key'],
                    "قبولی نهایی": ['Row', 'ExamCode', 'NationCode', 'key']
                }
                
                for sheet_name, headers in sheets_config.items():
                    ws = wb.create_sheet(title=sheet_name)
                    ws.append(headers)
                    # استایل‌دهی هدرها جهت زیبایی بیشتر
                    from openpyxl.styles import Font, PatternFill, Alignment
                    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 22
                
                wb.save(template_path)

        if os.path.exists(template_path):
            response = FileResponse(open(template_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="recruitment_import_template.xlsx"'
            return response
            
        return HttpResponse("فایل الگو یافت نشد.", status=404)



class FixedTemplateImportView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]
    template_name = 'historical_import/fixed_success.html'

    def post(self, request):
        from django.core.exceptions import ValidationError
        from .utils import import_fixed_template_excel

        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return render(request, 'historical_import/upload.html', {'error': 'لطفاً فایل اکسل را بارگذاری کنید.'})
            
        try:
            stats = import_fixed_template_excel(excel_file, request.user)
            return render(request, self.template_name, {'stats': stats})
        except ValidationError as e:
            return render(request, 'historical_import/upload.html', {'error': e.message if hasattr(e, 'message') else str(e)})
        except Exception as e:
            return render(request, 'historical_import/upload.html', {'error': f'خطا در تحلیل و بارگذاری فایل اکسل: {str(e)}'})


class ResetDatabaseView(LoginRequiredMixin, RoleRequiredMixin, View):
    allowed_roles = [UserProfile.ROLE_ADMIN, UserProfile.ROLE_RECRUITMENT_DIRECTOR, UserProfile.ROLE_RECRUITMENT_SPECIALIST]
    template_name = 'historical_import/reset_db.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        confirm_text = request.POST.get('confirm_text', '').strip()
        if confirm_text != 'RESET DATABASE':
            return render(request, self.template_name, {'error': 'متن تاییدیه به درستی وارد نشده است.'})

        from django.db import transaction
        from apps.jobs.models import JobOpportunity, JobOpportunityStage, WorkflowTemplate, WorkflowStageTemplate, JobStageInterviewer, AssessmentCompetency
        from apps.candidates.models import (
            Candidate, CandidateEducation, CandidateExperience, JobApplication, ApplicationStageState,
            CandidateLanguage, CandidateSkill, CandidateCertificate, InterviewerScore,
            AssessorCompetencyScore, ExternalInterviewerScore, JobDefaultInterviewer
        )
        from apps.recruitment_planning.models import JobRecruitmentPlan, JobStagePlan
        from apps.historical_import.models import ImportSession, StagingJobOpportunity, StagingCandidate, ImportSessionLog
        from apps.core.models import AuditLog
        from django.contrib.auth.models import User

        try:
            with transaction.atomic():
                # Candidates scores
                InterviewerScore.all_objects.all().delete()
                AssessorCompetencyScore.all_objects.all().delete()
                ExternalInterviewerScore.all_objects.all().delete()
                JobDefaultInterviewer.all_objects.all().delete()
                
                # Candidate profile items
                CandidateEducation.all_objects.all().delete()
                CandidateExperience.all_objects.all().delete()
                CandidateLanguage.all_objects.all().delete()
                CandidateSkill.all_objects.all().delete()
                CandidateCertificate.all_objects.all().delete()
                
                # Candidate applications
                ApplicationStageState.all_objects.all().delete()
                JobApplication.all_objects.all().delete()
                
                # Delete candidates
                Candidate.all_objects.all().delete()

                # Job stages and plans
                JobStageInterviewer.all_objects.all().delete()
                AssessmentCompetency.all_objects.all().delete()
                JobStagePlan.all_objects.all().delete()
                JobRecruitmentPlan.all_objects.all().delete()
                JobOpportunityStage.all_objects.all().delete()
                JobOpportunity.all_objects.all().delete()
                
                # Workflow templates
                WorkflowStageTemplate.all_objects.all().delete()
                WorkflowTemplate.all_objects.all().delete()

                # Import sessions and staging
                ImportSessionLog.objects.all().delete()
                StagingCandidate.objects.all().delete()
                StagingJobOpportunity.objects.all().delete()
                ImportSession.objects.all().delete()
                
                # Logs
                AuditLog.objects.all().delete()

                # Candidate users (only delete users who are NOT staff and are ROLE_CANDIDATE)
                candidate_users = User.objects.filter(profile__role=UserProfile.ROLE_CANDIDATE, is_staff=False, is_superuser=False)
                candidate_users.delete()

            return render(request, self.template_name, {'success': True})
        except Exception as e:
            return render(request, self.template_name, {'error': f'خطا در پاکسازی دیتابیس: {str(e)}'})

