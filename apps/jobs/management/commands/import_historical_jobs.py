"""
Management Command: import_historical_jobs
==========================================
فاز اول ایمپورت تاریخی — فقط فرصت‌های شغلی

استفاده:
    python manage.py import_historical_jobs <path_to_excel>
                       [--sheet وضعیت]
                       [--dry-run]
                       [--conflict skip|replace]
"""

import re
import sys
import datetime
import openpyxl
import jdatetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


# ---------------------------------------------------------------------------
# ثوابت نگاشت
# ---------------------------------------------------------------------------

# نگاشت بخش‌های الگوی استخدام به stage_type
PATTERN_STAGE_MAP = {
    'کتبی':     ('آزمون کتبی',    'EXAM'),
    'مهارتی':   ('آزمون مهارتی',  'SKILL_TEST'),
    'مصاحبه':   ('مصاحبه',        'INTERVIEW'),
    'کانون':    ('کانون ارزیابی', 'ASSESSMENT'),
    'غربال':    ('غربالگری',      'SCREENING'),
    'ارزیابی':  ('کانون ارزیابی', 'ASSESSMENT'),
}

# نگاشت وضعیت فارسی اکسل → کد سیستم
STATUS_KEYWORD_MAP = [
    (['لغو', 'کنسل', 'cancel', 'لغو درخواست'],                                'CANCELLED'),
    (['توقف', 'متوقف'],                                                         'CANCELLED'),
    (['پایان فرآیند', 'اتمام', 'بسته', 'closed', 'تمام'],                     'CLOSED'),
    (['ثبت نمرات کانون', 'کانون', 'assessment'],                               'ASSESSMENT'),
    (['هماهنگی مصاحبه', 'مصاحبه', 'interview'],                               'INTERVIEW'),
    (['ثبت نمرات آزمون مهارتی', 'آزمون مهارتی', 'ارزیابی مهارتی', 'مهارتی'], 'EXAM'),
    (['هماهنگی آزمون کتبی', 'ثبت نمرات کتبی', 'آزمون کتبی', 'کتبی'],        'EXAM'),
    (['غربالگری', 'غربال', 'screening'],                                        'SCREENING'),
    (['پیش نویس آگهی', 'منتشر', 'ثبت‌نام', 'ثبت نام', 'published'],          'PUBLISHED'),
    (['برنامه‌ریزی', 'برنامه ریزی', 'planning'],                               'PLANNING'),
    (['انتخاب نهایی', 'final'],                                                 'FINAL_SELECTION'),
    (['خاتمه ثبت'],                                                             'REGISTRATION_CLOSED'),
    (['دریافت', 'در انتظار', 'ثبت درخواست'],                                   'RECEIVED'),
]

# نگاشت کلیدهای تاریخ به کلمات جستجوی هدر
DATE_COLUMN_KEYS = {
    'start_reg':        ['شروع ثبت‌نام', 'شروع ثبت نام'],
    'end_reg':          ['پایان ثبت‌نام', 'پایان ثبت نام'],
    'end_screening':    ['پایان غربالگری'],
    'exam':             ['آزمون کتبی'],
    'exam_result':      ['اعلام نتایج کتبی', 'نتایج کتبی'],
    'skill':            ['آزمون مهارتی'],
    'interview':        ['مصاحبه'],
    'kanoon_intro':     ['معرفی به کانون'],
    'kanoon_result':    ['اعلام نتیجه کانون'],
    'final_result':     ['اعلام نتیجه نهایی'],
}

# ستون‌های فیلدهای اصلی: (کلید‌های جستجو, نام متغیر داخلی)
JOB_FIELD_KEYS = {
    'job_code':         ['کد', 'کد فرصت', 'کد شغل', 'کد پست', 'code', 'شماره فرصت'],
    'title':            ['عنوان پست', 'عنوان شغل', 'عنوان', 'title', 'شغل'],
    'unit':             ['واحد متقاضی', 'واحد', 'دپارتمان', 'department', 'بخش'],
    'headcount':        ['تعداد مورد نیاز', 'تعداد', 'ظرفیت', 'headcount'],
    'current_status':   ['آخرین مرحله', 'وضعیت فعلی', 'وضعیت', 'status'],
    'workflow_pattern': ['مسیر پیشنهادی (عنوان)', 'مسیر پیشنهادی', 'الگوی استخدام', 'الگو', 'فرآیند', 'workflow', 'pattern'],
    'request_number':   ['شماره نامه درخواست', 'شماره نامه', 'نامه درخواست'],
    'job_category':     ['رده شغلی', 'رده', 'category'],
    'notes':            ['توضیحات', 'یادداشت', 'notes', 'description'],
}


# ---------------------------------------------------------------------------
# توابع کمکی
# ---------------------------------------------------------------------------

def normalize_digits(s):
    """تبدیل ارقام فارسی/عربی به انگلیسی"""
    if s is None:
        return ''
    s = str(s).strip()
    for fa, en in zip('۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩', '01234567890123456789'):
        s = s.replace(fa, en)
    return s


def parse_jalali_date(val):
    """تبدیل تاریخ شمسی (رشته یا عدد) به میلادی. در صورت شکست None برمی‌گرداند."""
    if val is None:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val

    if isinstance(val, (int, float)):
        try:
            s = str(int(val))
        except Exception:
            s = str(val)
    else:
        s = str(val).strip()

    s = normalize_digits(s).strip()
    if not s or s.lower() in ('none', 'null', '-', '', '*'):
        return None

    # فرمت YYYYMMDD بدون جداکننده (مانند 14040710)
    if s.isdigit() and len(s) == 8:
        y = int(s[0:4])
        mo = int(s[4:6])
        d = int(s[6:8])
        if 1300 <= y <= 1500 and 1 <= mo <= 12 and 1 <= d <= 31:
            try:
                return jdatetime.date(y, mo, d).togregorian()
            except Exception:
                pass

    # سریال عددی اکسل
    if s.isdigit() and len(s) in (4, 5):
        try:
            return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(s))
        except Exception:
            pass

    # فرمت YYYY/MM/DD یا YYYY-MM-DD
    m = re.match(r'^(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})', s)
    if m:
        y, mo, d = map(int, m.groups())
        if y < 100:
            y += 1300 if y > 50 else 1400
        try:
            return jdatetime.date(y, mo, d).togregorian()
        except Exception:
            pass

    # فرمت YYYY/MM
    m2 = re.match(r'^(\d{4})[/\-](\d{1,2})$', s)
    if m2:
        y, mo = map(int, m2.groups())
        try:
            return jdatetime.date(y, mo, 1).togregorian()
        except Exception:
            pass

    return None


def map_status(raw_val):
    """نگاشت متن فارسی وضعیت به کد سیستم"""
    if not raw_val:
        return 'RECEIVED'
    raw_lower = str(raw_val).strip().lower()
    for keywords, code in STATUS_KEYWORD_MAP:
        for kw in keywords:
            if kw.lower() in raw_lower:
                return code
    return 'RECEIVED'


def find_col_idx(headers, keywords):
    """یافتن ایندکس ستون بر اساس کلمات کلیدی (اولین تطابق)"""
    # اولویت اول: تطابق دقیق
    for kw in keywords:
        for i, h in enumerate(headers):
            if str(h).strip().lower() == kw.lower():
                return i
    # اولویت دوم: تطابق جزئی (پیش‌فرض قبلی)
    for kw in keywords:
        for i, h in enumerate(headers):
            if kw.lower() in str(h).strip().lower():
                return i
    return None


def detect_workflow_col(headers, rows_sample):
    """تشخیص ستون الگوی استخدام: ابتدا از هدر، سپس از مقادیر با کاراکتر +"""
    # بررسی هدرها
    for i, h in enumerate(headers):
        h_lower = str(h).strip().lower()
        for kw in ['مسیر پیشنهادی', 'الگوی استخدام', 'الگو', 'فرآیند', 'workflow', 'pattern']:
            if kw.lower() in h_lower:
                return i

    # بررسی مقادیر (ستونی که حاوی + و کلمات کلیدی باشد)
    for row in rows_sample[:20]:
        for ci, cell in enumerate(row):
            if cell and isinstance(cell, str) and '+' in cell:
                if any(kw in cell for kw in ['کتبی', 'مصاحبه', 'کانون', 'مهارتی']):
                    return ci

    # زاپاس: آخرین ستون
    return len(headers) - 1 if headers else None


def get_or_create_workflow(pattern_str, dry_run=False):
    """
    دریافت یا ایجاد WorkflowTemplate بر اساس متن الگوی استخدام.
    مثال: 'کتبی + مصاحبه + کانون'
    بازگشت: (workflow_template, created: bool)
    """
    from apps.jobs.models import WorkflowTemplate, WorkflowStageTemplate

    if not pattern_str or not pattern_str.strip():
        return None, False

    pattern_clean = pattern_str.strip()
    wf_name = f"تاریخی — {pattern_clean}"

    existing = WorkflowTemplate.objects.filter(name=wf_name, is_deleted=False).first()
    if existing:
        return existing, False

    if dry_run:
        return None, True  # نشان می‌دهد که ایجاد می‌شود

    wf = WorkflowTemplate.objects.create(
        name=wf_name,
        description=f"ایجاد شده خودکار توسط import_historical_jobs — الگو: {pattern_clean}"
    )

    parts = [p.strip() for p in pattern_clean.split('+')]
    for seq, part in enumerate(parts, start=1):
        s_type = 'OTHER'
        for kw, (_, stype) in PATTERN_STAGE_MAP.items():
            if kw in part:
                s_type = stype
                break
        WorkflowStageTemplate.objects.create(
            workflow=wf,
            name=part.strip(),
            sequence=seq,
            stage_type=s_type,
            default_weight=round(100 / len(parts)),
        )

    return wf, True


def save_stage_dates(job, plan, stages_dates, dry_run=False):
    """
    ذخیره تاریخ‌های مراحل در JobStagePlan.
    stages_dates: list of (stage_name, stage_type, start_date, end_date)
    """
    from apps.recruitment_planning.models import JobStagePlan
    from apps.jobs.models import JobOpportunityStage

    saved = 0
    for stage_name, stage_type, start_date, end_date in stages_dates:
        if not start_date:
            continue
        if end_date is None:
            end_date = start_date

        # یافتن مرحله متناظر در JobOpportunityStage
        job_stage = job.stages.filter(stage_type=stage_type, is_deleted=False).first()
        if not job_stage:
            job_stage = job.stages.filter(
                name__icontains=stage_name.replace('آزمون ', ''),
                is_deleted=False
            ).first()

        if not job_stage:
            # اگر مرحله‌ای در workflow تعریف نشده اما تاریخ دارد → نادیده بگیر (warning در caller)
            continue

        if not dry_run:
            JobStagePlan.objects.update_or_create(
                plan=plan,
                stage=job_stage,
                defaults=dict(
                    stage_type=stage_type,
                    planned_start_date=start_date,
                    planned_end_date=end_date,
                    sla_days=max((end_date - start_date).days, 1),
                )
            )
        saved += 1

    return saved


# ---------------------------------------------------------------------------
# Command اصلی
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = 'فاز اول ایمپورت تاریخی: ایجاد فرصت‌های شغلی از فایل اکسل (بدون متقاضی)'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='مسیر فایل اکسل'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default=None,
            help='نام شیت (پیش‌فرض: جستجوی خودکار شیت «وضعیت»)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='فقط تحلیل کن، هیچ رکوردی ایجاد نکن'
        )
        parser.add_argument(
            '--conflict',
            choices=['skip', 'replace'],
            default='skip',
            help='رفتار در صورت وجود فرصت شغلی تکراری (skip یا replace، پیش‌فرض: skip)'
        )

    def handle(self, *args, **options):
        excel_path = options['excel_file']
        dry_run = options['dry_run']
        conflict = options['conflict']
        sheet_name_arg = options.get('sheet')

        self.stdout.write('')
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(self.style.HTTP_INFO('  ایمپورت تاریخی فرصت‌های شغلی — فاز اول'))
        if dry_run:
            self.stdout.write(self.style.WARNING('  [DRY-RUN] هیچ تغییری در دیتابیس ایجاد نخواهد شد'))
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write('')

        # --- بارگذاری اکسل ---
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'فایل یافت نشد: {excel_path}')
        except Exception as e:
            raise CommandError(f'خطا در بارگذاری فایل: {e}')

        sheet_names = wb.sheetnames
        self.stdout.write(f'شیت‌های موجود: {", ".join(sheet_names)}')

        # --- انتخاب شیت ---
        if sheet_name_arg:
            if sheet_name_arg not in sheet_names:
                raise CommandError(f"شیت '{sheet_name_arg}' یافت نشد.")
            main_sheet = sheet_name_arg
        else:
            main_sheet = None
            for name in sheet_names:
                if 'وضعیت' in name or 'status' in name.lower() or 'فرصت' in name:
                    main_sheet = name
                    break
            if not main_sheet:
                main_sheet = sheet_names[-1]
                self.stdout.write(self.style.WARNING(
                    f"شیت «وضعیت» پیدا نشد. از آخرین شیت استفاده می‌شود: '{main_sheet}'"
                ))

        self.stdout.write(f'شیت انتخاب شده: {self.style.SUCCESS(main_sheet)}')
        self.stdout.write('')

        ws = wb[main_sheet]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise CommandError('شیت انتخاب شده خالی است.')

        headers = [str(h).strip() if h is not None else '' for h in all_rows[0]]
        data_rows = [r for r in all_rows[1:] if any(c is not None for c in r)]

        self.stdout.write(f'تعداد ستون‌ها: {len(headers)}')
        self.stdout.write(f'تعداد ردیف‌های داده: {len(data_rows)}')
        self.stdout.write('')

        # --- تشخیص ستون‌های اصلی ---
        col_map = {}
        for field, keywords in JOB_FIELD_KEYS.items():
            col_map[field] = find_col_idx(headers, keywords)

        # اگر workflow_pattern پیدا نشد، با scan مقادیر پیدا می‌کنیم
        if col_map.get('workflow_pattern') is None:
            col_map['workflow_pattern'] = detect_workflow_col(headers, data_rows)

        self.stdout.write('نگاشت ستون‌ها:')
        field_labels = {
            'job_code': 'کد فرصت شغلی',
            'title': 'عنوان شغل',
            'unit': 'واحد متقاضی',
            'headcount': 'تعداد مورد نیاز',
            'current_status': 'وضعیت فعلی',
            'workflow_pattern': 'الگوی استخدام',
        }
        for field, label in field_labels.items():
            idx = col_map.get(field)
            col_name = f'"{headers[idx]}"' if idx is not None else '— یافت نشد'
            symbol = '✓' if idx is not None else '✗'
            style = self.style.SUCCESS if idx is not None else self.style.WARNING
            self.stdout.write(f'  {symbol} {label}: {style(col_name)}')

        # --- تشخیص ستون‌های تاریخ مراحل ---
        date_col_map = {}  # key -> col_idx
        self.stdout.write('\nستون‌های تاریخ مراحل:')
        for key, keywords in DATE_COLUMN_KEYS.items():
            idx = find_col_idx(headers, keywords)
            if idx is not None:
                date_col_map[key] = idx
                self.stdout.write(f'  ✓ {key}: {self.style.SUCCESS(f"{chr(34)}{headers[idx]}{chr(34)}")}')
            else:
                self.stdout.write(f'  - {key}: {self.style.WARNING("— یافت نشد")}')

        self.stdout.write('')

        if col_map.get('job_code') is None or col_map.get('title') is None:
            raise CommandError('ستون کد فرصت شغلی یا عنوان شغل یافت نشد. لطفاً نام شیت یا ستون‌ها را بررسی کنید.')

        # --- اجرای ایمپورت ---
        self._run_import(
            data_rows=data_rows,
            headers=headers,
            col_map=col_map,
            date_col_map=date_col_map,
            dry_run=dry_run,
            conflict=conflict,
            main_sheet=main_sheet,
        )

    @transaction.atomic
    def _run_import(self, data_rows, headers, col_map, date_col_map,
                    dry_run, conflict, main_sheet):
        from apps.jobs.models import JobOpportunity
        from apps.recruitment_planning.models import JobRecruitmentPlan

        stats = {
            'jobs_created': 0,
            'jobs_skipped': 0,
            'jobs_replaced': 0,
            'workflows_created': 0,
            'stage_dates_saved': 0,
            'warnings': [],
            'errors': [],
        }

        # Cache Workflows ایجاد شده در این اجرا
        workflow_cache = {}

        def get_cell(row, field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        for row_num, row in enumerate(data_rows, start=2):
            # خواندن فیلدهای اصلی
            raw_code = get_cell(row, 'job_code')
            raw_title = get_cell(row, 'title')
            raw_unit = get_cell(row, 'unit')
            raw_headcount = get_cell(row, 'headcount')
            raw_status = get_cell(row, 'current_status')
            raw_workflow = get_cell(row, 'workflow_pattern')

            job_code = normalize_digits(raw_code).strip() if raw_code is not None else ''
            title = str(raw_title).strip() if raw_title is not None else ''
            unit = str(raw_unit).strip() if raw_unit is not None else ''
            workflow_pattern = str(raw_workflow).strip() if raw_workflow else ''

            # اعتبارسنجی
            if not job_code:
                stats['warnings'].append(f'ردیف {row_num}: کد فرصت شغلی خالی است — نادیده گرفته شد')
                continue
            if not title:
                stats['warnings'].append(f'ردیف {row_num}: عنوان شغل خالی است (کد: {job_code}) — نادیده گرفته شد')
                continue

            # تعداد
            headcount = 1
            if raw_headcount is not None:
                try:
                    headcount = int(float(normalize_digits(raw_headcount)))
                    if headcount < 1:
                        headcount = 1
                except (ValueError, TypeError):
                    stats['warnings'].append(f'ردیف {row_num} ({job_code}): تعداد نامعتبر "{raw_headcount}" — مقدار ۱ جایگزین شد')

            # وضعیت
            mapped_status = map_status(raw_status)

            # خواندن تاریخ‌های مراحل و بررسی خطاهای پارسینگ
            for key, idx in date_col_map.items():
                if idx < len(row) and row[idx] is not None:
                    val_str = str(row[idx]).strip()
                    if val_str and val_str not in ('*', '-', 'None', 'null'):
                        parsed = parse_jalali_date(row[idx])
                        if not parsed:
                            stats['warnings'].append(
                                f'ردیف {row_num} ({job_code}): تاریخ ستون "{key}" = "{row[idx]}" قابل تبدیل نیست'
                            )

            def get_date_by_key(key):
                idx = date_col_map.get(key)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return parse_jalali_date(row[idx])
                return None

            stages_dates_raw = []

            # Screening stage
            s_start = get_date_by_key('start_reg')
            s_end = get_date_by_key('end_screening') or get_date_by_key('end_reg')
            if s_start or s_end:
                s_start = s_start or s_end
                s_end = s_end or s_start
                stages_dates_raw.append(('غربالگری', 'SCREENING', s_start, s_end))

            # Exam stage
            e_start = get_date_by_key('exam')
            e_end = get_date_by_key('exam_result')
            if e_start or e_end:
                e_start = e_start or e_end
                e_end = e_end or e_start
                stages_dates_raw.append(('آزمون کتبی', 'EXAM', e_start, e_end))

            # Skill test stage
            sk_start = get_date_by_key('skill')
            if sk_start:
                stages_dates_raw.append(('آزمون مهارتی', 'SKILL_TEST', sk_start, sk_start))

            # Interview stage
            i_start = get_date_by_key('interview')
            if i_start:
                stages_dates_raw.append(('مصاحبه', 'INTERVIEW', i_start, i_start))

            # Assessment stage
            a_start = get_date_by_key('kanoon_intro')
            a_end = get_date_by_key('kanoon_result')
            if a_start or a_end:
                a_start = a_start or a_end
                a_end = a_end or a_start
                stages_dates_raw.append(('کانون ارزیابی', 'ASSESSMENT', a_start, a_end))

            # محاسبه تاریخ شروع و پایان از مراحل
            all_dates = []
            for sd in stages_dates_raw:
                if sd[2]: all_dates.append(sd[2])
                if sd[3]: all_dates.append(sd[3])

            session_date = parse_jalali_date(row[8]) if len(row) > 8 and row[8] is not None else None
            plan_start = min(all_dates) if all_dates else (session_date or datetime.date.today())
            plan_end = max(all_dates) if all_dates else (session_date or datetime.date.today())

            # بررسی تکراری بودن
            existing = None
            try:
                existing = JobOpportunity.objects.get(code=job_code, is_deleted=False)
            except JobOpportunity.DoesNotExist:
                pass

            if existing:
                if conflict == 'skip':
                    stats['jobs_skipped'] += 1
                    stats['warnings'].append(
                        f'ردیف {row_num}: فرصت شغلی "{job_code}" از قبل وجود دارد — نادیده گرفته شد (conflict=skip)'
                    )
                    continue
                elif conflict == 'replace':
                    stats['warnings'].append(
                        f'ردیف {row_num}: فرصت شغلی "{job_code}" حذف و جایگزین می‌شود (conflict=replace)'
                    )
                    if not dry_run:
                        existing.hard_delete()
                    stats['jobs_replaced'] += 1
                    existing = None

            # Workflow
            if workflow_pattern not in workflow_cache:
                wf, wf_created = get_or_create_workflow(workflow_pattern, dry_run=dry_run)
                workflow_cache[workflow_pattern] = (wf, wf_created)
                if wf_created:
                    stats['workflows_created'] += 1
                    self.stdout.write(
                        f'  + Workflow ایجاد شد: {self.style.SUCCESS(workflow_pattern)}'
                    )
            wf_obj, _ = workflow_cache[workflow_pattern]

            if dry_run:
                stats['jobs_created'] += 1
                continue

            # خواندن فیلدهای جدید
            raw_req_num = get_cell(row, 'request_number')
            req_num_clean = str(raw_req_num).strip() if raw_req_num is not None else ''
            if req_num_clean and req_num_clean not in ('None', 'null', '', '-'):
                request_number = f"{req_num_clean} ({job_code})"
                if len(request_number) > 50:
                    space_left = 50 - len(f" ({job_code})")
                    request_number = f"{req_num_clean[:space_left]} ({job_code})"
            else:
                request_number = job_code

            raw_category = get_cell(row, 'job_category')
            job_category = None
            if raw_category is not None:
                cat_val = str(raw_category).strip()
                if 'کارشناس مسئول' in cat_val:
                    job_category = 'کارشناس مسئول'
                elif 'کارشناس مدیریت' in cat_val:
                    job_category = 'کارشناس مدیریت'
                elif 'کارشناس' in cat_val:
                    job_category = 'کارشناس'
                elif 'کاردان مسئول' in cat_val:
                    job_category = 'کاردان مسئول'
                elif 'کاردان' in cat_val or 'تکنسین' in cat_val:
                    job_category = 'کاردان'
                elif 'اپراتور' in cat_val or 'تعمیرکار' in cat_val:
                    job_category = 'اپراتور - تعمیرکار'
                else:
                    job_category = cat_val

            raw_notes = get_cell(row, 'notes')
            notes = str(raw_notes).strip() if raw_notes is not None else ''

            # ایجاد فرصت شغلی
            job = JobOpportunity.objects.create(
                code=job_code,
                request_number=request_number,
                title=title,
                department=unit,
                unit=unit,
                job_category=job_category,
                headcount=headcount,
                workflow=wf_obj,
                status=mapped_status,
                source=JobOpportunity.SOURCE_IMPORT,
                description='',
                requirements='',
                start_date=plan_start,
                end_date=plan_end,
                notes=notes,
            )
            stats['jobs_created'] += 1

            # ایجاد JobRecruitmentPlan
            plan, _ = JobRecruitmentPlan.objects.update_or_create(
                job=job,
                defaults=dict(
                    start_date=plan_start,
                    predicted_end_date=plan_end,
                    status='COMPLETED' if mapped_status in ('CLOSED', 'CANCELLED') else 'ACTIVE',
                )
            )

            # ذخیره تاریخ مراحل
            saved_dates = save_stage_dates(job, plan, stages_dates_raw, dry_run=dry_run)
            stats['stage_dates_saved'] += saved_dates

            self.stdout.write(
                f'  ✓ [{row_num}] {job_code} — {title[:40]}'
                + (f' ({len(stages_dates_raw)} تاریخ مرحله)' if stages_dates_raw else '')
            )

        # --- اگر dry-run، rollback ---
        if dry_run:
            transaction.set_rollback(True)

        # --- گزارش نهایی ---
        self.stdout.write('')
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(self.style.HTTP_INFO('  گزارش نهایی ایمپورت'))
        self.stdout.write(self.style.HTTP_INFO('═' * 60))
        self.stdout.write(f'  فرصت‌های شغلی ایجاد شده    : {self.style.SUCCESS(str(stats["jobs_created"]))}')
        self.stdout.write(f'  فرصت‌های شغلی نادیده گرفته : {stats["jobs_skipped"]}')
        self.stdout.write(f'  فرصت‌های شغلی جایگزین شده  : {stats["jobs_replaced"]}')
        self.stdout.write(f'  Workflow های ایجاد شده      : {self.style.SUCCESS(str(stats["workflows_created"]))}')
        self.stdout.write(f'  تاریخ‌های مرحله ذخیره شده   : {stats["stage_dates_saved"]}')
        self.stdout.write(f'  هشدارها                     : {self.style.WARNING(str(len(stats["warnings"])))}')
        self.stdout.write(f'  خطاهای بحرانی               : {self.style.ERROR(str(len(stats["errors"])))}')

        if stats['warnings']:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING('هشدارها:'))
            for w in stats['warnings']:
                self.stdout.write(f'  ⚠  {w}')

        if stats['errors']:
            self.stdout.write('')
            self.stdout.write(self.style.ERROR('خطاهای بحرانی:'))
            for e in stats['errors']:
                self.stdout.write(f'  ✗  {e}')

        self.stdout.write('')
        if dry_run:
            self.stdout.write(self.style.WARNING('[DRY-RUN] هیچ تغییری اعمال نشد. برای اجرای واقعی --dry-run را حذف کنید.'))
        else:
            self.stdout.write(self.style.SUCCESS('ایمپورت با موفقیت به پایان رسید.'))
        self.stdout.write('')
