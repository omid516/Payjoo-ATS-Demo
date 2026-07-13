import openpyxl
from django.db import transaction
from django.utils import timezone
from .models import CentralCompetency, JobOpportunityStage, AssessmentCompetency, JobOpportunityCompetency

def normalize_persian_digits(s):
    if not s:
        return ''
    s = str(s).strip()
    for fa, en in zip('۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩', '01234567890123456789'):
        s = s.replace(fa, en)
    return s

def clean_cell_value(val):
    if val is None:
        return ''
    return str(val).strip()

def parse_competencies_excel(file_path):
    """
    Parses شایستگی ها.xlsx and imports/updates the CentralCompetency table.
    Returns a dict with import statistics.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if 'Result' not in wb.sheetnames:
        raise ValueError("شیت با نام 'Result' در فایل اکسل یافت نشد.")
    
    ws = wb['Result']
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("فایل اکسل خالی است.")
        
    headers = [clean_cell_value(h) for h in rows[0]]
    
    # Helper to find column index (case and space insensitive, handles common typos)
    def find_col(names):
        for name in names:
            name_clean = name.replace(" ", "").replace("ی", "ي").replace("ک", "ك")
            for idx, h in enumerate(headers):
                h_clean = h.replace(" ", "").replace("ی", "ي").replace("ک", "ك")
                if h_clean == name_clean:
                    return idx
        return -1

    col_map = {
        'post_code': find_col(['کد پست']),
        'post_title': find_col(['پست']),
        'code': find_col(['کد شایستگی', 'کد شايستگي']),
        'old_code': find_col(['کد شایستگی قدیم', 'کد شايستگي قديم']),
        'title': find_col(['شایستگی', 'شايستگي']),
        'category_raw': find_col(['طبقه']),
        'cluster_raw': find_col(['خوشه']),
        'importance_raw': find_col(['اهمیت شایستگی', 'اهميت شايستگي']),
        'level_raw': find_col(['سطح شایستگی', 'سطح شايستگي']),
        'management_code': find_col(['کد مدیریت', 'کد  مديريت', 'کد مديريت']),
        'management_name': find_col(['مدیریت', 'مديريت']),
        'vice_president_code': find_col(['کد معاونت']),
        'vice_president_name': find_col(['معاونت']),
        'section_code': find_col(['کد قسمت']),
        'section_name': find_col(['قسمت']),
        'cost_center_code': find_col(['کد مرکز هزینه', 'کد مرکز هزينه']),
        'cost_center_name': find_col(['مرکز هزینه', 'مرکز هزينه']),
    }
    
    # Required columns validation
    required = ['post_code', 'code', 'title', 'category_raw']
    for req in required:
        if col_map[req] == -1:
            raise ValueError(f"ستون حیاتی '{req}' (یا معادل فارسی آن) در اکسل پیدا نشد.")

    # License limits check for posts
    from apps.core.license import get_system_license_limits
    limits = get_system_license_limits()
    max_posts = limits['max_posts']
    active_post_codes = set(CentralCompetency.objects.filter(is_deleted=False).values_list('post_code', flat=True).distinct())

    created_count = 0
    updated_count = 0
    skipped_count = 0
    seen_keys = set()
    
    with transaction.atomic():
        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or not any(row):
                continue
                
            def get_val(key):
                idx = col_map[key]
                if idx == -1 or idx >= len(row):
                    return ''
                return clean_cell_value(row[idx])
                
            post_code = normalize_persian_digits(get_val('post_code'))
            code = normalize_persian_digits(get_val('code'))
            title = get_val('title')
            category_raw = get_val('category_raw')
            
            if not post_code or not code or not title or not category_raw:
                skipped_count += 1
                continue
                
            # Check limits on active post codes
            if post_code not in active_post_codes:
                if len(active_post_codes) >= max_posts:
                    raise ValueError(f"سقف مجاز تعداد پست‌های بانک شایستگی نسخه جاری تکمیل شده است (حداکثر {int(max_posts)} پست). جهت ارتقا لایسنس با مدیر سیستم تماس بگیرید.")
                active_post_codes.add(post_code)
                
            # Parse competency type from category_raw (first two letters, e.g. KN, SK, AB...)
            comp_type = category_raw[:2].upper()
            valid_types = ['KN', 'SK', 'AB', 'GE', 'ST', 'PR', 'CQ', 'IN']
            if comp_type not in valid_types:
                # Try fallback from code prefix
                code_prefix = code[:2].upper()
                if code_prefix in valid_types:
                    comp_type = code_prefix
                else:
                    comp_type = 'GE' # fallback to general
                    
            # Parse importance
            importance_raw = get_val('importance_raw')
            importance = 3 # default: minimal
            if '1' in importance_raw or 'محوری' in importance_raw:
                importance = 1
            elif '2' in importance_raw or 'تکلیف' in importance_raw:
                importance = 2
            elif '3' in importance_raw or 'حداقلی' in importance_raw:
                importance = 3
                
            # Parse level
            level_raw = get_val('level_raw')
            level = 1 # default: familiarity
            if '3' in level_raw or 'تسلط' in level_raw:
                level = 3
            elif '2' in level_raw or 'توانایی' in level_raw or 'توانايي' in level_raw:
                level = 2
            elif '1' in level_raw or 'آشنایی' in level_raw or 'آشنايي' in level_raw:
                level = 1

            key = (post_code, code)
            seen_keys.add(key)
            
            # Find existing active or deleted competency with same post_code and code
            comp = CentralCompetency.all_objects.filter(post_code=post_code, code=code).first()
            
            data = {
                'post_title': get_val('post_title'),
                'old_code': normalize_persian_digits(get_val('old_code')),
                'title': title,
                'competency_type': comp_type,
                'category_raw': category_raw,
                'cluster_raw': get_val('cluster_raw'),
                'importance': importance,
                'level': level,
                'management_code': normalize_persian_digits(get_val('management_code')),
                'management_name': get_val('management_name'),
                'vice_president_code': normalize_persian_digits(get_val('vice_president_code')),
                'vice_president_name': get_val('vice_president_name'),
                'section_code': normalize_persian_digits(get_val('section_code')),
                'section_name': get_val('section_name'),
                'cost_center_code': normalize_persian_digits(get_val('cost_center_code')),
                'cost_center_name': get_val('cost_center_name'),
                'is_deleted': False,
                'deleted_at': None
            }
            
            if comp:
                # Update
                changed = False
                for field, val in data.items():
                    if getattr(comp, field) != val:
                        setattr(comp, field, val)
                        changed = True
                if changed:
                    comp.save()
                    updated_count += 1
            else:
                # Create
                CentralCompetency.objects.create(
                    post_code=post_code,
                    code=code,
                    **data
                )
                created_count += 1
                
        # Soft delete competencies that are NOT in the uploaded file
        # Retrieve all currently active competencies
        active_comps = CentralCompetency.objects.filter(is_deleted=False)
        deleted_count = 0
        for comp in active_comps:
            if (comp.post_code, comp.code) not in seen_keys:
                comp.delete()
                deleted_count += 1
                
    return {
        'created': created_count,
        'updated': updated_count,
        'deleted': deleted_count,
        'skipped': skipped_count
    }


def adjust_weights_to_step(weights, step=5, target_sum=100, limits=None):
    """
    Rounds a dictionary of {key: float_weight} to the nearest multiple of 'step'
    such that the sum is exactly 'target_sum', while respecting limits {key: (min, max)}.
    """
    # 1. Round to nearest step
    rounded = {}
    for k, val in weights.items():
        min_limit, max_limit = (0, 100) if not limits or k not in limits else limits[k]
        r_val = int(round(val / float(step))) * step
        # Clamp to limits if provided (must be multiples of step)
        r_val = max(min_limit, min(max_limit, r_val))
        rounded[k] = r_val
        
    # 2. Adjust if sum is not target_sum
    current_sum = sum(rounded.values())
    diff = target_sum - current_sum
    
    if diff != 0:
        # Number of steps to adjust
        num_adjustments = int(abs(diff) / step)
        direction = 1 if diff > 0 else -1
        
        # Calculate rounding error for each key: (original - rounded) * direction
        errors = {k: (val - rounded[k]) * direction for k, val in weights.items()}
        
        # We can only adjust keys that won't violate their limits
        for _ in range(num_adjustments):
            candidates = []
            for k in weights.keys():
                min_limit, max_limit = (0, 100) if not limits or k not in limits else limits[k]
                new_val = rounded[k] + (direction * step)
                if min_limit <= new_val <= max_limit:
                    candidates.append(k)
            
            if not candidates:
                # If no candidates satisfy limits, fallback to any key that has space
                candidates = []
                for k in weights.keys():
                    min_limit, max_limit = (0, 100) if not limits or k not in limits else limits[k]
                    if direction > 0 and rounded[k] < max_limit:
                        candidates.append(k)
                    elif direction < 0 and rounded[k] > min_limit:
                        candidates.append(k)
                if not candidates:
                    candidates = list(weights.keys())
                
            # Pick candidate with largest error (most deserving of adjustment in that direction)
            best_k = max(candidates, key=lambda k: errors[k])
            rounded[best_k] += direction * step
            # Update error for that key
            errors[best_k] -= step
            
    return rounded


def calculate_assessment_plan(competencies, custom_weights=None, custom_passing_scores=None, round_to_five=False, active_stages=None, deactivated_stages=None, bypass_limits=False):
    """
    Takes a list/queryset of JobOpportunityCompetency and calculates:
    - required stages and their weights (summing to 100%)
    - competencies mapped to each stage with their relative weights within the stage
    - cutoff score for each stage
    
    Returns:
    - dict: {
        'stages': {
            'EXAM': {'name': 'آزمون کتبی', 'weight': percent, 'min_limit': 20, 'max_limit': 50, 'passing_score': 60, 'competencies': [...], 'is_active': True},
            ...
        },
        'errors': [],
        'warnings': []
      }
    """
    # 1. Define Stage Constraints
    # Stage limits: (min, max)
    if bypass_limits:
        STAGE_LIMITS = {
            'SCREENING': (0, 0),
            'EXAM': (0, 100),
            'SKILL_TEST': (0, 100),
            'INTERVIEW': (0, 100),
            'ASSESSMENT': (0, 100)
        }
    else:
        STAGE_LIMITS = {
            'SCREENING': (0, 0),
            'EXAM': (20, 50),
            'SKILL_TEST': (20, 40),
            'INTERVIEW': (10, 25),
            'ASSESSMENT': (15, 40)
        }
    
    STAGE_NAMES = {
        'SCREENING': 'غربالگری اولیه',
        'EXAM': 'آزمون کتبی',
        'SKILL_TEST': 'آزمون مهارتی',
        'INTERVIEW': 'مصاحبه تخصصی',
        'ASSESSMENT': 'کانون ارزیابی'
    }

    # 2. Filter competencies by type and calculate individual competency weights
    valid_competencies = []
    has_kn = False
    has_sk_ab = False
    has_ge_st = False
    
    for comp in competencies:
        ctype = comp.competency_type
        if ctype in ['PR', 'CQ', 'IN']:
            continue
            
        imp_weight = 3 if comp.importance == 1 else (2 if comp.importance == 2 else 1)
        prof_weight = comp.level
        weight = imp_weight * prof_weight
        
        valid_competencies.append({
            'code': comp.code,
            'title': comp.title,
            'type': ctype,
            'weight': weight,
            'level': comp.level
        })
        
        if ctype == 'KN':
            has_kn = True
        elif ctype in ['SK', 'AB']:
            has_sk_ab = True
        elif ctype in ['GE', 'ST']:
            has_ge_st = True

    # 3. Determine recommended active stages based on rules
    recommended_stages = {'SCREENING'}
    if has_kn:
        recommended_stages.add('EXAM')
    if has_sk_ab:
        recommended_stages.add('SKILL_TEST')
        recommended_stages.add('INTERVIEW')
    if has_ge_st:
        recommended_stages.add('ASSESSMENT')
        
    # If active_stages is None, we default to recommended stages
    if active_stages is None:
        active_stages = recommended_stages
    else:
        active_stages = set(active_stages)
        
    if deactivated_stages:
        active_stages = active_stages - set(deactivated_stages)
        
    active_stages.add('SCREENING') # SCREENING is always active

    # 4. Aggregate weights to stages
    stage_raw_scores = {k: 0.0 for k in STAGE_LIMITS.keys()}
    stage_competencies = {k: [] for k in STAGE_LIMITS.keys()}
    
    for comp in valid_competencies:
        ctype = comp['type']
        if ctype == 'KN':
            stage_raw_scores['EXAM'] += comp['weight']
            stage_competencies['EXAM'].append(comp)
        elif ctype in ['SK', 'AB']:
            stage_raw_scores['SKILL_TEST'] += comp['weight']
            stage_competencies['SKILL_TEST'].append(comp)
            stage_raw_scores['INTERVIEW'] += comp['weight']
            stage_competencies['INTERVIEW'].append(comp)
        elif ctype in ['GE', 'ST']:
            stage_raw_scores['ASSESSMENT'] += comp['weight']
            stage_competencies['ASSESSMENT'].append(comp)

    errors = []
    warnings = []
    int_w = {k: 0 for k in STAGE_LIMITS.keys()}

    # 5. Calculate weights
    # Only active stages (excluding SCREENING) get non-zero weights
    active_keys_for_calc = [k for k in active_stages if k != 'SCREENING']
    
    if not active_keys_for_calc:
        errors.append("حداقل یک مرحله ارزیابی باید فعال باشد.")
    else:
        if custom_weights:
            for k in active_keys_for_calc:
                val = custom_weights.get(k)
                if val is not None and str(val).strip() != '':
                    try:
                        val_int = int(val)
                    except (ValueError, TypeError):
                        errors.append(f"وزن وارد شده برای مرحله {STAGE_NAMES[k]} نامعتبر است.")
                        val_int = 0
                    
                    if val_int < 0 or val_int > 100:
                        errors.append(f"وزن مرحله {STAGE_NAMES[k]} باید بین ۰ تا ۱۰۰ باشد.")
                        val_int = 0
                    
                    min_val, max_val = STAGE_LIMITS[k]
                    if val_int < min_val:
                        errors.append(f"وزن مرحله {STAGE_NAMES[k]} نمی‌تواند کمتر از {min_val}٪ باشد.")
                    elif val_int > max_val:
                        errors.append(f"وزن مرحله {STAGE_NAMES[k]} نمی‌تواند بیشتر از {max_val}٪ باشد.")
                    
                    int_w[k] = val_int
                else:
                    errors.append(f"وزن مرحله {STAGE_NAMES[k]} وارد نشده است.")
                    int_w[k] = 0
            
            total_custom = sum(int_w[k] for k in active_keys_for_calc)
            if not errors and total_custom != 100:
                errors.append(f"مجموع اوزان مراحل ارزیابی باید دقیقاً ۱۰۰٪ باشد. (مجموع فعلی: {total_custom}٪)")
        else:
            w = {k: stage_raw_scores.get(k, 0.0) for k in active_keys_for_calc}
            total_raw = sum(w.values())
            
            if total_raw == 0:
                w = {k: 100.0 / len(active_keys_for_calc) for k in active_keys_for_calc}
            else:
                w = {k: (val / total_raw) * 100.0 for k, val in w.items()}
                
            if round_to_five:
                int_w_non = adjust_weights_to_step(w, step=5, target_sum=100, limits={k: STAGE_LIMITS[k] for k in active_keys_for_calc})
                for k in active_keys_for_calc:
                    int_w[k] = int_w_non.get(k, 0)
            else:
                # Iterative clamping
                for _ in range(10):
                    clamped = {}
                    for k in active_keys_for_calc:
                        min_val, max_val = STAGE_LIMITS[k]
                        clamped[k] = max(min_val, min(max_val, w[k]))
                        
                    clamped_sum = sum(clamped.values())
                    diff = 100.0 - clamped_sum
                    
                    if abs(diff) < 0.01:
                        w = clamped
                        break
                        
                    if diff > 0:
                        adjustable = {k: STAGE_LIMITS[k][1] - clamped[k] for k in active_keys_for_calc if clamped[k] < STAGE_LIMITS[k][1]}
                        total_adj = sum(adjustable.values())
                        if total_adj == 0:
                            w = {k: (val / clamped_sum) * 100.0 for k, val in clamped.items()}
                            break
                        w = {k: clamped[k] + (adjustable.get(k, 0.0) / total_adj) * diff for k in active_keys_for_calc}
                    else:
                        adjustable = {k: clamped[k] - STAGE_LIMITS[k][0] for k in active_keys_for_calc if clamped[k] > STAGE_LIMITS[k][0]}
                        total_adj = sum(adjustable.values())
                        if total_adj == 0:
                            w = {k: (val / clamped_sum) * 100.0 for k, val in clamped.items()}
                            break
                        w = {k: clamped[k] + (adjustable.get(k, 0.0) / total_adj) * diff for k in active_keys_for_calc}

                # Integer rounding
                for k in active_keys_for_calc:
                    int_w[k] = int(round(w.get(k, 0)))
                
                non_screening_sum = sum(int_w[k] for k in active_keys_for_calc)
                diff = 100 - non_screening_sum
                if diff != 0:
                    candidates = []
                    for k in active_keys_for_calc:
                        min_val, max_val = STAGE_LIMITS[k]
                        new_val = int_w[k] + diff
                        if min_val <= new_val <= max_val:
                            candidates.append(k)
                    
                    if candidates:
                        best_candidate = max(candidates, key=lambda k: w[k] - int_w[k] if diff > 0 else int_w[k] - w[k])
                        int_w[best_candidate] += diff
                    else:
                        fallback_stage = max(active_keys_for_calc, key=lambda k: w[k])
                        int_w[fallback_stage] += diff

    # 6. Build final output stages
    result_stages = {}
    ordered_keys = ['SCREENING', 'EXAM', 'SKILL_TEST', 'INTERVIEW', 'ASSESSMENT']
    for k in ordered_keys:
        is_active = k in active_stages
        s_comps = stage_competencies.get(k, [])
        computed_comps = []
        
        if k != 'SCREENING' and s_comps:
            s_total_comp_w = sum(c['weight'] for c in s_comps)
            for c in s_comps:
                c_rel_w = int(round((c['weight'] / s_total_comp_w) * 100)) if s_total_comp_w > 0 else 100
                computed_comps.append({
                    'code': c['code'],
                    'title': c['title'],
                    'type': c['type'],
                    'weight': c_rel_w,
                    'level': c['level']
                })
                
            # Rounding check for competencies within the stage
            if round_to_five and computed_comps:
                weights_c = {cc['code']: cc['weight'] for cc in computed_comps}
                rounded_c = adjust_weights_to_step(weights_c, step=5, target_sum=100)
                for cc in computed_comps:
                    cc['weight'] = rounded_c[cc['code']]
            else:
                comp_diff = 100 - sum(cc['weight'] for cc in computed_comps)
                if comp_diff != 0 and computed_comps:
                    computed_comps[0]['weight'] += comp_diff
            
        min_val, max_val = STAGE_LIMITS[k]
        
        # Calculate passing score
        if k == 'SCREENING':
            passing_score = 0
        else:
            avg_level = sum(c['level'] for c in s_comps) / len(s_comps) if s_comps else 1.0
            default_score = int(round(50 + avg_level * 10))
            
            passing_score = default_score
            if custom_passing_scores and k in custom_passing_scores:
                val = custom_passing_scores[k]
                if val is not None and str(val).strip() != '':
                    try:
                        val_int = int(val)
                        passing_score = val_int
                        if not (0 <= val_int <= 100):
                            errors.append(f"حد نصاب قبولی برای مرحله {STAGE_NAMES[k]} باید بین ۰ و ۱۰۰ باشد.")
                    except (ValueError, TypeError):
                        errors.append(f"حد نصاب قبولی وارد شده برای مرحله {STAGE_NAMES[k]} نامعتبر است.")
                        passing_score = val

        result_stages[k] = {
            'name': STAGE_NAMES[k],
            'weight': int_w[k] if is_active else 0,
            'min_limit': min_val,
            'max_limit': max_val,
            'passing_score': passing_score if is_active else 0,
            'competencies': computed_comps,
            'is_active': is_active
        }
        
    return {
        'stages': result_stages,
        'errors': errors,
        'warnings': warnings
    }


def suggest_workflow_templates(active_stages):
    """
    Given a set or list of active stage types (e.g. {'EXAM', 'INTERVIEW'}),
    find and return the active WorkflowTemplates, ranked by how well they match.
    Returns:
        List of dicts: [
            {
                'template': WorkflowTemplate instance,
                'match_percentage': float (0 to 100),
                'reasons': [str, ...],
                'is_perfect_match': bool
            },
            ...
        ]
    """
    from apps.jobs.models import WorkflowTemplate
    
    templates = WorkflowTemplate.objects.filter(is_deleted=False).prefetch_related('stages')
    suggestions = []
    
    # Ignore SCREENING and OTHER in active_stages for template matching
    active_set = set(s for s in active_stages if s not in ['SCREENING', 'OTHER'])
    
    STAGE_NAMES = {
        'EXAM': 'آزمون کتبی',
        'SKILL_TEST': 'آزمون مهارتی',
        'INTERVIEW': 'مصاحبه تخصصی',
        'ASSESSMENT': 'کانون ارزیابی'
    }
    
    for t in templates:
        t_stages = t.stages.filter(is_deleted=False)
        # Get set of stage types in the template, ignoring SCREENING and OTHER
        t_set = set(s.stage_type for s in t_stages if s.stage_type not in ['SCREENING', 'OTHER'])
        
        # Calculate intersection and union
        intersection = active_set.intersection(t_set)
        union = active_set.union(t_set)
        
        if not union:
            match_percent = 0.0
        else:
            match_percent = (len(intersection) / len(union)) * 100.0
            
        reasons = []
        for s_type in intersection:
            name = STAGE_NAMES.get(s_type, s_type)
            reasons.append(f"دارای {name}")
            
        missing = active_set - t_set
        for s_type in missing:
            name = STAGE_NAMES.get(s_type, s_type)
            reasons.append(f"فاقد {name}")
            
        extra = t_set - active_set
        for s_type in extra:
            name = STAGE_NAMES.get(s_type, s_type)
            reasons.append(f"اضافه: {name}")
            
        is_perfect_match = (active_set == t_set)
        
        suggestions.append({
            'template': t,
            'match_percentage': int(match_percent),
            'is_perfect_match': is_perfect_match,
            'reasons': reasons
        })
        
    # Sort suggestions by perfect match (True first), then by match percentage (desc)
    suggestions.sort(key=lambda x: (x['is_perfect_match'], x['match_percentage']), reverse=True)
    return suggestions
