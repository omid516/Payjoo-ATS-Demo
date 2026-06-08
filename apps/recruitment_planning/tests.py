import datetime
import jdatetime
from django.test import TestCase
from django.contrib.auth.models import User
from django.urls import reverse

from apps.accounts.models import UserProfile
from apps.jobs.models import JobOpportunity, JobOpportunityStage, WorkflowTemplate, WorkflowStageTemplate
from apps.recruitment_planning.models import StageTypeConfiguration, Holiday, JobRecruitmentPlan, JobStagePlan
from apps.recruitment_planning.utils import (
    add_working_days, get_next_working_day, 
    parse_jalali_to_gregorian, to_jalali_string,
    calculate_recruitment_schedule, get_jalali_month_range
)

class RecruitmentPlanningTests(TestCase):

    def setUp(self):
        # Create administrative users
        self.admin_user = User.objects.create_user(username='admin_planning', password='testpassword123')
        self.admin_user.profile.role = UserProfile.ROLE_ADMIN
        self.admin_user.profile.save()

        # Create stages type config
        self.config_screening = StageTypeConfiguration.objects.create(
            stage_type='SCREENING',
            default_sla_days=5,
            monthly_capacity=50
        )
        self.config_exam = StageTypeConfiguration.objects.create(
            stage_type='EXAM',
            default_sla_days=10,
            monthly_capacity=20
        )

        # Create workflow template
        self.workflow = WorkflowTemplate.objects.create(name='Planning Workflow')
        self.stage_template_1 = WorkflowStageTemplate.objects.create(
            workflow=self.workflow,
            name='غربالگری اولیه',
            sequence=1,
            stage_type='SCREENING'
        )
        self.stage_template_2 = WorkflowStageTemplate.objects.create(
            workflow=self.workflow,
            name='آزمون کتبی',
            sequence=2,
            stage_type='EXAM'
        )

        # Create Job Opportunities
        self.job1 = JobOpportunity.objects.create(
            title='برنامه‌نویس پایتون',
            code='PYTHON-01',
            request_number='REQ-PLAN-01',
            department='فنی',
            headcount=15,
            workflow=self.workflow
        )
        self.job2 = JobOpportunity.objects.create(
            title='طراح رابط کاربری',
            code='UI-01',
            request_number='REQ-PLAN-02',
            department='هنری',
            headcount=10,
            workflow=self.workflow
        )

    def test_working_days_and_holidays_calculation(self):
        """تست عبور از جمعه‌ها و روزهای تعطیل ثبت‌شده در دیتابیس"""
        # 1405/03/16 is Friday (2026-06-06)
        # 1405/03/17 is Saturday (2026-06-07)
        # 1405/03/18 is Sunday (2026-06-08)
        
        start_date = datetime.date(2026, 6, 5) # Friday (1405/03/15)
        
        # Define a custom holiday on Sunday 2026-06-07 (1405/03/17)
        Holiday.objects.create(date=datetime.date(2026, 6, 7), title='تست تعطیلی')
        
        # Add 2 working days starting from Friday:
        # Day 1: Saturday (2026-06-06) is a working day
        # Sunday (2026-06-07) is a holiday -> skipped
        # Day 2: Monday (2026-06-08) is a working day
        # Result should be Monday (2026-06-08)
        end_date = add_working_days(start_date, 2)
        self.assertEqual(end_date, datetime.date(2026, 6, 8))

    def test_capacity_overflow_triggers_monthly_shift(self):
        """تست انتقال خودکار برنامه به ماه بعد در صورت تکمیل ظرفیت ماه جاری"""
        # Set monthly capacity limit of EXAM to 20
        # Job 1 has headcount 15 (occupies 15 slots of EXAM capacity in Mehr 1405)
        # Job 2 has headcount 10 (needs 10 slots of EXAM. 15 + 10 = 25 > 20 -> should trigger shift!)
        
        start_date = datetime.date(2026, 9, 23) # 1405/07/01 (Mehr start)
        
        # 1. Schedule Job 1
        schedule1 = calculate_recruitment_schedule(self.job1, start_date)
        # Save Job 1 plan to DB so it consumes capacity
        plan1 = JobRecruitmentPlan.objects.create(
            job=self.job1,
            start_date=start_date,
            predicted_end_date=schedule1[-1]['planned_end_date'],
            status=JobRecruitmentPlan.STATUS_ACTIVE
        )
        for s in schedule1:
            JobStagePlan.objects.create(
                plan=plan1,
                stage=s['stage'],
                stage_type=s['stage_type'],
                planned_start_date=s['planned_start_date'],
                planned_end_date=s['planned_end_date'],
                sla_days=s['sla_days'],
                capacity_shifted=s['capacity_shifted']
            )
            
        # 2. Schedule Job 2 starting at same date
        schedule2 = calculate_recruitment_schedule(self.job2, start_date)
        
        # The screening stage for Job 2 is OK (capacity screening is 50, consumes 10 + 15 = 25 <= 50)
        # The exam stage for Job 2 has capacity limit 20. Consumed is 15. Adding 10 exceeds 20.
        # So Job 2's exam stage should have capacity_shifted=True and be moved to Aban month (1405/08/01 onwards)
        exam_stage_plan = next(s for s in schedule2 if s['stage_type'] == 'EXAM')
        self.assertTrue(exam_stage_plan['capacity_shifted'])
        
        # Check that its planned_start_date is in Aban 1405 (Gregorian after 2026-10-23)
        j_planned_start = jdatetime.date.fromgregorian(date=exam_stage_plan['planned_start_date'])
        self.assertEqual(j_planned_start.month, 8) # Aban (month 8)

    def test_planning_views_and_dashboards(self):
        """تست عملکرد صفحات و ارسال فرم‌ها در ماژول برنامه‌ریزی"""
        self.client.login(username='admin_planning', password='testpassword123')

        # 1. Access Dashboard
        response = self.client.get(reverse('planning_dashboard'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'برنامه‌ریزی و ظرفیت‌سنجی جذب')

        # 2. Preview schedule for Job 1 via HTMX
        url = reverse('job_planning', kwargs={'job_id': self.job1.id})
        post_data = {
            'action': 'preview',
            'start_date': '1405/02/01'
        }
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'پیش‌بینی اتمام فرآیند جذب')

        # 3. Save schedule plan (Standard request - redirects with 302)
        post_data['action'] = 'save'
        response = self.client.post(url, post_data)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('planning_dashboard'))

        # Check plan is saved in DB
        self.assertTrue(JobRecruitmentPlan.objects.filter(job=self.job1, status='ACTIVE').exists())

        # Test HTMX save (HTMX request - returns 200 with HX-Redirect header)
        JobRecruitmentPlan.objects.filter(job=self.job1).delete()
        response_htmx = self.client.post(url, post_data, HTTP_HX_REQUEST='true')
        self.assertEqual(response_htmx.status_code, 200)
        self.assertEqual(response_htmx['HX-Redirect'], reverse('planning_dashboard'))
        self.assertTrue(JobRecruitmentPlan.objects.filter(job=self.job1, status='ACTIVE').exists())

        # 4. Access and Save configs view
        config_url = reverse('planning_config')
        response = self.client.get(config_url)
        self.assertEqual(response.status_code, 200)

        # Save config update post
        post_config = {
            'action': 'save_configs',
            'sla_SCREENING': '6',
            'capacity_SCREENING': '60'
        }
        response = self.client.post(config_url, post_config)
        self.assertEqual(response.status_code, 302) # Redirects back to config view
        
        self.config_screening.refresh_from_db()
        self.assertEqual(self.config_screening.default_sla_days, 6)
        self.assertEqual(self.config_screening.monthly_capacity, 60)

    def test_planning_export_excel(self):
        """تست خروجی اکسل برنامه‌های جذب"""
        self.client.login(username='admin_planning', password='testpassword123')
        
        # Save a plan first
        url = reverse('job_planning', kwargs={'job_id': self.job1.id})
        post_data = {
            'action': 'save',
            'start_date': '1405/02/01'
        }
        self.client.post(url, post_data)
        
        # Test export view
        export_url = reverse('planning_export_excel')
        response = self.client.get(export_url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertTrue(len(response.content) > 0)

        # Parse Excel bytes and verify that the planned job is present in the sheet
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = wb.active
        
        row_values = list(sheet.iter_rows(values_only=True))
        found = False
        for row in row_values:
            if self.job1.title in row or self.job1.code in row:
                found = True
                break
        self.assertTrue(found, "The planned job details were not found in the Excel export rows.")

    def test_planning_calendar_view(self):
        """تست دسترسی به نمای تقویم شمسی ماهانه"""
        self.client.login(username='admin_planning', password='testpassword123')
        calendar_url = reverse('planning_calendar')
        response = self.client.get(calendar_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'تقویم برنامه‌ریزی جذب')
        self.assertContains(response, 'شنبه')
        self.assertContains(response, 'جمعه')

    def test_weekly_agenda_excel_export(self):
        """تست خروجی اکسل دستور کار هفتگی"""
        self.client.login(username='admin_planning', password='testpassword123')
        
        # Save a plan so that we have events scheduled for the next week
        # (Start date: 1405/03/18 which corresponds to 2026-06-08)
        planning_url = reverse('job_planning', kwargs={'job_id': self.job1.id})
        post_data = {
            'action': 'save',
            'start_date': '1405/03/18'
        }
        self.client.post(planning_url, post_data)
        
        export_url = reverse('planning_agenda_export_excel')
        response = self.client.get(export_url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = wb.active
        
        row_values = list(sheet.iter_rows(values_only=True))
        # Headings verify
        self.assertIn("تاریخ شمسی", row_values[0])
        self.assertIn("روز هفته", row_values[0])
        self.assertIn("نوع رویداد", row_values[0])
        
        # Find if our saved job's stage plan is exported
        found = False
        for row in row_values:
            if self.job1.title in row:
                found = True
                break
        self.assertTrue(found, "The weekly agenda job events were not found in the Excel export.")

    def test_weekly_agenda_print_view(self):
        """تست نمای چاپی شکیل دستور کار هفتگی"""
        self.client.login(username='admin_planning', password='testpassword123')
        print_url = reverse('planning_agenda_print')
        response = self.client.get(print_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'گزارش دستور کار و برنامه‌های جذب هفته آینده')
        self.assertContains(response, 'چاپ گزارش / ذخیره PDF')

    def test_job_creation_redirect_flow(self):
        """تست هدایت خودکار ایجاد شغل به برنامه‌ریزی و سپس سند آزمون"""
        self.client.login(username='admin_planning', password='testpassword123')
        
        job_add_url = reverse('job_add')
        form_data = {
            'request_number': 'REQ-TEST-NEW-01',
            'title': 'کارشناس امنیت شبکه',
            'code': 'SEC-01',
            'department': 'فناوری اطلاعات',
            'headcount': '2',
            'recruitment_type': 'EXTERNAL',
            'status': 'PLANNING',
            'start_date': '1405/03/18',
            'end_date': '1405/04/18',
            'description': 'شرح وظایف کارشناس امنیت',
            'stages-TOTAL_FORMS': '1',
            'stages-INITIAL_FORMS': '0',
            'stages-MIN_NUM_FORMS': '0',
            'stages-MAX_NUM_FORMS': '1000',
            'stages-0-name': 'غربالگری اولیه',
            'stages-0-weight': '100',
            'stages-0-sequence': '1',
            'stages-0-stage_type': 'SCREENING',
        }
        
        response = self.client.post(job_add_url, form_data)
        
        # Check that it redirected to the job planning page with ?next=print_doc
        new_job = JobOpportunity.objects.get(code='SEC-01')
        expected_redirect_url = reverse('job_planning', kwargs={'job_id': new_job.id}) + '?next=print_doc'
        self.assertEqual(response.status_code, 302)
        self.assertIn(expected_redirect_url, response.url)
        
        # Load the job planning view with ?next=print_doc
        plan_page_url = reverse('job_planning', kwargs={'job_id': new_job.id}) + '?next=print_doc'
        response_plan_page = self.client.get(plan_page_url)
        self.assertEqual(response_plan_page.status_code, 200)
        
        # Verify the skip link is present
        skip_url = reverse('job_print_doc', kwargs={'pk': new_job.id})
        self.assertContains(response_plan_page, 'رد کردن و مشاهده سند آزمون')
        self.assertContains(response_plan_page, skip_url)
        
        # Submit the planning form with next=print_doc
        post_data = {
            'action': 'save',
            'start_date': '1405/03/18',
            'next': 'print_doc'
        }
        
        # Standard request
        response_save = self.client.post(plan_page_url, post_data)
        self.assertEqual(response_save.status_code, 302)
        self.assertEqual(response_save.url, reverse('job_print_doc', kwargs={'pk': new_job.id}))
        
        # HTMX request
        response_save_htmx = self.client.post(plan_page_url, post_data, HTTP_HX_REQUEST='true')
        self.assertEqual(response_save_htmx.status_code, 200)
        self.assertEqual(response_save_htmx['HX-Redirect'], reverse('job_print_doc', kwargs={'pk': new_job.id}))

    def test_job_planning_suggestions_view(self):
        """تست نمای دستیار هوشمند پیشنهاد تاریخ شروع و ظرفیت"""
        self.client.login(username='admin_planning', password='testpassword123')
        suggestions_url = reverse('job_planning_suggestions', kwargs={'job_id': self.job1.id})
        response = self.client.get(suggestions_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'دستیار هوشمند تاریخ و ظرفیت ماه‌ها')
        self.assertContains(response, 'تاریخ‌های شروع پیشنهادی')
        self.assertContains(response, 'پایش ظرفیت تفصیلی ۶ ماهه')


