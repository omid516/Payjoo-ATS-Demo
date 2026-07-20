import openpyxl

wb = openpyxl.load_workbook('/Users/omidsalehi/Desktop/ATS/1-3.xlsx', data_only=True)
print("Sheet Names:", wb.sheetnames)
ws = wb['قبولی نهایی']
print("Active Sheet Title:", ws.title)

# Let's inspect the first row (headers)
rows = list(ws.iter_rows(values_only=True))
print("Total rows:", len(rows))
if rows:
    headers = rows[0]
    print("\nHeaders:")
    for idx, h in enumerate(headers):
        print(f"  Col {idx}: {h}")
    
    print("\nFirst 5 rows of data:")
    for r in rows[1:6]:
        print(" ", r)

