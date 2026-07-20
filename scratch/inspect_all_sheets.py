import openpyxl

wb = openpyxl.load_workbook('/Users/omidsalehi/Desktop/ATS/1-3.xlsx', read_only=True)
print("Sheets in 1-3.xlsx:")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    # read first row only
    first_row = next(ws.iter_rows(values_only=True), None)
    print(f"\nSheet: {sheet}")
    if first_row:
        print("  Headers:", [str(h).strip() if h is not None else "" for h in first_row])
    else:
        print("  Empty sheet")
