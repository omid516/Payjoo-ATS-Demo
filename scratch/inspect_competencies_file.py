import openpyxl

wb = openpyxl.load_workbook('/Users/omidsalehi/Desktop/ATS/شایستگی ها.xlsx', read_only=True)
print("Sheets in شایستگی ها.xlsx:")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    first_row = next(ws.iter_rows(values_only=True), None)
    print(f"\nSheet: {sheet}")
    if first_row:
        print("  Headers:", [str(h).strip() if h is not None else "" for h in first_row])
        # Print first few rows
        row_iter = ws.iter_rows(values_only=True)
        next(row_iter) # skip header
        print("  Rows:")
        for _ in range(5):
            row = next(row_iter, None)
            if row:
                print("    ", row)
    else:
        print("  Empty sheet")
