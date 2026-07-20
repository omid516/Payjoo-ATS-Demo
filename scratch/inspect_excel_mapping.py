import openpyxl

wb = openpyxl.load_workbook('/Users/omidsalehi/Desktop/ATS/شایستگی ها.xlsx', read_only=True)
ws = wb['Result']

headers = [str(h).strip() if h is not None else "" for h in next(ws.iter_rows(values_only=True))]
print("Headers:", headers)

col_map = {
    'post_code': headers.index('کد پست'),
    'code': headers.index('کد شايستگي'),
    'title': headers.index('شايستگي'),
    'category_raw': headers.index('طبقه'),
}

category_raw_values = set()
invalid_mappings = []

row_count = 0
for row in ws.iter_rows(values_only=True):
    if row_count == 0:
        row_count += 1
        continue
    
    post_code = row[col_map['post_code']]
    code = row[col_map['code']]
    title = row[col_map['title']]
    category_raw = row[col_map['category_raw']]
    
    if not post_code or not code or not title or not category_raw:
        continue
        
    category_raw_values.add(str(category_raw))
    
    # Check parser mapping logic
    comp_type = str(category_raw)[:2].upper()
    valid_types = ['KN', 'SK', 'AB', 'GE', 'ST', 'PR', 'CQ', 'IN']
    if comp_type not in valid_types:
        code_prefix = str(code)[:2].upper()
        if code_prefix not in valid_types:
            invalid_mappings.append((post_code, code, category_raw, title))
            
    row_count += 1

print(f"\nTotal rows processed: {row_count}")
print("\nUnique category_raw values:")
for val in sorted(category_raw_values):
    print(f"  {val}")

print(f"\nNumber of invalid/fallback mappings: {len(invalid_mappings)}")
if invalid_mappings:
    print("Some examples of invalid/fallback mappings (fallback to GE):")
    for item in invalid_mappings[:15]:
        print(f"  Post Code: {item[0]}, Code: {item[1]}, Category Raw: {item[2]}, Title: {item[3]}")
